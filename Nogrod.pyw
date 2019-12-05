# -*- coding: utf-8 -*-

################################################################################################
##                                                                                            ##
##                              Nogrod 1.1.2. Based on Angrist 1.2                              ##
##                              --------------------------------                              ##
##                                                                                            ##
##  Programmer: Martin Wettstein,                                                             ##
##              University of Zürich                                                          ##
##              m.wettstein@ikmz.uzh.ch                                                       ##
##                                                                                            ##
##  For detailed references, see:                                                             ##
##              http://www.ikmz.uzh.ch/en/Abteilungen/Medienpsychologie/Recource/Angrist.html ##
##                                                                                            ##
##                                                                                            ##
################################################################################################

from __future__ import print_function

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    xlsx = 1
except:
    xlsx = 0

try: ##if Python Version 3.x
    from tkinter import *
    from tkinter import messagebox
    from tkinter import filedialog
    from tkinter import simpledialog
    import re
    py_version = 3
except: ##If Python Version 2.7
    from Tkinter import *
    import tkMessageBox
    import tkFileDialog
    py_version = 2
import time
from datetime import datetime
from datetime import timedelta
import math
import random
import unicodedata
import os



#print(os.getcwd())

class CMD: #Auxilliary function for callbacks using parameters. Syntax: CMD(function, argument1, argument2, ...)
    def __init__(s1, func, *args):
        s1.func = func
        s1.args = args
    def __call__(s1, *args):
        args = s1.args+args
        try:
            s1.func(*args)
        except:
            apply(s1.func,args)
        
class Anzeige(Frame):
    def __init__(self, master=None):        
        Frame.__init__(self,master)
        top=self.winfo_toplevel() #Flexible Toplevel of the window
        top.rowconfigure(0, weight=1)
        top.columnconfigure(0, weight=1)
        self.grid(sticky=N+S+W+E)
        self.rowconfigure(6, weight=1) #Only expandable row in the grid is row 6 (above bottomline frame)
        self.columnconfigure(3, weight=1) #Only expandable column in the grid is column 15 (text field)
        self.bind_all("<F1>",self.debug_on)
        self.bind_all("<F2>",self.show_path)
        self.bind_all("<F3>",self.show_verb)
        self.bind_all("<F4>",self.show_storage)
        
        #root.title("NOGROD 1.1(b): Utility for data reshaping and -analysis")
        root.title("Nogrod 1.1(b): Running under Python "+str(settings['Python_Version']))

        self.fuellen()
        
    def fuellen(self):
        global codebook #codebook, entsprechend der Datei 'codebook.ini'
        global storage #Variabeltupel, das angibt, welche Variable wie belegt ist
        global dta_pos #Aktuelle Position im Codierbaum. Ein Tupel der Form: (Artikel, Development, Element)
        dta_pos = 'NO DATA YET'
        global prog_pos #Aktuelle Position im codebook. Ein String, der angibt, bei welcher Variable das Programm steht.
        prog_pos = 'NO POSITION YET'
        global def_val #Default-Werte für einzelne Variablen in einem Directory
        def_val = {}
        global farbton_text #Farbton für die Frames, in denen Fragen gestellt werden. Ist von Windows-Version abhängig.
        farbton_text = "SystemMenu" #Windows System Color
        global settings #Dictionary for flexible settings and entries.

        #Just trying to remove everything on the screen. Only useful in the case
        #of critical aborts that are redirected to fuellen()
        try:
            self.f_review.destroy()
            self.f_location.destroy()
            self.f_explanation.destroy()
            self.f_questions.destroy()
            self.f_bottomline.destroy()
            self.Artikel.destroy()
            self.etui.destroy()
            self.yscroller.destroy()
        except:
            verb("Start of the session.",1)

        #Just a spacer to keep the window size constant
        self.spacer = Text(self,width=1,height=40, bg=farbton_text, relief=FLAT,takefocus = 0)
        self.spacer.grid(row=0,column=0,rowspan=8)

        ##
        ## Below this point all global variables and settings are defined
        ##
        #All settings values set below are overwritten as soon as a
        #file for coder settings is found to specify other settings.


        settings['First_Page'] = 'methode'

            
        ##
        ## Below external sources are included to load additional settings
        ##

        global cini
        if available('Settings'):
            try:
                cini = get_codebook(settings['Settings'])
            except:
                cini = {}
        else:
            cini = {}

        if 'Coder-Settings' in cini.keys():
            for i in range(0,len(cini['Coder-Settings'][3])):
                cod = cini['Coder-Settings'][3][i]
                val = cini['Coder-Settings'][2][i]
                settings[cod] = val
        else:
            cini['Coder-Settings']=['\n','\n',[],[],'\n']
        if 'Default-Values' in cini.keys():
            dv = self.load_cset(cini['Default-Values'])
            for i in range(0,len(dv[3])):
                def_val[dv[3][i]] = dv[2][i]
            verb('Default_Values found for: '+str(def_val.keys()))
        else:
            cini['Default-Values']=['\n','\n',[],[],'\n']

        if available('Styleset'):
            define_styleset(settings['Styleset'])
        else:
            define_styleset()
          
        codebook = {}
        if available('Codebook'):
            path = resource_path(settings['Codebook'])
            codebook = get_codebook(path) #settings['Codebook'])
        else:
            codebook = get_codebook('a_codebook.ini')
        
        for v in sorted(codebook.keys()):
            verb(v+'; ',nl=0)
        verb('\n'+str(len(codebook.keys()))+'Variables in Codebook.')

        ## Add Excel functionality if openpyxl was loaded correctly
        if settings['Excel']==1:
            codebook["In_Sep"][3].append("5")
            codebook["In_Sep"][2].append("Excel Spreadsheet")
            codebook["Out_Sep"][3].append("5")
            codebook["Out_Sep"][2].append("Excel Spreadsheet")

            

        storage = {}
        storage['Breaks'] = 0
        storage['Helptexts'] = 0
        storage['Backs'] = 0
        storage['Remove_item'] = 0        
        storage['ID'] = ''
        verb('Coder: '+settings['Coder'])
        verb('Font: '+settings['Font'])

        prog_pos = settings['First_Page']
        dta_pos = ['-','-','-','-']

        self.set_window() #Create the Dialog Window.


###################################
#                                 #
#    ASK-Function                 #
#                                 #
###################################

       
    def ask(self):
        global prog_pos
        global dta_pos
        global codebook
        global settings
        global storage
        log('\nCalling Function: ASK')
        #The MCP function is the core of this interface. In this function all pages of the questionnaire are defined.
        #A variety of self.question...-functions may be used in the questionnaire
        #   self.question_dd(variabel, Position): Dropdown selection
        #   self.question_cb(variabel, Position): Checkbox (Multiple item selection possible)
        #   self.question_txt(variabel, Position): Text input (one line)
        #   self.question_txt2(variabel, Position[, width][, height]): Text input (multiple lines)
        #   self.question_rb(variabel, Position): Radiobutton selection (Single item selection)
        #   self.question_rat(variabel, Position): Multiple Items with 1-5 Scale
        #   self.question_rat2(variabel, Position): Multiple Items with 1-2 Scale
        #   self.question_rat7(variabel, Position): Multiple Items with 1-7 Scale
        #   self.question_ls(variabel, listenvariabel): List selection
        #   self.question_lseek(variable,listvariable): Searchable list selection
        #   self.question_ladd(variable,listvariable): Creating a sub-list of a large list
        #   self.question_bt(variabel, Position): Up to four buttons to press (each one has to be defined in this function)
        
        settings['Page_History'].append(prog_pos)
        self.intronase()
        self.hide_review()
        self.buttons()
        settings['Curr_Page'] = [['',''],['',''],['','']]

        self.cini_schreiben() ###Kann nach Debugging auskommentiert werden

        if prog_pos == 'methode': #First page if no ID was found in any to-do list
            self.ausblenden()
            self.buttons(0,0,0,0)
            #self.question_dd('Methode',1)
            self.question_menu('Methode_Menu',1)
            
        elif prog_pos == 'Cooc_In':
            self.question_file('Cooc_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'Cooc_Spec':
            self.question_dd('Cooc_Case',1)
            self.question_dd('Cooc_MN',2)
            self.question_dd('Cooc_Mode',3)

        elif prog_pos == 'Cooc_Minanz':
            self.question_dd('Cooc_MinCas',1)
            self.question_dd('Cooc_MinAnz',2)
            self.question_dd('Cooc_Margin',3)

        elif prog_pos == 'Cooc_Out':
            self.question_file('Cooc_Output',1,'save')
            self.question_dd('Out_Sep',2)
            if storage['Cooc_Mode'][1] in ['inv_prob','sokal','eukl']:
                self.question_dd('Cooc_SSA',3)

        elif prog_pos == 'cd_in':
            self.question_file('CD_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'cd_spec':
            self.question_ladd('CD_Varlist','CD_Varlist',1)
            self.question_dd('Cooc_Mode',3)

        elif prog_pos == 'cd_margins':
            self.question_dd('Cooc_Margin',2)

        elif prog_pos == 'Dummy_In':
            self.question_file('Dummy_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'Dummy_Spec':
            self.question_ladd('Dummy_Case','Dummy_Case',1)
            self.question_dd('Dummy_MN',3)

        elif prog_pos == 'Dummy_Minanz':
            self.question_dd('Dummy_Mode',1)
            self.question_txt('Dummy_MinCas',2)
            self.question_txt('Dummy_MinAnz',3)

        elif prog_pos == 'Dummy_Out':
            self.question_file('Dummy_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'v_in':
            self.question_file('Visone_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'v_so':
            self.question_dd('Visone_Subject',1)
            self.question_dd('Visone_Object',2)
            self.question_dd('Visone_SubjObj',3)

        elif prog_pos == 'v_rel':
            self.question_dd('Visone_Minanz',1)
            self.question_dd('Visone_Relation',2)
            self.question_dd('Visone_Methode',3)

        elif prog_pos == 'v_out':
            self.question_file('Visone_Output',1,'save')
            self.question_file('Visone_Out2',2,'save')
            if storage['Visone_Methode'][1] == 'all':
                self.question_file('Visone_Out3',3,'save')

        elif prog_pos == 'm_haupt':
            self.question_file('Merge_Input1',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'm_schl':
            self.question_file('Merge_Input2',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'm_key1':
            self.question_ladd('Merge_Key1','Merge_Key1')

        elif prog_pos == 'm_key2':
            self.question_ladd('Merge_Key2','Merge_Key2')

        elif prog_pos == 'm_add':
            self.question_ladd('Merge_Add','Merge_Add')

        elif prog_pos == 'm_out':
            self.question_file('Merge_Out',1,'save')

        elif prog_pos == 'm_dir':
            self.question_file('Merge_Dir',1,'folder')

        elif prog_pos == 'm_files':
            self.question_ladd('Merge_Files','Merge_Files')

        elif prog_pos == 'mdir_out':
            self.question_file('Elong_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'elong_first':
            self.question_file('Elong_Input1',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'elong_second':
            self.question_file('Elong_Input2',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'elong_vars':
            self.question_ladd('Elong_Vars','Elong_Vars',1)

        elif prog_pos == 'elong_out':
            self.question_file('Elong_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'nino_input':
            self.question_file('Nino_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)
                        
        elif prog_pos == 'nino_spec':
            self.question_dd('Nino_Apptype',1)
            self.question_dd('Nino_Methode',2)
            self.question_file('Nino_Out',3,'save')

        elif prog_pos == 'nino_reshape':
            self.question_file('Nino_Input',1)
            self.question_file('Nino_Appeal',2)
            self.question_file('Nino_Outtable',3)

        elif prog_pos == 'agg_input':
            self.question_file('Agg_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)
                        
        elif prog_pos == 'agg_key':
            self.question_ladd('Agg_Group','Agg_Group',1)

        elif prog_pos == 'agg_var':
            self.question_ladd('Agg_Var','Agg_Var',1)
            self.question_dd('Agg_Method',3)

        elif prog_pos == 'agg_weigh':
            self.question_lseek('Agg_Weight','Agg_Var',1)

        elif prog_pos == 'agg_out':
            self.question_file('Agg_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'sub_input':
            self.question_file('Sub_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'sub_group':
            self.question_ladd('Sub_Group','Sub_Group',1)

        elif prog_pos == 'sub_values':
            self.question_ladd('Sub_Val','Sub_Val',1)

        elif prog_pos == 'sub_out':
            self.question_file('Sub_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'sub_split':
            self.question_file('Sub_Split',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'subr_input':
            self.question_file('Sub_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'subr_number':
            self.question_txt('Sub_Random_Number',2)

        elif prog_pos == 'subr_out':
            self.question_file('Sub_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'sv_input':
            self.question_file('Sub_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'sv_vars':
            self.question_ladd('Sub_Vars','Sub_Vars',1)

        elif prog_pos == 'sv_out':
            self.question_file('Sub_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'calc_input':
            self.question_file('Calc_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'calc_vars':
            self.question_ladd('Calc_Vars','Calc_Vars',1)
            self.question_dd('Calc_Method',3)

        elif prog_pos == 'calc_new':
            self.question_txt('Calc_New',1)

        elif prog_pos == 'calc_out':
            self.question_file('Calc_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'entropy_input':
            self.question_file('Entropy_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'entropy_vars':
            self.question_ladd('Entropy_Group','Entropy_Group')

        elif prog_pos == 'entropy_vars2':
            self.question_ladd('Entropy_Var','Entropy_Var')
            self.question_dd('Entropy_Mode',3)

        elif prog_pos == 'entropy_out':
            self.question_file('Entropy_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'resh_input':
            self.question_file('Resh_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'resh_vars':
            self.question_ladd('Resh_Vars','Resh_Vars',1)
            self.question_dd('Resh_Type',3)

        elif prog_pos == 'resh_out':
            self.question_file('Resh_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'ts_input':
            self.question_file('TS_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'ts_vars':
            self.question_dd('TS_Var',1)
            self.question_txt('TS_Nvar',2)

        elif prog_pos == 'ts_format':
            self.question_rbopen('TS_Informat',1)
            self.question_rbopen('TS_Outformat',2)

        elif prog_pos == 'ts_integer':
            self.question_dd('TS_Int',2)

        elif prog_pos == 'ts_out':
            self.question_file('TS_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'group_input':
            self.question_file('Group_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'group_vars':
            self.question_dd('Group_Var',1)
            self.question_txt('Group_NVar',2)

        elif prog_pos == 'group_mode':
            self.question_rb('Group_Mode',1)

        elif prog_pos == 'group_out':
            self.question_file('Group_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'sort_input':
            self.question_file('Sort_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'sort_vars':
            self.question_ladd('Sort_Var','Sort_Var')
            self.question_dd('Sort_Dir',3)

        elif prog_pos == 'sort_out':
            self.question_file('Sort_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

#rear_input

        elif prog_pos == 'rear_input':
            self.question_file('Rear_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'rear_vars':
            self.question_ladd('Rear_Var','Rear_Var')
            
        elif prog_pos == 'rear_out':
            self.question_file('Rear_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


#ren_input

        elif prog_pos == 'ren_input':
            self.question_file('Ren_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'ren_vars':
            self.question_dd('Ren_Var',1)
            self.question_txt('Ren_New',2)

        elif prog_pos == 'ren_other':
            self.question_bt('Ren_Other',1)
            
        elif prog_pos == 'ren_out':
            self.question_file('Ren_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'peak_input':
            self.question_file('Peak_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'peak_vars':
            self.question_dd('Peak_Time',1)
            self.question_dd('Peak_Series',2)
            self.question_bt('TS_Show',3)

        elif prog_pos == 'peak_method':
            self.question_dd('Peak_Direction',1)
            self.question_dd('Peak_Threshold',2)
            self.question_txt('Peak_Var',3)

        elif prog_pos == 'peak_out':
            self.question_file('Peak_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'flat_input':
            self.question_file('Flat_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'flat_invars':
            self.question_dd('Flat_Time',1)
            self.question_dd('Flat_Series',2)
            self.question_bt('TS_Show',3)

        elif prog_pos == 'flat_outvars':
            self.question_txt('Flat_Window',1)
            self.question_txt('Flat_Var',2)
            self.question_txt('Flat_Peaks',3)

        elif prog_pos == 'flat_out':
            self.question_file('Flat_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)
           

        elif prog_pos == 'glide_input':
            self.question_file('Glide_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'glide_units':
            self.question_dd('Glide_Time',1)
            self.question_dd('Glide_Units',2)

        elif prog_pos == 'glide_var':
            self.question_dd('Glide_Position',1)
            self.question_txt('Glide_Var',2)

        elif prog_pos == 'glide_out':
            self.question_file('Glide_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'gaps_input':
            self.question_file('Gap_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'gaps_var':
            self.question_dd('Gap_Tvar',1)
            self.question_dd('Gap_Gvar',2)
            self.question_txt('Gap_Length',3)

        elif prog_pos == 'gaps_opt':
            self.question_rb('Gap_Sort',1)
            self.question_txt('Gap_Store',2)

            
        elif prog_pos == 'gaps_out':
            self.question_file('Gap_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

            

        elif prog_pos == 'nts_input':
            self.question_file('NTS_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'nts_var':
            self.question_dd('NTS_Tvar',1)
            self.question_dd('NTS_Gvar',2)
            self.question_dd('NTS_Zero',3)

        elif prog_pos == 'nts_dur':
            self.question_txt('NTS_Duration',1)

        elif prog_pos == 'nts_vars':
            self.question_ladd('NTS_Vars','NTS_Vars')
            self.question_dd('NTS_Method',3)

        elif prog_pos == 'nts_out':
            self.question_file('NTS_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)
            


        elif prog_pos == 'mpatd_input':
            self.question_file('PD_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'mpatd_pattern':
            self.question_file('PD_Patterns',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)
           
        elif prog_pos == 'mpatd_var':
            self.question_ladd('PD_Series_mult','PD_Series_mult')
            self.question_txt('PD_Var',3)

        elif prog_pos == 'mpatd_var2':
            self.question_ladd('PD_Pattern_mult','PD_Pattern_mult')

        elif prog_pos == 'mpatd_len':
            self.question_txt('PD_Minlen',1)
            self.question_txt('PD_Maxlen',2)

        elif prog_pos == 'patd_opt':
            self.question_rb('PD_Method',1)
            #self.question_txt('PD_Cutoff',2)

        elif prog_pos == 'patd_out':
            self.question_file('PD_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'syn_input':
            self.question_file('Syn_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'syn_var':
            self.question_ladd('Syn_Var','Syn_Var')

        elif prog_pos == 'syn_opt':
            self.question_txt('Syn_Frame',1)
            self.question_dd('Syn_Meas',2)
            self.question_dd('Syn_TS',3)

        elif prog_pos == 'syn_output':
            self.question_file('Syn_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

           
 
        elif prog_pos == 'seq_input':
            self.question_file('Seq_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'seq_vars':
            self.question_dd('Seq_Svar',1)
            self.question_dd('Seq_Tvar',2)
            self.question_dd('Seq_Gvar',3)

        elif prog_pos == 'seq_length':
            self.question_dd('Seq_Length',1)
            self.question_dd('Seq_Omit',2)
            self.question_dd('Seq_Mode',3)


        elif prog_pos == 'seq_out':
            self.question_file('Seq_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'tpat_input':
            self.question_file('Tpat_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'tpat_vars':
            self.question_ladd('Tpat_Vars','Tpat_Vars')
            self.question_dd('Tpat_Time',3)

        elif prog_pos == 'tpat_options':
            if not 'Tpat_Level' in def_val.keys():
                def_val['Tpat_Level'] = '0.05'
            self.question_txt('Tpat_Level',1)
            self.question_dd('Tpat_Group',2)
            self.question_dd('Tpat_Long',3)

        elif prog_pos == 'tpat_out':
            self.question_file('Tpat_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)


        elif prog_pos == 'gind_input':
            self.question_file('Gind_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'gind_vars':
            self.question_dd('Gind_Symbol',1)
            self.question_dd('Gind_Group',2)
            self.question_dd('Gind_Time',3)

        elif prog_pos == 'gind_opt':
            #def_val['Gind_Glitch']='1'
            self.question_txt('Gind_Glitch',1)
            self.question_dd('Gind_Rep',2)
            self.question_dd('Gind_Subs',3)

        elif prog_pos == 'gind_repet':
            #def_val['Gind_Rep_Min']='1'
            #def_val['Gind_Rep_Max']='5'
            self.question_txt('Gind_Rep_Min',1)
            self.question_txt('Gind_Rep_Max',2)


        elif prog_pos == 'gind_len':
            #def_val['Gind_Len_Min']='3'
            #def_val['Gind_Len_Max']='7'
            #def_val['Gind_Eta']='0.9'
            self.question_txt('Gind_Len_Min',1)
            self.question_txt('Gind_Len_Max',2)
            self.question_txt('Gind_Eta',3)

        elif prog_pos == 'gind_out':
            self.question_file('Gind_Out',1,'save')
            

        elif prog_pos == 'cluster_input':
            self.question_file('Cluster_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'cluster_vars':
            self.question_ladd('Cluster_Vars','Cluster_Vars',1)
            self.question_cb('Cluster_Std',3)

        elif prog_pos == 'cluster_out':
            self.question_file('Cluster_Output',1,'save')
            self.question_cb('Cluster_Add',2)


        elif prog_pos == 'kcluster_input':
            self.question_file('Kcluster_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'kcluster_opt':
            if storage['Methode'] == 'An4':
                self.question_dd('Kcluster_Group',1)
            self.question_dd('Kcluster_Direction',2)
            self.question_txt('Kcluster_Anz',3)

        elif prog_pos == 'kcluster_vars':
            self.question_ladd('Kcluster_Vars','Kcluster_Vars',1)
            self.question_dd('Kcluster_Stand',3)

        elif prog_pos == 'kcluster_out':
            self.question_file('Kcluster_Out',1,'save')
            


        elif prog_pos == 'anent_input':
            self.question_file('Anent_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'anent_vars':
            self.question_ladd('Anent_Multi','Anent_Multi',1)
            self.question_dd('Anent_Group',3)

        elif prog_pos == 'anent_out':
            self.question_cb('Anent_Option',1)
            self.question_file('Anent_Output',2,'save')


        elif prog_pos == 'focus_input':
            self.question_file('Focus_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'focus_date':
            self.question_dd('Focus_Date',1)
            self.question_rb('Focus_Dformat',2)

        elif prog_pos == 'focus_text':
            self.question_dd('Focus_Ntext',1)
            self.question_dd('Focus_Weight',2)

        elif prog_pos == 'focus_issue':
            self.question_ladd('Focus_Issue','Focus_Issue')

        elif prog_pos == 'focus_actor':
            self.question_ladd('Focus_Actor','Focus_Actor')


        elif prog_pos == 'focus_window':
            self.question_dd('Focus_Window',1)
            self.question_dd('Focus_Direction',2)

        elif prog_pos == 'focus_out':
            self.question_file('Focus_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)
 

        elif prog_pos == 'heat_input':
            self.question_file('Heat_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'heat_vars':
            self.question_ladd('Heat_Vars','Heat_Vars',1)
            self.question_dd('Heat_Sort',3)

        elif prog_pos == 'heat_display':
            self.buttons(1,1,1,0)
            self.f_bottomline.b_abort["text"]="Finished"
            self.question_dd('Heat_Maxsize',1)
            self.question_dd('Heat_Color',2)
            self.question_dd('Heat_Log',3)


        elif prog_pos == 'visu_input':
            self.question_file('Visu_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'visu_plot':
            self.buttons(0,0,1,0)
            self.question_dd('Visu_X',1)
            self.question_dd('Visu_Y',2)
            self.question_bt('Visu_Plot',3)


        elif prog_pos == 'rt_input':
            self.question_file('Reltest_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'rt_var1':
            self.question_dd('Reltest_Unit',1)
            self.question_dd('Reltest_Coder',2)

        elif prog_pos == 'rt_var2':
            self.question_ladd('Reltest_Var','Reltest_Var')

        elif prog_pos == 'rt_units':
            self.question_ladd('Reltest_Units','Reltest_Units')

        elif prog_pos == 'rt_coders':
            self.question_ladd('Reltest_Coders','Reltest_Coders')

        elif prog_pos == 'rt_set':
            self.question_dd('Reltest_Core',1)
            self.question_cb('Reltest_Method',2)
            self.question_cb('Reltest_Options',3)

        elif prog_pos == 'rt_out':
            self.question_file('Reltest_Out',1,'save')
            self.question_file('Reltest_Report',2,'save')

### Text analysis methods

########################## Universe
            
        elif prog_pos == 'univ_in':
            self.question_file('Univ_Input',1,'folder')
            self.question_dd('Univ_Subdir',2)
            self.question_dd('Encoding',3)

        elif prog_pos == 'univ_mode':
            self.question_dd('Univ_Lang',1)
            self.question_dd('Univ_Length',2)
            self.question_dd('Univ_Sparse',3)

        elif prog_pos == 'univ_out':
            self.question_file('Univ_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

############################ Corpus
            
        elif prog_pos == 'corpus_in':
            self.question_file('Corpus_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'corpus_dir':
            def_val['Corpus_Indic']='path provided'
            self.question_file('Corpus_Indic',1,'folder')
            self.question_dd('Encoding',2)
            self.question_dd('Corpus_ID',3)

        elif prog_pos == 'corpus_nvar':
            self.question_txt('Corpus_Outvar',2)

        elif prog_pos == 'corpus_out':
            self.question_file('Corpus_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

        elif prog_pos == 'corpust_in':
            self.question_file('Corpus_Indic',1,'folder')
            self.question_dd('Encoding',2)
            self.question_dd('Corpus_Subdir',3)

        elif prog_pos == 'corpust_nvar':
            self.question_txt('Corpus_OutID',1)
            self.question_txt('Corpus_Outvar',2)

        elif prog_pos == 'corpust_out':
            self.question_file('Corpus_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)

######################### Inspect Corpus

        elif prog_pos == 'insp_in':
            self.question_file('RE_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'insp_var':
            self.question_dd('RE_Fulltext',1)
            self.question_dd('RE_ID',2)

        elif prog_pos == 'insp_find':
            self.question_txt('RE_Expression',1)
            self.question_rb('RE_Case',2)



######################### SVM
            
        elif prog_pos == 'svm_in':
            self.question_file('SVM_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'svm_vars':
            self.question_dd('SVM_Textvar',1)
            self.question_dd('SVM_Classvar',2)

        elif prog_pos == 'svm_opt':
            self.question_dd('Univ_Lang',1)
            self.question_dd('Univ_Length',2)
            self.question_dd('SVM_Sparse',3)

        elif prog_pos == 'svm_out':
            self.question_file('SVM_Out',1,'save')


        elif prog_pos == 'svmt_in':
            self.question_file('SVM_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'svmt_vars':
            self.question_dd('SVM_Textvar',1)
            self.question_dd('SVM_Classvar',2)

        elif prog_pos == 'svmt_opt':
            self.question_dd('Univ_Lang',1)
            self.question_dd('Univ_Length',2)


        elif prog_pos == 'svmt_model':
            self.question_file('SVM_Model',1,defext='.json')
            self.question_dd('SVM_Type',2)
            self.question_txt('SVM_Newvar',3)

        elif prog_pos == 'svmt_try':
            self.buttons(0,0,0,0)
            self.question_txt('SVM_Adjust',1)
            self.question_bt('SVM_Tryout',2)

        elif prog_pos == 'svmt_out':
            self.question_file('SVM_Outtable',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)            


        elif prog_pos == 'svma_in1':
            self.question_file('SVMA_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'svma_var':
            self.question_dd('SVM_Textvar',1)

        elif prog_pos == 'svma_in2':
            self.question_file('SVMA_Folder',1,'folder')
            self.question_dd('Encoding',2)
            self.question_dd('Corpus_Subdir',3)

        elif prog_pos == 'svma_opt':
            self.question_dd('Univ_Lang',1)
            self.question_dd('Univ_Length',2)

        elif prog_pos == 'svma_model':
            self.question_file('SVM_Model',1,defext='.json')
            self.question_dd('SVM_Type',2)
            self.question_txt('SVM_Newvar',3)

        elif prog_pos == 'svma_out':
            self.question_file('SVM_Outtable',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)       


######################### NBC


        elif prog_pos == 'nbc_in':
            self.question_file('SVM_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'nbc_vars':
            self.question_dd('SVM_Textvar',1)
            self.question_dd('SVM_Classvar',2)

        elif prog_pos == 'nbc_opt':
            self.question_dd('Univ_Lang',1)
            self.question_dd('Univ_Length',2)
            self.question_dd('SVM_Sparse',3)

        elif prog_pos == 'svm_out':
            self.question_file('NBC_Out',1,'save')


        elif prog_pos == 'nbct_in':
            self.question_file('SVM_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'nbct_vars':
            self.question_dd('SVM_Textvar',1)
            self.question_dd('SVM_Classvar',2)

        elif prog_pos == 'nbct_opt':
            self.question_dd('Univ_Lang',1)
            self.question_dd('Univ_Length',2)


        elif prog_pos == 'nbct_model':
            self.question_file('NBC_Model',1,defext='.json')
            self.question_dd('NBC_Type',2)
            self.question_txt('SVM_Newvar',3)



##        elif prog_pos == 'svmt_try':
##            self.buttons(0,0,0,0)
##            self.question_txt('SVM_Adjust',1)
##            self.question_bt('SVM_Tryout',2)
##
##        elif prog_pos == 'svmt_out':
##            self.question_file('SVM_Outtable',1,'save')
##            self.question_dd('Out_Header',2)
##            self.question_dd('Out_Sep',3)            
##
##
##        elif prog_pos == 'svma_in1':
##            self.question_file('SVMA_Input',1)
##            self.question_dd('In_Sep',2)
##            self.question_dd('In_Header',3)
##
##        elif prog_pos == 'svma_var':
##            self.question_dd('SVM_Textvar',1)
##
##        elif prog_pos == 'svma_in2':
##            self.question_file('SVMA_Folder',1,'folder')
##            self.question_dd('Encoding',2)
##            self.question_dd('Corpus_Subdir',3)
##
##        elif prog_pos == 'svma_opt':
##            self.question_dd('Univ_Lang',1)
##            self.question_dd('Univ_Length',2)
##
##        elif prog_pos == 'svma_model':
##            self.question_file('SVM_Model',1,defext='.json')
##            self.question_dd('SVM_Type',2)
##            self.question_txt('SVM_Newvar',3)
##
##        elif prog_pos == 'svma_out':
##            self.question_file('SVM_Outtable',1,'save')
##            self.question_dd('Out_Header',2)
##            self.question_dd('Out_Sep',3)       
                


############## Duplicate analysis (N-Gram Shingling)

        elif prog_pos == 'dupli_in1':
            self.question_file('NGS_Input',1)
            self.question_dd('In_Sep',2)
            self.question_dd('In_Header',3)

        elif prog_pos == 'dupli_var':
            self.question_dd('NGS_Fulltext',1)
            self.question_dd('NGS_Tid',2)
            
        elif prog_pos == 'dupli_in2':
            self.question_file('Corpus_Indic',1,'folder')
            self.question_dd('Encoding',2)
            self.question_dd('Corpus_Subdir',3)
        

        elif prog_pos == 'dupli_opt':
            self.question_dd('NGS_Nglen',1)
            #self.question_txt('NGS_Overlap',2)

        elif prog_pos == 'dupli_dec':
            self.question_txt('NGS_Minover',1)
            self.question_dd('NGS_Sym',2)
            self.question_dd('NGS_Share',3)

        elif prog_pos == 'dupli_out':
            self.question_file('NGS_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)            


############################## NCCR Populism
            
        elif prog_pos == 'pop_input':
            self.question_file('Populism_Input',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'pop_out':
            self.question_file('Populism_Output',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)
           

        elif prog_pos == 'angrist_hours':
            def_val['Hours_Output']='_test_stunden.txt'
            self.question_file('Hours_Input',1)
            self.question_file('Hours_Output',2,'save')

        elif prog_pos == 'match_in':
            self.question_file('Match_Content',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

        elif prog_pos == 'match_vnames':
            self.question_dd('Match_Mdate',1)
            self.question_dd('Match_Mdate_Trans',2)
            self.question_dd('Match_Mweight',3)

        elif prog_pos == 'match_vars':
            self.question_ladd('Match_Mvars','Match_Mvars')
            self.question_dd('Match_Aggmode',3)

        
        elif prog_pos == 'match_resp':
            self.question_file('Match_Respondents',1)
            self.question_dd('In_Header',2)
            self.question_dd('In_Sep',3)

     
        elif prog_pos == 'match_survey':
            self.question_dd('Match_Gweight',1)
            self.question_dd('Match_Lweight',2)

        elif prog_pos == 'match_date':
            self.question_dd('Match_Sdate',1)
            self.question_dd('Match_Wdate',2)
            self.question_dd('Match_Calcmode',3)

        elif prog_pos == 'match_out':
            self.question_file('Match_Out',1,'save')
            self.question_dd('Out_Header',2)
            self.question_dd('Out_Sep',3)
        
        
        elif prog_pos == 'ggcrisi':
            def_val['GGCRISI_Output']='_Summary.txt'
            self.question_file('GGCRISI_Input',1,'folder')
            self.question_file('GGCRISI_Output',2,'save')


        elif prog_pos == 'otherart':
            self.buttons(0,0,1,1)
            self.question_bt('Otherart',1)
            self.f_questions.bu1_1["command"] = self.submit
            self.f_questions.bu1_2["command"] = self.abort
            self.f_questions.bu1_1.bind('<Return>',self.submit)
            self.f_questions.bu1_1.focus()

            if available('Settings'):
                verb('Writing Coder information')
                self.cini_schreiben()        
            
        elif prog_pos == 'ende':
            self.buttons(0,0,0,0)
            self.ausblenden()
            self.f_questions.Frage2.insert(INSERT,self.namegetter('Location','End'), 'fett')

        else:
            verb('ERROR: Unknown program position for MCP: ' + prog_pos)
                


############################################
##                                        ##
##       SUBMIT - FUNCTION                ##
##                                        ##
############################################
        
    def submit(self,overspill=0):
        global prog_pos
        global dta_pos
        global settings
        global storage
        global codebook
        log('\nCalling Function: SUBMIT')
        #The Submit-Function takes all entries as soon as the Check-Button is klicked by the user (or any other button directing
        #to this function).
        #All Values are stored to the central dictionary 'storage'. This may be done using the function self.store_var.
        #If any invalid entry was made, the submit-function is not executed. In this case, an error message will be shown to the coder.
        #At the end of storage and cleaning up, a new program position has to be defined and the MCP-function has to be called again.
        #
        #If necessary, the submit-Function will take one argument (overspill) which may be used in the handling of entries.
        #The overspill-parameter will also catch the event when binding a widget-event to the submit function.

        accept_entry = 0 #Only if the entry is valid, acceptance is set to 1
        accept_entry = self.check_entries() #Checking for any invalid entries. Returning 1 if everything is OK
        try:
            self.f_bottomline.b_check["state"] = DISABLED
        except:
            verb('No Check button')

        if accept_entry == 1:

            if prog_pos == 'methode':
                self.clean_up_all()
                settings['Datasets'] = {}
                m = overspill
                storage['Methode']=m
                
                if m == 'R1':
                    verb('----Choice: Dummy Table')
                    self.anzeigen()
                    verbout('New Procedure: Dummy Variables','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'Dummy_In'
                    self.ask()                                       
                elif m == 'C1':
                    verb('----Choice: co-occurrence')
                    self.anzeigen()
                    verbout('New Procedure: Co-occurrence Matrix','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'Cooc_In'
                    self.ask()
                elif m == 'C2':
                    verb('----Choice: co-occurrence from dummy table')
                    self.anzeigen()
                    verbout('New Procedure: Co-occurrence Matrix','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'cd_in'
                    self.ask()
                elif m == 'V1':
                    verb('----Choice: Visone')
                    self.anzeigen()
                    verbout('New Procedure: Visualization of Social Networks','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'v_in'
                    self.ask()
                elif m == 'V2':
                    verb('----Choice: NINO')
                    self.anzeigen()
                    verbout('New Procedure: Social Network from Ninos Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'nino_input'
                    self.ask()
                elif m == 'V3':
                    verb('----Choice: NINO')
                    self.anzeigen()
                    verbout('New Procedure: Reshaping Ninos Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'nino_reshape'
                    self.ask()
                elif m == 'M1':
                    verb('----Choice: Merge')
                    self.anzeigen()
                    verbout('New Procedure: Merging of two tables (add variables)','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'm_haupt'
                    self.ask()
                elif m == 'M2':
                    verb('----Choice: Elongate')
                    self.anzeigen()
                    verbout('New Procedure: Merging of two tables (add cases)','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'elong_first'
                    self.ask()
                elif m == 'M3':
                    verb('----Choice: Elongate massive')
                    self.anzeigen()
                    verbout('New Procedure: Merging of many tables (add cases)','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'm_dir'
                    self.ask()
                elif m == 'A1':
                    verb('----Choice: Aggregate')
                    self.anzeigen()
                    verbout('New Procedure: Aggregation of Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'agg_input'
                    self.ask()
                elif m == 'S1':
                    verb('----Choice: Subset')
                    self.anzeigen()
                    verbout('New Procedure: Subset from Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'sub_input'
                    storage['Sub_Method'] = 'select'
                    self.ask()
                elif m == 'S4':
                    verb('----Choice: Subsets')
                    self.anzeigen()
                    verbout('New Procedure: Several subsets from Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'sub_input'
                    storage['Sub_Method'] = 'split'
                    self.ask()
                elif m == 'S3':
                    verb('----Choice: Random Subset')
                    self.anzeigen()
                    verbout('New Procedure: Random Subset from Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'subr_input'
                    self.ask()
                elif m == 'S2':
                    verb('----Choice: Subset of Variables')
                    self.anzeigen()
                    verbout('New Procedure: Select Variables','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'sv_input'
                    self.ask()
                elif m == 'A2':
                    verb('----Choice: Easycalc')
                    self.anzeigen()
                    verbout('New Procedure: Aggregate Variables: Easy Calculation','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'calc_input'
                    self.ask()
                elif m == 'A3':
                    verb('----Choice: Entropy')
                    self.anzeigen()
                    verbout('New Procedure: Calculate Entropy','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'entropy_input'
                    self.ask()
                elif m == 'An1':
                    verb('----Choice: HECANE Cluster Analysis')
                    self.anzeigen()
                    verbout('New Procedure: Cluster Analysis of count data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'cluster_input'
                    self.ask()
                elif m == 'An3':
                    verb('----Choice: k-means Cluster Analysis')
                    self.anzeigen()
                    verbout('New Procedure: k-means Cluster Analysis','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'kcluster_input'
                    self.ask()
                elif m == 'An4':
                    verb('----Choice: k-means Cluster Analysis')
                    self.anzeigen()
                    verbout('New Procedure: Multigroup Cluster Analysis','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'kcluster_input'
                    self.ask()
                elif m == 'An2':
                    verb('----Choice: Analysis of Entropy')
                    self.anzeigen()
                    verbout('New Procedure: Analysis of Entropy','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'anent_input'
                    self.ask()                
                elif m == 'An5':
                    verb('----Choice: Analysis of Attention and Focus')
                    self.anzeigen()
                    verbout('New Procedure: Analysis of Attention and Focus','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'focus_input'
                    self.ask()                
                elif m == 'Vi1':
                    verb('----Choice: Heat Map')
                    self.anzeigen()
                    verbout('New Procedure: Display Heat Map','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'heat_input'
                    self.ask()
                elif m == 'Vi2':
                    verb('----Choice: X-Y Plot')
                    self.anzeigen()
                    verbout('New Procedure: Display X-Y Plot','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'visu_input'
                    self.ask()
                elif m == 'R2':
                    verb('----Choice: Reshape Slim')
                    self.anzeigen()
                    verbout('New Procedure: Reshape Dummy to slim table','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'resh_input'
                    self.ask()
                elif m == 'T1':
                    verb('----Choice: Timestamps')
                    self.anzeigen()
                    verbout('New Procedure: Transformation of Timestamps','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'ts_input'
                    self.ask()
                elif m == 'T2':
                    verb('----Choice: Group')
                    self.anzeigen()
                    verbout('New Procedure: Transform scale to groups','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'group_input'
                    self.ask()
                elif m == 'So1':
                    verb('----Choice: Sort Cases')
                    self.anzeigen()
                    verbout('New Procedure: Sorting cases','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'sort_input'
                    self.ask()                    
                elif m == 'So2':
                    verb('----Choice: Rearrange Variables')
                    self.anzeigen()
                    verbout('New Procedure: Rearrange variables','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'rear_input'
                    self.ask()                    
                elif m == 'So3':
                    verb('----Choice: Rename Variables')
                    self.anzeigen()
                    verbout('New Procedure: Rename Variables','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'ren_input'
                    self.ask()                    
                elif m == 'Ts1':
                    verb('----Choice: Peak Detext')
                    self.anzeigen()
                    verbout('New Procedure: Detection of Peaks','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'peak_input'
                    self.ask()
                elif m == 'Ts2':
                    verb('----Choice: Flatten Moving Average')
                    self.anzeigen()
                    verbout('New Procedure: Flatten Moving Average','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'flat_input'
                    self.ask()
                elif m == 'Ts3':
                    verb('----Choice: Gliding Window')
                    self.anzeigen()
                    verbout('New Procedure: Creating gliding window','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'glide_input'
                    self.ask()
                elif m == 'Ts4':
                    verb('----Choice: Detect Gaps')
                    self.anzeigen()
                    verbout('New Procedure: Detect gaps in timeseries','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'gaps_input'
                    self.ask()
                elif m == 'Ts7':
                    verb('----Choice: Normalize Time series')
                    self.anzeigen()
                    verbout('New Procedure: Normalize Timeseries to force equidistant measurements','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'nts_input'
                    self.ask()                
                elif m == 'Ts5':
                    verb('----Choice: Parallel Pattern Detection')
                    self.anzeigen()
                    verbout('New Procedure: Detecting parallel patterns in Timeseries Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'mpatd_input'
                    self.ask()
                elif m == 'Ts6':
                    verb('----Choice: Synchronize Event Data')
                    self.anzeigen()
                    verbout('New Procedure: Synchronize Event data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'syn_input'
                    self.ask()
                elif m == 'Se1':
                    verb('----Choice: Find Sequences')
                    self.anzeigen()
                    verbout('New Procedure: Finding common sequences','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'seq_input'
                    self.ask()
                elif m == 'Se2':
                    verb('----Choice: Find T-Patterns')
                    self.anzeigen()
                    verbout('New Procedure: Finding T-Patterns','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'tpat_input'
                    self.ask()
                elif m == 'Se3':
                    verb('----Choice: Grammar Inductiuon')
                    self.anzeigen()
                    verbout('New Procedure: Grammar Induction','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'gind_input'
                    self.ask()
                elif m=='RT1':
                    verb('---Choice: Reliability Testing')
                    self.anzeigen()
                    verbout('New Procedure: Reliability Testing','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'rt_input'
                    self.ask()
                elif m == 'U':
                    verb('----Choice: Get Universe')
                    self.anzeigen()
                    self.verbout('New Procedure: Universe of n-grams\n---------------------\n')
                    prog_pos = 'univ_in'
                    self.ask()                                       
                elif m == 'C':
                    verb('----Choice: Corpus')
                    self.anzeigen()
                    self.verbout('New Procedure: Corpus\n-------------------------\n')
                    prog_pos = 'corpus_in'
                    self.ask()
                elif m == 'CT':
                    verb('----Choice: Corpus from folder')
                    self.anzeigen()
                    self.verbout('New Procedure: Corpus from folder\n-------------------------\n')
                    prog_pos = 'corpust_in'
                    self.ask()
                elif m in ['Train','Tcorr','Tcheck']:
                    verb('----Choice: Training')
                    self.anzeigen()
                    self.verbout('New Procedure: Training\n-------------------------\n')
                    prog_pos = 'train_in'
                    self.ask()
                elif m == 'T1':
                    verb('----Choice: Test')
                    self.anzeigen()
                    self.verbout('New Procedure: Test\n-------------------------\n')
                    prog_pos = 'test_in'
                    self.ask()                  
                elif m == 'T2':
                    verb('----Choice: Reltest')
                    self.anzeigen()
                    self.verbout('New Procedure: Reliability Test\n-------------------------\n')
                    prog_pos = 'rtest_in'
                    self.ask()                  
                elif m == 'T4':
                    verb('----Choice: Reltest 10')
                    self.anzeigen()
                    self.verbout('New Procedure: Reliability Test with half corpora\n-------------------------\n')
                    prog_pos = 'rtest_in'
                    self.ask()                  
                elif m == 'T3':
                    verb('----Choice: Batch Reltest')
                    self.anzeigen()
                    self.verbout('New Procedure: Reliability Test\n-------------------------\n')
                    prog_pos = 'btest_in'
                    self.ask()                  
                elif m == 'P1':
                    verb('----Choice: Predict Values')
                    self.anzeigen()
                    self.verbout('New Procedure: Predict values\n-------------------------\n')
                    prog_pos = 'pred_in'
                    storage['Pred_Mode'] = 1
                    self.ask()                  
                elif m == 'P2':
                    verb('----Choice: Predict Values')
                    self.anzeigen()
                    self.verbout('New Procedure: Predict probabilities\n-------------------------\n')
                    prog_pos = 'pred_in'
                    storage['Pred_Mode'] = 2
                    self.ask()
                elif m == 'In1':
                    self.anzeigen()
                    self.verbout('New Procedure: Inspect Corpus\n-------------------------\n')
                    prog_pos = 'insp_in'
                    self.ask()                    
                elif m == 'SVM_Train':
                    verb('----Choice: Train SVM')
                    self.anzeigen()
                    self.verbout('New Procedure: Train Support Vector Machine\n-------------------------\n')
                    prog_pos = 'svm_in'
                    self.ask()                  
                elif m == 'SVM_Test':
                    verb('----Choice: Train SVM')
                    self.anzeigen()
                    self.verbout('New Procedure: Test Support Vector Machine\n-------------------------\n')
                    prog_pos = 'svmt_in'
                    self.ask()                  
                elif m == 'SVM_Apply1':
                    verb('----Choice: Train SVM')
                    self.anzeigen()
                    self.verbout('New Procedure: Apply Support Vector Machine to Corpus\n-------------------------\n')
                    prog_pos = 'svma_in1'
                    self.ask()                  
                elif m == 'SVM_Apply2':
                    verb('----Choice: Train SVM')
                    self.anzeigen()
                    self.verbout('New Procedure: Apply Support Vector Machine to Directory\n-------------------------\n')
                    prog_pos = 'svma_in2'
                    self.ask()                  
                elif m == 'Tool1':
                    verb('----Choice: TDM')
                    self.anzeigen()
                    self.verbout('New Procedure: Term-Document-Matrix\n-------------------------\n')
                    prog_pos = 'tdm_in'
                    self.ask()                  
                elif m == 'Tool2':
                    verb('----Choice: Term Co-Occurrence')
                    self.anzeigen()
                    self.verbout('New Procedure: Term Co-Occurrence with Keyword\n-------------------------\n')
                    prog_pos = 'tco_in'
                    self.ask()                  
                elif m == 'Tool3':
                    verb('----Choice: Split Files')
                    self.anzeigen()
                    self.verbout('New Procedure: Split long textfile to single texts\n-------------------------\n')
                    prog_pos = 'split_in'
                    self.ask()                  
                elif m == 'Tool4':
                    verb('----Choice: Language Settings')
                    self.anzeigen()
                    self.verbout('New Procedure: Test Language/Encoding settings\n-------------------------\n')
                    prog_pos = 'langtest_in'
                    self.ask()
                elif m in ['Dict1','Dict2']:
                    verb('----Choice: Dictionary based annotation')
                    self.anzeigen()
                    self.verbout('New Procedure: Dictionary based annotation')
                    prog_pos = 'dict_in'
                    self.ask()
                elif m in ['Dupli1']:
                    verb('----Choice: Finding duplicates')
                    self.anzeigen()
                    self.verbout('New Procedure: Duplicate analysis using n-gram shingles')
                    prog_pos = 'dupli_in1'
                    self.ask()
                elif m in ['Dupli2']:
                    verb('----Choice: Finding duplicates')
                    self.anzeigen()
                    self.verbout('New Procedure: Duplicate analysis of folder using n-gram shingles')
                    prog_pos = 'dupli_in2'
                    self.ask()
                elif m == 'NCCR1':
                    verb('----Choice: Adding Populism')
                    self.anzeigen()
                    verbout('New Procedure: Creating Populism Variables','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'pop_input'
                    self.ask()
                elif m == 'NCCR2':
                    verb('----Choice: Calculating Working Hours')
                    self.anzeigen()
                    verbout('New Procedure: Calculating Hours','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'angrist_hours'
                    self.ask()
                elif m in ['NCCR3','NCCR4']:
                    verb('----Choice: Matching')
                    self.anzeigen()
                    verbout('New Procedure: Matching Content and Survey','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'match_in'
                    self.ask()
                elif m == 'GGCRISI':
                    verb('----Choice: GGCRISI aggregate')
                    self.anzeigen()
                    verbout('New Procedure: Aggregating all GGCRISI Data','title',master=self)
                    verbout('\n\n',master=self)
                    prog_pos = 'ggcrisi'
                    self.ask()
                
                    
                else:
                    verb('ERROR: Option '+m+' is not defined')



######################################################## Co-Occurrence Analysis

            elif prog_pos == 'Cooc_In':
                fname = make_fname(self.store_var('Cooc_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Cooc_Case',dset[1],excludes=[])
                    add_varlist('Cooc_MN',dset[1],excludes=[])
                    
                    self.clean_up_all()
                    prog_pos = 'Cooc_Spec'
                    self.ask()

            elif prog_pos == 'Cooc_Spec':
                self.store_var_all()
                self.clean_up_all()

                tmp = storage['Data']
                tmp_a = get_unique(tmp[storage['Cooc_Case'][1]])
                tmp_b = get_unique(tmp[storage['Cooc_MN'][1]])

                verbout('\n\nGroup-Variable: ',master=self)
                verbout(storage['Cooc_Case'][1],'table',master=self)
                verbout('\n' + str(len(tmp_a)) + ' Groups in this File:\n',master=self)
                verbout(str(tmp_a[:20]),'table',master=self)
                if len(tmp_a)>20:
                    verbout(' (trunctuated list. First 20 elements)',master=self)
                verbout('\n\nElement-Variable: ',master=self)
                verbout(storage['Cooc_MN'][1],'table',master=self)
                verbout('\n' +str(len(tmp_b)) + ' Elements in this File:\n',master=self)
                verbout(str(tmp_b[:20]),'table',master=self)
                if len(tmp_b)>20:
                    verbout(' (trunctuated list. First 20 elements)',master=self)
                verbout('\n',master=self)
                
                prog_pos = 'Cooc_Minanz'
                self.ask()
            
            elif prog_pos == 'Cooc_Minanz':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'Cooc_Out'
                self.ask()

            elif prog_pos == 'Cooc_Out':
                self.store_var_all()
                outfile = make_fname(storage['Cooc_Output'])
                method = storage['Cooc_Mode'][1]
                margin = storage['Cooc_Margin'][1]
                if 'Cooc_SSA' in storage.keys(): #Safety Override
                    if storage['Cooc_SSA'][1] == '1':
                        margin = '0'
                header = 1
                sep = get_sep(storage['Out_Sep'][1])
                        
                if storage['Methode']=='C1':
                    data = storage['Data']
                    cases = storage['Cooc_Case'][1]
                    keyvar = storage['Cooc_MN'][1]
                    prefix = storage['Cooc_MN'][1]+'_'
                    min_case = int(storage['Cooc_MinCas'][1])
                    min_anz = int(storage['Cooc_MinAnz'][1])
                    dmode = 'nominal'
                    out_data = co_occurrence(data,dmode,method,margin,prefix,cases=cases,keyvar=keyvar,min_case=min_case,
                                         min_anz=min_anz,master=self)
                    
                elif storage['Methode']=='C2':
                    dummylist = storage['CD_Varlist'][1]
                    data = storage['Dummy_Data']
                    prefix = ''
                    dmode = 'dummy'
                    out_data = co_occurrence(data,dmode,method,margin,prefix,dummylist=dummylist,master=self)


                t = write_dataset(out_data,outfile,header,sep)
                verbout('\n'+t[0],master=self)
                if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                if 'Cooc_SSA' in storage.keys():
                    verbout('\nPreparing SSA-script: ',master=self)
                    if storage['Cooc_SSA'][1] == '1':
                        ssafile = outfile[:-4]+'_SSA.R'
                        verbout(ssafile,master=self)
                        write_ssa(out_mat,ssafile)
                
                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()


            elif prog_pos == 'cd_in':
                fname = make_fname(self.store_var('CD_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_vars = []
                    verbout('\nSuitability of variables:\nFor each variable which may be used as dummy variable (1/0; value/missing), the number of true (1 or value) and false (0 or missing) values are indicated below.\nInvariant variables are regarded as unsuitable for co-occurrence analyses.\n\n',master=self)
                    for var in v:
                        if cd_proof(var,master=self):
                            add_vars.append(var)
                    add_varlist('CD_Varlist',add_vars,excludes=[])
                    
                    self.clean_up_all()
                    prog_pos = 'cd_spec'
                    self.ask()

            elif prog_pos == 'cd_spec':
                self.store_var_all()
                l = self.store_var('CD_Varlist')
                if len(l) > 0:
                    self.clean_up_all()
                    prog_pos = 'cd_margins'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'cd_margins':
                self.store_var_all()
                self.clean_up_all()
                if storage['Methode']=='C1':
                    prog_pos = 'Cooc_Minanz'
                elif storage['Methode']=='C2':
                    prog_pos = 'Cooc_Out'
                self.ask()

######################################################## Reshape Dummy Table

            elif prog_pos == 'Dummy_In':
                fname = make_fname(self.store_var('Dummy_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Dummy_Case',v,excludes=[]) ##Liste von Variablen hinzufügen
                    add_varlist('Dummy_MN',v,excludes=[])

                    self.clean_up_all()
                    prog_pos = 'Dummy_Spec'
                    self.ask()

            elif prog_pos == 'Dummy_Spec':
                self.store_var_all()
                self.clean_up_all()

                tmp = storage['Data']
                cv = storage['Dummy_Case'][1]
                if not cv == []:
                    casdic = {}
                    for i in range(len(tmp[cv[0]])):
                        case = ''
                        for v in cv:
                            case = case + tmp[v][i]
                        casdic[case]=1
                    tmp_a = list(casdic.keys())
                else:
                    verbout('\n',master=self)
                    verbout('Information: No case variable selected.\nEach row will be treated as a case.\n','warning',master=self)
                    tmp_a = tmp[storage['Dummy_MN'][1]]
                tmp_b = get_unique(tmp[storage['Dummy_MN'][1]])

                if not cv == []:
                    verbout('\nCase-Variable:\n',master=self)
                    verbout(str(storage['Dummy_Case'][1]),'table',master=self)
                    verbout('\n' + str(len(tmp_a)) + ' Cases in this File:\n',master=self)
                    verbout(str(tmp_a[:20]),'table',master=self)
                    if len(tmp_a)>20:
                        verbout(' (trunctuated list. More than 20 cases)',master=self)
                    verbout('\n',master=self)
                
                verbout('\nValue-Variable: \n',master=self)
                verbout(str(storage['Dummy_MN'][1]),'table',master=self)
                verbout('\n' +str(len(tmp_b)) + ' Values for multinomial variable in this File: \n',master=self)
                verbout(str(tmp_b[:20]),'table',master=self)
                if len(tmp_b)>20:
                    verbout(' (trunctuated list. More than 20 values)',master=self)
                verbout('\n',master=self)
                
                prog_pos = 'Dummy_Minanz'
                self.ask()               

            elif prog_pos == 'Dummy_Minanz':
                self.store_var_all()
                try:
                    storage['Dummy_MinCas'] = int(storage['Dummy_MinCas'])
                    storage['Dummy_MinAnz'] = int(storage['Dummy_MinAnz'])
                    self.clean_up_all()
                    prog_pos = 'Dummy_Out'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')


            elif prog_pos == 'Dummy_Out':
                self.store_var_all()
                data = storage['Data']
                cases = storage['Dummy_Case'][1]
                keyvar = storage['Dummy_MN'][1]
                outfile = make_fname(storage['Dummy_Output'])
                header = int(storage['Out_Header'][1])
                min_case = storage['Dummy_MinCas']
                min_anz = storage['Dummy_MinAnz']
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nCalculating Dummy Table. Please be patient.',master=self)
                method = storage['Dummy_Mode'][1]

                out_data = dummytab(data,cases,keyvar,method,min_case,min_anz,master=self)
                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

######################################################## Visone

            elif prog_pos == 'v_in':                
                fname = make_fname(self.store_var('Visone_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Visone_Subject',v,excludes=[])
                    add_varlist('Visone_Object',v,excludes=[])
                    add_varlist('Visone_Relation',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'v_so'
                    self.ask()

            elif prog_pos == 'v_so':
                self.store_var_all()
                self.clean_up_all()
                subj = storage['Visone_Subject'][1]
                obj = storage['Visone_Object'][1]
                equal = storage['Visone_SubjObj'][1]

                l1 = storage['Data'][subj]
                l2 = storage['Data'][obj]

                if equal == '1':
                    liste = get_unique(l1+l2)
                    verbout('\n\n' +str(len(liste)) +' Cases for Subject/Object:\n',master=self)
                    verbout(str(liste[:20]),'table',master=self)
                    if len(liste)>20:verbout(' (trunctuated list. First 20 cases)',master=self)
                else:
                    verbout('\n\n' +str(len(get_unique(l1))) +' Cases for Subject:\n',master=self)
                    verbout(str(get_unique(l1)[:20]),'table',master=self)
                    if len(l1)>20:verbout(' (trunctuated list. First 20 cases)',master=self)
                    verbout('\n' +str(len(get_unique(l2))) +' Cases for Object:\n',master=self)
                    verbout(str(get_unique(l2)[:20]),'table',master=self)
                    if len(l2)>20:verbout(' (trunctuated list. First 20 cases)',master=self)
                    
                prog_pos = 'v_rel'
                self.ask()                    


            elif prog_pos == 'v_rel':
                self.store_var_all()
                accept = 0
                if storage['Visone_Relation'][1] == '9':
                    if storage['Visone_Methode'][1] in ['anz','dicho','entf']:
                        accept = 1
                else:
                    accept = 1

                if accept == 1:
                    self.clean_up_all()
                    prog_pos = 'v_out'
                    self.ask()
                else:
                    self.message('Method Fail')


            elif prog_pos == 'v_out':
                self.store_var_all()
                self.clean_up_all()
                adj = make_fname(storage['Visone_Output'])
                att = make_fname(storage['Visone_Out2'])
                if 'Visone_Out3' in storage.keys():
                    lin = make_fname(storage['Visone_Out3'])
                else:
                    lin = ''
                subj = storage['Data'][storage['Visone_Subject'][1]]
                obj = storage['Data'][storage['Visone_Object'][1]]
                methode = storage['Visone_Methode'][1]
                if storage['Visone_Relation'][1] == '9':
                    rel = []#subj ###Spielt keine Rolle, da ja nur Anzahl oder Dichotom gemessen wird.
                else:
                    rel = storage['Data'][storage['Visone_Relation'][1]]
                min_anz = int(storage['Visone_Minanz'][1])
                obj_is_subj = int(storage['Visone_SubjObj'][1])

                v_data = create_visone(subj,obj,rel,methode,min_anz,obj_is_subj,master=self)

                t = write_dataset(v_data['Adjacency'],adj) ##Write adjacency matrix
                verbout(t[0],master=self)
                if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                t = write_dataset(v_data['Nodes'],att) ##Write Nodes table
                verbout(t[0],master=self)
                if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                if methode == 'all':
                    t = write_dataset(v_data['Links'],lin) ##Write adjacency matrix
                    verbout(t[0],master=self)
                    if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)                
                
                prog_pos = 'otherart'
                self.ask()

######################################################## Merge datasets

            elif prog_pos == 'm_haupt':
                fname = make_fname(self.store_var('Merge_Input1'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading Main Table: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Haupt','H_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Merge_Key1',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'm_schl'
                    self.ask()

            elif prog_pos == 'm_schl':
                fname = make_fname(self.store_var('Merge_Input2'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading Key Table: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Schl','S_Var','Key Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Merge_Key2',v,excludes=[])
                    
                    self.clean_up_all()
                    prog_pos = 'm_key1'
                    self.ask()

            elif prog_pos == 'm_key1':
                vlist = self.store_var('Merge_Key1')[1]
                verbout('\nKey Variables in Main Table:\n',master=self)
                if len(vlist) > 0:
                    for v in vlist:
                        verbout(v+'\n',master=self)
                    self.clean_up_all()
                    prog_pos = 'm_key2'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')

            elif prog_pos == 'm_key2':
                vlist = self.store_var('Merge_Key2')[1]
                accept = 1
                if len(vlist) < 0:
                    accept = 0
                    self.message('Invalid-Selection01')
                if not len(vlist) == len(storage['Merge_Key1'][1]):
                    accept = 0
                    self.message('Invalid-Selection05')

                if accept == 1:
                    add_varlist('Merge_Add',codebook['Merge_Key2'][2],excludes=vlist) ##Update codebook: Merge_Add
                    storage['Schl_Dic'] = create_keydic(storage['Schl'],vlist,master=self)
                    self.clean_up_all()
                    prog_pos = 'm_add'
                    self.ask()
                    

            elif prog_pos == 'm_add':
                addvar = self.store_var('Merge_Add')
                if len(addvar[1])>0:
                    self.clean_up_all()
                    prog_pos = 'm_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')

            elif prog_pos == 'm_out':
                self.store_var_all()
                outfile = make_fname(storage['Merge_Out'])
                data = storage['Haupt']
                schldic = storage['Schl_Dic']
                vorder = storage['H_Var']
                keyvars = storage['Merge_Key1'][1]
                vadd = storage['Merge_Add'][1]

                out_data = merge_data(data,keyvars,schldic,vadd,vretain=vorder,master=self)
                    
                t = write_dataset(out_data,outfile)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

            elif prog_pos == 'm_dir':
                fname = self.store_var('Merge_Dir')
                
                verbout('\nFolder: '+fname,master=self)

                flist = get_directory(fname,'txt')
                if len(flist)>0:
                    self.clean_up_all()
                    add_varlist('Merge_Files',flist)
                    prog_pos = 'm_files'
                    self.ask()
                else:
                    verbout('\nERROR: No textfiles in this folder','warning',master=self)


            elif prog_pos == 'm_files':
                fl = self.store_var('Merge_Files')[1]
                if len(fl)==0:
                    self.message('Invalid-Selection01')
                else:
                    combi = merge_files(fl,master=self)
                    
                    storage['Data'] = combi[0]
                    storage['Vlist'] = combi[1]
                    settings['Datasets']['Combined Table'] = {}
                    settings['Datasets']['Combined Table']['Data'] = 'Data'
                    settings['Datasets']['Combined Table']['Var'] = 'Vlist'

                    self.clean_up_all()
                    prog_pos = 'mdir_out'
                    self.ask()

            elif prog_pos == 'mdir_out':
                self.store_var_all()
                
                outfile = make_fname(storage['Elong_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\n\nWriting table',master=self)
                
                t = write_data(storage['Data'],storage['Vlist'],outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

            elif prog_pos == 'elong_first':
                fname = make_fname(self.store_var('Elong_Input1'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading First Table: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'D_First','V_First','First Table')
                if not dset == 0:                 
                    self.clean_up_all()
                    prog_pos = 'elong_second'
                    self.ask()
                

            elif prog_pos == 'elong_second':
                fname = make_fname(self.store_var('Elong_Input2'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading Second Table: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'D_Second','V_Second','Second Table')
                if not dset == 0:                      
                    self.clean_up_all()
                    v_haupt = storage['V_First']
                    v_sec = storage['V_Second']
                    vdic = {}
                    for v in (v_haupt+v_sec):
                        vdic[v] = 0
                    for v in v_haupt:
                        vdic[v] = vdic[v] + 1
                    for v in v_sec:
                        vdic[v] = vdic[v] + 2

                    comm = 0
                    tab1 = 0
                    tab2 = 0

                    for v in sorted(vdic.keys()):
                        codebook['Elong_Vars'][3].append(v)
                        if vdic[v] == 1:
                            codebook['Elong_Vars'][2].append(v+' (only in table 1)')
                            tab1 = tab1 + 1
                        elif vdic[v] == 2:
                            codebook['Elong_Vars'][2].append(v+' (only in table 2)')
                            tab2 = tab2 + 1
                        elif vdic[v] == 3:
                            codebook['Elong_Vars'][2].append(v+' (in both tables)')
                            comm = comm+1

                    verbout('\nFound '+str(len(vdic.keys()))+' variables.\n'+
                                 str(comm)+' found in both tables\n'+
                                 str(tab1)+ ' found only in table 1\n'+
                                 str(tab2)+ ' found only in table 2.\n\n',master=self)

                    prog_pos = 'elong_vars'
                    self.ask()
               
            elif prog_pos == 'elong_vars':
                addvar = self.store_var('Elong_Vars')
                if len(addvar[1])>0:
                    self.clean_up_all()
                    prog_pos = 'elong_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')                

            elif prog_pos == 'elong_out':
                self.store_var_all()
                d1 = storage['D_First']
                d2 = storage['D_Second']
                v1 = storage['V_First']
                v2 = storage['V_Second']
                v_list = storage['Elong_Vars'][1]
                
                outfile = make_fname(storage['Elong_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                out_data = merge_elong(d1,d2,v1,v2,v_list,master=self)

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

######################################################## Aggregate
               

            elif prog_pos == 'agg_input':          
                fname = make_fname(self.store_var('Agg_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Agg_Group',v,excludes=[])
                    add_varlist('Agg_Var',v,excludes=[])
                
                    self.clean_up_all()
                    prog_pos = 'agg_key'
                    self.ask()


            elif prog_pos == 'agg_key':
                self.store_var_all()
                key = storage['Agg_Group'][1]
                if len(key) > 0:
                    data = storage['Data']
                    self.clean_up_all()
                    grpdic = {}
                    for i in range(len(data[key[0]])):
                        agg_key = []
                        for v in key:
                            agg_key.append(data[v][i])
                        agg_key = str(agg_key)
                        grpdic[agg_key]=1
                    verbout('\n\nGroup Variables: ',master=self)
                    verbout(str(key),'table',master=self)
                    verbout('\nAggregating to '+str(len(grpdic.keys()))+'Groups:\n',master=self)
                    verbout(str(sorted(grpdic.keys())[:20]),'table',master=self)
                    if len(grpdic.keys())>20:
                        verbout(' (only first 20 values)',master=self)
                    verbout('\n',master=self)
                    prog_pos = 'agg_var'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'agg_var':
                self.store_var_all()
                l = self.store_var('Agg_Var')
                if len(l) > 0:
                    self.clean_up_all()
                    if storage['Agg_Method'][1] in ['wsum','wmean','wall']:
                        prog_pos = 'agg_weigh'
                    else:
                        prog_pos = 'agg_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')

            elif prog_pos == 'agg_weigh':
                wv = self.store_var('Agg_Weight')[1][0]
                weight = storage['Data'][wv]
                storage['Weight'] = []
                hit = 0
                miss = 0
                for e in weight:
                    try:
                        storage['Weight'].append(float(e))
                        hit = hit + 1
                    except:
                        storage['Weight'].append(0.0)
                        miss = miss + 1
                verbout('\nValid weighting factor found in '+str(hit)+' cases. '+str(miss)+' cases had non-numeric values. These cases are ignored (weight set to 0.0)',master=self)
                self.clean_up_all()
                prog_pos = 'agg_out'
                self.ask()


            elif prog_pos == 'agg_out':
                self.store_var_all()
                data = storage['Data']
                keyvar = storage['Agg_Group'][1]
                aggvar = storage['Agg_Var'][1]
                method = storage['Agg_Method'][1]
                if 'Weight' in storage.keys():
                    weight = storage['Weight']
                else:
                    weight = 0
                    
                outfile = make_fname(storage['Agg_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                
                out_table = aggregate(data,keyvar,aggvar,method,weight,master=self)

                t = write_dataset(out_table,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

######################################################## Subsets


            elif prog_pos == 'sub_input':                
                fname = make_fname(self.store_var('Sub_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Sub_Group',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'sub_group'
                    self.ask()
                

            elif prog_pos == 'sub_group':
                varlist = self.store_var('Sub_Group')[1]
                if len(varlist) > 0:
                    data = storage['Data']
                    valdic = {}
                    vallist = []

                    for i in range(len(data[varlist[0]])):
                        val = ''
                        for v in varlist:
                            if len(val) > 1:
                                val = val + '; '
                            val = val + v + '='+data[v][i]            
                        if not val in valdic.keys():
                            valdic[val] = []
                        valdic[val].append(i)

                    storage['Valdic'] = valdic
                    vallist = sorted(valdic.keys())
                    codebook['Sub_Val'][2]=vallist
                    codebook['Sub_Val'][3]=vallist

                    verbout('\nFound '+str(len(vallist))+' values for the variables ', master=self)
                    verbout(str(varlist),'table', master=self)
                    verbout('\n', master=self)
                    verbout(str(vallist[:20]),'table', master=self)
                    if len(vallist)>20:
                        verbout(' (trunctuated list)', master=self)
                    verbout('\n', master=self)

                    self.clean_up_all()
                    if storage['Sub_Method'] == 'select':
                        prog_pos = 'sub_values'
                    else:
                        prog_pos = 'sub_split'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')

            elif prog_pos == 'sub_values':
                sv = self.store_var('Sub_Val')
                if len(sv) > 0:
                    self.clean_up_all()
                    prog_pos = 'sub_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'sub_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                valdic = storage['Valdic']
                vallist = storage['Sub_Val'][1]
                outfile = make_fname(storage['Sub_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\n\nCalculating Subset.',master=self)
                
                outdic = {}
                for v in varlist:
                    outdic[v] = []

                for val in vallist:
                    for i in valdic[val]:
                        for v in varlist:
                            outdic[v].append(data[v][i])

                t = write_data(outdic,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


            elif prog_pos == 'sub_split':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                selvars = storage['Sub_Group'][1]
                
                outfile = make_fname(storage['Sub_Split'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nCreating Subsets.',master=self)

                outfile_ns = outfile[:-4]
                outfile_ext = outfile[-4:]

                verbout('\nPreparing datasets..',master=self)

                outf_dic = {}
                for i in range(len(data[varlist[0]])):
                    val = ''
                    for v in selvars:
                        val = val + data[v][i] + '_'
                    val = val[:-1]
                    outf_dic[val] = {'Fname':outfile_ns+'_'+val+outfile_ext,'Data':{}}
                    for v in varlist:
                        outf_dic[val]['Data'][v] = []

                for o in sorted(outf_dic.keys()):
                    verbout('\n'+outf_dic[o]['Fname'],master=self)

                verbout('\n\nFilling datasets...',master=self)
                for i in range(len(data[varlist[0]])):
                    val = ''
                    for v in selvars:
                        val = val + data[v][i] + '_'
                    val = val[:-1]
                    for v in varlist:
                        outf_dic[val]['Data'][v].append(data[v][i])


                verbout('\n\nWriting datasets...',master=self)
                for o in sorted(outf_dic.keys()):
                    t = write_data(outf_dic[o]['Data'],varlist,outf_dic[o]['Fname'],header,sep)
                    verbout(t[0],master=self)
                    if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()


            elif prog_pos == 'subr_input':                
                fname = make_fname(self.store_var('Sub_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Sub_Group',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'subr_number'
                    self.ask()

            elif prog_pos == 'subr_number':
                n = self.store_var('Sub_Random_Number')
                
                length = len(storage['Data'][storage['D_Var'][0]])
                
                try:
                    number = int(n)
                    if number > length:
                        number = length
                        verbout('\n',master=self)
                        verbout('Warning: The number you entered is higher that the total count of cases in your table. You will get your complete table in randomized order as output','warning',master=self)
                    storage['Sub_Random_Number'] = number
                    
                    population = range(0,length)
                    rsamp = random.sample(population,number)
                    verbout('\nSelected cases: \n',master=self)
                    verbout(str(rsamp)+'\n','table',master=self)

                    storage['Sub_Sample']=rsamp
                    self.clean_up_all()
                    prog_pos = 'subr_out'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')


            elif prog_pos == 'subr_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                rsamp = storage['Sub_Sample']
                outfile = make_fname(storage['Sub_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\n\nPreparing random subset for output.',master=self)
                
                outdic = {}
                for v in varlist:
                    outdic[v] = []

                for i in rsamp:
                    for v in varlist:
                        outdic[v].append(data[v][i])
                
                t = write_data(outdic,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

            elif prog_pos == 'sv_input':                
                fname = make_fname(self.store_var('Sub_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Sub_Vars',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'sv_vars'
                    self.ask()

            elif prog_pos == 'sv_vars':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'sv_out'
                self.ask()

            elif prog_pos == 'sv_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['Sub_Vars'][1]
                outfile = make_fname(storage['Sub_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nWriting Subset.',master=self)
                
                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Variable calculation


            elif prog_pos == 'calc_input':                
                fname = make_fname(self.store_var('Calc_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname,master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Calc_Vars',v,excludes=[])
            
                    self.clean_up_all()
                    prog_pos = 'calc_vars'
                    self.ask()
               

            elif prog_pos == 'calc_vars':
                varlist = self.store_var('Calc_Vars')[1]
                self.store_var_all()              
                if len(varlist) > 0:
                    verbout('\nSelected Variables:\n',master=self)
                    verbout(str(varlist)+'\n','table',master=self)
                    self.clean_up_all()
                    prog_pos = 'calc_new'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')

            elif prog_pos == 'calc_new':
                nv = self.store_var('Calc_New',store=0)
                if len(nv) > 0:
                    vlab = nv
                    nr = 1
                    while nv in storage['D_Var']:
                        nv = vlab + "{0:02}".format(nr)
                        nr = nr + 1            
                    storage['Calc_New'] = nv
                    if nr > 1:
                        verbout('\n\n',master=self)
                        verbout('Attention. The name "'+vlab+'" was already used for a variable. The new variable is called: '+nv,'warning',master=self)
    
                    self.clean_up_all()
                    prog_pos = 'calc_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection06')

            elif prog_pos == 'calc_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                cvar = storage['Calc_Vars'][1]
                method = storage['Calc_Method'][1]
                nvar = storage['Calc_New']
                outfile = make_fname(storage['Calc_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\n\nCalculating new variable.',master=self)

                vectors = []
                for v in cvar:
                    vectors.append(data[v])

                data[nvar] = calculate_vectors(vectors,method,master=self)
                varlist.append(nvar)

                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()
    

######################################################## Calculate Entropy

            elif prog_pos == 'entropy_input':                
                fname = make_fname(self.store_var('Entropy_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Entropy_Group',v,excludes=[])
                    add_varlist('Entropy_Var',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'entropy_vars'
                    self.ask()
                
            elif prog_pos == 'entropy_vars':
                self.store_var_all()              
                self.clean_up_all()
                prog_pos = 'entropy_vars2'
                self.ask()
                
            elif prog_pos == 'entropy_vars2':
                self.store_var_all()              
                self.clean_up_all()
                prog_pos = 'entropy_out'
                self.ask()

            elif prog_pos == 'entropy_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                group = storage['Entropy_Group'][1]
                mvar= storage['Entropy_Var'][1]
                mode = storage['Entropy_Mode'][1]
                outfile = make_fname(storage['Entropy_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                
                out_table = agg_entropy(data,group,mvar,mode,master=self)

                t = write_dataset(out_table,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Reshape

            elif prog_pos == 'resh_input':                
                fname = make_fname(self.store_var('Resh_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Resh_Vars',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'resh_vars'
                    self.ask()
                

            elif prog_pos == 'resh_vars':
                self.store_var_all()              
                self.clean_up_all()
                prog_pos = 'resh_out'
                self.ask()

            elif prog_pos == 'resh_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                rvars = storage['Resh_Vars'][1]
                rtype = storage['Resh_Type'][1]
                
                outfile = make_fname(storage['Resh_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                out_data = dummy_reshape(data,varlist,rvars,rtype,master=self)

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Transform Timestamps

            elif prog_pos == 'ts_input':                
                fname = make_fname(self.store_var('TS_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    d = dset[0]
                    codebook['TS_Var'][2] = []
                    codebook['TS_Var'][3] = []
                
                    for var in v:
                        element = ''
                        i = 0
                        while element=='':
                            element=d[var][i]
                            i = i + 1
                            if i == len(d[var]): element = 'NO ELEMENTS'
                        verbout('\nVariable: "'+var+'" First Value: "'+element+'"',master=self)
                        codebook['TS_Var'][2].append(var)
                        codebook['TS_Var'][3].append(var)                        
                    self.clean_up_all()
                    prog_pos = 'ts_vars'
                    self.ask()
                

            elif prog_pos == 'ts_vars':
                self.store_var_all()
                a = self.store_var('TS_Nvar')
                b = self.store_var('TS_Var')[1]
                if len(a)>0:
                    self.clean_up_all()
                    prog_pos = 'ts_format'
                    storage['TStamps']=storage['Data'][b]
                    self.ask()
                else:
                    self.message('Invalid-Selection06')

            elif prog_pos == 'ts_format':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'ts_out'
                if storage['TS_Outformat'] in ['ex','pyn']:
                    prog_pos = 'ts_integer' ##Option for float output
                self.ask()

            elif prog_pos == 'ts_integer':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'ts_out'
                self.ask()


            elif prog_pos == 'ts_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                nvar = storage['TS_Nvar']
                ts = storage['TStamps']
                informat = storage['TS_Informat']
                outformat = storage['TS_Outformat']

                if type(informat) == tuple:
                    informat = informat[1]
                if type(outformat) == tuple:
                    outformat = outformat[1]
                    
                if 'TS_Int' in storage.keys():
                    numformat = storage['TS_Int'][1]
                else:
                    numformat = 'dec'
                outfile = make_fname(storage['TS_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\n\nTransforming Timestamps.',master=self)
                
                data[nvar] = []
                for e in ts:
                    trans = tts(e,informat,outformat,numformat)
                    data[nvar].append(trans)
                varlist.append(nvar)

                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Transform variable to groups

            elif prog_pos == 'group_input':                
                fname = make_fname(self.store_var('Group_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Group_Var',v,excludes=[])
                                      
                    self.clean_up_all()
                    prog_pos = 'group_vars'
                    self.ask()
               

            elif prog_pos == 'group_vars':
                self.store_var_all()
                a = self.store_var('Group_NVar')
                b = self.store_var('Group_Var')[1]
                if len(a)>0:
                    self.clean_up_all()
                    prog_pos = 'group_mode'
                    self.ask()
                else:
                    self.message('Invalid-Selection06')

            elif prog_pos == 'group_mode':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'group_out'
                self.ask()

            elif prog_pos == 'group_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                gvar = storage['Group_Var'][1]
                nvar = storage['Group_NVar']
                mode = storage['Group_Mode'][1]
                if mode == 'equal':
                    param = storage['Group_Equal'][1]
                elif mode == 'fixed':
                    param = storage['Group_Fixed']
                elif mode == 'tails':
                    param = storage['Group_Tails'][1]
                    
                outfile = make_fname(storage['Group_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                
                data[nvar] = group_variable(data[gvar],mode,param,master=self)
                varlist.append(nvar)

                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()
              

######################################################## Sort table

            elif prog_pos == 'sort_input':                
                fname = make_fname(self.store_var('Sort_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Sort_Var',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'sort_vars'
                    self.ask()
               

            elif prog_pos == 'sort_vars':
                self.store_var_all()
                a = storage['Sort_Var'][1]
                if len(a)>0:
                    self.clean_up_all()
                    prog_pos = 'sort_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'sort_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                svar = list(storage['Sort_Var'][1])
                mode = storage['Sort_Dir'][1]
                if mode == '2':
                    mode = True
                else:
                    mode = False
                
                outfile = make_fname(storage['Sort_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
           
                out_data = sort_dataset([data,varlist],svar,mode,master=self)
                
                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

######################################################## Rearrange Dataset

            elif prog_pos == 'rear_input':                
                fname = make_fname(self.store_var('Rear_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                def_val['Rear_Output'] = fname

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Rear_Var',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'rear_vars'
                    self.ask()
               

            elif prog_pos == 'rear_vars':
                self.store_var_all()
                a = storage['Rear_Var'][1]
                if len(a)>0:
                    self.clean_up_all()
                    prog_pos = 'rear_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'rear_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['Rear_Var'][1]                    
                outfile = make_fname(storage['Rear_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nWriting Dataset.',master=self)
                
                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                    self.clean_up_all() ## In this special case, missing and duplicate variables are acceptable
                    prog_pos = 'otherart'
                    self.ask()
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Rename


            elif prog_pos == 'ren_input':                
                fname = make_fname(self.store_var('Ren_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                def_val['Ren_Output'] = fname

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Ren_Var',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'ren_vars'
                    self.ask()
               

            elif prog_pos == 'ren_vars':
                self.store_var_all()
                a = storage['Ren_Var'][1]
                b = storage['Ren_New']

                if b in storage['D_Var']:
                    verbout('\n\n',master=self)
                    verbout('ERROR: Variable name "'+b+'" already taken\n','warning',master=self)
                elif len(b) < 1:
                    verbout('\n\n',master=self)
                    verbout('ERROR: No name entered\n','warning',master=self)
                else:
                    vidx = -1
                    for i in range(0,len(storage['D_Var'])):
                        if storage['D_Var'][i]==a:
                            vidx = i
                    if vidx > -1:
                        storage['D_Var'][vidx] = b
                        storage['Data'][b] = storage['Data'][a]
                        add_varlist('Ren_Var',storage['D_Var'],excludes=[])
                        verbout('\nRenamed variable "'+a+'" to "'+b+'"',master=self)
                        self.clean_up_all()
                        prog_pos = 'ren_other'
                        self.ask()
                        

            elif prog_pos == 'ren_other':
                if overspill == '1':
                    prog_pos = 'ren_vars'
                    self.clean_up_all()
                else:
                    prog_pos = 'ren_out'
                    self.clean_up_all()
                self.ask()
                

            elif prog_pos == 'ren_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']                  
                outfile = make_fname(storage['Ren_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nWriting Dataset.',master=self)
                
                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Peak Detection


            elif prog_pos == 'peak_input':                
                fname = make_fname(self.store_var('Peak_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Peak_Time',v,excludes=[])
                    add_varlist('Peak_Series',v,excludes=[])
                               
                    self.clean_up_all()
                    prog_pos = 'peak_vars'
                    self.ask()
               

            elif prog_pos == 'peak_vars':
                if overspill == '1':
                    xv = self.store_var('Peak_Time',store=0)[1]
                    yv = self.store_var('Peak_Series',store=0)[1]
                    x = storage['Data'][xv]
                    y = storage['Data'][yv]
                    self.display_line_plot({'X':x,'Y':y,'Title':'Time Series Preview'})
                    
                else:
                    self.store_var_all()
                    self.clean_up_all()
                    prog_pos = 'peak_method'
                    self.ask()

            elif prog_pos == 'peak_method':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'peak_out'
                self.ask()

            elif prog_pos == 'peak_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                timevar = storage['Peak_Time'][1]
                svar = storage['Peak_Series'][1]
                nvar = storage['Peak_Var']
                pdir = storage['Peak_Direction'][1]
                pthres = storage['Peak_Threshold'][1]
                outfile = make_fname(storage['Peak_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                data[nvar] = find_peaks(data[svar],data[timevar],pdir,pthres,master=self)
                varlist.append(nvar)             

                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Flatten Curve


            elif prog_pos == 'flat_input':                
                fname = make_fname(self.store_var('Flat_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Flat_Time',v,excludes=[])
                    add_varlist('Flat_Series',v,excludes=[])
                                     
                    self.clean_up_all()
                    prog_pos = 'flat_invars'
                    self.ask()
                
            elif prog_pos == 'flat_invars':
                if overspill == '1':
                    xv = self.store_var('Flat_Time',store=0)[1]
                    yv = self.store_var('Flat_Series',store=0)[1]
                    x = storage['Data'][xv]
                    y = storage['Data'][yv]
                    self.display_line_plot({'X':x,'Y':y,'Title':'Time Series Preview'})
                    
                else:
                    self.store_var_all()
                    self.clean_up_all()
                    prog_pos = 'flat_outvars'
                    self.ask()
                

            elif prog_pos == 'flat_outvars':
                self.store_var_all()
                try:
                    a = float(storage['Flat_Window'])
                    self.clean_up_all()
                    prog_pos = 'flat_out'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')

            elif prog_pos == 'flat_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                timevar = storage['Flat_Time'][1]
                svar = storage['Flat_Series'][1]
                window = float(storage['Flat_Window'])
                nvar1 = storage['Flat_Var']
                nvar2 = storage['Flat_Peaks']
                outfile = make_fname(storage['Flat_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                nv = flatten_curve(data[timevar],data[svar],window,master=self)
                data[nvar1]=nv[0]
                data[nvar2]=nv[1]
                
                varlist.append(nvar1)
                varlist.append(nvar2)   

                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Gliding Window


            elif prog_pos == 'glide_input':                
                fname = make_fname(self.store_var('Glide_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Glide_Time',v,excludes=[])
                
                    self.clean_up_all()
                    prog_pos = 'glide_units'
                    self.ask()

            elif prog_pos == 'glide_units':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'glide_var'
                self.ask()
                
            elif prog_pos == 'glide_var':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'glide_out'
                self.ask()

            elif prog_pos == 'glide_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                gvar = storage['Glide_Time'][1]
                nvar = storage['Glide_Var']
                units = int(storage['Glide_Units'][1])
                pos = storage['Glide_Position'][1]
                outfile = make_fname(storage['Glide_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                out_data = create_window([data,varlist],gvar,nvar,units,pos,master=self)        

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Gap Detection
                    
            elif prog_pos == 'gaps_input':                
                fname = make_fname(self.store_var('Gap_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Gap_Tvar',v,excludes=[])
                    add_varlist('Gap_Gvar',v,excludes=[],retain=1)
                
                    self.clean_up_all()
                    prog_pos = 'gaps_var'
                    self.ask()


            elif prog_pos == 'gaps_var':
                self.store_var_all()
                try:
                    storage['Gap_Length'] = float(storage['Gap_Length'])
                    self.clean_up_all()
                    prog_pos = 'gaps_opt'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')
                
            elif prog_pos == 'gaps_opt':
                self.store_var_all()
                self.clean_up_all()

                if storage['Gap_Store'] in storage['D_Var']:
                    verbout('\n',master=self)
                    verbout('Warning: The variable "'+storage['Gap_Store']+'" already exists in the dataset. It will be overwritten!¨','warning',master=self)
                    verbout('\n',master=self)
                prog_pos = 'gaps_out'
                self.ask()


            elif prog_pos == 'gaps_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                gvar = storage['Gap_Gvar'][1]
                tvar = storage['Gap_Tvar'][1]
                nvar = storage['Gap_Store']
                sorting = int(storage['Gap_Sort'][1])
                length = int(storage['Gap_Length'])

                outfile = make_fname(storage['Gap_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                out_data = detect_gaps([data,varlist],tvar,nvar,length,gvar,sorting,master=self)        

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()


######################################################## Normalize Time series


##        elif prog_pos == 'nts_input':
##            self.question_file('NTS_Input',1)
##            self.question_dd('In_Header',2)
##            self.question_dd('In_Sep',3)
##
##        elif prog_pos == 'nts_var':
##            self.question_dd('NTS_Tvar',1)
##            self.question_dd('NTS_Gvar',2)
##
##        elif prog_pos == 'nts_dur':
##            self.question_txt('NTS_Duration',1)
##
##        elif prog_pos == 'nts_vars':
##            self.question_ladd('NTS_Vars',1)
##            self.question_dd('NTS_Method',3)
##
##        elif prog_pos == 'gaps_out':
##            self.question_file('NTS_Out',1,'save')
##            self.question_dd('Out_Header',2)
##            self.question_dd('Out_Sep',3)
            
                  
            elif prog_pos == 'nts_input':                
                fname = make_fname(self.store_var('NTS_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('NTS_Tvar',v,excludes=[])
                    add_varlist('NTS_Gvar',v,excludes=[],retain=1)
                    add_varlist('NTS_Vars',v,excludes=[])
                                    
                    self.clean_up_all()
                    prog_pos = 'nts_var'
                    self.ask()


            elif prog_pos == 'nts_var':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'nts_dur'

                ts = transform_float(storage['Data'][storage['NTS_Tvar'][1]])
                storage['Data'][storage['NTS_Tvar'][1]] = ts

                if storage['NTS_Gvar'][1] == 'res_nogroup':
                    grp = ['1']*len(ts)
                    storage['Data']['res_nogroup']=grp
                    storage['D_Var'].append('res_nogroup')
                else:
                    grp = storage['Data'][storage['NTS_Gvar'][1]]

                storage['Data'] = sort_table(storage['Data'],[storage['NTS_Tvar'][1],storage['NTS_Gvar'][1]])

                sequences = {}
                for i in range(len(ts)):
                    if not grp[i] in sequences.keys():
                        sequences[grp[i]] = []
                    sequences[grp[i]].append(ts[i])

                verbout('\n\nSummary of time series distributions:',master=self)

                for s in sequences.keys():
                    if len(sequences[s])>2:
                        ds = stat_desc(sequences[s])
                        verbout('\n\nGroup: "'+str(s)+'":\n-----------------',master=self)
                        verbout('\nM={0:.3f}; SD={1:.3f}; N={2}; Range={3:.3f}'.format(ds['M'],ds['SD'],ds['N_Total'],ds['Range']),'table',master=self)
                        steps = []
                        for i in range(1,len(ds['Val'])):
                            steps.append(ds['Val'][i]-ds['Val'][i-1])
                        ds2 = stat_desc(steps)
                        verbout('\nMean step={0:.3f}; Smallest step: {1:.3f}; Largest step: {2:.3f}'.format(ds2['M'],ds2['Min'],ds2['Max']),'table',master=self)
                    else:
                        verbout('\n\nGroup: "'+str(s)+'":\n-----------------\nToo small to do statistics (N='+str(len(sequences[s]))+')',master=self)
                    
                ## Compute and display session lengths, min, max and mean time stamp gaps
                self.ask()


            elif prog_pos == 'nts_dur':
                self.store_var_all()
                try:
                    storage['NTS_Duration'] = float(storage['NTS_Duration'])
                    self.clean_up_all()
                    prog_pos = 'nts_vars'
                    self.ask()
                except Exception as f:
                    print(f)
                    self.message('Invalid-Selection07')
            

            elif prog_pos == 'nts_vars':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'nts_out'
                self.ask()


            elif prog_pos == 'nts_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                gvar = storage['NTS_Gvar'][1]
                tvar = storage['NTS_Tvar'][1]
                addvars = storage['NTS_Vars'][1]
                method = storage['NTS_Method'][1]
                relts = storage['NTS_Zero'][1]
                length = int(storage['NTS_Duration'])

                outfile = make_fname(storage['NTS_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                out_data = normalize_ts([data,varlist],tvar,gvar,length,addvars,method,relts,master=self)        

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()



######################################################## Pattern detection

## padt_input is deprecated. Always use mpatd
                
            elif prog_pos == 'mpatd_input':
                fname = make_fname(self.store_var('PD_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('PD_Series_mult',v,excludes=[])
                                
                    self.clean_up_all()
                    prog_pos = 'mpatd_pattern'
                    self.ask()
                
            elif prog_pos == 'mpatd_pattern':                
                fname = make_fname(self.store_var('PD_Patterns'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                
                v = get_varnames(fname,header,sep)
                d = get_data(fname,header,sep)[0]

                storage['Patterns'] = {}
                for var in v:
                    storage['Patterns'][var] = []
                    for e in d[var]:
                        try:
                            storage['Patterns'][var].append(float(e))
                        except:
                            verb('Not a valid character: '+e)
                    if storage['Patterns'][var] == []:
                        del storage['Patterns'][var]

                if len(storage['Patterns'].keys()) > 0:
                    verbout('\n\nPatterns loaded:\n',master=self)
                    verbout(baum_schreiben(storage['Patterns']),'table',master=self)
                    for var in storage['Patterns'].keys():
                        codebook['PD_Pattern_mult'][2].append(var)
                        codebook['PD_Pattern_mult'][3].append(storage['Patterns'][var])                        
                    self.clean_up_all()
                    prog_pos = 'mpatd_var'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')
                    verbout('No Valid file for patterns. No patterns found',master=self)

                
            elif prog_pos == 'mpatd_var':
                self.store_var_all()
                self.clean_up_all()
                storage['Series']={}
                nr = 0
                for var in storage['PD_Series_mult'][1]:
                    valid = 0
                    missing = 0
                    storage['Series'][nr] = []
                    for e in storage['Data'][var]: 
                        try:
                            storage['Series'][nr].append(float(e))
                            valid = valid + 1
                        except:
                            storage['Series'][nr].append('')
                            missing = missing + 1
                    verbout('\n\nSeries "'+var+'" loaded.\n'+str(valid)+' valid cases, '+str(missing)+' missings',master=self)
                    verbout('\nNumbers in series ranging from '+str(min(storage['Series'][nr]))+' to '+str(max(storage['Series'][nr]))+'',master=self)
                    nr = nr +1
                prog_pos = 'mpatd_var2'
                self.ask()

            elif prog_pos == 'mpatd_var2':
                self.store_var_all()
                storage['Pat']={}
                nr = 0
                for pat in storage['PD_Pattern_mult'][1]:
                    storage['Pat'][nr] = pat
                    nr = nr + 1
                prog_pos = 'mpatd_len'

                if len(storage['Series'].keys())==len(storage['Pat'].keys()):
                    verbout('\n\nPatterns selected.',master=self)
                    for i in range(len(storage['Pat'].keys())):
                        var = storage['PD_Series_mult'][1][i]
                        pat = storage['Pat'][i]
                        verbout('\nScanning series "'+var+'" for pattern: ',master=self)
                        verbout(str(pat),'table',master=self) 
                    self.clean_up_all()
                    self.ask()
                else:
                    verbout('\n\nERROR: Number of patterns and number of sequences do not match',master=self)
                
            elif prog_pos == 'mpatd_len':
                minlen = self.store_var('PD_Minlen')
                maxlen = self.store_var('PD_Maxlen')
                accept = 0
                try:
                    minlen = int(minlen)
                    maxlen = int(maxlen)
                    accept = 1
                except:
                    self.message('Invalid-Selection07')

                if accept == 1:
                    pdresult = mpdetection(storage['Series'],storage['Pat'],minlen,maxlen,master=self)
                    descriptives = stat_desc(pdresult[0])
                    verbout('\n\nPattern detection complete.\nMean correlation: '+"{0:0.3}".format(float(descriptives['M'])),master=self)
                    verbout('\nSD: '+"{0:0.3}".format(float(descriptives['SD'])),master=self)
                    verbout('\nMin: '+"{0:0.3}".format(float(descriptives['Min'])),master=self)
                    verbout('\nMax: '+"{0:0.3}".format(float(descriptives['Max'])),master=self)
                    storage['Pdresult']=pdresult
                    self.clean_up_all()
                    prog_pos = 'patd_opt'
                    self.ask()
                

            elif prog_pos == 'patd_opt':
                self.store_var_all()
                accept = 0
                if 'PD_Cutoff' in storage.keys():
                    try:
                        storage['PD_Cutoff'] = float(storage['PD_Cutoff'])
                        accept = 1
                    except:
                        self.message('Invalid-Selection07')
                else:
                    accept = 1

                if accept ==1:
                    self.clean_up_all()
                    prog_pos = 'patd_out'
                    self.ask()
                
            elif prog_pos == 'patd_out':
                self.store_var_all()
                corrlist = storage['Pdresult'][0]
                altlist = storage['Pdresult'][1]
                alt2list = storage['Pdresult'][2]
                lenlist = storage['Pdresult'][3]
                data = storage['Data']
                varlist = storage['D_Var']
                if 'PD_Cutoff' in storage.keys():
                    cutoff = storage['PD_Cutoff']
                outfile = make_fname(storage['PD_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\n\nGenerating output...',master=self)

                label = storage['PD_Var']

                if storage['PD_Method'][1] == '5':
                    vlab1 = label+'_Corr'
                    vlab2 = label+'_AMean'
                    vlab3 = label+'_GMean'
                    vlab4 = label+'_Length'
                    vnum = 0
                    while vlab1 in varlist or vlab2 in varlist or vlab3 in varlist:
                        vnum = vnum + 1
                        vlab1 = label+'_Corr' + "{0:02}".format(vnum)
                        vlab2 = label+'_AMean' + "{0:02}".format(vnum)
                        vlab3 = label+'_GMean' + "{0:02}".format(vnum)
                        vlab4 = label+'_Length' + "{0:02}".format(vnum)
                    varlist = varlist + [vlab1,vlab2,vlab3,vlab4]    
                else:
                    vlab = label
                    vnum = 0
                    while vlab in varlist:
                        vnum = vnum + 1
                        vlab = label + "{0:02}".format(vnum)
                    varlist.append(vlab)

                if storage['PD_Method'][1] == '1':
                    data[vlab] = corrlist
                elif storage['PD_Method'][1] == '5':
                    data[vlab1] = corrlist
                    data[vlab2] = altlist
                    data[vlab3] = alt2list
                    data[vlab4] = lenlist
                else:
                    dicholist = []
                    for i in range(len(corrlist)):  
                        if corrlist[i] > cutoff:
                            dicholist.append(1)
                        else:
                            dicholist.append(0)

                    if storage['PD_Method'][1] == '2':
                        data[vlab] = dicholist
                    elif storage['PD_Method'][1] == '3':
                        comblist = []
                        for i in range(len(corrlist)):
                            try:
                                comblist.append(dicholist[i]*altlist[i])
                            except:
                                comblist.append('')
                        data[vlab] = comblist
                    elif storage['PD_Method'][1] == '4':
                        comblist = []
                        for i in range(len(corrlist)):
                            try:
                                comblist.append(dicholist[i]*alt2list[i])
                            except:
                                comblist.append('')
                        data[vlab] = comblist                        
                
                t = write_data(data,varlist,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()



######################################################## Synchronize Event Data

##        elif prog_pos == 'syn_input':
##            self.question_file('Syn_Input',1)
##            self.question_dd('In_Header',2)
##            self.question_dd('In_Sep',3)
##
##        elif prog_pos == 'syn_var':
##            self.question_ladd('Syn_Var','Syn_Var')
##
##        elif prog_pos == 'syn_opt':
##            self.question_txt('Syn_Frame',1)
##            self.question_dd('Syn_Meas',2)
##            self.question_dd('Syn_Cur',3)
##
##        elif prog_pos == 'syn_output':
##            self.question_file('Syn_Out',1,'save')
##            self.question_dd('Out_Header',2)
##            self.question_dd('Out_Sep',3)
                    


            elif prog_pos == 'syn_input':             
                fname = make_fname(self.store_var('Syn_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    dvars = []
                    for v in dset[1]:
                        if cd_proof(v,master=self): dvars.append(v)                       
    
                    add_varlist('Syn_Var',dvars,excludes=[])
                    add_varlist('Syn_TS',dset[1],retain=1)

                if len(dvars) > 0:                                     
                    self.clean_up_all()
                    prog_pos = 'syn_var'
                    self.ask()

                else:
                    verbout('\n\nERROR: The table does not contain any dummy variables','warning',master=self)
                    

            elif prog_pos == 'syn_var':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'syn_opt'
                self.ask()          

            elif prog_pos == 'syn_opt':
                self.store_var_all()

                try:
                    storage['Syn_Frame'] = int(storage['Syn_Frame'])
                    self.clean_up_all()
                    prog_pos = 'syn_output'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')

            elif prog_pos == 'syn_output':
                self.store_var_all()
                cvars = storage['Syn_Var'][1]
                frame = storage['Syn_Frame']
                meas = storage['Syn_Meas'][1]
                tsvar = storage['Syn_TS'][1]

                outfile = make_fname(storage['Syn_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                out_data = synch_events(storage['Dummy_Data'],cvars,frame,meas,master=self)

                if not tsvar == 'notime':
                    ts_vec = storage['Data'][tsvar]
                    diff = data_dim(out_data)[1]-len(ts_vec)
                    ts_vec = ts_vec + diff*['']
                    out_data[0][tsvar] = ts_vec
                    out_data[1] = [tsvar] + out_data[1]

                #write_dataset(out_data,outfile,header,sep)

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()
                


######################################################## Sequence Analysis


            elif prog_pos == 'seq_input':                
                fname = make_fname(self.store_var('Seq_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Seq_Svar',v,excludes=[])
                    add_varlist('Seq_Tvar',v,excludes=[])
                    add_varlist('Seq_Gvar',v,excludes=[])
                    
                    self.clean_up_all()
                    prog_pos = 'seq_vars'
                    self.ask()

            elif prog_pos == 'seq_vars':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'seq_length'
                self.ask()
                
            elif prog_pos == 'seq_length':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'seq_out'
                self.ask()

            elif prog_pos == 'seq_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                svar = storage['Seq_Svar'][1]
                tvar = storage['Seq_Tvar'][1]
                gvar = storage['Seq_Gvar'][1]
                slen = storage['Seq_Length'][1]
                somit = storage['Seq_Omit'][1]
                mode = storage['Seq_Mode'][1]
                
                outfile = make_fname(storage['Seq_Output'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nFinding Sequences. Please be patient.',master=self)

                out_data = find_sequence(data,svar,tvar,gvar,slen,somit,mode,master=self)

                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

######################################################## T-Pattern


            elif prog_pos == 'tpat_input':                
                fname = make_fname(self.store_var('Tpat_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Tpat_Vars',v,excludes=[])
                    add_varlist('Tpat_Time',v,excludes=[])
                    add_varlist('Tpat_Group',v,excludes=[])
                
                    self.clean_up_all()
                    prog_pos = 'tpat_vars'
                    self.ask()

            elif prog_pos == 'tpat_vars':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'tpat_options'
                if storage['Tpat_Time'][1] == 'notime':
                    ##No time variable. Create new one.
                    storage['Data']['notime'] = []
                    for i in range(len(storage['Data'][storage['D_Var'][0]])):
                        storage['Data']['notime'].append(i)
                    timeser_string = storage['Data']['notime']
                timeser_string = storage['Data'][storage['Tpat_Time'][1]]
                timeser = []
                for t in timeser_string:
                    try:
                        timeser.append(float(t))
                    except:
                        verbout('\nWarning: Timestamp "'+t+'" ist not a number',master=self)
                timeser = sorted(timeser)
                mindist = timeser[1]-timeser[0]+1
                for i in range(1,len(timeser)):
                    if timeser[i]-timeser[i-1] < mindist and timeser[i]-timeser[i-1] > 0:
                        mindist = timeser[i]-timeser[i-1]

                if mindist < 1:
                    factor = 1/mindist
                    if factor > 1000:
                        tf = factor
                    elif factor > 100:
                        tf = 1000
                    elif factor > 10:
                        tf = 100
                    else:
                        tf = 10
                    verbout('\n\nWarning: Your timeseries is scaled too narrowly (distances < 1). It is rescaled by a factor of '+str(tf),master=master)
                else:
                    tf = 1

                timeser_new = []

                for t in timeser_string:
                    try:
                        timeser_new.append(float(t)*tf)
                    except:
                        timeser_new.append('')

                storage['Data'][storage['Tpat_Time'][1]] = timeser_new
                self.ask()


            elif prog_pos == 'tpat_options':
                accept = 1
                self.store_var_all()
                try:
                    storage['Tpat_Level'] = float(self.store_var('Tpat_Level',store=0))
                except:
                    accept = 0
                    self.message('Invalid-Selection07')
                if accept == 1:
                    self.clean_up_all()
                    if storage['Tpat_Time'][1] == 'notime':
                        storage['Data']['notime']=range(0,len(storage['Data'][storage['D_Var'][0]]))
                    if storage['Tpat_Group'][1] == 'nogroup':
                        storage['Data']['nogroup'] = []
                        for i in range(0,len(storage['Data'][storage['D_Var'][0]])):
                            storage['Data']['nogroup'].append('Group')
                    prog_pos = 'tpat_options'
                    prog_pos = 'tpat_out'
                    self.ask()

            elif prog_pos == 'tpat_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                varlist = storage['Tpat_Vars'][1]
                tvar = storage['Tpat_Time'][1]
                gvar = storage['Tpat_Group'][1]
                p_level = storage['Tpat_Level']
                outfile = make_fname(storage['Tpat_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                verbout('\nFinding Patterns. Please be patient.',master=self)

                tpattern = find_tpats(data,tvar,gvar,varlist,p_level,master=self)

                out_data = {'T_Pattern':[], 'Elements':[], 'Total_Count':[]}
                outvars = ['T_Pattern', 'Elements', 'Total_Count']
                groups = []

                for g in get_unique(data[gvar]):
                    out_data[g] = []
                    groups.append(g)
                for p in sorted(tpattern.keys()):
                    out_data['T_Pattern'].append(str(tpattern[p]['Tree']))
                    out_data['Elements'].append(str(tpattern[p]['List']))
                    out_data['Total_Count'].append(0)
                    for g in groups:
                        out_data[g].append(0)
                    for s in tpattern[p]['Start']:
                        out_data[s[1]][-1] = out_data[s[1]][-1]+1
                        out_data['Total_Count'][-1] = out_data['Total_Count'][-1]+1

                t = write_data(out_data,outvars+groups,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()

######################################################## Grammar induction


            elif prog_pos == 'gind_input':                
                fname = make_fname(self.store_var('Gind_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Gind_Symbol',v,excludes=[])
                    add_varlist('Gind_Group',v,excludes=[])
                    add_varlist('Gind_Time',v,excludes=[])

                    self.clean_up_all()
                    prog_pos = 'gind_vars'
                    self.ask()

            elif prog_pos == 'gind_vars':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'gind_opt'
                self.ask()

            elif prog_pos == 'gind_opt':
                self.store_var_all()
                try:
                    storage['Gind_Glitch'] = int(storage['Gind_Glitch'])
                    self.clean_up_all()
                    if storage['Gind_Rep'][1] == '0':
                        prog_pos = 'gind_out'
                    else:
                        prog_pos = 'gind_repet'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')


            elif prog_pos == 'gind_repet':
                self.store_var_all()
                try:
                    storage['Gind_Rep_Min'] = int(storage['Gind_Rep_Min'])
                    storage['Gind_Rep_Max'] = int(storage['Gind_Rep_Max'])
                    self.clean_up_all()
                    prog_pos = 'gind_len'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')


            elif prog_pos == 'gind_len':
                self.store_var_all()
                try:
                    storage['Gind_Len_Min'] = int(storage['Gind_Len_Min'])
                    storage['Gind_Len_Max'] = int(storage['Gind_Len_Max'])
                    storage['Gind_Eta'] = float(storage['Gind_Eta'])
                    self.clean_up_all()
                    if storage['Gind_Len_Min'] > storage['Gind_Len_Max']:
                        storage['Gind_Len_Max'] = storage['Gind_Len_Max'] + 1
                        verbout('\nMaximal length was below minimal length. Was corrected to: '+str(storage['Gind_Rep_Max']),master=self)
                    if storage['Gind_Eta'] > 0.9 or storage['Gind_Eta'] < 0:
                        storage['Gind_Eta'] = 0.9
                        verbout('\nEta was out of boundaries. Corrected to 0.9 (default value)',master=self)
                                                     
                    prog_pos = 'gind_out'
                    self.ask()
                except:
                    self.message('Invalid-Selection07')
            

            elif prog_pos == 'gind_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']
                symb = storage['Gind_Symbol'][1]
                gvar = storage['Gind_Group'][1]
                tvar = storage['Gind_Time'][1]
                mini = storage['Gind_Glitch']
                rep = storage['Gind_Rep'][1]
                subst = int(storage['Gind_Subs'][1])
                if rep == '1':
                    minrep = storage['Gind_Rep_Min']
                    maxrep = storage['Gind_Rep_Max']
                else:
                    minrep=1
                    maxrep=0
                rep = (minrep,maxrep)

                len1 = storage['Gind_Len_Min']
                len2 = storage['Gind_Len_Max']
                eta = storage['Gind_Eta']
                
                outfile = make_fname(storage['Gind_Out'])

                if tvar == 'notime': tvar = ''
                if gvar == 'nogroup': gvar = ''

                output_string = calc_adios(data,symb,gvar,tvar,mini,rep,(len1,len2),eta,subst,master=self)

                outf = open(outfile,'w')
                outf.write(output_string)
                outf.close()
                verbout('\n\nFile '+outfile+' successfully created.\n',master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()

                
######################################################## Cluster analysis HECANE


            elif prog_pos == 'cluster_input':             
                fname = make_fname(self.store_var('Cluster_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Cluster_Vars',v,excludes=[])
                
                    self.clean_up_all()
                    prog_pos = 'cluster_vars'
                    self.ask()

            elif prog_pos == 'cluster_vars':
                varlist = self.store_var('Cluster_Vars')[1]
                self.store_var_all()              
                if len(varlist) > 0:
                    verbout('\nSelected Variables (first 50):\n',master=self)
                    verbout(str(varlist[:50]),'table',master=self)
                    self.clean_up_all()
                    prog_pos = 'cluster_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')

            elif prog_pos == 'cluster_out':
                self.store_var_all()
                data = storage['Data']
                dvar = storage['D_Var']
                varlist = storage['Cluster_Vars'][1]
                cvar = storage['Cluster_Vars'][1]
                std = storage['Cluster_Std']
                outfile = make_fname(storage['Cluster_Output'])
                add_outputs = storage['Cluster_Add']
##                specials = []
##                for a in ['ssa','dendro','hist','dist','vector']:
##                    specials.append(add[a])

                verbout('\n\nCalculating Cluster solution.\n',master=self)
                
                if std['row']==1:
                    rs = 1
                else:
                    rs = 0
                if std['table']==1:
                    ts = 1
                else:
                    ts = 0

                find_cluster([data,dvar],varlist,outfile,add_outputs,rs,ts,master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()

######################################################## Cluster analysis Kmeans


            elif prog_pos == 'kcluster_input':             
                fname = make_fname(self.store_var('Kcluster_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Kcluster_Group',v,excludes=[])
                    add_varlist('Kcluster_Vars',v,excludes=[])
                    
                    self.clean_up_all()
                    prog_pos = 'kcluster_opt'
                    self.ask()

            elif prog_pos == 'kcluster_opt':
                self.store_var_all()
                if storage['Methode']=='An4':
##                    v = storage['D_Var']
##                    add_varlist('Kcluster_Vars',v,excludes=[storage['Kcluster_Group'][1]])
                    storage['Kcluster_Group'] = storage['Kcluster_Group'][1]
                else:
                    storage['Kcluster_Group'] = 0

                accept = 1
                numrange = storage['Kcluster_Anz']
                if numrange == '':
                    self.message('Invalid-Selection06')
                    accept = 0
                elif '-' in numrange:
                    nl = numrange.split('-')
                    try:
                        r = [int(nl[0]),int(nl[1])]
                    except:
                        self.message('Invalid-Selection07')
                        accept = 0
                else:
                    try:
                        r = [int(numrange),int(numrange)]
                    except:
                        self.message('Invalid-Selection07')
                        accept = 0
                if accept == 1:
                    if r[0] < 2:
                        r[0] = 2
                    storage['Kcluster_Anz'] = r
                    verbout('\nNumber of clusters to be extracted: Range '+str(r),master=self)
                    self.clean_up_all()
                    prog_pos = 'kcluster_vars'
                    self.ask()
                            

            elif prog_pos == 'kcluster_vars':
                varlist = self.store_var('Kcluster_Vars')[1]
                self.store_var_all()              
                if len(varlist) > 0:
                    verbout('\nSelected Variables (first 50):\n',master=self)
                    verbout(str(varlist[:50]),'table',master=self)
                    self.clean_up_all()

                    create_cluster_table(storage['Data'], varlist, storage['Kcluster_Stand'][1], storage['Kcluster_Group'],master=self)
                    
                    prog_pos = 'kcluster_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'kcluster_out':
                self.store_var_all()
                data = storage['CData']
                dvar = storage['CVars']
                varlist = storage['Kcluster_Vars'][1]
                direction = storage['Kcluster_Direction'][1]
                outfile = make_fname(storage['Kcluster_Out'])
                numrange = storage['Kcluster_Anz']
                gv = storage['Kcluster_Group']
                
                verbout('\n\nCalculating Cluster solution.\n',master=self)

                icenters = {}
                nsolution = {}
                for num in range(numrange[0],numrange[1]+1):
                    #verbout('\nk-means: Vars: '+str(varlist)+'; Direction: '+str(direction)+'; Anz: '+str(num),'table',master=self)
                    solution = kmeans(data,varlist,direction,num,master=self)
                    nsolution[num] = solution
                    icenters[num] = {}
                    for c in solution.keys():
                        icenters[num][c] = solution[c]['Center']

                if storage['Methode']=='An4':
                    msolution = multi_kmeans(data, dvar, varlist, direction, range(numrange[0],numrange[1]+1), 0, icenters, gv,master=self)

                    ###Assign memberships

                    is_member = {}
                    v_label = {}
                    for i in range(len(storage['Data'][varlist[0]])):
                        is_member[i] = {}
                        for num in range(numrange[0],numrange[1]+1):
                            is_member[i][num] = {'#Pooled':-100,'#Group':-100}
                            v_label[num] = {}
                            v_label[num]['#Pooled'] = 'Cluster_n'+str(num)+'_pooled'
                            v_label[num]['#Group'] = 'Cluster_n'+str(num)+'_group'

                    for num in msolution[1]:
                        for g in msolution[1][num].keys():
                            for c in range(num):
                                for entity in msolution[1][num][g][c]:
                                    if g == '#Pooled':
                                        is_member[entity][num]['#Pooled'] = c
                                    else:
                                        is_member[entity][num]['#Group'] = c

                    outdata = storage['Data']
                    outvars = storage['D_Var']
                    for num in sorted(v_label.keys()):
                        outvars.append(v_label[num]['#Pooled'])
                        outvars.append(v_label[num]['#Group'])
                        outdata[v_label[num]['#Pooled']] = []
                        outdata[v_label[num]['#Group']] = []

                    for i in range(len(outdata[outvars[0]])):
                        for num in v_label.keys():
                            outdata[v_label[num]['#Pooled']].append(is_member[i][num]['#Pooled']+1)
                            outdata[v_label[num]['#Group']].append(is_member[i][num]['#Group']+1)
                            
                    
                    t = write_data(outdata,outvars,outfile)
                    verbout(t[0],master=self)
                    if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                    verbout('\n\n\n------------------------------\nREPORT:\n------------------------------\n',master=self)
                    verbout(msolution[0],'table',master=self)

                    verbout('\n\nContingency analysis of pooled and groupwise cluster assignment:\nGxx: Group Clusters\nPxx: Pooled Clusters',master=self)
                    for num in sorted(v_label.keys()):
                        verbout('\n\n'+str(num)+'-Cluster solution:',master=self)
                        crosstab = {}
                        cmargin = {}
                        rmargin = {}
                        nmargin = {}
                        ntotal = 0
                        for n in range(num):
                            r = 'G'+"{0:02}".format(n+1)
                            crosstab[r] = {}
                            rmargin[r] = 0
                            nmargin[r] = 0
                            for n in range(num):
                                c = 'P'+"{0:02}".format(n+1)
                                crosstab[r][c] = 0
                                cmargin[c] = 0


                        hits = 0
                        miss = 0                        
                        for i in  range(len(outdata[outvars[0]])):
                            p = 'P'+ "{0:02}".format(outdata[v_label[num]['#Pooled']][i])
                            g = 'G'+ "{0:02}".format(outdata[v_label[num]['#Group']][i])
                            try:
                                crosstab[g][p] = crosstab[g][p] + 1
                                rmargin[g] = rmargin[g] + 1
                                cmargin[p] = cmargin[p] + 1
                                if outdata[v_label[num]['#Pooled']][i] == outdata[v_label[num]['#Group']][i]:
                                    hits = hits + 1
                                else:
                                    miss = miss + 1
                            except:
                                pass
                        verbout('\n'+display_table(crosstab),'table',master=self)
                        if hits + miss > 0:
                            pa = float(hits)/(hits+miss)
                            pc = 0.0
                            ntotal = 2*(hits + miss)
                            for n in range(num):
                                rm = rmargin['G'+ "{0:02}".format(n+1)]
                                cm = cmargin['P'+ "{0:02}".format(n+1)]
                                pc = pc + (float(rm+cm)/ntotal)**2
                            if not pc == 1:
                                kappa = (pa-pc)/(1-pc)
                            else:
                                kappa = 1
                        else:
                            pa = '-'
                            kappa = '-'
                        verbout('\nAgreement: '+"{0:.2f}".format(pa*100)+'%',master=self)
                        verbout('\nCohens Kappa: '+"{0:.3f}".format(kappa),master=self)
                        verbout('\n',master=self)
                                            
                    
                else:
                    is_member = {}
                    v_label = {}
                    for i in range(len(storage['Data'][varlist[0]])):
                        is_member[i] = {}
                        for num in range(numrange[0],numrange[1]+1):
                            is_member[i][num] = -100
                            v_label[num] = 'Cluster_n'+str(num)

                    for num in nsolution.keys():
                        for c in range(num):
                            for entity in nsolution[num][c]['Members']:
                                is_member[entity][num] = c
                    
                    outdata = storage['Data']
                    outvars = storage['D_Var']
                    for num in sorted(v_label.keys()):
                        outvars.append(v_label[num])
                        outdata[v_label[num]] = []

                    for i in range(len(outdata[outvars[0]])):
                        for num in v_label.keys():
                            outdata[v_label[num]].append(is_member[i][num]+1)

                    t = write_data(outdata,outvars,outfile)
                    verbout(t[0],master=self)
                    if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)
                    
                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()


######################################################## Analysis of Entropy
                

            elif prog_pos == 'anent_input':             
                fname = make_fname(self.store_var('Anent_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Anent_Multi',v,excludes=[])
                    add_varlist('Anent_Group',v,excludes=[])
                
                    self.clean_up_all()
                    prog_pos = 'anent_vars'
                    self.ask()


            elif prog_pos == 'anent_vars':
                varlist = self.store_var('Anent_Multi')[1]
                self.store_var_all()
                if len(varlist) > 0:
                    verbout('\n\nSelected Variables:\n',master=self)
                    verbout(str(varlist),'table',master=self)
                    verbout('\n',master=self)
                    self.clean_up_all()
                    prog_pos = 'anent_out'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'anent_out':
                self.store_var_all()
                data = storage['Data']
                dvar = storage['D_Var']
                varlist = storage['Anent_Multi'][1]
                group = storage['Anent_Group'][1]
                opt = storage['Anent_Option']
                
                outfile = make_fname(storage['Anent_Output'])

                report = 'Analysis of Entropy.\n--------------\n\nInput File: '+storage['Anent_Input']
                report = report + '\nList of Variables: '+str(varlist)
                report = report + '\nGrouping Variables: '+str(group)
                report = report + '\nOptions: '+str(opt)
                report = report + '\n-----------------------------------------------------------\n'

                verbout('\nCalculating Entropies.\n',master=self)
                
                report = analyze_entropy(data,group,varlist,opt,report,master=self)

                verbout('\n\nWriting results to: '+outfile,master=self)
                repfile = open(outfile,'w')
                repfile.write(report)
                repfile.close()

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()

######################################################## Analysis Focus and Attention


            elif prog_pos == 'focus_input':             
                fname = make_fname(self.store_var('Focus_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Focus_Date',v,excludes=[],retain=1)
                    add_varlist('Focus_Ntext',v,excludes=[],retain=1)
                    add_varlist('Focus_Weight',v,excludes=[],retain=1)
                    add_varlist('Focus_Issue',v,excludes=[])
                    add_varlist('Focus_Actor',v,excludes=[])
                    
                
                    self.clean_up_all()
                    prog_pos = 'focus_date'
                    self.ask()


            elif prog_pos == 'focus_date':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'focus_text'
                self.ask()          

            elif prog_pos == 'focus_text':
                self.store_var_all()
                self.clean_up_all()

                storage['Data']['res_Textcount'] = []
                storage['Data']['res_Weighting'] = []
                if not 'res_Textcount' in storage['D_Var']:
                    storage['D_Var'].append('res_Textcount')
                if not 'res_Weighting' in storage['D_Var']:
                    storage['D_Var'].append('res_Weighting')

                valid = 0
                invalid = 0
                
                if storage['Focus_Ntext'][1] == 'count_em':
                    for e in storage['Data'][storage['D_Var'][0]]:
                        storage['Data']['res_Textcount'].append(1)
                        valid = valid+1
                else:
                    for e in storage['Data'][storage['Focus_Ntext'][1]]:
                        try:
                            storage['Data']['res_Textcount'].append(int(e))
                            valid = valid + 1
                        except:
                            storage['Data']['res_Textcount'].append('')
                            invalid = invalid + 1

                if storage['Focus_Weight'][1] == 'no_weight':
                    for e in storage['Data'][storage['D_Var'][0]]:
                        storage['Data']['res_Weighting'].append(1)
                else:
                    now = 0
                    for e in storage['Data'][storage['Focus_Weight'][1]]:
                        try:
                            storage['Data']['res_Weighting'].append(float(e))
                        except:
                            storage['Data']['res_Weighting'].append(0.0)
                            now = now + 1
                            valid = valid-1
                    verbout('\nWeighting applied. There were '+str(now)+' invalid values in the weighting variable.',master=self)
                       

                verbout('\n\nCounted texts.\n'+str(valid)+' data rows contain valid number of texts.',master=self)
                if invalid > 0:
                    verbout('\n',master=self)
                    verbout('There were '+str(invalid)+' invalid counts in the variable',master=self)

                                        
                prog_pos = 'focus_issue'
                self.ask()          


            elif prog_pos == 'focus_issue':
                varlist = self.store_var('Focus_Issue')[1]
                if len(varlist) > 1:
                    self.clean_up_all()
                    prog_pos = 'focus_actor'
                    self.ask()   
                else:
                    self.message('Invalid-Selection01')
                

            elif prog_pos == 'focus_actor':
                varlist = self.store_var('Focus_Actor')[1]
                if len(varlist) > 1:
                    self.clean_up_all()
                    prog_pos = 'focus_window'
                    self.ask()   
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'focus_window':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'focus_out'
                self.ask()          


            elif prog_pos == 'focus_out':
                self.store_var_all()
                data = storage['Data']
                varlist = storage['D_Var']

                issvar = storage['Focus_Issue'][1]
                actvar = storage['Focus_Actor'][1]

                windir = storage['Focus_Direction'][1]
                winlen = int(storage['Focus_Window'][1])

                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])                
                
                outfile = make_fname(storage['Focus_Out'])
                
                verbout('\n\nCalculating time series.',master=self)
                
                out_data = focus_timeseries(data,issvar,actvar,windir,winlen,master=self)
                
                t = write_dataset(out_data,outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                ##Automated detection of peaks:

                tsdata = {0:outdata[0]['Volume'],
                          1:outdata[0]['Issue_Focus'],
                          2:outdata[0]['Actor_Focus']}

                eventpat = {0:[0.0, 0.0, 0.7, 1.0, 1.0, 1.0, 0.7, 0.0, 0.0],
                            1: [1.0, 1.0, 0.3, 0.0, 0.0, 0.0, 0.3, 1.0, 1.0],
                            2: [1.0, 1.0, 0.3, 0.0, 0.0, 0.0, 0.3, 1.0, 1.0]}

                mediapat = {0:[0.0, 0.0, 0.7, 1.0, 1.0, 1.0, 0.7, 0.0, 0.0],
                            1: [1.0, 1.0, 0.3, 0.0, 0.0, 0.0, 0.3, 1.0, 1.0],
                            2: [0.8, 0.9, 1.0, 0.8, 0.6, 0.4, 0.2, 0.0, 0.0]}

                debatpat = {0:[0.0, 0.0, 0.7, 1.0, 1.0, 1.0, 0.7, 0.0, 0.0],
                            1:[0.8, 0.9, 1.0, 0.8, 0.6, 0.4, 0.2, 0.0, 0.0],
                            2: [1.0, 1.0, 0.3, 0.0, 0.0, 0.0, 0.3, 1.0, 1.0]}

                expectpat = {0:[0.0, 0.0, 0.7, 1.0, 1.0, 1.0, 0.7, 0.0, 0.0],
                            1:[1.0, 0.0, 0.3, 0.5, 0.7, 0.9, 1.0, 1.0, 1.0],
                            2:[1.0, 1.0, 0.3, 0.0, 0.0, 0.0, 0.3, 1.0, 1.0]}

                verbout('\n\nScanning for surprising events',master=self)
                pdresult_event = mpdetection(tsdata,eventpat,7,25,master=self)

                verbout('\n\nScanning for expected events',master=self)
                pdresult_expect = mpdetection(tsdata,expectpat,7,25,master=self)

                verbout('\n\nScanning for mediatized/staged events',master=self)
                pdresult_media = mpdetection(tsdata,mediapat,7,25,master=self)

                verbout('\n\nScanning for political debates',master=self)
                pdresult_debate = mpdetection(tsdata,debatpat,7,25,master=self)

                verbout('\n\nScanning for other attention peaks',master=self)
                pdresult_other = mpdetection({0:outdata[0]['Volume']},{0:[0.0, 0.0, 0.7, 1.0, 1.0, 1.0, 0.7, 0.0, 0.0]},7,25,master=self)

                #0: Correlation
                #1: Artithm Mean
                #2: Geom. Mean
                #3: Length

                outdata[0]['Media_Hype'] = []
                outdata[1].append('Media_Hype')
                for i in range(len(outdata[0]['Date'])):
                    outdata[0]['Media_Hype'].append('')

                i = 0
                while i < len(outdata[0]['Date']):
                    highest_cor = max(pdresult_event[0][i],pdresult_media[0][i],pdresult_debate[0][i],pdresult_expect[0][i])
                    if highest_cor > .7:
                        if pdresult_event[0][i] > max(pdresult_media[0][i],pdresult_debate[0][i],pdresult_expect[0][i]):
                            verbout('\nGenuine Event found at: '+outdata[0]['Date'][i],master=self)
                            verbout('\n -Correlation: '+str(pdresult_event[0][i]),master=self)
                            verbout(' / Amplitude: '+str(pdresult_event[2][i]),master=self)
                            verbout(' / Duration: '+str(pdresult_event[3][i]),master=self)
                            verbout('\n',master=self)
                            outdata[0]['Media_Hype'][i] = 'Genuine Event (Cor='+str(pdresult_event[0][i])+'/ Amp='+str(pdresult_event[2][i])+' / Dur='+str(pdresult_event[3][i])+')'
                            i = i + int(pdresult_event[3][i]/3)
                            
                        elif pdresult_debate[0][i] > max(pdresult_media[0][i],pdresult_event[0][i],pdresult_expect[0][i]):
                            verbout('\nPolitical Debate found at: '+outdata[0]['Date'][i],master=self)
                            verbout('\n -Correlation: '+str(pdresult_debate[0][i]),master=self)
                            verbout(' / Amplitude: '+str(pdresult_debate[2][i]),master=self)
                            verbout(' / Duration: '+str(pdresult_debate[3][i]),master=self)
                            verbout('\n',master=self)
                            outdata[0]['Media_Hype'][i] = 'Political Debate (Cor='+str(pdresult_debate[0][i])+'/ Amp='+str(pdresult_debate[2][i])+' / Dur='+str(pdresult_debate[3][i])+')'
                            i = i + int(pdresult_debate[3][i]/3)
                            
                        elif pdresult_media[0][i] > max(pdresult_debate[0][i],pdresult_event[0][i],pdresult_expect[0][i]):
                            verbout('\nStaged Event found at: '+outdata[0]['Date'][i],master=self)
                            verbout('\n -Correlation: '+str(pdresult_media[0][i]),master=self)
                            verbout(' / Amplitude: '+str(pdresult_media[2][i]),master=self)
                            verbout(' / Duration: '+str(pdresult_media[3][i]),master=self)
                            verbout('\n',master=self)
                            outdata[0]['Media_Hype'][i] = 'Staged Event (Cor='+str(pdresult_media[0][i])+'/ Amp='+str(pdresult_media[2][i])+' / Dur='+str(pdresult_media[3][i])+')'
                            i = i + int(pdresult_media[3][i]/3)
                            
                        elif pdresult_expect[0][i] > max(pdresult_debate[0][i],pdresult_event[0][i],pdresult_media[0][i]):
                            verbout('\nExpected Event found at: '+outdata[0]['Date'][i],master=self)
                            verbout('\n -Correlation: '+str(pdresult_expect[0][i]),master=self)
                            verbout(' / Amplitude: '+str(pdresult_expect[2][i]),master=self)
                            verbout(' / Duration: '+str(pdresult_expect[3][i]),master=self)
                            verbout('\n',master=self)
                            outdata[0]['Media_Hype'][i] = 'nExpected Event (Cor='+str(pdresult_expect[0][i])+'/ Amp='+str(pdresult_expect[2][i])+' / Dur='+str(pdresult_expect[3][i])+')'
                            i = i + int(pdresult_expect[3][i]/3)
                            
##                    elif pdresult_other[0][i] > .8:
##                        verbout('\nUnknown event found at: '+outdata[0]['Date'][i],master=self)
##                        verbout('\nCorrelation: '+str(pdresult_other[0][i]),master=self)
##                        verbout('\nAmplitude: '+str(pdresult_other[2][i]),master=self)
##                        verbout('\nDuration: '+str(pdresult_other[3][i]),master=self)
##                        verbout('\n',master=self)
##                        outdata[0]['Media_Hype'][i] = 'Political Debate (Cor='+str(pdresult_other[0][i])+'/ Amp='+str(pdresult_other[2][i])+' / Dur='+str(pdresult_other[3][i])+')'
                                               

                    i = i + 1

                t = write_data(outdata[0],outdata[1],outfile,header,sep)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()
            




######################################################## Heat Map



            elif prog_pos == 'heat_input':             
                fname = make_fname(self.store_var('Heat_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Heat_Vars',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'heat_vars'
                    self.ask()

            

            elif prog_pos == 'heat_vars':
                varlist = self.store_var('Heat_Vars')[1]
                self.store_var_all()              
                if len(varlist) > 0:
                    verbout('\nSelected Variables (first 50): '+str(varlist[:50]),master=self)
                    d = storage['Data']
                    gridlist = []
                    misval = 0
                    for i in range(len(d[varlist[0]])):
                        zeile = []
                        for v in varlist:
                            try:
                                zeile.append(float(d[v][i]))
                            except:
                                zeile.append(0.0)
                                misval = misval + 1
                        gridlist.append(zeile)

                    verbout('\n\nTable loaded.\nNumber of rows: '+str(len(gridlist)),master=self)
                    verbout('\nNumber of columns: '+str(len(gridlist[0])),master=self)
                    if misval > 0:
                        verbout('\nThere were '+str(misval)+' missing values. They were set to 0.0',master=self)

                    if storage['Heat_Sort'][1] == '1':
                        rowsum = []
                        colsum = []
                        for l in gridlist:
                            rowsum.append(sum(l))
                        for i in range(len(gridlist[0])):
                            col = []
                            for l in gridlist:
                                col.append(l[i])
                            colsum.append(sum(col))

                        rown = range(0,len(rowsum))
                        coln = range(0,len(colsum))

                        rsort = list(zip(*sorted(zip(rowsum, rown),reverse=True)))[1]
                        csort = list(zip(*sorted(zip(colsum, coln),reverse=True)))[1]

                        outlist = []
                        for r in rsort:
                            line = []
                            for c in csort:
                                line.append(gridlist[r][c])
                            outlist.append(line)
                        gridlist = list(outlist)

                        nvarlist = []
                        for c in csort:
                            nvarlist.append(varlist[c])

                        verbout('\n\nSorted Variables in dataset. New order: '+str(nvarlist),master=self)

                    storage['Glist'] = gridlist                                         
                    
                    self.clean_up_all()
                    prog_pos = 'heat_display'
                    self.ask()
                else:
                    self.message('Invalid-Selection01')


            elif prog_pos == 'heat_display':
                self.store_var_all()
                gridlist = storage['Glist']
                size_set = storage['Heat_Maxsize'][1]
                if size_set == '1':
                    size = (400,300)
                elif size_set == '2':
                    size = (640,480)
                elif size_set == '3':
                    size = (900,600)
                elif size_set == '4':
                    size = (1200,900)
                elif size_set == '5':
                    size = (1500,1500)
                elif size_set == '6':
                    size = (2000,2000)
                
                mode = storage['Heat_Color'][1]
                vlog = storage['Heat_Log'][1]

                gw = len(gridlist[0])
                gh = len(gridlist)

                if vlog == '1':
                    outlist = []
                    for l in gridlist:
                        outl = []
                        for c in l:
                            try:
                                outl.append(math.log(c))
                            except:
                                outl.append(0)
                        outlist.append(outl)
                    gridlist = list(outlist)

                
                scale = 0
                pxsize = 1
                while pxsize < 2:
                    scale = scale + 1
                    pxx = int(float(size[0]*scale)/gw)
                    pxy = int(float(size[1]*scale)/gh)
                    if pxx > pxy:
                        pxsize = pxy
                    else:
                        pxsize = pxx

                verbout('\nScaling graph by factor '+str(scale),master=self)

                py = pxsize
                px = pxsize
                xscale = scale

                while px*gw/xscale < 140:
                    if xscale > 1:
                        xscale = xscale -1
                    else:
                        px = px + 1
                
                width = int(gw/xscale)
                height = int(gh/scale)
                self.f_bottomline.b_check["state"] = NORMAL

                self.display_heat_map(gridlist,w=width,h=height,ph=py,pw=px,legend=1,mode=mode,verbose=1)               



######################################################## X-Y Plot

            elif prog_pos == 'visu_input':             
                fname = make_fname(self.store_var('Visu_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Visu_X',v,excludes=[])
                    add_varlist('Visu_Y',v,excludes=[])
                        
                    self.clean_up_all()
                    prog_pos = 'visu_plot'
                    self.ask()

            elif prog_pos == 'visu_plot':
                if overspill == 'line':
                    ptype = 'Line'
                elif overspill == 'scat':
                    ptype = 'Scatter'
                
                xv = self.store_var('Visu_X',store=0)[1]
                yv = self.store_var('Visu_Y',store=0)[1]
                x = storage['Data'][xv]
                y = storage['Data'][yv]
                self.display_line_plot({'X':x,'Y':y,'Title':'Visualization: '+yv+' by '+xv,'Type':ptype})
 

######################################################## Reliability Testing



            elif prog_pos == 'rt_input':                
                fname = make_fname(self.store_var('Reltest_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    v = dset[1]
                    add_varlist('Reltest_Unit',v,excludes=[])
                    add_varlist('Reltest_Coder',v,excludes=[])
                    
                    self.clean_up_all()
                    prog_pos = 'rt_var1'
                    self.ask()

            elif prog_pos == 'rt_var1':
                self.store_var_all()
                self.clean_up_all()
                cvar = storage['Reltest_Coder'][1]
                uvar = storage['Reltest_Unit'][1]
                crosstab = {}
                for c in get_unique(storage['Data'][cvar]):
                    crosstab[c] = {}
                    for u in get_unique(storage['Data'][uvar]):
                        crosstab[c][u] = 0
                for i in range(len(storage['Data'][cvar])):
                    c = storage['Data'][cvar][i]
                    u = storage['Data'][uvar][i]
                    crosstab[c][u] = crosstab[c][u] + 1
                for c in crosstab.keys():
                    for u in crosstab[c].keys():
                        if crosstab[c][u] > 1: crosstab[c][u] = '*'+str(crosstab[c][u])+'*'

                #ct = display_table(crosstab)
                verbout('\nUnits coded by each coder:\n\n',master=self)
                verbout(display_table(crosstab),'table',master=self)
                

                for var in storage['D_Var']:
                    if not var in [cvar,uvar]:
                        codebook['Reltest_Var'][2].append(var)
                        codebook['Reltest_Var'][3].append(var)                    

                codebook['Reltest_Units'][2] = []
                codebook['Reltest_Units'][3] = []

                for u in get_unique(storage['Data'][uvar]):
                    codebook['Reltest_Units'][2].append(u)
                    codebook['Reltest_Units'][3].append(u)

                codebook['Reltest_Core'][2] = codebook['Reltest_Core'][2][:1]
                codebook['Reltest_Core'][3] = codebook['Reltest_Core'][3][:1]
                codebook['Reltest_Coders'][2] = []
                codebook['Reltest_Coders'][3] = []

                for c in get_unique(storage['Data'][cvar]):
                    codebook['Reltest_Core'][2].append(c)
                    codebook['Reltest_Core'][3].append(c)
                    codebook['Reltest_Coders'][2].append(c)
                    codebook['Reltest_Coders'][3].append(c) 

                
                prog_pos = 'rt_var2'
                self.ask()

            elif prog_pos == 'rt_var2':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'rt_units'
                self.ask()

            elif prog_pos == 'rt_units':
                self.store_var_all()
                
                storage['Coding_Dic'] = create_coding_dic(storage['Data'],storage['Reltest_Unit'][1],
                                                          storage['Reltest_Coder'][1],storage['Reltest_Var'][1],
                                                          storage['Reltest_Units'][1],master=self)
                self.clean_up_all()
                prog_pos = 'rt_coders'
                self.ask()

            elif prog_pos == 'rt_coders':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'rt_set'
                self.ask()


            elif prog_pos == 'rt_set':
                self.store_var_all()

                options = []
                for o in storage['Reltest_Options'].keys():
                    if storage['Reltest_Options'][o] == 1:
                        options.append(o)
                        
                storage['Reltest_O'] = options


                methods = []
                for m in storage['Reltest_Method'].keys():
                    if storage['Reltest_Method'][m] == 1:
                        methods.append(m)

                if 'PRF' in methods and storage['Reltest_Core'][1] == 'no_core':
                    verbout('\n')
                    verbout('ERROR: You can not caluclate precision/recall without core coder','warning',master=self)
                elif len(methods) > 0:
                    storage['Reltest_M'] = methods
                    self.clean_up_all()
                    prog_pos = 'rt_out'
                    self.ask()
                else:
                    verbout('\nNo Method selected. Please select a method (coefficient) to continue',master=self)

            elif prog_pos == 'rt_out':
                self.store_var_all()
                codings = storage['Coding_Dic']
                outfile1 = make_fname(storage['Reltest_Out'])
                outfile2 = make_fname(storage['Reltest_Report'])

                tvars = storage['Reltest_Var'][1]
                units = storage['Reltest_Units'][1]
                coders = storage['Reltest_Coders'][1]
                kerncod = storage['Reltest_Core'][1]
                methods = storage['Reltest_M']
                opt = storage['Reltest_O']

                if 'rico' in opt:
                    opt.remove('rico')
                    rico = 1
                else:
                    rico = 0

                verbout('\n\nCalculating Interrater Reliability.\n',master=self)

                rt_output = reltest(codings,tvars,units,coders,kerncod,methods,opt,master=self)
                header = 'Reliability test conducted on: '+time.ctime()
                header = header + '\nInput file: '+storage['Reltest_Input']
                header = header + '\nVariables (N='+str(len(tvars))+'): '+str(tvars)
                header = header + '\nUnits of Analysis (N='+str(len(units))+'): '+str(units)
                header = header + '\nCoders (N='+str(len(coders))+'): '+str(coders)
                header = header + '\nCore coder: '+str(kerncod)
                header = header + '\nCoefficients: '+str(methods)
                header = header + '\nOptions: '+str(opt)
                header = header + '\n----------------\n\n'
                text1_out = rt_output[0]
                text2_out = rt_output[1]

                if rico == 1:
                    for r in coders:
                        rcod = []
                        for c in coders:
                            if not c == r: rcod.append(c)
                        rt_output = reltest(codings,tvars,units,rcod,kerncod,methods,opt,master=self)
                        text1_out = text1_out + '\n\nRICO: Omitting coder: '+r+'\n--------------------\n\n'+rt_output[0]
                        text2_out = text2_out + '\n\nRICO2: Omitting coder: '+r+'\n--------------------\n\n'+rt_output[1]
                            
                
                out1 = open(outfile1,'w')
                out1.write(header + text1_out)
                out1.close()
                verbout('\nShort summary of results written to: '+outfile1,master=self)

                out2 = open(outfile2,'w')
                out2.write(header + text2_out)
                out2.close()
                verbout('\n\nDetailed report of reliability test written to: '+outfile1+'\n',master=self)
                
                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()


######################################################## Get UNIVERSE

            elif prog_pos == 'univ_in':
                fname = self.store_var('Univ_Input')
                sd = self.store_var('Univ_Subdir')
                self.store_var('Encoding')
                self.clean_up_all()
                verbout('\nFolder: '+fname,master=self)

                if sd[1]=='1':
                    os.system('dir "'+fname.replace('/','\\')+'\\*.txt" /s /b >textfiles.txt')
                    inlist = open('textfiles.txt','r')
                    inf = inlist.readlines()
                    inlist.close()
                    flist = []
                    for e in inf:
                        flist.append(e[:-1])
                else:
                    os.system('dir "'+fname.replace('/','\\')+'\\*.txt" /b >textfiles.txt')
                    inlist = open('textfiles.txt','r')
                    inf = inlist.readlines()
                    inlist.close()
                    flist = []
                    for e in inf:
                        flist.append(fname.replace('/','\\')+'\\'+e[:-1])

                storage['Files'] = flist
                verbout('\nFound '+str(len(flist))+' Text-Files in this folder. First 10:\n',master=self)
                for f in flist[:10]:
                    verbout(f+'\n','table',master=self)

                prog_pos = 'univ_mode'
                self.ask()

            elif prog_pos == 'univ_mode':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'univ_out'
                self.ask()


            elif prog_pos == 'univ_out':
                self.store_var_all()
                flist = storage['Files']
                lang = storage['Univ_Lang'][1]
                ngram = int(storage['Univ_Length'][1])
                spar = int(storage['Univ_Sparse'][1])
                outfile = make_fname(storage['Univ_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                encod = storage['Encoding'][1]

                sparsity = [spar/100,1-spar/100]
                
                verbout('\nCounting Words in all textfiles. This may take some minutes...\n',master=self)
                
                unidic = get_univ(flist,lang,ngram,sparsity,encod,master=self)
                outdic = {'Ngram':[],'Share':[]}
                for n in sorted(unidic.keys()):
                    outdic['Ngram'].append(n)
                    outdic['Share'].append("{0:.3f}".format(unidic[n]))
                
                verbout('\n\n'+write_data(outdic,['Ngram','Share'],outfile,header,sep)[0],master=self)
 
                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()


            elif prog_pos == 'corpus_in':
                fname = self.store_var('Corpus_Input')
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])
                verbout('\nCodesheet location: '+fname, master=self)

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    add_varlist('Corpus_ID',dset[1],excludes=[])
                    self.clean_up_all()
                    prog_pos = 'corpus_dir'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'corpus_dir':
                self.store_var_all()
                self.clean_up_all()
                if storage['Corpus_Indic'] == 'path provided':
                    storage['Corpus_Indic'] = ''
                prog_pos = 'corpus_nvar'
                self.ask()

            elif prog_pos == 'corpus_nvar':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'corpus_out'
                self.ask()

            elif prog_pos == 'corpus_out':
                self.store_var_all()
                path = storage['Corpus_Indic']
                data = storage['Data']
                vlist = storage['D_Var']
                idvar = storage['Corpus_ID'][1]
                nvar = storage['Corpus_Outvar']
                outfile = make_fname(storage['Corpus_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                encod = storage['Encoding'][1]

                if len(path) > 1:
                    path = path + '/'

                self.verbout('\nAttaching text to dataset. This may take some minutes...\n')

                outdset = create_corpus(data,path,idvar,nvar,encod,master=self)
                vlist = vlist+[nvar]
                outdset[1] = vlist
                t = write_dataset(outdset,outfile,header,sep)
                verbout('\n'+t[0],master=self)
 
                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()               


            elif prog_pos == 'corpust_in':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'corpust_nvar'
                self.ask()

            elif prog_pos == 'corpust_nvar':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'corpust_out'
                self.ask()

            elif prog_pos == 'corpust_out':
                self.store_var_all()
                path = storage['Corpus_Indic']
                idvar = storage['Corpus_OutID']
                nvar = storage['Corpus_Outvar']
                outfile = make_fname(storage['Corpus_Out'])
                subd = storage['Corpus_Subdir'][1]
                encod = storage['Encoding'][1]
                
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                self.verbout('\nCreating corpus from textfiles. This may take some minutes...\n')
                path = path.replace('/','\\')

                if subd == '0':
                    comm = 'dir "'+path+'\\*.txt" /b >temp_dir.txt'
                    fpath = path+'\\'
                else:
                    comm = 'dir "'+path+'\\*.txt" /b /s >temp_dir.txt'
                    fpath = ''
                os.system(comm)

                liste = "not ready"
                while liste == "not ready":
                    try:
                        inf = open("temp_dir.txt","r")
                        liste = inf.readlines()
                        inf.close()
                        os.system("del temp_dir.txt")
                    except:
                        liste = "not ready"

                corpus = {idvar:[]}
                for l in liste:
                    corpus[idvar].append(l[:-1])

                outdset = create_corpus([corpus,[idvar]],fpath,idvar,nvar,encod,master=self)

                t = write_dataset(outdset,outfile,header,sep)
                self.verbout('\n'+t[0])
                
                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()                

##        elif prog_pos == 'insp_in':
##            self.question_file('RE_Input',1)
##            self.question_dd('In_Sep',2)
##            self.question_dd('In_Header',3)
##
##        elif prog_pos == 'insp_var':
##            self.question_dd('RE_Fulltext',2)
##
##        elif prog_pos == 'insp_find':
##            self.question_dd('RE_Expression',2)


            elif prog_pos == 'insp_in':
                fname = self.store_var('RE_Input')
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])
                verbout('\nCorpus location: '+fname, master=self)

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    add_varlist('RE_Fulltext',dset[1],excludes=[])
                    add_varlist('RE_ID',dset[1],excludes=[],retain=1)
                    self.clean_up_all()
                    prog_pos = 'insp_var'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'insp_var':
                self.store_var_all()
                self.clean_up_all()

                storage['Text_Vector'] = storage['Data'][storage['RE_Fulltext'][1]]
                if storage['RE_ID'][1] == 'res_none':
                    storage['ID_Vector'] = list(range(len(storage['Text_Vector'])))
                else:
                    storage['ID_Vector'] = storage['Data'][storage['RE_ID'][1]]
                
                prog_pos = 'insp_find'
                self.ask()

            elif prog_pos == 'insp_find':
                self.store_var_all()
                self.clean_up_all()

                expression = storage['RE_Expression']

                accept = 0
                try:
                    p = re.compile(expression)
                    accept = 1
                except:
                    self.message("Invalid-Selection08")

                if accept == 1:
                    verbout("\n\nSearching for Pattern: "+str([expression]),'title',master=self)
                    verbout("\nNumber of texts: "+str(len(storage['Text_Vector']))+'\n (Capping at 200 matches)\n',master=self)
                    nmatch = 0
                    for i in range(len(storage['Text_Vector'])):
                        con = context(storage['Text_Vector'][i], expression, span = 15, case=storage['RE_Case'][1])

                        for c in con:
                            nmatch+=1
                            if nmatch<200:
                                verbout('\n'+c[0],'text',master=self)
                                verbout(c[1],'warning',master=self)
                                verbout(c[2],'text',master=self)
                                verbout(' ({0})'.format(storage['ID_Vector'][i]),master=self)

                    verbout("\n\n  Found: "+str(nmatch)+' Results.\n',master=self)
                        
                self.ask()



            elif prog_pos == 'svm_in':
                fname = self.store_var('SVM_Input')
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])
                verbout('\nCorpus location: '+fname, master=self)

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    add_varlist('SVM_Textvar',dset[1],excludes=[])
                    add_varlist('SVM_Classvar',dset[1],excludes=[])
                    self.clean_up_all()
                    prog_pos = 'svm_vars'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'svm_vars':
                self.store_var_all()
                self.clean_up_all()
                classvec = []
                for e in storage['Data'][storage['SVM_Classvar'][1]]:
                    if e in [1,'1']:
                        classvec.append(1)
                    else:
                        classvec.append(0)
                verbout('\nClassifications: 1: '+str(sum(classvec))+' / 0: '+str(len(classvec)-sum(classvec)),master=self)
                storage['Classvector'] = classvec
                
                prog_pos = 'svm_opt'
                self.ask()

            elif prog_pos == 'svm_opt':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'svm_out'
                self.ask()


            elif prog_pos == 'svm_out':
                self.store_var_all()
                data = storage['Data']
                vlist = storage['D_Var']
                lang = storage['Univ_Lang'][1]
                ngrams = int(storage['Univ_Length'][1])
                spars = int(storage['SVM_Sparse'][1])
                sparsity = [float(spars)/100,1-float(spars)/100]
                textvar = storage['SVM_Textvar'][1]
                classvec = storage['Classvector']
                outfile = make_fname(storage['SVM_Out'],ext='json')

                tdm = generate_tdm(data[textvar],lang=lang,ngrams=ngrams,sparsity=sparsity,master=self)
                del tdm['res_Document'] ##Remove document name from feature list
                svm_result = train_svm(tdm,classvec,master=self)
                hyperplane = svm_result[0]
                quality = svm_result[1]

                scores = svm_scores(tdm,hyperplane,verbose=1)
                prf = svm_prf(scores,classvec)

                verbout('\n\nTraining completed. Optimal result applied to training corpus, yields this accuracy:',master=self)
                verbout('\nPrecision: '+"{0:.3f}".format(prf[0]),'table',master=self)
                verbout('\nRecall:    '+"{0:.3f}".format(prf[1]),'table',master=self)
                verbout('\nF-Score:   '+"{0:.3f}".format(prf[2]),'table',master=self)
                verbout('\n',master=self)
                
                strain = open(outfile,"w")
                strain.write(str(hyperplane))
                strain.close()
                verbout('\nHyperplane Json written to: '+outfile+'\n',master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()                
   


            elif prog_pos == 'svmt_in':
                fname = self.store_var('SVM_Input')
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])
                verbout('\nCorpus location: '+fname, master=self)

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    add_varlist('SVM_Textvar',dset[1],excludes=[])
                    add_varlist('SVM_Classvar',dset[1],excludes=[])
                    self.clean_up_all()
                    prog_pos = 'svmt_vars'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'svmt_vars':
                self.store_var_all()
                self.clean_up_all()
                classvec = []
                for e in storage['Data'][storage['SVM_Classvar'][1]]:
                    if e in [1,'1']:
                        classvec.append(1)
                    else:
                        classvec.append(0)
                verbout('\nClassifications: 1: '+str(sum(classvec))+' / 0: '+str(len(classvec)-sum(classvec)),master=self)
                storage['Classvector'] = classvec
                
                prog_pos = 'svmt_opt'
                self.ask()

            elif prog_pos == 'svmt_opt':
                self.store_var_all()
                self.clean_up_all()
                prog_pos = 'svmt_model'
                self.ask()

            elif prog_pos == 'svmt_model':
                self.store_var_all()
                strain = open(storage['SVM_Model'],'r')
                sl = strain.readlines()
                strain.close()
                sl = " ".join(sl)
                sl = sl.replace('\n','')
                try:
                    rvec = eval(sl)
                except:
                    rvec = {}

                if len(rvec.keys()) > 1:
                    storage['SVM_Hyperplane'] = rvec
                    verbout('\n\nModel loaded: ',master=self)
                    for k in sorted(rvec.keys()):
                        verbout('\n   '+k+': '+"{0:.4f}".format(rvec[k]),'table',self)
                    verbout('\n',master=self)
                    self.clean_up_all()
                    def_val["SVM_Adjust"]=rvec["Intercept"]

                    verbout("\n\nGenerating Term-Document Matrix and visualizing test results",master=self)

                    storage['TDM']=generate_tdm(storage['Data'][storage['SVM_Textvar'][1]],
                                                lang=storage['Univ_Lang'][1],
                                                sparsity=0,
                                                ngrams=int(storage['Univ_Length'][1]),
                                                universe=list(rvec.keys()),master=self)
                    
                    curve = svm_prf_curve(storage['TDM'], rvec, storage['Classvector']) ##Creates a dataset with PRF values for all possible intercepts.
                    self.display_line_plot({'X':curve[0]['Intercept'],'Y':list(zip(curve[0]['Precision'],curve[0]['Recall'],curve[0]['F_Score'])),
                                            'Title':'Precision (black), Recall (blue) and F (red) in relation to Intercept','Type':'Line'})

                    prog_pos = 'svmt_try'
                    self.ask()
                else:
                    verbout('\nERROR: not a valid SVM model','warning',self)

            elif prog_pos == 'svmt_try':
                self.store_var_all()
                accept = 0
                try:
                    storage['SVM_Adjust'] = float(storage['SVM_Adjust'])
                    accept = 1
                except:
                    self.message("Invalid-Selection07")


                if accept == 1:
                    tdm = storage['TDM']
                    classvec = storage['Classvector']
                    rvec = storage['SVM_Hyperplane']
                    rvec['Intercept']=storage['SVM_Adjust']                    

                    if overspill == "try":
                        scores = svm_scores(tdm, rvec)
                        prf = svm_prf(scores,classvec)

                        verbout('\n\nTrial completed. Accuracy for test corpus with intercept='+str(storage['SVM_Adjust'])+':',master=self)
                        verbout('\nPrecision: '+"{0:.3f}".format(prf[0]),'table',master=self)
                        verbout('\nRecall:    '+"{0:.3f}".format(prf[1]),'table',master=self)
                        verbout('\nF-Score:   '+"{0:.3f}".format(prf[2]),'table',master=self)
                        verbout('\n',master=self)
                    elif overspill == "test":
                        storage['SVM_Hyperplane']['Intercept']=storage['SVM_Adjust']
                        strain = open(storage['SVM_Model'],"w")
                        strain.write(str(storage['SVM_Hyperplane']))
                        strain.close()
                        verbout('\nHyperplane with adjusted intercept written: '+storage['SVM_Model']+'\n',master=self)

                        self.clean_up_all()
                        prog_pos = 'svmt_out'        
                        self.ask()


            elif prog_pos == 'svmt_out':
                self.store_var_all()
                data = storage['Data']
                vlist = storage['D_Var']
                classvec = storage['Classvector']
                newvar = storage['SVM_Newvar']
                rvec = storage['SVM_Hyperplane']

                outfile = make_fname(storage['SVM_Outtable'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                tdm = storage['TDM']

                scores = svm_scores(tdm, rvec)
                prf = svm_prf(scores,classvec)

                verbout('\n\nTest completed. Accuracy for test corpus:',master=self)
                verbout('\nPrecision: '+"{0:.3f}".format(prf[0]),'table',master=self)
                verbout('\nRecall:    '+"{0:.3f}".format(prf[1]),'table',master=self)
                verbout('\nF-Score:   '+"{0:.3f}".format(prf[2]),'table',master=self)
                verbout('\n',master=self)

                if storage['SVM_Type'][1] == '2': #Make Dichotomous
                    dscore = []
                    for s in scores:
                        if s<0:
                            dscore.append(0)
                        else:
                            dscore.append(1)
                    scores = dscore

                data[newvar] = scores
                vlist.append(newvar)
                t = write_data(data,vlist,outfile,header,sep)
                
                verbout('\n'+t[0],master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()                


##        elif prog_pos == 'svma_in1':
##            self.question_file('SVMA_Input',1)
##            self.question_dd('In_Sep',2)
##            self.question_dd('In_Header',3)
##
##        elif prog_pos == 'svma_var':
##            self.question_dd('SVM_Textvar',1)
##
##        elif prog_pos == 'svma_in2':
##            self.question_file('SVMA_Folder',1,'folder')
##            self.question_dd('Encoding',2)
##            self.question_dd('Corpus_Subdir',3)
##
##        elif prog_pos == 'svma_opt':
##            self.question_dd('Univ_Lang',1)
##            self.question_dd('Univ_Length',2)
##
##        elif prog_pos == 'svma_model':
##            self.question_file('SVM_Model',1,defext='.json')
##            self.question_dd('SVM_Type',2)
##            self.question_txt('SVM_Newvar',3)
##
##        elif prog_pos == 'svmt_out':
##            self.question_file('SVM_Outtable',1,'save')
##            self.question_dd('Out_Header',2)
##            self.question_dd('Out_Sep',3)       
##                


            elif prog_pos == 'svma_in1':
                fname = self.store_var('SVMA_Input')
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])
                verbout('\nCorpus location: '+fname, master=self)

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    add_varlist('SVM_Textvar',dset[1],excludes=[])
                    self.clean_up_all()
                    prog_pos = 'svma_var'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'svma_var':
                self.store_var_all()
                self.clean_up_all()                
                prog_pos = 'svma_opt'
                self.ask()

            elif prog_pos == 'svma_in2':
                self.store_var_all()
                self.clean_up_all()       

                path = storage['SVMA_Folder']
                subd = storage['Corpus_Subdir'][1]
                encod = storage['Encoding'][1]

                self.verbout('\nCreating corpus from textfiles. This may take some minutes...\n')
                path = path.replace('/','\\')

                if subd == '0':
                    comm = 'dir "'+path+'\\*.txt" /b >temp_dir.txt'
                    fpath = path+'\\'
                else:
                    comm = 'dir "'+path+'\\*.txt" /b /s >temp_dir.txt'
                    fpath = ''
                os.system(comm)

                liste = "not ready"
                while liste == "not ready":
                    try:
                        inf = open("temp_dir.txt","r")
                        liste = inf.readlines()
                        inf.close()
                        os.system("del temp_dir.txt")
                    except:
                        liste = "not ready"

                idvar = "Text_ID"
                nvar = "Fulltext"

                corpus = {idvar:[]}
                for l in liste:
                    corpus[idvar].append(l[:-1])

                outdset = create_corpus([corpus,[idvar]],fpath,idvar,nvar,encod,master=self)

                storage["Data"] = outdset[0]
                storage["D_Var"] = outdset[1]
                storage["SVM_Textvar"] = (nvar,nvar)
                
                self.clean_up_all()
                prog_pos = 'svma_opt'
                self.ask()

            
            elif prog_pos == 'svma_opt':
                self.store_var_all()
                self.clean_up_all()                
                prog_pos = 'svma_model'
                self.ask()


            elif prog_pos == 'svma_model':
                self.store_var_all()
                strain = open(storage['SVM_Model'],'r')
                sl = strain.readlines()
                strain.close()
                sl = " ".join(sl)
                sl = sl.replace('\n','')
                try:
                    rvec = eval(sl)
                except:
                    rvec = {}

                if len(rvec.keys()) > 1:
                    storage['SVM_Hyperplane'] = rvec
                    verbout('\n\nModel loaded: ',master=self)
                    for k in sorted(rvec.keys()):
                        verbout('\n   '+k+': '+"{0:.4f}".format(rvec[k]),'table',self)
                    verbout('\n',master=self)
                    self.clean_up_all()

                    storage['TDM']=generate_tdm(storage['Data'][storage['SVM_Textvar'][1]],
                                                lang=storage['Univ_Lang'][1],
                                                sparsity=0,
                                                ngrams=int(storage['Univ_Length'][1]),
                                                universe=list(rvec.keys()),master=self)
                    
                    prog_pos = 'svma_out'
                    self.ask()
                else:
                    verbout('\nERROR: not a valid SVM model','warning',self)


            elif prog_pos == 'svma_out':
                self.store_var_all()
                data = storage['Data']
                vlist = storage['D_Var']
                newvar = storage['SVM_Newvar']
                rvec = storage['SVM_Hyperplane']

                outfile = make_fname(storage['SVM_Outtable'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                tdm = storage['TDM']

                scores = svm_scores(tdm, rvec)
                if storage['SVM_Type'][1] == '2': #Make Dichotomous
                    dscore = []
                    for s in scores:
                        if s<0:
                            dscore.append(0)
                        else:
                            dscore.append(1)
                    scores = dscore

                data[newvar] = scores
                vlist.append(newvar)
                t = write_data(data,vlist,outfile,header,sep)
                
                verbout('\n'+t[0],master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()                




################ Duplicate analysis (N-Gram Shingling)


            elif prog_pos == 'dupli_in1':
                fname = self.store_var('NGS_Input')
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])
                verbout('\nCorpus location: '+fname, master=self)

                verbout('\nLoading file: '+fname, master=self)
                dset = self.load_dset(fname,header,sep,'Data','D_Var','Main_Table')
                if not dset == 0:
                    add_varlist('NGS_Fulltext',dset[1],excludes=[])
                    add_varlist('NGS_Tid',dset[1],excludes=[])
                    self.clean_up_all()
                    prog_pos = 'dupli_var'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'dupli_var':
                self.store_var_all()
                self.clean_up_all()                
                prog_pos = 'dupli_opt'
                self.ask()

            elif prog_pos == 'dupli_in2':
                self.store_var_all()
                self.clean_up_all()       

                path = storage['Corpus_Indic']
                subd = storage['Corpus_Subdir'][1]
                encod = storage['Encoding'][1]

                self.verbout('\nCreating corpus from textfiles. This may take some minutes...\n')
                path = path.replace('/','\\')

                if subd == '0':
                    comm = 'dir "'+path+'\\*.txt" /b >temp_dir.txt'
                    fpath = path+'\\'
                else:
                    comm = 'dir "'+path+'\\*.txt" /b /s >temp_dir.txt'
                    fpath = ''
                os.system(comm)

                liste = "not ready"
                while liste == "not ready":
                    try:
                        inf = open("temp_dir.txt","r")
                        liste = inf.readlines()
                        inf.close()
                        os.system("del temp_dir.txt")
                    except:
                        liste = "not ready"

                idvar = "Text_ID"
                nvar = "Fulltext"

                corpus = {idvar:[]}
                for l in liste:
                    corpus[idvar].append(l[:-1])

                outdset = create_corpus([corpus,[idvar]],fpath,idvar,nvar,encod,master=self)

                storage["Data"] = outdset[0]
                storage["D_Var"] = outdset[1]
                storage["NGS_Tid"] = (idvar,idvar)
                storage["NGS_Fulltext"] = (nvar,nvar)
                
                self.clean_up_all()
                prog_pos = 'dupli_opt'
                self.ask()                

            elif prog_pos == 'dupli_opt':
                self.store_var_all()

                if 'NGS_Overlap' in storage.keys():
                    accept = 0
                    try:
                        storage['NGS_Overlap']=int(storage['NGS_Overlap'])
                        accept = 1
                    except:
                        self.message("Invalid-Selection07")
                else:
                    accept = 1
                if accept == 1:
                    self.clean_up_all()
                    storage["Text_Matrix"] = duplicate_shingling([storage["Data"],storage["D_Var"]],
                                                                 storage["NGS_Tid"][1],
                                                                 storage["NGS_Fulltext"][1],
                                                                 int(storage["NGS_Nglen"][1]),master=self)
                    prog_pos = 'dupli_dec'
                    self.ask()

            elif prog_pos == 'dupli_dec':
                self.store_var_all()

                try:
                    storage["NGS_Minover"] = int(storage["NGS_Minover"])
                    self.clean_up_all()
                    prog_pos = 'dupli_out'
                    self.ask()
                except:
                    self.message("Invalid-Selection07")
                

            elif prog_pos == 'dupli_out':
                self.store_var_all()
                tm = storage["Text_Matrix"]
                
                minover = storage["NGS_Minover"]
                sym = storage['NGS_Sym'][1]
                sh = storage['NGS_Share'][1]
                
                outfile = make_fname(storage['NGS_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])

                
                outvars=["Text_1","Text_2","Overlap"]
                if sh == '2':
                    outvars += ["Share_1","Share_2"]
                    tid = {}
                    for i in range(len(storage["Data"][storage["NGS_Tid"][1]])): ##indexing text numbers
                        tid[storage["Data"][storage["NGS_Tid"][1]][i]] = i
                        
                outdic = {}
                for v in outvars:
                    outdic[v] = []

                texts = sorted(list(tm.keys()))

                max1 = len(texts)
                if sym=="1":max1=max1-1
                for ti1 in range(max1):
                    min2=0
                    if sym=="1": min2=ti1+1
                    for ti2 in range(min2,len(texts)):
                        if tm[texts[ti1]][texts[ti2]]>minover:
                            outdic["Text_1"].append(texts[ti1])
                            outdic["Text_2"].append(texts[ti2])
                            outdic["Overlap"].append(tm[texts[ti1]][texts[ti2]])

                            if sh == '2':
                                i1 = tid[texts[ti1]]
                                i2 = tid[texts[ti2]]
                                l1 = len(naive_tokenizer(storage["Data"][storage["NGS_Fulltext"][1]][i1]))
                                l2 = len(naive_tokenizer(storage["Data"][storage["NGS_Fulltext"][1]][i2]))
                                print(i1,i2,l1,l2)

                                outdic["Share_1"].append(tm[texts[ti1]][texts[ti2]]/l1)
                                outdic["Share_2"].append(tm[texts[ti1]][texts[ti2]]/l2)

                t = write_data(outdic,outvars,outfile,header,sep)
                verbout('\n'+t[0],master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()
                

######################################################## Add Populism               

            elif prog_pos == 'pop_input':                
                fname = make_fname(self.store_var('Populism_Input'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                                
                v = get_varnames(fname,header,sep)
                d = get_data(fname,header,sep)[0]

                accept = 1
                impvar = ['STRAT_ShiftingBlame','STRAT_Virtues','STRAT_Denouncing','STRAT_Sovereignty','Monolith','STRAT_Closeness']
                for iv in impvar:
                    if not iv in v:
                        accept = 0
                        verbout('\nVARIABLE MISSING: '+iv,master=self)
                        
                if accept == 1:                                
                    storage['Data'] = d
                    storage['D_Var'] = v
                    check_data(d,v,master=self)
                    self.clean_up_all()
                    prog_pos = 'pop_out'
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'pop_out':
                self.store_var_all()
                data = storage['Data']
                dvar = storage['D_Var']
                outfile = make_fname(storage['Populism_Output'])

                verbout('\n\nCalculating Populism Variables.\n',master=self)
                
                if not outfile[-4] == '.':
                    outfile=outfile+'.txt'

                dta = add_populism(data,master=self)

                verbout(write_data(dta[0],dvar+dta[1],outfile)[0],master=self)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()

            elif prog_pos == 'angrist_hours':
                fname = self.store_var('Hours_Input')
                outfile = self.store_var('Hours_Output')
                if outfile[-4] == '.':
                    outfile = outfile[:-4]
                outfile1 = outfile+'_detail.txt'
                outfile2 = outfile+'.txt'

                data = get_data(fname)[0]
                data['Coding_Day'] = []
                for timestamp in data['#TS']:
                    data['Coding_Day'].append(tts(timestamp,'pyn','eng','fl'))

                data['Hours_Worked'] = data['T_H']

                agg1 = aggregate(data,['Coder','Coding_Day'],['T_Brutto','T_Break','T_Netto','Hours_Worked'],'sum',master=self)
                t = write_data(agg1[0],agg1[1],outfile1)
                verbout(t[0],master=self)
                if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                data = agg1[0]

                data['Count_Texts'] = data['Number_of_Cases']
                del data['Number_of_Cases']
                data['Coding_Month'] = []
                for timestamp in data['Coding_Day']:
                    if len(timestamp) == 10:
                        data['Coding_Month'].append(timestamp[:3]+timestamp[-2:])
                    else:
                        verbout('\nERROR: Could not convert '+timestamp,master=self)

                agg2 = aggregate(data,['Coding_Month','Coder'],['Count_Texts','T_Brutto','T_Break','T_Netto','Hours_Worked'],'sum',master=self)

                data = agg2[0]
                vlist = agg2[1]

                for i in range(len(data['T_Netto'])):
                    ntext = data['Count_Texts'][i]
                    bonus = ntext * 2 ##Sekunden pro Text
                    data['T_Netto'][i] = data['T_Netto'][i]+bonus
                    data['Hours_Worked'][i] = data['T_Netto'][i]/3600
                                                            

                data['HH_MM_SS'] = []
                for timestamp in data['T_Netto']:
                    data['HH_MM_SS'].append(tts(timestamp,'pyn','time','fl'))

                for i in range(len(data['Hours_Worked'])):
                    h = float(data['Hours_Worked'][i])
                    if h > 23:
                        stunden = int(math.floor(h))
                        data['HH_MM_SS'][i] = str(stunden)+data['HH_MM_SS'][i][2:]

                data['Mean_Duration'] = []
                for i in range(len(data['T_Netto'])):
                    mdur = float(data['T_Netto'][i]) / int(data['Count_Texts'][i])
                    data['Mean_Duration'].append(tts(str(mdur),'pyn','time','fl'))

                data['Days_Active'] = data['Number_of_Cases']
                vlist[2] = 'Days_Active'

                vlist = vlist + ['HH_MM_SS','Mean_Duration']

                t = write_data(data,vlist,outfile2)
                verbout(t[0],master=self)
                if len(t[1])>2:
                    verbout('\n'+t[1]+'\n','warning',master=self)
                else:   
                    self.clean_up_all()
                    prog_pos = 'otherart'
                    self.ask()               

            elif prog_pos == 'ggcrisi':
                self.store_var_all()
                accept = 1
                path = storage['GGCRISI_Input']
                outfile = make_fname(storage['GGCRISI_Output'])
                files = ['Text','Author','Event','EType','EActor','Attribution','Reason']
                data = {}
                varlist = {}
                for f in files:
                    try:
                        filename = path+'/_'+f+'.txt'
                        verbout('\nLoading data: '+filename+' ... ',master=self)
                        data[f] = get_data(filename)[0]
                        varlist[f] = get_varnames(filename)
                        verbout('OK',master=self)
                    except:
                        verbout('File not found. Could not aggregate data',master=self)
                        accept = 0

                if accept == 1:

                    ##Text level

                    text_data = data['Text']
                    text_vars = varlist['Text']
                    agg_author = aggregate(data['Author'],['Text_ID'],varlist['Author'][3:],'broad',master=self)

                    text_data = merge_ggcrisi(text_data,agg_author[0],['Text_ID'],['Text_ID'],master=self)
                    text_vars = text_vars[:-6] + agg_author[1][1:]

                    ##Event Level

                    event_data = data['Event']
                    event_vars = varlist['Event']

                    agg_type = aggregate(data['EType'],['Text_ID','Event_ID'],varlist['EType'][5:],'broad',master=self)
                    agg_act = aggregate(data['EActor'],['Text_ID','Event_ID'],varlist['EActor'][5:],'broad',master=self)

                    event_data = merge_ggcrisi(event_data,agg_type[0],['Text_ID','Event_ID'],['Text_ID','Event_ID'],master=self)
                    event_data = merge_ggcrisi(event_data,agg_act[0],['Text_ID','Event_ID'],['Text_ID','Event_ID'],master=self)

                    event_vars = event_vars + agg_type[1][2:]
                    event_vars = event_vars + agg_act[1][2:]

                    ##Att Level

                    att_data = data['Attribution']
                    att_vars = varlist['Attribution']

                    agg_reason = aggregate(data['Reason'],['Text_ID','Attrib_ID'],varlist['Reason'][5:],'broad',master=self)

                    att_data = merge_ggcrisi(att_data,agg_reason[0],['Text_ID','Attrib_ID'],['Text_ID','Attrib_ID'],master=self)
                    att_vars = att_vars + agg_reason[1][2:]

                    verbout('\n\nAdding Events\n',master=self)

                    att_data = merge_ggcrisi(att_data,event_data,['Text_ID','LINK_EVENT'],['Text_ID','Event_ID'],master=self)
                    if event_vars[2] in att_data.keys():
                        att_vars = event_vars + att_vars
                    else:
                        verbout('\nNo coded events found. Event variables are omitted from the dataset',master=self)

                    verbout('\n\nAdding Texts\n',master=self)

                    att_data = merge_ggcrisi(att_data,text_data,['Text_ID'],['Text_ID'],master=self)
                    att_vars = text_vars + att_vars

                    verbout('\n\nEvents without attributions\n',master=self)
                    event_data = merge_ggcrisi(event_data,text_data,['Text_ID'],['Text_ID'],master=self)

                    event_dic = {}
                    for i in range(len(event_data['Text_ID'])):
                        evlab = event_data['Text_ID'][i]+event_data['Event_ID'][i]
                        event_dic[evlab] = {}
                        for v in event_data.keys():
                            event_dic[evlab][v] = event_data[v][i]

                    for i in range(len(att_data['Text_ID'])):
                        evlab = att_data['Text_ID'][i]+att_data['Event_ID'][i]
                        try:
                            del event_dic[evlab]
                        except:
                            evlab = ''

                    verbout('\nFound '+str(len(event_dic.keys()))+' Events without attribution. Adding them to file',master=self)

                    for e in sorted(event_dic.keys()):
                        for v in att_data.keys():
                            if v in event_dic[e].keys():
                                att_data[v].append(event_dic[e][v])
                            else:
                                att_data[v].append('')       

                    out_vars = []
                    for a in att_vars:
                        if not a in out_vars:
                            out_vars.append(a)

                    verbout('\nTable complete',master=self)
                    
                    ##Output

                    t = write_data(att_data,out_vars,outfile)
                    verbout(t[0],master=self)
                    if len(t[1])>2:
                        verbout('\n'+t[1]+'\n','warning',master=self)
                    else:   
                        self.clean_up_all()
                        prog_pos = 'otherart'
                        self.ask()


            elif prog_pos == 'match_in':             
                fname = make_fname(self.store_var('Match_Content'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                
                v = get_varnames(fname,header,sep)
                d = get_data(fname,header,sep)[0]
                
                storage['MData'] = d
                storage['M_Var'] = v

                settings['Datasets']['Content Analysis'] = {}
                settings['Datasets']['Content Analysis']['Data'] = 'MData'
                settings['Datasets']['Content Analysis']['Var'] = 'M_Var'
                
                if check_data(d,v) == 1:
                    if 'Medium' in v:
                        verbout('\nValid content data. Found the following media codes: '+str(get_unique(d['Medium']))+'\n',master=self)
                        verb('Variables in file: '+str(v))

                        add_varlist('Match_Mdate',v,excludes=[])
                        add_varlist('Match_Mweight',v,excludes=[])
                        add_varlist('Match_Mvars',v,excludes=[])
                        
                        if 'Date' in v:
                            def_val['Match_Mdate']='Date'
                        else:
                            def_val['Match_Mdate']='Art_Date'
                        if 'Weight' in v:
                            def_val['Match_Mweight']='Weight'
                            
                        self.clean_up_all()
                        prog_pos = 'match_vnames'
                        self.ask()
                    else:
                        verbout('\nERROR: Invalid file. The file specified does not contain a variable called "Medium" containing the medium',master=self)
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'match_vnames':
                self.store_var_all()

                total_length = len(storage['MData'][storage['M_Var'][0]])

                if storage['Match_Mdate'][1] == 'nodate':
                    storage['MData']['res_Date'] = []
                    for i in range(total_length):
                        storage['MData']['res_Date'].append(1)
                else:
                    olddate = storage['MData'][storage['Match_Mdate'][1]]
                    storage['MData']['res_Date'] = []
                    for i in range(total_length):
                        if storage['Match_Mdate_Trans'][1] == 'no':
                            storage['MData']['res_Date'].append(tts(olddate[i],'ger','ex','ic'))
                        elif storage['Match_Mdate_Trans'][1] == 'week':
                            storage['MData']['res_Date'].append(tts(olddate[i],'ger','ex7','ic'))
                        elif storage['Match_Mdate_Trans'][1] == 'month':
                            storage['MData']['res_Date'].append(tts(olddate[i],'ger','ex30','ic'))
                        
                if storage['Match_Mweight'][1] == 'noweight':
                    storage['MData']['res_Weight'] = []
                    for i in range(total_length):
                        storage['MData']['res_Weight'].append(1)
                else:
                    oldweight = storage['MData'][storage['Match_Mweight'][1]]
                    storage['MData']['res_Weight'] = []
                    for i in range(total_length):
                        try:
                            w = float(oldweight[i])
                        except:
                            w = ''
                        storage['MData']['res_Weight'].append(w)

                storage['MData'] = delete_missing(storage['MData'],['res_Date','res_Weight'],0)

                new_length = len(storage['MData'][storage['M_Var'][0]])

                if new_length < total_length:
                    verbout('\n\nRemoved cases with missing date or weight. '+str(new_length)+' of '+str(total_length)+' remaining.\n',master=self)
                else:
                    verbout('\n\nChecked for cases with missing date or weight. Found none. '+str(total_length)+' cases remaining.\n',master=self)

                verbout('\nEarliest date: '+tts(min(storage['MData']['res_Date']),'ex','pys')+'\nLatest date: '+tts(max(storage['MData']['res_Date']),'ex','pys')+'\n',master=self)
                verbout('Weighting Factors ranging from '+str(min(storage['MData']['res_Weight']))+' to '+str(max(storage['MData']['res_Weight']))+'\n',master=self)

                self.clean_up_all()
                prog_pos = 'match_vars'
                self.ask()


            elif prog_pos == 'match_vars':
                self.store_var_all()

                verbout('\n\nReformatting Media Data',master=self)

                data = storage['MData']

                vlist = storage['Match_Mvars'][1]
                mode = storage['Match_Aggmode'][1]
                dates = get_unique(data['res_Date'])
                media = get_unique(data['Medium'])

                data['N_Texts'] = []
                for i in range(len(data[vlist[0]])):
                    data['N_Texts'].append(1)
                vlist.append('N_Texts')
                
                aggdic = {}

                for d in dates:
                    aggdic[d] = {}
                    for m in media:
                        aggdic[d][m] = {}
                        for v in vlist:
                            aggdic[d][m][v] = []
                        aggdic[d][m]['res_Weight']=[]
                

                step = int(len(data['Medium'])/40)
                if step<1: step = 1
                verbout('\nAggregating data: \n0%-------25%-------50%-------75%-------100%\n','progress',master=self)

                for i in range(len(data['Medium'])):
                    d = data['res_Date'][i]
                    w = data['res_Weight'][i]
                    m = data['Medium'][i]
                    for v in vlist:
                        aggdic[d][m][v].append(data[v][i])
                    aggdic[d][m]['res_Weight'].append(data['res_Weight'][i])
                    if i %step == 0:
                        verbout('.','progress',master=self)

                anz_calc = 0
                for d in dates:
                    for m in media:
                        weights = aggdic[d][m]['res_Weight']
                        if len(weights)>0:anz_calc = anz_calc+1
                        for v in vlist:
                            vallist = aggdic[d][m][v]
                            #if not type(vallist) == list:
                            wert = ''
                            if len(vallist)>0:
                                if mode in ['sum','mean']:
                                    numlist = []
                                    tot_weight = 0
                                    weights = aggdic[d][m]['res_Weight']
                                    for i in range(len(vallist)):
                                        try:
                                            numlist.append(float(vallist[i])*weights[i])
                                            tot_weight = tot_weight + weights[i]
                                        except:
                                            val = ''
                                    if mode == 'sum':
                                        if len(numlist) > 0:
                                            wert = sum(numlist)#/(tot_weight/len(numlist))
                                    else:
                                        if len(numlist) > 0:
                                            wert = sum(numlist)/tot_weight
                                elif mode == 'first':
                                    wert = vallist[0]
                                elif mode == 'mode':
                                    moddic = {}
                                    for val in vallist:
                                        if val in moddic.keys():
                                            moddic[val] = moddic[val] + 1
                                        else:
                                            moddic[val] = 1
                            aggdic[d][m][v] = wert
                            
                storage['Aggregated_Media'] = aggdic
                verbout('\nScores calculated for '+str(anz_calc)+' media days\n',master=self)

                #tempout(aggdic)

                self.clean_up_all()
                prog_pos = 'match_resp'
                self.ask()
               


            elif prog_pos == 'match_resp':             
                fname = make_fname(self.store_var('Match_Respondents'))
                header = self.store_var('In_Header')[1]
                sep = get_sep(self.store_var('In_Sep')[1])

                verbout('\nLoading file: '+fname, master=self)
                
                v = get_varnames(fname,header,sep)
                d = get_data(fname,header,sep)[0]

                accept = check_data(d,v,master=self)

                if storage['Methode'] == 'NCCR3':
                    d,result = reformat_mediause(d)
                elif storage['Methode'] == 'NCCR4':
                    d,result = reformat_mediause_css(d)
                    
                accept2 = result[0]
                panel = result[1]
                css = result[2]

                if panel == 1:
                    verbout('\nPanel survey data prepared.\n\n',master=self)
                elif css == 1:
                    verbout('\nData from CSS prepared.\n\n',master=self)
                else:
                    accept = 0
                    verbout('\nERROR: no valid data\n',master=self)
                
                storage['SData'] = d
                storage['S_Var'] = v

                settings['Datasets']['Survey Data'] = {}
                settings['Datasets']['Survey Data']['Data'] = 'SData'
                settings['Datasets']['Survey Data']['Var'] = 'S_Var'
                
                if accept == 1:
                    verb('Variables in file: '+str(v))
                    for var in v:
                        #verbout('\n'+var+ ': '+self.get_dta_level(d[var]),master=self)
                        codebook['Match_Gweight'][2].append(var)
                        codebook['Match_Gweight'][3].append(var)
                        codebook['Match_Sdate'][2].append(var)
                        codebook['Match_Sdate'][3].append(var)

                    if 'interview_startL2' in v:
                        def_val['Match_Sdate']='interview_startL2'
                        
                    self.clean_up_all()
                    prog_pos = 'match_survey'
                    self.cini_schreiben()
                    self.ask()
                else:
                    verb('ERROR: Invalid File')

            elif prog_pos == 'match_survey':
                self.store_var_all()
                self.clean_up_all()

                sdata = storage['SData']

                warray = []
                wcount = 0
                for i in range(data_dim(sdata)[1]):
                    if storage['Match_Gweight'][1] == '99':
                        warray.append(1)
                    else:
                        try:
                            warray.append(float(sdata[storage['Match_Gweight'][1]][i]))
                            wcount = wcount + 1
                        except:
                            warray.append('')

                if storage['Match_Gweight'][1] == '99':
                    verbout('\nNo global weighting',master=self)
                else:
                    verbout('\nFound valid global matching variable for '+str(wcount)+' cases (of '+str(i+1)+')',master=self)

                storage['SData']['res_GWeight'] = warray

                prog_pos = 'match_date'
                self.ask()

            elif prog_pos == 'match_date':
                self.store_var_all()
                self.clean_up_all()

                sdata = storage['SData']

                darray = []
                dcount = 0

                for i in range(data_dim(sdata)[1]):
                    if storage['Match_Sdate'][1] == '99':
                        darray.append(1)
                    else:
                        try:
                            datum = time.strptime(sdata[storage['Match_Sdate'][1]][i],"%m/%d/%Y")
                            ts = time.mktime(datum)
                            ts = ts / 24 / 3600
                            ts = int(math.floor(ts + 25569))
                            dcount = dcount + 1
                        except:
                            ts = ''
                            
                        darray.append(ts)

                if storage['Match_Sdate'][1] == '99':
                    verbout('\nNo survey date',master=self)
                else:
                    verbout('\nFound valid date variable for '+str(dcount)+' cases (of '+str(i+1)+')',master=self)


                storage['SData']['res_SDate'] = darray

                prog_pos = 'match_out'
                self.ask()


            elif prog_pos == 'match_out':
                self.store_var_all()
                sdata = storage['SData']
                svars = storage['S_Var']
                mdata = storage['Aggregated_Media']
                
                lweight = storage['Match_Lweight'][1]
                wdate = storage['Match_Wdate'][1]
                mode = storage['Match_Calcmode'][1]
                
                ##Variables                
                outfile = make_fname(storage['Match_Out'])
                header = int(storage['Out_Header'][1])
                sep = get_sep(storage['Out_Sep'][1])
                
                verbout('\n\nMatching data. This could take some moments...\n\nRemoving cases without valid global weighting or date.',master=self)

                orig_length=len(sdata[svars[0]])
                sdata = delete_missing(sdata,['res_SDate','res_GWeight'],0)
                new_length=len(sdata[svars[0]])
                verbout(str(new_length)+' valid survey cases (of '+str(orig_length)+') remaining.',master=self)

                if wdate in ['combo'] and mode in ['combo','combo2']:
                    out_data = sdata
                    if mode == 'combo':
                        sel_modes = ['sum_sum','mean_mean']
                    else:
                        sel_modes = ['sum_sum','sum_mean','mean_sum','mean_mean']
                    verbout('\n\nYou requested a large set of combinations. Doing them in the following order:\n',master=self)
                    for wdate_s in ['before','1mbefore','7d']:
                        for mode_s in sel_modes:
                            verbout('\nCombination: Date: '+wdate_s+' Calculation: '+mode_s,master=self)
                            suffix = '_'+wdate_s+'_'+mode_s
                            verbout('\nAttaching suffix to variables: '+suffix,master=self)
                            
                            out_data, addvars = match_nccr(out_data,mdata,lweight,wdate_s,mode_s,suffix=suffix,master=self)
                            svars = svars + addvars
                            verbout('\nNew variables: '+str(addvars),master=self)
                            verbout('\n--------------\n',master=self)
                elif wdate in ['combo']:
                    out_data = sdata
                    for wdate_s in ['before','1mbefore','7d']:   
                        verbout('\nMatching version: Date: '+wdate_s,master=self)
                        suffix = '_'+wdate_s
                        verbout('\nAttaching suffix to variables: '+suffix,master=self)
                        out_data, addvars = match_nccr(out_data,mdata,lweight,wdate_s,mode,suffix=suffix,master=self)
                        svars = svars + addvars
                        verbout('\nNew variables: '+str(addvars),master=self)
                        verbout('\n--------------\n',master=self)
                elif mode == 'combo':
                    out_data = sdata
                    for mode_s in ['sum_sum','mean_mean']:   
                        verbout('\nMatching version: Calculation: '+mode_s,master=self)
                        suffix = '_'+mode_s
                        verbout('\nAttaching suffix to variables: '+suffix,master=self)
                        out_data, addvars = match_nccr(out_data,mdata,lweight,wdate,mode_s,suffix=suffix,master=self)
                        svars = svars + addvars
                        verbout('\nNew variables: '+str(addvars),master=self)
                        verbout('\n--------------\n',master=self)
                elif mode == 'combo2':
                    out_data = sdata
                    for mode_s in ['sum_sum','sum_mean','mean_sum','mean_mean']:   
                        verbout('\nMatching version: Calculation: '+mode_s,master=self)
                        suffix = '_'+mode_s
                        verbout('\nAttaching suffix to variables: '+suffix,master=self)
                        out_data, addvars = match_nccr(out_data,mdata,lweight,wdate,mode_s,suffix=suffix,master=self)
                        svars = svars + addvars
                        verbout('\nNew variables: '+str(addvars),master=self)
                        verbout('\n--------------\n',master=self)                    
                else:
                    out_data, addvars = match_nccr(sdata,mdata,lweight,wdate,mode,suffix='',master=self)
                    svars = svars+addvars

                t = write_data(out_data,svars,outfile,header,sep)
                verbout('\n'+t[0],master=self)
                #if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=self)

                logfile = '_logger_matching_nccr.txt'

                verbout('Writing logfile to: '+logfile,master=self)
                log_settings(['Match_Content','Match_Mdate','Match_Mweight',
                              'Match_Mvars','Match_Aggmode','Match_Respondents',
                              'Match_Gweight','Match_Lweight','Match_Sdate','Match_Wdate',
                              'Match_Calcmode','Match_Out'],logfile)

                self.clean_up_all()
                prog_pos = 'otherart'
                self.ask()


            
            elif prog_pos == 'otherart':
                self.fuellen()                     
            else:
                verb("ERROR: The program position '"+prog_pos+"' is not defined for SUBMIT")
        else:
            verb('Invalid Entries')
        
        try:
            self.f_bottomline.b_check["state"] = NORMAL
        except:
            verb('No Check button')


############################################
##                                        ##
##       Individual Functions             ##
##       for each project                 ##
##                                        ##
############################################


    def abort(self):
        #This function is called by the abort-button. For each page on which the abort-button is enabled
        #this function has to be defined.
        #The abort-function may also be called by other buttons such as questions defined in the MCP function      
        global prog_pos
        global dta_pos
        log('Calling Function: Abort')

        if prog_pos in ['otherart','art_otherart']:
            self.clean_up_all()
            prog_pos = 'ende'
            self.ask()
        elif prog_pos in ['art_issue','artspez1']:
            if self.message('Caution02',3)==1:
                self.clean_up_all()
                storage['Bemerkungen']='Does not belong to the sample'
                prog_pos = 'otherart'
                self.ask()
        elif prog_pos == 'heat_display':
            self.clean_up_all()
            prog_pos = 'otherart'
            self.ask()
            
        else:
            verb('ERROR: Abort function not defined for this page')


    def back(self):
        #This function is called by the back-button. For each page on which the back-button is enabled
        #this function has to be defined. The standard definition (go back one page) is default for all pages.
        #In some cases, this default will not hold, however, as parts of the data structure also has to
        #be changed. For example, the data level (data_pos) has to be changed or some entry has to be
        #deleted when going back from certain pages. In these cases a manual definition must be provided
        #in order to prevent errors.
        global prog_pos
        global dta_pos
        log('Calling Function: Back')

        if len(settings['Out_Track']) > 0:
            tf = open(settings['Out_Track'], 'a')
            tf.write('back-button\n')
            tf.close()

        storage['Backs'] = storage['Backs'] + 1
        if settings['Page_History'][-2] == 'methode': ## Jumping back to start
            if self.message('Caution01', 3) == 1:
                self.fuellen()        
        elif len(settings['Page_History']) > 1:
            if prog_pos == 'some page':
                pass
            else:
                prog_pos = settings['Page_History'][len(settings['Page_History'])-2] #Go back one step

            self.clean_up_all()
            settings['Page_History'].pop(len(settings['Page_History'])-1) #remove the two most recent pages in history.
            settings['Page_History'].pop(len(settings['Page_History'])-1)
            self.hide_review()
            self.ask()
        else:
            if self.message('Caution01',3) == 1: #Warning when trying to go back to the very beginning.
                self.fuellen() #Delete everyting on this text and start anew


    def rb_tamper(self):
        #This function will be called whenever the value of a Radiobutton is changed. It may be used to
        #Change the display according to the current selection.
        #This function has to be defined for all pages on which it is needed.  
        global prog_pos
        global storage
        log('Calling Function: RB-Tamper')
        #print(settings['Curr_Page'])

        if prog_pos == 'ts_format':
            try:
                self.clean_up(pos=3)
            except:
                verb('No TS_Int found')
            rbpos1 = self.store_var('TS_Informat',store=0)
            if type(rbpos1)==tuple:
                rbpos1 = rbpos1[1]

            if settings['Curr_Page'][1][1]=='TS_Outformat':
                rbpos2 = self.store_var('TS_Outformat',store=0)
                if type(rbpos2)==tuple:
                    rbpos2 = rbpos2[1]
            else:
                rbpos2 = 'pys'
                
            excalc = ''
            invalue = ''
            i = 0
            while invalue == '':
                invalue = storage['TStamps'][i]
                i = i + 1
            excalc = excalc + invalue
            outval = tts(invalue,rbpos1,rbpos2)
            excalc = excalc + ' -> '+str(outval)

            if outval == '':
                verb('ERROR: Conversion not possible: '+rbpos1+' > '+rbpos2+' for value "'+str(invalue)+'"')
                verbout('\nInput and output formats not valid')
            else:
                verbout('\n\nInput='+rbpos1+'\nOutput='+rbpos2+'\nExample: '+excalc,master=self)

        elif prog_pos == 'group_mode':
            try:
                self.clean_up(pos=2)
            except:
                verb('Nothing to clean')
            a = self.store_var('Group_Mode')[1]
            if a == 'equal':
                self.question_dd('Group_Equal',2)
            elif a == 'fixed':
                self.question_txt('Group_Fixed',2)
            elif a == 'tails':
                self.question_dd('Group_Tails',2)

        elif prog_pos == 'patd_opt':
            try:
                self.clean_up(pos=2)
            except:
                verb('Nothing to clean')
            a = self.store_var('PD_Method')[1]

            if a in ['2','3','4']:
                self.question_txt('PD_Cutoff',2)

        elif prog_pos == 'focus_date':
            try:
                v = self.store_var('Focus_Date')[1]
                f = self.store_var('Focus_Dformat')[1]

                outdat = {}
                invalids = []
                valids = []
                newcol = []

                for ts in storage['Data'][v]:
                    if f == 'lfdn':
                        try:
                            outvalue = int(ts)
                        except:
                            outvalue = ''
                    else:
                        outvalue = tts(ts,f,'ex')
                        
                    if outvalue == '':
                        invalids.append(ts)
                    else:
                        valids.append(ts)
                        outdat[outvalue]=1
                    
                    newcol.append(outvalue)
                                                        
                verbout('\n\nAnalysis of available timestamps:\nNumber of different dates: '+str(len(outdat.keys())),master=self)
                verbout('\nNumber of valid cases: '+str(len(valids)),master=self)
                verbout('\nNumber of invalid cases: '+str(len(invalids)),master=self)
                if len(invalids)>0:
                    verbout('\n',master=self)
                    verbout('-First 20 invalid cases: '+str(invalids[:20]),'warning',master=self)


                storage['Data']['res_Day'] = newcol
                if not 'res_Day' in storage['D_Var']:
                    storage['D_Var'].append('res_Day')
                
            except Exception as fehler:
                verb('No valid time format yet ('+v+'/'+f+')')
                verb(str(fehler))
            
        else:
            verb('Error: RB-Tamper not defined for this page.')
        

        

            
############################################
##                                        ##
##       Question-Functions               ##
##                                        ##
############################################


    def question_dd(self, cb_var, question_pos,width=40): #Dropdown-Question: Up to three dropdown selections may be displayed per page
        global settings
        log('--Question: Dropdown. Variable: '+cb_var+'; Position: '+str(question_pos),pos=0)
        #### Setting the default value according to text statistics using AEGLOS:
        if settings['AEGLOS'] == '1':
            wert = -1
            pred = acabc.predict_short(cb_var,settings['Fulltext'])
            wert = self.namegetter(cb_var,pred)
            if not wert == '':
                if not cb_var in def_val.keys():
                    def_val[cb_var] = wert
                    if settings['Verbose'] == '1':
                        verb('Automated determination of the value')
                        verb('Best prediction: '+str(pred)+' Value: '+str(wert))
                else:
                    if settings['Verbose'] == '1': verb('Another default value has been set already. No changes made')

        settings['Curr_Page'][question_pos-1] = ['dd',cb_var]
        settings['Input'][question_pos-1] = ''
        curr_tree=curr()

        verb('Looking for previously set values..')
        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = '' ###Has to have the same data type as set default values or previous codings

        if type(previous_coding) == tuple:
            previous_coding = previous_coding[0]
        elif type(previous_coding) == str:
            if len(self.namegetter(cb_var,previous_coding)) >0:
                previous_coding = self.namegetter(cb_var,previous_coding)
            elif previous_coding in codebook[cb_var][2]:
                previous_coding = str(previous_coding)
            else:
                previous_coding = ''
                verb('No valid coding found')
        else:
            previous_coding = ''
            verb('No valid coding found')

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]
         
        settings['Input'][question_pos-1] = StringVar()
        if previous_coding in namelist:
            settings['Input'][question_pos-1].set(previous_coding)
        else:
            settings['Input'][question_pos-1].set(namelist[0])

        if question_pos == 1:
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Display question
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Display additional information
            #Display dropdown
            try:
                self.f_questions.dd1 = OptionMenu(*(self.f_questions, settings['Input'][question_pos-1]) + tuple(namelist))
            except:
                self.f_questions.dd1 = apply(OptionMenu, (self.f_questions, settings['Input'][question_pos-1]) + tuple(namelist))
            self.f_questions.dd1["width"] = width
            self.f_questions.dd1["takefocus"] = 1
            self.f_questions.dd1.grid(row=3, column=1, sticky=W+E)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            if settings['Insecure']=='1':
                self.f_questions.ins1 = Label(self.f_questions, text="unsich.",fg="#ee0000")
                self.f_questions.ins1.grid(row=4, column=3, sticky=E)
                self.f_questions.ins1.bind('<Button-1>', CMD(self.insecure, cb_var))               
        if question_pos == 2:
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Display question
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Display additional information
            #Display Dropdown
            try:
                self.f_questions.dd2 = OptionMenu(*(self.f_questions, settings['Input'][question_pos-1]) + tuple(namelist))
            except:
                self.f_questions.dd2 = apply(OptionMenu, (self.f_questions, settings['Input'][question_pos-1]) + tuple(namelist))
            self.f_questions.dd2["width"] = width
            self.f_questions.dd2["takefocus"] = 1
            self.f_questions.dd2.grid(row=7, column=1, sticky=W+E)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            if settings['Insecure']=='1':
                self.f_questions.ins2 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins2.grid(row=8, column=3, sticky=E)
                self.f_questions.ins2.bind('<Button-1>', CMD(self.insecure, cb_var))               
        if question_pos == 3:
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Display Question
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Display additional information
            #Display Dropdown
            try:
                self.f_questions.dd3 = OptionMenu(*(self.f_questions, settings['Input'][question_pos-1]) + tuple(namelist))
            except:
                self.f_questions.dd3 = apply(OptionMenu, (self.f_questions, settings['Input'][question_pos-1]) + tuple(namelist))
            self.f_questions.dd3["width"] = width
            self.f_questions.dd3["takefocus"] = 1
            self.f_questions.dd3.grid(row=11, column=1, sticky=W+E)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            if settings['Insecure']=='1':
                self.f_questions.ins3 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins3.grid(row=12, column=3, sticky=E)
                self.f_questions.ins3.bind('<Button-1>', CMD(self.insecure, cb_var))               

    def question_txt(self, cb_var, question_pos, width=40): #Textfield-Entry. Display a one-line text entry.
        log('--Question: TXT(line). Variable: '+cb_var+'; Position: '+str(question_pos),pos=0)
        
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = '' ###Has to have the same data type as set default values or previous codings

        if question_pos == 1:
            settings['Curr_Page'][0] = ['txt',cb_var]
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt1 = Entry(self.f_questions, width=width)
            self.f_questions.txt1.grid(row=3,column=0,columnspan=3,sticky=W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.txt1.insert(END,str(previous_coding))
                
        if question_pos == 2:
            settings['Curr_Page'][1] = ['txt',cb_var]
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt2 = Entry(self.f_questions, width=width)
            self.f_questions.txt2.grid(row=7,column=0,columnspan=3,sticky=W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.txt2.insert(END,str(previous_coding))
                
        if question_pos == 3:
            settings['Curr_Page'][2] = ['txt',cb_var]
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt3 = Entry(self.f_questions, width=width)
            self.f_questions.txt3.grid(row=11,column=0,columnspan=3,sticky=W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.txt3.insert(END,str(previous_coding))


    def opendialog(self,mode='load',question_pos=1,defaultextension='.txt'):
        fname = ''

        extlist = [('Text File','.txt'),('Data File','.dat'),('Excel Spreadsheet','.xlsx'),('All Files','.*')]

        if not defaultextension == extlist[0][1]:
            extlist = [('*'+defaultextension,defaultextension)]+extlist
        
        if mode == 'save':
            try: ##Python 3
                fname = filedialog.asksaveasfilename(**{'defaultextension':defaultextension,
                                                      'filetypes':extlist})
            except:
                fname = tkFileDialog.asksaveasfilename(**{'defaultextension':defaultextension,
                                                          'filetypes':extlist})
        elif mode == 'load':
            try: ##Python 3
                fname = filedialog.askopenfilename(**{'defaultextension':defaultextension,
                                                          'filetypes':extlist})
            except:
                fname = tkFileDialog.askopenfilename(**{'defaultextension':defaultextension,
                                                          'filetypes':extlist})    


        elif mode == 'folder':
            try:
                fname = filedialog.askdirectory()
            except:
                fname = tkFileDialog.askdirectory()

        if question_pos == 1:
            self.f_questions.txt1.delete('0',END)
            self.f_questions.txt1.insert(END,fname)
        if question_pos == 2:
            self.f_questions.txt2.delete('0',END)
            self.f_questions.txt2.insert(END,fname)
        if question_pos == 3:
            self.f_questions.txt3.delete('0',END)
            self.f_questions.txt3.insert(END,fname)
            

    def question_file(self, cb_var, question_pos, mode='load',defext='.txt'):
        log('--Question: Filename. Variable: '+cb_var+'; Position: '+str(question_pos),pos=0)
        width=50
        
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = '' ###Has to have the same data type as set default values or previous codings

        if question_pos == 1:
            settings['Curr_Page'][0] = ['file',cb_var]
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt1 = Entry(self.f_questions, width=width)
            self.f_questions.txt1.grid(row=3,column=0,columnspan=2,sticky=E+W)
            self.f_questions.getselect1 = Button(self.f_questions,text='Browse...',command=CMD(self.opendialog,mode,question_pos,defext))
            self.f_questions.getselect1.grid(row=3,column=2,sticky=W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.txt1.insert(END,str(previous_coding))
                
        if question_pos == 2:
            settings['Curr_Page'][1] = ['file',cb_var]
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt2 = Entry(self.f_questions, width=width)
            self.f_questions.txt2.grid(row=7,column=0,columnspan=2,sticky=E+W)
            self.f_questions.getselect2 = Button(self.f_questions,text='Browse...',command=CMD(self.opendialog,mode,question_pos,defext))
            self.f_questions.getselect2.grid(row=7,column=2,sticky=W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.txt2.insert(END,str(previous_coding))
                
        if question_pos == 3:
            settings['Curr_Page'][2] = ['file',cb_var]
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt3 = Entry(self.f_questions, width=width)
            self.f_questions.txt3.grid(row=11,column=0,columnspan=2,sticky=E+W)
            self.f_questions.getselect3 = Button(self.f_questions,text='Browse...',command=CMD(self.opendialog,mode,question_pos,defext))
            self.f_questions.getselect3.grid(row=11,column=2,sticky=W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.txt3.insert(END,str(previous_coding))


    def getselect(self,textfeld): #Import selected text to the current entry form
        log('Calling Function: Getselect')
        try:
            select = bereinigen(self.Artikel.get(SEL_FIRST,SEL_LAST))
        except:
            verb('No Text selected')
            select = ""
        if textfeld == 1:
            self.f_questions.txt1.delete('1.0',END)
            self.f_questions.txt1.insert(END, select)
        elif textfeld == 2:
            self.f_questions.txt2.delete('1.0',END)
            self.f_questions.txt2.insert(END, select)
        elif textfeld == 3:
            self.f_questions.txt3.delete('1.0',END)
            self.f_questions.txt3.insert(END, select)


    def question_txt2(self, cb_var, question_pos, width=40, height=3, getselect=0): #Text Entry for input of multiple lines of text.
        log('--Question: TXT2(multiline). Variable: '+cb_var+'; Position: '+str(question_pos)+'; Getselect:'+str(getselect),pos=0)
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = '' ###Has to have the same data type as set default values or previous codings

        if question_pos == 1:
            settings['Curr_Page'][0] = ['txt2',cb_var]
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt1 = Text(self.f_questions, width=width, height=height, wrap=WORD, relief=RIDGE, font = (settings['Font'], "9"))
            self.f_questions.txt1.grid(row=3,column=0,columnspan=3,sticky=W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.getselect1 = Button(self.f_questions, text = "Get Selection", command=CMD(self.getselect,1))
            self.f_questions.getselect1.grid(row=3,column=3)
            if getselect == 0:
                self.f_questions.getselect1.destroy()
            self.f_questions.txt1.insert(END,str(previous_coding))
                
        if question_pos == 2:
            settings['Curr_Page'][1] = ['txt2',cb_var]
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Coder Information
            self.f_questions.txt2 = Text(self.f_questions, width=width, height=height, wrap=WORD, relief=RIDGE, font = (settings['Font'], "9"))
            self.f_questions.txt2.grid(row=7,column=0,columnspan=3,sticky=W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.getselect2 = Button(self.f_questions, text = "Get Selection", command=CMD(self.getselect,2))
            self.f_questions.getselect2.grid(row=7,column=3)
            if getselect == 0:
                self.f_questions.getselect2.destroy()
            self.f_questions.txt2.insert(END,str(previous_coding))
                
        if question_pos == 3:
            settings['Curr_Page'][2] = ['txt2',cb_var]
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Coder information
            self.f_questions.txt3 = Text(self.f_questions, width=width, height=height, wrap=WORD, relief=RIDGE, font = (settings['Font'], "9"))
            self.f_questions.txt3.grid(row=11,column=0,columnspan=3,sticky=W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.getselect3 = Button(self.f_questions, text = "Get Selection", command=CMD(self.getselect,3))
            self.f_questions.getselect3.grid(row=11,column=3)
            if getselect == 0:
                self.f_questions.getselect3.destroy()
            self.f_questions.txt3.insert(END,str(previous_coding))
            
    def question_cb(self, cb_var, question_pos,layout="hor",defval=0): #Checkbox-question. Up to 14 checkboxes may be defined in the codebook
        log('--Question: Checkbox. Variable: '+cb_var+'; Position: '+str(question_pos)+'; Layout: '+layout,pos=0)
        global settings
        if settings['AEGLOS'] == '1':
            verb('Automated detection')
            predliste = []
            for kat in codebook[cb_var][3]:
                vname = cb_var + '_' + kat
                wert = acabc.predict_short(vname,settings['Fulltext'])
                if vname in def_val.keys():
                    predliste.append(str(def_val[vname]))
                    verb('Previous Default-Value set: '+str(def_val[vname]))
                else:
                    predliste.append(str(wert))
                    verb('New Default-Value set: '+str(wert))
        else:
            verb('Searching Default-Values')
            predliste = []
            for kat in codebook[cb_var][3]:
                vname = cb_var + '_' + kat
                if vname in def_val.keys():
                    predliste.append(str(def_val[vname]))
                    if settings['Verbose'] == '1': verb('Old Default-Value set: '+str(def_val[vname]))
                else:
                    predliste.append('0')

        settings['Curr_Page'][question_pos-1] = ['cb',cb_var]
        settings['Input'][question_pos-1] = []
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = {} ###Has to have the same data type as set default values or previous codings


        if question_pos == 1:
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist1 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist1.grid(row=3, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist1
            if settings['Insecure']=='1':
                self.f_questions.ins1 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins1.grid(row=12, column=3, sticky=E)
                self.f_questions.ins1.bind('<Button-1>', CMD(self.insecure, cb_var))
        elif question_pos == 2:
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist2 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist2.grid(row=7, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist2
            if settings['Insecure']=='1':
                self.f_questions.ins2 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins2.grid(row=12, column=3, sticky=E)
                self.f_questions.ins2.bind('<Button-1>', CMD(self.insecure, cb_var))               
        elif question_pos == 3:
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist3 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist3.grid(row=11, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist3
            if settings['Insecure']=='1':
                self.f_questions.ins3 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins3.grid(row=12, column=3, sticky=E)
                self.f_questions.ins3.bind('<Button-1>', CMD(self.insecure, cb_var))               

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]

        if layout == 'vert':
            lpos = [(1,1),(2,1),(3,1),(4,1),(5,1),(6,1),(7,1),(1,3),(2,3),(3,3),(4,3),(5,3),(6,3),(7,3)]
        if layout == 'hor':
            lpos = [(1,1),(1,3),(2,1),(2,3),(3,1),(3,3),(4,1),(4,3),(5,1),(5,3),(6,1),(6,3),(7,1),(7,3)]

        for k in range(len(namelist)):
            v = IntVar()
            if codelist[k] in previous_coding.keys():
                v.set(int(previous_coding[codelist[k]]))
            else:
                v.set(defval)
            settings['Input'][question_pos-1].append(v)          

        buttons = []
        labels = []      
        for k in range(len(namelist)):
            lab = Label(f,text=namelist[k])
            labels.append(lab)
            lab.grid(row=lpos[k][0],column=lpos[k][1],sticky=W)
            but = Checkbutton(f,variable=settings['Input'][question_pos-1][k])
            buttons.append(but)
            but.grid(row=lpos[k][0],column=lpos[k][1]-1,sticky=E)

             
    def question_rb(self, cb_var, question_pos, layout='vert', defval='98'): #Radiobutton-Question: Up to 7 Radiobuttons may be defined in the codebook.
        log('--Question: Radiobutton. Variable: '+cb_var+'; Position: '+str(question_pos)+'; Layout: '+layout,pos=0)
        global settings
        #### Automated content analysis:
        if settings['AEGLOS'] == '1':
            wert = -1
            pred = acabc.predict_short(cb_var,settings['Fulltext'])
            wert = self.namegetter(cb_var,pred)
            if not wert == '':
                if not cb_var in def_val.keys():
                    def_val[cb_var] = pred
                    if settings['Verbose'] == '1':
                        verb('Automated value determination')
                        verb('Best prediction: '+str(pred)+ 'Value: '+str(wert))
                else:
                    if settings['Verbose'] == '1': verb('Another default value has been set already. No changes made')

        settings['Curr_Page'][question_pos-1] = ['rb',cb_var]
        settings['Input'][question_pos-1] = ''
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = '' ###Has to have the same data type as set default values or previous codings

        if type(previous_coding) == tuple:
            previous_coding = previous_coding[1]
        elif type(previous_coding) == str:
            if len(self.codegetter(cb_var,previous_coding)) >0:
                previous_coding = self.codegetter(cb_var,previous_coding)
            elif previous_coding in codebook[cb_var][3]:
                previous_coding = str(previous_coding)
            else:
                previous_coding = ''
                verb('No valid coding found')
        else:
            previous_coding = ''
            verb('No valid coding found')


        if question_pos == 1:
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist1 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist1.grid(row=3, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist1
            if settings['Insecure']=='1':
                self.f_questions.ins1 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins1.grid(row=12, column=3, sticky=E)
                self.f_questions.ins1.bind('<Button-1>', CMD(self.insecure, cb_var))
        elif question_pos == 2:
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist2 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist2.grid(row=7, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist2
            if settings['Insecure']=='1':
                self.f_questions.ins2 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins2.grid(row=12, column=3, sticky=E)
                self.f_questions.ins2.bind('<Button-1>', CMD(self.insecure, cb_var))               
        elif question_pos == 3:
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist3 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist3.grid(row=11, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist3
            if settings['Insecure']=='1':
                self.f_questions.ins3 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins3.grid(row=12, column=3, sticky=E)
                self.f_questions.ins3.bind('<Button-1>', CMD(self.insecure, cb_var))               

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]

        if layout == 'vert':
            lpos = [(1,1),(2,1),(3,1),(4,1),(5,1),(6,1),(7,1),(1,3),(2,3),(3,3),(4,3),(5,3),(6,3),(7,3)]
        if layout == 'hor':
            lpos = [(1,1),(1,3),(2,1),(2,3),(3,1),(3,3),(4,1),(4,3),(5,1),(5,3),(6,1),(6,3),(7,1),(7,3)]
         
        settings['Input'][question_pos-1] = StringVar()
        if previous_coding == '':
            settings['Input'][question_pos-1].set(defval)
        else:
            settings['Input'][question_pos-1].set(previous_coding)

        buttons = []
        labels = []      
        for k in range(len(namelist)):
            lab = Label(f,text=namelist[k])
            labels.append(lab)
            lab.grid(row=lpos[k][0],column=lpos[k][1],sticky=W)
            but = Radiobutton(f,variable=settings['Input'][question_pos-1],value=codelist[k],command=self.rb_tamper)
            buttons.append(but)
            but.grid(row=lpos[k][0],column=lpos[k][1]-1,sticky=W)

        self.rb_tamper()


    def question_rbopen(self, cb_var, question_pos, layout='vert', defval='98'): #Radiobutton-Question: Up to 7 Radiobuttons may be defined in the codebook.
        log('--Question: Radiobutton with open ended box. Variable: '+cb_var+'; Position: '+str(question_pos)+'; Layout: '+layout,pos=0)
        global settings
        #### Automated content analysis:
        if settings['AEGLOS'] == '1':
            wert = -1
            pred = acabc.predict_short(cb_var,settings['Fulltext'])
            wert = self.namegetter(cb_var,pred)
            if not wert == '':
                if not cb_var in def_val.keys():
                    def_val[cb_var] = pred
                    if settings['Verbose'] == '1':
                        verb('Automated value determination')
                        verb('Best prediction: '+str(pred)+ 'Value: '+str(wert))
                else:
                    if settings['Verbose'] == '1': verb('Another default value has been set already. No changes made')

        settings['Curr_Page'][question_pos-1] = ['rbopen',cb_var]
        settings['Input'][question_pos-1] = ''
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = '' ###Has to have the same data type as set default values or previous codings

        previous_text = ''
        if type(previous_coding) == tuple:
            previous_coding = previous_coding[1]
        elif type(previous_coding) == str:
            if len(self.codegetter(cb_var,previous_coding)) >0:
                previous_coding = self.codegetter(cb_var,previous_coding)
            elif previous_coding in codebook[cb_var][3]:
                previous_coding = str(previous_coding)
            else:
                previous_text = previous_coding   
                previous_coding = 'Open Answer'
##                previous_coding = ''
##                verb('No valid coding found')
        else:
            previous_coding = ''
            verb('No valid coding found')


        if question_pos == 1:
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist1 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist1.grid(row=3, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist1
            self.t1 = Entry(f, width=20)
            if not previous_text == '': self.t1.insert(0,previous_text)
            if settings['Insecure']=='1':
                self.f_questions.ins1 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins1.grid(row=12, column=3, sticky=E)
                self.f_questions.ins1.bind('<Button-1>', CMD(self.insecure, cb_var))
        elif question_pos == 2:
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist2 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist2.grid(row=7, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist2
            self.t2 = Entry(f, width=20)
            if not previous_text == '': self.t2.insert(0,previous_text)
            if settings['Insecure']=='1':
                self.f_questions.ins2 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins2.grid(row=12, column=3, sticky=E)
                self.f_questions.ins2.bind('<Button-1>', CMD(self.insecure, cb_var))               
        elif question_pos == 3:
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist3 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist3.grid(row=11, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist3
            self.t3 = Entry(f, width=20)
            if not previous_text == '': self.t3.insert(0,previous_text)
            if settings['Insecure']=='1':
                self.f_questions.ins3 = Label(self.f_questions, text="unsich.", fg="#ee0000")
                self.f_questions.ins3.grid(row=12, column=3, sticky=E)
                self.f_questions.ins3.bind('<Button-1>', CMD(self.insecure, cb_var))               

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]+['Open Answer']

        if layout == 'vert':
            lpos = [(1,1),(2,1),(3,1),(4,1),(5,1),(6,1),(7,1),(1,3),(2,3),(3,3),(4,3),(5,3),(6,3),(7,3)]
        if layout == 'hor':
            lpos = [(1,1),(1,3),(2,1),(2,3),(3,1),(3,3),(4,1),(4,3),(5,1),(5,3),(6,1),(6,3),(7,1),(7,3)]
         
        settings['Input'][question_pos-1] = StringVar()
        if previous_coding == '':
            settings['Input'][question_pos-1].set(defval)
        else:
            settings['Input'][question_pos-1].set(previous_coding)

        buttons = []
        labels = []
        maxlen = 0
        for k in range(len(namelist)):
            lab = Label(f,text=namelist[k])
            if len(namelist[k])>maxlen:maxlen=len(namelist[k])
            labels.append(lab)
            lab.grid(row=lpos[k][0],column=lpos[k][1],sticky=W)
            but = Radiobutton(f,variable=settings['Input'][question_pos-1],value=codelist[k],command=self.rb_tamper)
            buttons.append(but)
            but.grid(row=lpos[k][0],column=lpos[k][1]-1,sticky=W)

        if question_pos == 1:
            self.t1.grid(row=lpos[k+1][0],column=lpos[k+1][1],sticky=W)
        elif question_pos == 2:
            self.t2.grid(row=lpos[k+1][0],column=lpos[k+1][1],sticky=W)
        elif question_pos == 3:
            self.t3.grid(row=lpos[k+1][0],column=lpos[k+1][1],sticky=W)
        but = Radiobutton(f,variable=settings['Input'][question_pos-1],value=codelist[k+1],command=self.rb_tamper)
        buttons.append(but)
        but.grid(row=lpos[k+1][0],column=lpos[k+1][1]-1,sticky=W)

        self.rb_tamper()


          

    def question_sd(self, cb_var, question_pos,points=5,defval=0): #Semantic differential question
        log('--Question: Semantic Differential. Variable: '+cb_var+'; Position: '+str(question_pos),pos=0)
        global settings
        settings['Curr_Page'][question_pos-1] = ['sd',cb_var]
        settings['Input'][question_pos-1] = []
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = {} ###Has to have the same data type as set default values or previous codings

        if question_pos == 1:
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist1 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist1.grid(row=3, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist1
        elif question_pos == 2:
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist2 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist2.grid(row=7, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist2
        elif question_pos == 3:
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist3 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist3.grid(row=11, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist3

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]
        for k in range(len(namelist)):
            v = StringVar()
            if codelist[k] in previous_coding.keys():
                v.set(previous_coding[codelist[k]])
            else:
                v.set(defval)
            settings['Input'][question_pos-1].append(v)          

        scale = range(points)
        buttons = []
        labels = []
        for k in range(len(namelist)):
            lab = Label(f,text=namelist[k])
            labels.append(lab)
            lab.grid(row=k+1,column=points+2,sticky=W)
            lab2 = Label(f,text=codelist[k])
            labels.append(lab2)
            lab2.grid(row=k+1,column=0,sticky=E)

        for i in range(points):
            for k in range(len(namelist)):
                rb = Radiobutton(f,variable=settings['Input'][question_pos-1][k], value=i)
                buttons.append(rb)
                rb.grid(row=k+1,column=i+1)
         
    def question_rating(self,cb_var,question_pos,scalelist=['disagree','','','','agree'],valuelist=['1','2','3','4','5'],defval='1'):
        ##Rating question. For each item you may rate from scalelist[0] to scalelist[-1] setting the codes from valuelist[0] to valuelist[-1]. The length of these lists has to be identical.
        log('--Question: Rating. Variable: '+cb_var+'; Position: '+str(question_pos)+'; Number of Points: '+str(len(scalelist)),pos=0)
        global settings
        settings['Curr_Page'][question_pos-1] = ['rating',cb_var]
        settings['Input'][question_pos-1] = []
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = {} ###Has to have the same data type as set default values or previous codings

        if question_pos == 1:
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist1 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist1.grid(row=3, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist1
        elif question_pos == 2:
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist2 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist2.grid(row=7, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist2
        elif question_pos == 3:
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Frage
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Codieranweisung
            self.f_questions.rblist3 = Frame(self.f_questions, borderwidth=2, bg=farbton_text)
            self.f_questions.rblist3.grid(row=11, column=0, columnspan=3, sticky=E+W)
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            f = self.f_questions.rblist3

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]
        for k in range(len(namelist)):
            v = StringVar()
            if codelist[k] in previous_coding.keys():
                v.set(previous_coding[codelist[k]])
            else:
                v.set(defval)
            settings['Input'][question_pos-1].append(v)           
        anzahl = len(codebook[cb_var][2])
        scale = []
        labels = []
        buttons = []
        for i in range(len(scalelist)):
            sc = Label(f, text=scalelist[i])
            scale.append(sc)
            sc.grid(row=0,column=i+1)

        for k in range(len(namelist)):
            lab = Label(f,text=namelist[k])
            labels.append(lab)
            lab.grid(row=k+1,column=0,sticky=W)

        for i in range(len(scalelist)):
            for k in range(len(namelist)):
                rb = Radiobutton(f,variable=settings['Input'][question_pos-1][k], value=valuelist[i], command=self.rb_tamper)
                buttons.append(rb)
                rb.grid(row=k+1,column=i+1)


    def question_bt(self, cb_var, question_pos): #Button-Question: Up to four buttons may be defined in the codebook
        log('--Question: Buttons. Variable: '+cb_var+'; Position: '+str(question_pos),pos=0)
        auswliste = codebook[cb_var][2]
        codeliste = codebook[cb_var][3]
        anzahl = len(codebook[cb_var][2])
        if question_pos == 1:
            settings['Curr_Page'][0] = ['bt',cb_var]
            self.f_questions.Frage1.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage1.insert(END, codebook[cb_var][1]) #Coder information
            settings[cb_var] = StringVar(self)            
            if anzahl > 0:
                self.f_questions.bu1_1 = Button(self.f_questions, text=auswliste[0], width=20, command=CMD(self.submit,codeliste[0]))
                self.f_questions.bu1_1.grid(row=3,column=1, sticky = W)
            if anzahl > 1:
                self.f_questions.bu1_2 = Button(self.f_questions, text=auswliste[1], width=20, command=CMD(self.submit,codeliste[1]))
                self.f_questions.bu1_2.grid(row=3,column=2, sticky = W)
            if anzahl > 2:
                self.f_questions.bu1_3 = Button(self.f_questions, text=auswliste[2], width=20, command=CMD(self.submit,codeliste[2]))
                self.f_questions.bu1_3.grid(row=3,column=3, sticky = W)

            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=4, sticky=W)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))

        if question_pos == 2:
            settings['Curr_Page'][1] = ['bt',cb_var]
            self.f_questions.Frage2.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage2.insert(END, codebook[cb_var][1]) #Coder Information
            settings[cb_var] = StringVar(self)
                        
            if anzahl > 0:
                self.f_questions.bu2_1 = Button(self.f_questions, text=auswliste[0], width=20, command=CMD(self.submit,codeliste[0]))
                self.f_questions.bu2_1.grid(row=7,column=1, sticky = W)
            if anzahl > 1:
                self.f_questions.bu2_2 = Button(self.f_questions, text=auswliste[1], width=20, command=CMD(self.submit,codeliste[1]))
                self.f_questions.bu2_2.grid(row=7,column=2, sticky = W)
            if anzahl > 2:
                self.f_questions.bu2_3 = Button(self.f_questions, text=auswliste[2], width=20, command=CMD(self.submit,codeliste[2]))
                self.f_questions.bu2_3.grid(row=7,column=3, sticky = W)

            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=4, sticky=W)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))

        if question_pos == 3:
            settings['Curr_Page'][2] = ['bt',cb_var]
            self.f_questions.Frage3.insert(INSERT,codebook[cb_var][0], 'fett') #Question
            self.f_questions.Frage3.insert(END, codebook[cb_var][1]) #Coder Information
            settings[cb_var] = StringVar(self)
            
            if anzahl > 0:
                self.f_questions.bu3_1 = Button(self.f_questions, text=auswliste[0], width=20, command=CMD(self.submit,codeliste[0]))
                self.f_questions.bu3_1.grid(row=11, column=1, sticky = W)
            if anzahl > 1:
                self.f_questions.bu3_2 = Button(self.f_questions, text=auswliste[1], width=20, command=CMD(self.submit,codeliste[1]))
                self.f_questions.bu3_2.grid(row=11, column=2, sticky = W)
            if anzahl > 2:
                self.f_questions.bu3_3 = Button(self.f_questions, text=auswliste[2], width=20, command=CMD(self.submit,codeliste[2]))
                self.f_questions.bu3_3.grid(row=11, column=3, sticky = W)

            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=4, sticky=W)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))


    def question_ls(self,cb_var,liste,multi=1):
        log('--Question: Listselection. Variable: '+cb_var+'; List: '+liste,pos=0)
        global prog_pos
        predlist = []
        if settings['AEGLOS'] == '1':
            wert = -1
            pred = acabc.predict_short(cb_var,settings['Fulltext'],top=5)
            for item in pred:
                wert = self.namegetter(liste,item)
                predlist.append(wert)
            if len(predlist) > 0:
                if not cb_var in def_val.keys():
                    def_val[cb_var] = predlist
                    if settings['Verbose'] == '1':
                        verb('Automated value detection')
                        verb('Best prediction: '+str(predlist))
                else:
                    verb('Another default value has been set already. No changes made')
  
        settings['Curr_Page'][0] = ['list',cb_var,liste]
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = [] ###Has to have the same data type as set default values or previous codings

        if type(previous_coding) == tuple:
            previous_coding = previous_coding[0]            

        if type(previous_coding) == list:
            outlist = []
            for element in previous_coding:
                if element in codebook[liste][2]:
                    outlist.append(element)
                elif len(self.namegetter(liste,element)) > 0:
                    outlist.append(self.namegetter(liste,element))
                else:
                    verb('ERROR: Value '+element+' not found in list')
            previous_coding = outlist
        elif type(previous_coding) == str:
            if len(self.namegetter(liste,previous_coding)) >0:
                previous_coding = [self.namegetter(liste,previous_coding)]
            elif previous_coding in codebook[liste][2]:
                previous_coding = [str(previous_coding)]
            else:
                previous_coding = []
                verb('No valid coding found')
        else:
            previous_coding = []
            verb('No valid coding found')
            
        verb('Previous coding: '+str(previous_coding))

        self.f_questions.Frage1.insert(INSERT, codebook[cb_var][0],'fett') #Question
        self.f_questions.Frage1.insert(INSERT, codebook[cb_var][1])
        if multi == 1:
            self.f_questions.Aspliste = Listbox(self.f_questions,height=10,width=80, selectmode=MULTIPLE)
        else:
            self.f_questions.Aspliste = Listbox(self.f_questions,height=10,width=80, selectmode=BROWSE)
        self.f_questions.Aspliste.grid(row=3, column=0, columnspan=5, sticky=W+E)
        self.f_questions.Aspliste.focus()
        self.f_questions.scroll_AspListe = Scrollbar(self.f_questions, orient=VERTICAL, command=self.f_questions.Aspliste.yview)
        self.f_questions.scroll_AspListe.grid(row=3, column=5, sticky=W+N+S)
        self.f_questions.Aspliste["yscrollcommand"] = self.f_questions.scroll_AspListe.set
        for element in previous_coding:
            self.f_questions.Aspliste.insert(END,element)
        if len(previous_coding) > 0:
            self.f_questions.Aspliste.insert(END,'****')
        if len(previous_coding) == 1:
            self.f_questions.Aspliste.selection_set(0)
        for element in codebook[liste][2]:
            self.f_questions.Aspliste.insert(END,element)
        self.f_questions.h_Aspliste = Label(self.f_questions, text="?")
        self.f_questions.h_Aspliste.grid(row=4, column=5, sticky=N+E)
        self.f_questions.h_Aspliste.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
        self.f_bottomline.b_check.grid()
        
    def cut_entry(self,wtup):
        log('Calling Function: Cut_Entry')
        ##Fragmentation of a string using punctuation and spaces as cut-points
        outliste = []
        seg = ''
        for i in range(0,len(wtup)):
            if wtup[i] in [' ','\t',',',';']:
                if len(seg) > 0:
                    outliste.append(seg)
                seg = ''
            else:
                seg = seg + wtup[i]
        outliste.append(seg)
        return outliste
    
    def cutback_list(self,liste,auffang=0,broadseek=0): #Reducing the list according to input
        log('Calling Function: Cutback_List')
        if type(auffang) == str:
            suchstring = auffang
        elif auffang.char == chr(8): ##Erase-Key
            suchstring = bereinigen(self.f_questions.seektext.get()[:-1])
        else:
            suchstring = bereinigen(self.f_questions.seektext.get()) + auffang.char
        self.f_questions.Aspliste.delete(0,END)
        sl = self.cut_entry(suchstring)
        for element in codebook[liste][2]:
            inside = 1
            for s in sl:
                if not bereinigen(s,1) in bereinigen(element,1):
                    inside = 0
            if inside == 1:
                self.f_questions.Aspliste.insert(END,element)

        if self.f_questions.Aspliste.size() <1 and broadseek==1:
            for element in codebook[liste][2]:
                inside = 0
                for s in sl:
                    if bereinigen(s,1) in bereinigen(element,1):
                        inside = 1
                if inside == 1:
                    self.f_questions.Aspliste.insert(END,element)

        if self.f_questions.Aspliste.size() == 1:
            self.f_questions.Aspliste.selection_set(0)
            

    def question_lseek(self,cb_var,liste,multi=0):
        log('--Question: List Seek. Variable: '+cb_var+'; List: '+liste,pos=0)
        global prog_pos
        predlist = []
        if settings['AEGLOS'] == '1':
            wert = -1
            pred = acabc.predict_short(cb_var,settings['Fulltext'],top=5)
            for item in pred:
                wert = self.namegetter(liste,item)
                predlist.append(wert)
            if len(predlist) > 0:
                if not cb_var in def_val.keys():
                    def_val[cb_var] = predlist
                    if settings['Verbose'] == '1':
                        verb('Automated value detection')
                        verb('Best prediction: '+str(predlist))
                else:
                    verb('Another default value has been set already. No changes made')
        
        settings['Curr_Page'][0] = ['listseek',cb_var,liste]
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var][0]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+': '+str(previous_coding))
        else:
            previous_coding = [] ###Has to have the same data type as set default values or previous codings

        if type(previous_coding) == tuple:
            previous_coding = previous_coding[0]
            
        if type(previous_coding) == list:
            outlist = []
            for element in previous_coding:
                if element in codebook[liste][2]:
                    outlist.append(element)
                elif len(self.namegetter(liste,element)) > 0:
                    outlist.append(self.namegetter(liste,element))
                else:
                    verb('ERROR: Value '+element+' not found in list')
            previous_coding = outlist
        elif type(previous_coding) == str:
            if len(self.namegetter(liste,previous_coding)) >0:
                previous_coding = [self.namegetter(liste,previous_coding)]
            elif previous_coding in codebook[liste][2]:
                previous_coding = [str(previous_coding)]
            else:
                previous_coding = []
                verb('No valid coding found')
        else:
            previous_coding = []
            verb('No valid coding found')
            
        self.f_questions.Frage1.insert(INSERT, codebook[cb_var][0],'fett') #Question
        self.f_questions.Frage1.insert(INSERT, codebook[cb_var][1])

        self.f_questions.seektext = Entry(self.f_questions,width=80)
        self.f_questions.seektext.grid(row=2,column=0,columnspan=5,sticky=W+E)
        self.f_questions.seektext.bind('<Key>', CMD(self.cutback_list, liste))
        
        if multi == 1:
            self.f_questions.Aspliste = Listbox(self.f_questions,height=10,width=80, selectmode=MULTIPLE)
        else:
            self.f_questions.Aspliste = Listbox(self.f_questions,height=10,width=80, selectmode=BROWSE)
        self.f_questions.Aspliste.grid(row=3, column=0, columnspan=5, sticky=W+E)
        self.f_questions.scroll_AspListe = Scrollbar(self.f_questions, orient=VERTICAL, command=self.f_questions.Aspliste.yview)
        self.f_questions.scroll_AspListe.grid(row=3, column=5, sticky=W+N+S)
        self.f_questions.Aspliste["yscrollcommand"] = self.f_questions.scroll_AspListe.set
        for element in previous_coding:
            self.f_questions.Aspliste.insert(END,element)
        if len(previous_coding) > 0:
            self.f_questions.Aspliste.insert(END,'****')
        if len(previous_coding) == 1:
            self.f_questions.Aspliste.selection_set(0)
        for element in codebook[liste][2]:
            self.f_questions.Aspliste.insert(END,element)
        self.f_questions.h_Aspliste = Label(self.f_questions, text="?")
        self.f_questions.h_Aspliste.grid(row=4, column=5, sticky=N+E)
        self.f_questions.h_Aspliste.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
        self.f_questions.seektext.focus()
        self.f_bottomline.b_check.grid()

    def list_add(self,spillover=0):
        log('Calling Function: List Add')
        if spillover == 1:
            for element in self.f_questions.Aspliste.get(0,END):
                self.f_questions.Itmliste.insert(END,element)
        else:
            listsel = self.f_questions.Aspliste.curselection()
            if len(listsel)>0:
                for selection in listsel:
                    self.f_questions.Itmliste.insert(END,self.f_questions.Aspliste.get(selection))
                    self.f_questions.Aspliste.selection_clear(selection)
                self.f_questions.seektext.delete(0,END)
            elif len(bereinigen(self.f_questions.seektext.get())) > 0:
                self.f_questions.Itmliste.insert(END,bereinigen(self.f_questions.seektext.get()))
                self.f_questions.seektext.delete(0,END)
            else:
                self.message('Invalid-Selection01')

    def list_rem(self,spillover=0):
        log('Calling Function: List Remove')
        if spillover == 1:
            self.f_questions.Itmliste.delete(0,END)
        else:
            listsel = self.f_questions.Itmliste.curselection()
            if len(listsel)==1:
                self.f_questions.Itmliste.delete(listsel[0])
            else:
                self.message('Invalid-Selection01')
        
    def question_ladd(self,cb_var,liste,multi=1):
        log('--Question: List Add. Variable: '+cb_var+'; List: '+liste,pos=0)
        global prog_pos
        predlist = []
        if settings['AEGLOS'] == '1':
            wert = -1
            pred = acabc.predict_short(cb_var,settings['Fulltext'],top=5)
            for item in pred:
                wert = self.namegetter(liste,item)
                predlist.append(wert)
            if len(predlist) > 0:
                if not cb_var in def_val.keys():
                    def_val[cb_var] = predlist
                    if settings['Verbose'] == '1':
                        verb('Automated value detection')
                        verb('Best prediction:'+str(predlist))
                else:
                    verb('Another default value has been set already. No changes made')
        
        settings['Curr_Page'][0] = ['listadd',cb_var,liste]
        curr_tree = curr()

        if cb_var in curr_tree.keys():
            previous_coding = curr_tree[cb_var][0]
            verb('Previous coding found for '+cb_var+': '+str(previous_coding))
        elif cb_var in def_val.keys():
            previous_coding = def_val[cb_var]
            verb('Set default values found for '+cb_var+' (Type: '+str(type(previous_coding))+'): '+str(previous_coding))
        else:
            previous_coding = [] ###Has to have the same data type as set default values or previous codings

        if type(previous_coding) == tuple:
            previous_coding = previous_coding[0]
            
        if type(previous_coding) == list:
            outlist = []
            for element in previous_coding:
                if element in codebook[liste][2]:
                    outlist.append(element)
                elif len(self.namegetter(liste,element)) > 0:
                    outlist.append(self.namegetter(liste,element))
                else:
                    verb('ERROR: Value '+element+' not found in list')
            previous_coding = outlist
        elif type(previous_coding) == str:
            if len(self.namegetter(liste,previous_coding)) >0:
                previous_coding = [self.namegetter(liste,previous_coding)]
            elif previous_coding in codebook[liste][2]:
                previous_coding = [str(previous_coding)]
            else:
                previous_coding = []
                verb('No valid coding found')
        else:
            previous_coding = []
            verb('No valid coding found')

        self.f_questions.Frage1.insert(INSERT, codebook[cb_var][0],'fett') #Question
        self.f_questions.Frage1.insert(INSERT, codebook[cb_var][1])

        self.f_questions.seektext = Entry(self.f_questions,width=80)
        self.f_questions.seektext.grid(row=2,column=0,columnspan=5,sticky=W+E)
        self.f_questions.seektext.bind('<Key>', CMD(self.cutback_list, liste))
        self.f_questions.seektext.bind('<Return>', CMD(self.list_add))
        
        if multi == 1:
            self.f_questions.Aspliste = Listbox(self.f_questions,height=7,width=80, selectmode=MULTIPLE)
        else:
            self.f_questions.Aspliste = Listbox(self.f_questions,height=7,width=80, selectmode=BROWSE)
        self.f_questions.Aspliste.grid(row=3, column=0, columnspan=5, sticky=W+E)
        self.f_questions.scroll_AspListe = Scrollbar(self.f_questions, orient=VERTICAL, command=self.f_questions.Aspliste.yview)
        self.f_questions.scroll_AspListe.grid(row=3, column=5, sticky=W+N+S)
        self.f_questions.Aspliste["yscrollcommand"] = self.f_questions.scroll_AspListe.set

        self.f_questions.adb = Button(self.f_questions,text='Add Item',command=self.list_add)
        self.f_questions.adb.grid(row=4,column=0)
        self.f_questions.adb2 = Button(self.f_questions,text='Add All',command=CMD(self.list_add,1))
        self.f_questions.adb2.grid(row=4,column=1)
        self.f_questions.rb = Button(self.f_questions,text='Remove Item',command=self.list_rem)
        self.f_questions.rb.grid(row=4,column=2)
        self.f_questions.rb2 = Button(self.f_questions,text='Remove All',command=CMD(self.list_rem,1))
        self.f_questions.rb2.grid(row=4,column=3)

        self.f_questions.Itmliste = Listbox(self.f_questions,height=7,width=80, selectmode=BROWSE)
        self.f_questions.Itmliste.grid(row=5, column=0, columnspan=5, sticky=W+E)
        self.f_questions.scroll_ItmListe = Scrollbar(self.f_questions, orient=VERTICAL, command=self.f_questions.Itmliste.yview)
        self.f_questions.scroll_ItmListe.grid(row=5, column=5, sticky=W+N+S)
        self.f_questions.Itmliste["yscrollcommand"] = self.f_questions.scroll_ItmListe.set

        for element in previous_coding:
            self.f_questions.Itmliste.insert(END,element)
        for element in codebook[liste][2]:
            self.f_questions.Aspliste.insert(END,element)

        self.f_questions.h_Aspliste = Label(self.f_questions, text="?")
        self.f_questions.h_Aspliste.grid(row=4, column=5, sticky=N+E)
        self.f_questions.h_Aspliste.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
        self.f_questions.seektext.focus()
        self.f_bottomline.b_check.grid()

    def question_menu(self, cb_var, question_pos): #Menu-Question: Up to three menu selections may be displayed per page
        global settings
        log('--Question: Menu. Variable: '+cb_var+'; Position: '+str(question_pos),pos=0)
        settings['Curr_Page'][question_pos-1] = ['menu',cb_var]
        settings['Input'][question_pos-1] = ''
        curr_tree=curr()

        namelist = codebook[cb_var][2]
        codelist = codebook[cb_var][3]
        
        settings['Input'][question_pos-1] = StringVar()
        settings['Input'][question_pos-1].set(namelist[0])

        if question_pos == 1:
            self.f_questions.Frage1.insert(END, codebook[cb_var][1],'fett') #Display additional information
            self.f_questions.help1 = Label(self.f_questions, text="?")
            self.f_questions.help1.grid(row=3, column=3, sticky=E)
            self.f_questions.help1.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.mb1 = Menubutton(self.f_questions, text=codebook[cb_var][0][:-1],relief=RAISED)
            self.f_questions.mb1.grid(row=3,column=1,sticky=W+E)
            self.f_questions.menu1 = Menu(self.f_questions.mb1,tearoff=0)
            self.f_questions.mb1["menu"] = self.f_questions.menu1
            m = self.f_questions.menu1
               
        if question_pos == 2:
            self.f_questions.Frage2.insert(END, codebook[cb_var][1],'fett') #Display additional information
            self.f_questions.help2 = Label(self.f_questions, text="?")
            self.f_questions.help2.grid(row=7, column=3, sticky=E)
            self.f_questions.help2.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.mb2 = Menubutton(self.f_questions, text=codebook[cb_var][0][:-1],relief=RAISED)
            self.f_questions.mb2.grid(row=7,column=1,sticky=W+E)
            self.f_questions.menu2 = Menu(self.f_questions.mb2,tearoff=0)
            self.f_questions.mb2["menu"] = self.f_questions.menu2
            m = self.f_questions.menu2

        if question_pos == 3:
            self.f_questions.Frage3.insert(END, codebook[cb_var][1],'fett') #Display additional information
            self.f_questions.help3 = Label(self.f_questions, text="?")
            self.f_questions.help3.grid(row=11, column=3, sticky=E)
            self.f_questions.help3.bind('<Button-1>', CMD(self.hilfe_zu, codebook[cb_var][4]))
            self.f_questions.mb3 = Menubutton(self.f_questions, text=codebook[cb_var][0][:-1],relief=RAISED)
            self.f_questions.mb3.grid(row=11,column=1,sticky=W+E)
            self.f_questions.menu3 = Menu(self.f_questions.mb3,tearoff=0)
            self.f_questions.mb3["menu"] = self.f_questions.menu3
            m = self.f_questions.menu3

        curcas = Menu(m,tearoff=0)
        curlab = ''
        lastone = -1
        for i in range(len(codelist)-1):
            if codelist[i][0] == '*':
                if settings['Debugging'] in [1,'1']:
                    codelist[i] = codelist[i][1:]
                    verb('New Code: '+str(codelist[i]))
                    
            if not codelist[i][0] == '*': 
                if codelist[i][0] == '-':
                    curcas.add_command(label=namelist[i],command=CMD(self.submit,codelist[i][1:]))
                elif codelist[i][0] == '#':
                    if len(curlab) > 1:
                        m.add_cascade(label=curlab,menu=curcas)
                        curlab = ''
                    m.add_separator()
                else:
                    if len(curlab) > 1:
                        m.add_cascade(label=curlab,menu=curcas)
                    if codelist[i+1][0] == '-' or (codelist[i+1][:2]=='*-' and settings['Debugging'] in [1,'1']):
                        curcas = Menu(m,tearoff=0)
                        curlab = namelist[i]
                    else:
                        m.add_command(label=namelist[i],command=CMD(self.submit,codelist[i]))
                        curlab = ''
                lastone = i
            else:
                verb('OMIT: '+str(codelist[i]))

        if codelist[lastone][0] == '-':
            curcas.add_command(label=namelist[lastone],command=CMD(self.submit,codelist[lastone][1:]))
            m.add_cascade(label=curlab,menu=curcas)
        else:
            if len(curlab) > 1:
                m.add_cascade(label=curlab,menu=curcas)
            m.add_command(label=namelist[lastone],command=CMD(self.submit,codelist[lastone]))


    def question_mark_units(self,cb_var,Einheit):
        log('--Question: Mark Units. Variable: '+cb_var+'; Level: '+Einheit,pos=0)
        self.question_bt(cb_var,1)
        self.f_questions.bu1_1["command"]=self.submit
        if Einheit in settings.keys():
            for bez in settings[Einheit].keys():
                if settings[Einheit][bez]['Done'] == 0:
                    self.Artikel.tag_add(settings[Einheit][bez]['Typ'],settings[Einheit][bez]['Start'],settings[Einheit][bez]['End'])
       
    def units_action(self,Einheit,action='none',sel_tags=[]):
        log('Calling Function: Units Action')
        global prog_pos
        global dta_pos
        if action == 'add':
            self.clean_up_all()
            prog_pos='s_markieren'
            self.ask()
        elif action == 'rem':
            selection = self.f_questions.Aspliste.curselection()[0]
            label = self.f_questions.Aspliste.get(selection)
            verb('List Selection: '+str(label))
            for fk in settings[Einheit].keys():
                if label == settings[Einheit][fk]['Label']:
                    del settings[Einheit][fk]
            self.clean_up_all()
            self.ask()
        elif action == 'mark':
            selection = self.f_questions.Aspliste.curselection()[0]
            label = self.f_questions.Aspliste.get(selection)
            verb('List Selection: '+str(label))
            for fk in settings[Einheit].keys():
                if label == settings[Einheit][fk]['Label']:
                    self.clean_all_tags(settings[Einheit][fk]['Typ'])
                    self.Artikel.tag_add(settings[Einheit][fk]['Typ'],settings[Einheit][fk]['Start'],settings[Einheit][fk]['End'])
        else:
            verb('Warning: No function tied to this button')


    def question_sel_units(self,var,Einheit):
        log('--Question: Select Units. Variable: '+var+'; Level: '+Einheit,pos=0)
        global prog_pos
        global dta_pos
        unit_dic = settings[Einheit]
        settings['Curr_Page'][0] = ['unit_auswahl',Einheit]
        verb('Available Units:'+str(sorted(unit_dic.keys())))

        self.f_questions.Frage1.insert(INSERT, codebook[var][0],'fett') #Asking for a list
        self.f_questions.Frage1.insert(INSERT, codebook[var][1]) 
        self.f_questions.Aspliste = Listbox(self.f_questions,height=5,width=80, selectmode=BROWSE)
        self.f_questions.Aspliste.grid(row=3, column=0, columnspan=5, sticky=W+E)
        self.f_questions.Aspliste.focus()
        self.f_questions.scroll_AspListe = Scrollbar(self.f_questions, orient=VERTICAL, command=self.f_questions.Aspliste.yview)
        self.f_questions.scroll_AspListe.grid(row=3, column=5, sticky=W+N+S)
        self.f_questions.Aspliste["yscrollcommand"] = self.f_questions.scroll_AspListe.set
        anz = 0
        for u in sorted(unit_dic.keys()):
            if unit_dic[u]['Done'] == 0:
                anz = anz + 1
                self.f_questions.Aspliste.insert(END,unit_dic[u]['Label'])
        self.f_questions.h_Aspliste = Label(self.f_questions, text="?")
        self.f_questions.h_Aspliste.grid(row=4, column=5, sticky=N+E)
        self.f_questions.h_Aspliste.bind('<Button-1>', CMD(self.hilfe_zu, codebook[var][4]))
        self.f_bottomline.b_check.grid()
        
        self.f_questions.Aspliste.selection_set('0')

        self.f_questions.fk_hinzu = Button(self.f_questions, text = codebook[var][2][0], width=20,command=CMD(self.units_action,Einheit,'add'))
        self.f_questions.fk_hinzu.grid(row=4,column=1,sticky=W+E)
        self.f_questions.fk_weg = Button(self.f_questions, text = codebook[var][2][1], width=20,command=CMD(self.units_action,Einheit,'rem'))
        self.f_questions.fk_weg.grid(row=4,column=2,sticky=W+E)
        self.f_questions.fk_markieren = Button(self.f_questions, text = codebook[var][2][2], width=20,command=CMD(self.units_action,Einheit,'mark'))
        self.f_questions.fk_markieren.grid(row=4,column=3,sticky=W+E)

        if anz == 0:
            self.clean_up_all()
            prog_pos = 'last_review'
            self.ask()




############################################
##                                        ##
##       Speicherroutinen                 ##
##                                        ##
############################################

    def export_data(self, dta_pos_all, varlist, filename,debug=0):
        try:
            a = open(filename,'r')
            a.close()
        except:
            verb('No storage file yet. Creating new file')
            a = open(filename,'w')
            a.write('Coder\tID\t')
            for d in dta_pos_all:
                a.write('Level\tUnit_ID\t')
            for v in varlist:
                if v in settings['Multi_Items']:
                    for code in codebook[v][3]:
                        a.write(v+'_'+code+'\t')
                else:
                    a.write(v+'\t')
            a.write('\n')
            a.close()
            
        log('Calling Function: Export_Data of all elements in: '+ str(dta_pos_all))
        if len(dta_pos_all) == 3:
            verb('Fourth Level of Analysis')
            if dta_pos_all[0] in storage.keys():
                for i in storage[dta_pos_all[0]].keys():
                    if dta_pos_all[1] in storage[dta_pos_all[0]][i].keys():
                        for k in storage[dta_pos_all[0]][i][dta_pos_all[1]].keys():
                            if dta_pos_all[2] in storage[dta_pos_all[0]][i][dta_pos_all[1]][k].keys():
                                for l in storage[dta_pos_all[0]][i][dta_pos_all[1]][k][dta_pos_all[2]].keys():
                                    direc = storage[dta_pos_all[0]][i][dta_pos_all[1]][k][dta_pos_all[2]][l]
                                    exp_file = open(filename, 'a')
                                    exp_file.write(settings['Coder'])
                                    exp_file.write('\t')
                                    exp_file.write(storage['ID'])
                                    exp_file.write('\t')
                                    exp_file.write(dta_pos_all[0])
                                    exp_file.write('\t')
                                    exp_file.write(i)
                                    exp_file.write('\t')
                                    exp_file.write(dta_pos_all[1])
                                    exp_file.write('\t')
                                    exp_file.write(k)
                                    exp_file.write('\t')                            
                                    exp_file.write(dta_pos_all[2])
                                    exp_file.write('\t')
                                    exp_file.write(l)
                                    exp_file.write('\t')                            
                                    for var in varlist:
                                        self.var_export(exp_file, direc,var,debug)
                                    exp_file.write('\n')
                                    exp_file.close()

        elif len(dta_pos_all) == 2:
            verb('Third Level of Analysis')
            if dta_pos_all[0] in storage.keys():
                for i in storage[dta_pos_all[0]].keys():
                    if dta_pos_all[1] in storage[dta_pos_all[0]][i].keys():
                        for k in storage[dta_pos_all[0]][i][dta_pos_all[1]].keys():
                            direc = storage[dta_pos_all[0]][i][dta_pos_all[1]][k]
                            exp_file = open(filename, 'a')
                            exp_file.write(settings['Coder'])
                            exp_file.write('\t')
                            exp_file.write(storage['ID'])
                            exp_file.write('\t')
                            exp_file.write(dta_pos_all[0])
                            exp_file.write('\t')
                            exp_file.write(i)
                            exp_file.write('\t')
                            exp_file.write(dta_pos_all[1])
                            exp_file.write('\t')
                            exp_file.write(k)
                            exp_file.write('\t')                            
                            for var in varlist:
                                self.var_export(exp_file, direc,var,debug)
                            exp_file.write('\n')
                            exp_file.close()
                            
        elif len(dta_pos_all) == 1:
            verb('Second Level of Analysis')
            if dta_pos_all[0] in storage.keys():
                for i in storage[dta_pos_all[0]].keys():
                    direc = storage[dta_pos_all[0]][i]
                    exp_file = open(filename, 'a')
                    exp_file.write(settings['Coder'])
                    exp_file.write('\t')
                    exp_file.write(storage['ID'])
                    exp_file.write('\t')
                    exp_file.write(dta_pos_all[0])
                    exp_file.write('\t')
                    exp_file.write(i)
                    exp_file.write('\t')
                    for var in varlist:
                        self.var_export(exp_file, direc,var,debug)
                    exp_file.write('\n')
                    exp_file.close()
            else:
                verb('Level of Analysis not found: '+str(dta_pos_all[0]))
                    
        elif len(dta_pos_all) == 0:
            verb('First Level of Analysis (root-Level)')
            direc = storage
            exp_file = open(filename, 'a')
            exp_file.write(settings['Coder'])
            exp_file.write('\t')
            exp_file.write(storage['ID'])
            exp_file.write('\t')
            for var in varlist:
                self.var_export(exp_file, direc,var,debug)
            dauer = time.time()-storage['#TS'][1]
            dauer_net = dauer - settings['Break_Time']
            dauer_h = dauer_net/3600
            exp_file.write(str(dauer))
            exp_file.write('\t')              
            exp_file.write(str(settings['Break_Time']))
            exp_file.write('\t')              
            exp_file.write(str(dauer_net))
            exp_file.write('\t')              
            exp_file.write(str(dauer_h))
            exp_file.write('\n')
            exp_file.close()

    def var_export(self, exp_file, dictionary, variabel,debug=0):
        global codebook
        global settings
        verb('--Var Export: '+variabel)
        if variabel in dictionary.keys():
            if type(dictionary[variabel]) == tuple:
                if debug == 1: exp_file.write(variabel+"=")
                if not dictionary[variabel][1] == '':
                    exp_file.write(str(dictionary[variabel][1]))
                else:
                    exp_file.write(bereinigen(dictionary[variabel][0]))
                exp_file.write('\t')
            elif type(dictionary[variabel]) == str or type(dictionary[variabel]) == unicode:
                if debug == 1: exp_file.write(variabel+"=")
                exp_file.write(bereinigen(dictionary[variabel]))
                exp_file.write('\t')
            elif type(dictionary[variabel]) == dict:
                for item in codebook[variabel][3]:
                    if debug == 1: exp_file.write(variabel+'_'+item+"=")
                    if type(dictionary[variabel][item]) == str:
                        exp_file.write(str(dictionary[variabel][item]))
                        exp_file.write('\t')
                    elif type(dictionary[variabel][item]) == int:
                        exp_file.write(str(dictionary[variabel][item]))
                        exp_file.write('\t')
                    elif type(dictionary[variabel][item]) == tuple:
                        if not dictionary[variabel][item][1] == '':
                            exp_file.write(str(dictionary[variabel][item][1]))
                        else:
                            exp_file.write(str(dictionary[variabel][item][0]))
                        exp_file.write('\t')
            elif type(dictionary[variabel]) == int:
                if debug == 1: exp_file.write(variabel+"=")
                exp_file.write(str(dictionary[variabel]))
                exp_file.write('\t')
            else:
                if debug == 1: exp_file.write(variabel+"=")
                verb('Unknown Type: '+str(type(dictionary[variabel])))
                exp_file.write('\t')
        else:
            if variabel in settings['Multi_Items']:
                laenge = len(codebook[variabel][3])
                if debug == 1: exp_file.write(variabel+"=")
                for i in range(1,laenge):
                    exp_file.write('\t')
            else:
                if debug == 1: exp_file.write(variabel+"=")               
            exp_file.write('\t')   


############################################
##                                        ##
##       Hilfsfunktionen                  ##
##                                        ##
############################################

    def set_window(self):
        log('Setting the Window')
        #Set the window in which the query form is set up. There are five elements:
        #   1) Review-Frame: May be used to display a list of previously coded elements on this level.
        #   2) Location-Frame: Contains a text-area which may be used to display additional information on the current position.
        #   3) Explanation-Frame: Spare frame beneath the location frame (usually not used)
        #   4) Question-Frame: Frame which contains empty spaces for three questions. Questions may be called by self.question..()-functions.
        #   5) Button-Frame: Is used to display four buttons: Check, Abort, Back and Break. Each has a defined target and may be
        #                    displayed and removed by calling the self.buttons()-Function
        
        if settings['Layout'] == 'Lefty':
            c1 = 1
            butt_order = [0,1,2,3]
            verb('Left-Handed Design')
        else:
            c1 = 5
            butt_order = [3,2,1,0]
            verb('Right-Handed Design')
            
        #Remove everything to prevent redundancy
        try:
            self.f_review.destroy()
            self.f_location.destroy()
            self.f_explanation.destroy()
            self.f_questions.destroy()
            self.f_bottomline.destroy()
            verb("Window reset")         
        except:
            verb("First Window")

        #Build:   
        self.f_review = Frame(self, borderwidth=2, bg=farbton_text, height=20, relief=FLAT)
        self.f_review.grid(row=1, column=c1, sticky=N+E+S+W)
        self.f_location = Frame(self, borderwidth=2, bg=farbton_text)
        self.f_location.grid(row=2, column=c1, sticky=N+E+S+W)
        self.f_explanation = Frame(self, borderwidth=2, bg=farbton_text)
        self.f_explanation.grid(row=3, column=c1, sticky=N+E+S+W)
        self.f_questions = Frame(self, borderwidth=2, bg=farbton_text)
        self.f_questions.grid(row=4, column=c1, sticky=N+E+S+W)
        self.f_bottomline = Frame(self, borderwidth=2, bg=farbton_text)
        self.f_bottomline.grid(row=6, column=c1, sticky=E+S+W)

        self.f_location.angabe = Text(self.f_location, width=80, height=2, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_location.angabe.grid(row=0, column=0, columnspan=4, sticky=N+S+E+W)
        self.f_location.angabe.tag_config('highlight', font = (settings['Font'], "9", "italic"))
        self.f_location.angabe.tag_config('fett', font = (settings['Font'], "9", "bold"))

        self.f_questions.Frage1 = Text(self.f_questions, width=80, height=4, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_questions.Frage1.grid(row=1, column=0, columnspan=4, sticky=N+S+E+W)
        self.f_questions.Frage1.tag_config('fett', font = (settings['Font'], "10", "bold"))
        self.f_questions.Frage2 = Text(self.f_questions, width=80, height=4, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_questions.Frage2.grid(row=5, column=0, columnspan=4, sticky=N+S+E+W)
        self.f_questions.Frage2.tag_config('fett', font = (settings['Font'], "10", "bold"))
        self.f_questions.Frage3 = Text(self.f_questions, width=80, height=4, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_questions.Frage3.grid(row=9, column=0, columnspan=4, sticky=N+S+E+W)
        self.f_questions.Frage3.tag_config('fett', font = (settings['Font'], "10", "bold"))

        self.f_questions.spacer1 = Text(self.f_questions, width=80, height=1, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_questions.spacer1.grid(row=4, column=0, columnspan=4, sticky=S+E+W)
        self.f_questions.spacer2 = Text(self.f_questions, width=80, height=1, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_questions.spacer2.grid(row=8, column=0, columnspan=4, sticky=S+E+W)
        self.f_questions.spacer3 = Text(self.f_questions, width=80, height=1, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg=farbton_text, takefocus = 0)
        self.f_questions.spacer3.grid(row=12, column=0, columnspan=4, sticky=S+E+W)

        for i in range(0,5):
            self.f_bottomline.columnconfigure(i, minsize=40)

        self.f_bottomline.b_check = Button(self.f_bottomline, text = "Check", width=20, command=self.submit)
        self.f_bottomline.b_check.grid(row=0, column=butt_order[0], sticky=N+S+E+W)
        self.f_bottomline.b_abort = Button(self.f_bottomline, text = "-", width=20, command=self.abort)
        self.f_bottomline.b_abort.grid(row=0, column=butt_order[1], sticky=N+S+E+W)
        self.f_bottomline.b_back = Button(self.f_bottomline, text = "Back", width=15, command=self.back, takefocus = 0)
        self.f_bottomline.b_back.grid(row=0, column=butt_order[2], sticky=N+S+W)

        self.f_bottomline.b_break = Button(self.f_bottomline, text = "Break", width=15, command=self.pause, takefocus = 0)
        self.f_bottomline.b_break.grid(row=0, column=butt_order[3], sticky=N+S+W)

        self.f_bottomline.b_check.grid_remove()
        self.f_bottomline.b_abort.grid_remove()
        self.f_bottomline.b_back.grid_remove()

        self.deb = Menubutton(self, text="?",relief=RAISED)
        self.deb.grid(row=0,column=1,sticky=N+W)
        self.deb.menu = Menu(self.deb, tearoff=0)
        self.deb["menu"] = self.deb.menu
        if settings['Debugging']=='0':
            self.deb.grid_remove()

        self.deb.menu.add_command(label="Where am I?",command=self.show_parameters)
        self.deb.menu.add_command(label="How did I get here?",command=self.show_path)
        self.deb.menu.add_command(label="What just happened?",command=self.show_verb)       
        self.deb.menu.add_command(label="What did I code?",command=self.show_storage)
        self.deb.menu.add_command(label="How are you set up?",command=self.show_settings)
        self.deb.menu.add_command(label="IT CRASHED! Take me back to somewhere safe",command=self.reset_coding)
        #self.deb.menu.add_command(label="test",command=self.display_text)


##        self.quick = Menubutton(self,text="Inspect data", relief=RAISED)
##        self.quick.grid(row=0,column=1,sticky=N+E+W)
##        self.quick.menu = Menu(self.quick, tearoff=0)
##        self.quick["menu"] = self.quick.menu
##
##        self.quick.menu.add_command(label="Variable overview",command=self.show_variables)
##        self.quick.menu.add_command(label="Descriptives of one variable",command=self.show_descriptives)
##        
                        
        self.ask()

    def test_styleset(self):
        verbout('\n',master=self)
        verbout('This is an example output','title',master=self)
        verbout('\n\nJust to test the styleset, all types are tested.\n\n','text',master=self)
        verbout('ERROR: This is a warning message.','warning',master=self)
        verbout('\n\nAnd this is a progress bar:\n\n',master=self)
        verbout('Checking for duplicates: \n0%-------25%-------50%-------75%-------100%\n..............\n','progress',master=self)
        verbout('\n\n\nAnd this is a table: \n\n',master=self)
        verbout('Col1     Col2      Col3\n  30     Blub     Hallo\n 100       AB      Welt\n','table',master=self)
        

    def debug_on(self,event=0):
        self.deb.grid()


    def select_sheet(self,sheetlist):
        self.sheet = Toplevel(self)
        self.sheet.title("Select Sheet within Excel Workbook")
        l = Label(self.sheet,text="Please select a sheet within this workbook")
        l.grid(row=0,column=0)

        b = Button(self.sheet,text="OK",command=self.confirm_sheet)
        b.grid(row=2,column=0)


    def confirm_sheet(self,event=0):
        global settings

        settings['Current_Sheet']="Hallo"
        self.sheet.destroy()


    def show_variables(self):
        self.infobox = Toplevel(self)
        self.infobox.rowconfigure(1, weight=1)
        self.infobox.columnconfigure(1, weight=1)
        self.infobox.title("List of currently loaded variables")
        self.infobox.ysc = Scrollbar(self.infobox, orient=VERTICAL)
        self.infobox.ysc.grid(row=1,column=2,sticky=W+N+S)
        self.infobox.info = Text(self.infobox,width=100,height=30,bg="#ffffcc",wrap=WORD, yscrollcommand=self.infobox.ysc.set,font=("Arial",9),takefocus=0)
        self.infobox.info.grid(row=1,column=1,sticky=N+E+S+W)
        self.infobox.ysc["command"]=self.infobox.info.yview
        self.infobox.info.tag_config('errormsg',background="#ffcccc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('functcall',foreground="#0000cc", font = ("Arial",9, "bold"))

        self.infobox.info.tag_config('errormsg',background="#ffcccc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('functcall',foreground="#0000cc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('important',foreground="#0000cc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('table',foreground="#000000", font = ("Courier",9, "bold"))
        self.infobox.info.tag_config('title',foreground="#000000", font = ("Arial",11, "bold"))

        ### Get all variables

        if len(settings['Datasets'].keys()) == 0:
            pass
        else:
            for ds in sorted(settings['Datasets'].keys()):
                self.infobox.info.insert(END,'\nVariables in Dataset: '+ds+'\n','title')     
                dset = storage[settings['Datasets'][ds]['Data']]
                dvar = storage[settings['Datasets'][ds]['Var']]
                vartable = {'Missing':{},'Strings':{},'Integers':{},'Decimals':{}}
                for v in dvar:
                    #self.infobox.info.insert(END,'\n'+v)
                    varinfo = stat_type(dset[v])
                    vartable['Missing'][v] = varinfo['Type_Missing']
                    vartable['Strings'][v] = varinfo['Type_String']
                    vartable['Integers'][v] = varinfo['Type_Int']
                    vartable['Decimals'][v] = varinfo['Type_Float']

                    cp = ['Missing','Strings','Integers','Decimals']
                    rp = dvar

                self.infobox.info.insert(END,'\n'+display_table(vartable,cols_pre=cp, rows_pre=rp),'table')
                    
 

    def show_descriptives(self):
        self.infobox = Toplevel(self)
        self.infobox.rowconfigure(1, weight=1)
        self.infobox.columnconfigure(1, weight=1)
        self.infobox.title("Descriptives of one variable")
        self.infobox.ysc = Scrollbar(self.infobox, orient=VERTICAL)
        self.infobox.ysc.grid(row=1,column=2,sticky=W+N+S)

        self.infobox.select = Menubutton(self.infobox,text="Select Dataset and Variable", relief=RAISED)
        self.infobox.select.grid(row=0,column=1,sticky=N+W+E)
        self.infobox.select.menu = Menu(self.infobox.select, tearoff=0)
        self.infobox.select["menu"] = self.infobox.select.menu

        ### Get all variables

        varlist = {}

        if len(settings['Datasets'].keys()) == 0:
            pass
        else:
            for ds in sorted(settings['Datasets'].keys()):
                varlist[ds] = []
                dset = storage[settings['Datasets'][ds]['Data']]
                dvar = storage[settings['Datasets'][ds]['Var']]
                for v in dvar:
                    if len(dset[v])>0:
                        varlist[ds].append(v)
                if len(varlist[ds]) == 0:
                    del varlist[ds]

        if varlist == {}:
            self.infobox.select.menu.add_command(label="No data loaded",command=self.show_descriptives)
        else:
            m = self.infobox.select.menu
            for ds in sorted(varlist.keys()):
                curcas = Menu(m,tearoff=0)
                for var in varlist[ds]:
                    curcas.add_command(label=var,command=CMD(self.display_descriptives_add,[ds,var]))
                self.infobox.select.menu.add_cascade(label=ds, menu=curcas)

        
        self.infobox.info = Text(self.infobox,width=100,height=30,bg="#ffffcc",wrap=WORD, yscrollcommand=self.infobox.ysc.set,font=("Arial",9),takefocus=0)
        self.infobox.info.grid(row=1,column=1,sticky=N+E+S+W)
        self.infobox.ysc["command"]=self.infobox.info.yview

        self.infobox.info.tag_config('errormsg',background="#ffcccc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('functcall',foreground="#0000cc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('important',foreground="#0000cc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('table',foreground="#000000", font = ("Courier",9, "bold"))
        self.infobox.info.tag_config('title',foreground="#000000", font = ("Arial",11, "bold"))

        content = '<select variable for descriptives above>'
        
        self.infobox.info.insert(END,content)
        pass
    

    def display_descriptives_add(self,optlist):
        self.infobox.info.delete('1.0',END)
        self.infobox.info.insert(END,'\nDataset: ')
        self.infobox.info.insert(END,optlist[0],'important')
        self.infobox.info.insert(END,' / Variable: ')
        self.infobox.info.insert(END,optlist[1],'important')
        self.infobox.info.insert(END,'\n----------------------------------------------------------------------------\n\n')

        varinfo = inspect_variable(storage[settings['Datasets'][optlist[0]]['Data']][optlist[1]])

        self.infobox.info.insert(END,'Values:','title')   
        self.infobox.info.insert(END,'\n\nTotal Number of Values: '+str(varinfo['N_Total']))
        self.infobox.info.insert(END,'\n  Number of Missings: '+str(varinfo['Type_Missing']))
        self.infobox.info.insert(END,'\n  Number of Strings: '+str(varinfo['Type_String']))
        self.infobox.info.insert(END,'\n  Number of Integers: '+str(varinfo['Type_Int']))
        self.infobox.info.insert(END,'\n  Number of Decimals: '+str(varinfo['Type_Float']))
        if len(varinfo['Uniques']) > 20:
            self.infobox.info.insert(END,'\n\n  '+str(len(varinfo['Uniques']))+' Unique Values (first 20): '+str(sorted(varinfo['Uniques'])[:20]))
        else:
            self.infobox.info.insert(END,'\n\n  '+str(len(varinfo['Uniques']))+' Unique Values: '+str(varinfo['Uniques']))

        if varinfo['N']>0:
            self.infobox.info.insert(END,'\n\n\nDescriptive Statistics:','title')
            self.infobox.info.insert(END,'\n\n  Numeric Values: '+str(varinfo['N']))
            self.infobox.info.insert(END,'\n  Minimum Value: '+str(varinfo['Min']))
            self.infobox.info.insert(END,'\n  Maximum Value: '+str(varinfo['Max']))
            self.infobox.info.insert(END,'\n  Mean: '+"{0:0.3}".format(varinfo['M']))
            self.infobox.info.insert(END,'\n  Std.Dev.: '+"{0:0.3}".format(varinfo['SD']))
            self.infobox.info.insert(END,'\n  Variance: '+"{0:0.3}".format(varinfo['Vari']))
        



        self.infobox.info.insert(END,'\n\n\nFrequencies:\n','title')
        self.infobox.info.insert(END,display_table(varinfo['Freq_Table']),'table')

    def export_output(self,overspill=0):
        text_content = self.Artikel.get(1.0,END)

        try: ##Python 3
            fname = filedialog.asksaveasfilename(**{'defaultextension':'.txt',
                                                  'filetypes':[('Text File','.txt'),('Data File','.dat'),('All Files','.*')]})
        except:
            fname = tkFileDialog.asksaveasfilename(**{'defaultextension':'.txt',
                                                      'filetypes':[('Text File','.txt'),('Data File','.dat'),('All Files','.*')]})
        outf = open(fname,'w')
        outf.write(text_content)
        outf.close()

        verbout('\n',master=self)
        verbout('Text output successfully stored to: '+fname,'warning',master=self)
        verbout('\n',master=self)
        


    def display_line_plot(self,plotdic,width=800,height=600, verbose=1):
        verb('Painting line plot')
        
        if 'X' in plotdic.keys() and 'Y' in plotdic.keys():            
            if len(plotdic['X']) == len(plotdic['Y']):
                xc = []
                yc = []
                yc2 = []
                for i in range(len(plotdic['X'])):
                    try:
                        x = float(plotdic['X'][i])
                        if type(plotdic['Y'][i]) == tuple:
                            y = []
                            for e in plotdic['Y'][i]:
                                y.append(float(e))
                                yc2.append(float(e))
                            y = tuple(y)
                        else:
                            y = float(plotdic['Y'][i])
                            yc2.append(y)
                        xc.append(x)
                        yc.append(y)
                    except:
                        x = 0
                if len(xc) > 1:
                    accept = 1
                else:
                    accept = 0
                    error = 'ERROR: Less than 2 valid pairs of X and Y'
            else:
                accept = 0
                error = 'ERROR: X and Y lengths differ.'
        else:
            accept = 0
            error = 'ERROR: Plotdic does not contain X and Y'

        if accept == 1:
            xrng = max(xc)-min(xc)
            yrng = max(yc2)-min(yc2)
            coord = sorted(list(zip(xc,yc)))
            xscale = float(width-150)/xrng
            yscale = float(height-150)/yrng

            nlines = 1
            if type(yc[0]) == tuple: nlines = len(yc[0])
            
            if 'Title' in plotdic.keys():
                title = str(plotdic['Title'])
            else:
                title = 'Line plot'

            self.infobox = Toplevel(self)
            self.infobox.rowconfigure(1, weight=1)
            self.infobox.columnconfigure(1, weight=1)
            self.infobox.title(title)
            self.infobox.ysc = Scrollbar(self.infobox, orient=VERTICAL)
            self.infobox.ysc.grid(row=1,column=2,sticky=N+S)
            self.infobox.xsc = Scrollbar(self.infobox, orient=HORIZONTAL)
            self.infobox.xsc.grid(row=2,column=1,sticky=E+W)
            self.infobox.plot = Canvas(self.infobox,bd=0,width=width,height=height,bg="#ffffff", scrollregion=(0, 0, width, height),
                                       yscrollcommand=self.infobox.ysc.set, xscrollcommand=self.infobox.xsc.set)
            self.infobox.plot.grid(row=1,column=1,sticky=N+E+S+W)
            self.infobox.ysc["command"]=self.infobox.plot.yview
            self.infobox.xsc["command"]=self.infobox.plot.xview

            pad = 10
            ax = [100-pad,height-100+pad,width-50+pad,height-100+pad]
            ay = [100-pad,height-100+pad,100-pad,50-pad]

            self.infobox.plot.create_line(ax[0],ax[1],ax[2],ax[3],fill="#000000")
            self.infobox.plot.create_line(ay[0],ay[1],ay[2],ay[3],fill="#000000")
            self.infobox.plot.create_text(ax[0],ax[1]+20,text=str(min(xc)),anchor=N)
            self.infobox.plot.create_text(ax[2],ax[3]+20,text=str(max(xc)),anchor=N)
            self.infobox.plot.create_text(ay[0]-50,ay[1],text=str(min(yc2)),anchor=NW)
            self.infobox.plot.create_text(ay[2]-50,ay[3],text=str(max(yc2)),anchor=NW)

            if 'Note' in plotdic.keys():
                addheight = 100
                if 'Noteheight' in plotdic.keys():
                    addheight = plotdic['Noteheight']
                self.infobox.plot['scrollregion'] = (0, 0, width, height+addheight)
                self.infobox.plot.create_rectangle(10,height-50,width-10,height+addheight-10,fill='#f3f3ff',width=0)
                self.infobox.plot.create_text(30,height-45,text='NOTES:', font=("Arial",10, "bold"), width=width-60, anchor=NW)
                self.infobox.plot.create_text(30,height-20,text=plotdic['Note'] , width=width-60, anchor=NW)

            if 'Color' in plotdic.keys():
                if nlines > 1:
                    lcol = []
                    for i in range(nlines):
                        lcol.append(plotdic['Color'])
                else:
                    lcol = plotdic['Color']
            else:
                palette = ['#000000','#0000ff','#ff0000','#00aa00','#aaaa40']
                if nlines > 1:
                    lcol = []
                    for i in range(nlines):
                        lcol.append(palette[i])
                else:
                    lcol = '#000080'

            tcoord = []
            for c in coord:
                cx = (c[0]-min(xc))*xscale + 100
                if nlines == 1:
                    cy = height-100-(c[1]-min(yc))*yscale
                    tcoord.append((cx,cy))
                else:
                    cy = []
                    for k in range(nlines):
                        cy.append(height-100-(c[1][k]-min(yc2))*yscale)
                    tcoord.append((cx,tuple(cy)))

            if 'Type' in plotdic.keys():
                ptype = plotdic['Type']
            else:
                ptype = 'Line'
            
            if ptype == 'Line':
                for i in range(len(tcoord)-1):
                    if nlines == 1:
                        self.infobox.plot.create_line(tcoord[i][0],tcoord[i][1],tcoord[i+1][0],tcoord[i+1][1],fill=lcol)
                    else:
                        for l in range(nlines):
                            self.infobox.plot.create_line(tcoord[i][0],tcoord[i][1][l],tcoord[i+1][0],tcoord[i+1][1][l],fill=lcol[l])
            elif ptype == 'Scatter':
                rad = 3
                for i in range(len(tcoord)):
                    if nlines == 1:
                        self.infobox.plot.create_oval(tcoord[i][0]-rad,tcoord[i][1]-rad,tcoord[i][0]+rad,tcoord[i][1]+rad,outline=lcol)
                    else:
                        for l in range(nlines):
                            self.infobox.plot.create_oval(tcoord[i][0]-rad,tcoord[i][1][l]-rad,tcoord[i][0]+rad,tcoord[i][1][l]+rad,outline=lcol[l])               
        else:
            verb(error)
        
        
                

    def display_heat_map(self, gridlist=[],w=0,h=0,ph=4,pw=4,legend=1,mode='bw',verbose=1):
        verb('Painting grid')
        gh = len(gridlist)
        gw = len(gridlist[0])

        verb('Grid width: '+str(gw))
        verb('Grid height: '+str(gh))
        
        if h==0:
            h=gh
        if w==0:
            w=gw


        verb('Plot width: '+str(w))
        verb('Plot height: '+str(h))
        verb('Pixel width: '+str(pw))
        verb('Pixel height: '+str(ph))
              
        outgrid = {}
        for x in range(0,w):
            outgrid[x] = {}
            for y in range(0,h):
                outgrid[x][y] = {}
                outgrid[x][y]['Values']=[]
                outgrid[x][y]['M']=0
                outgrid[x][y]['Col']='#ffffff'

        for gx in range(0,gw):
            for gy in range(0,gh):
                val = gridlist[gy][gx]
                xpos = int(float(gx*w)/gw)
                ypos = h-int(float(gy*h)/gh)-1
                outgrid[xpos][ypos]['Values'].append(val)

        allval = []
        misval = 0

        for x in range(0,w):
            for y in range(0,h):
                if len(outgrid[x][y]['Values']) > 0:
                    outgrid[x][y]['M'] = float(sum(outgrid[x][y]['Values']))/len(outgrid[x][y]['Values'])
                    allval.append(outgrid[x][y]['M'])
                else:
                    outgrid[x][y]['M']='-'
                    misval = misval + 1                

        minval = min(allval)
        maxval = max(allval)

        verb('Lowest value: '+str(minval))
        verb('Highest value: '+str(maxval))
        verb('Missing values: '+str(misval))
        
        for x in range(0,w):
            for y in range(0,h):
                val = outgrid[x][y]['M']
                if type(val) == float:
                    stdval = (val-minval)/(maxval-minval)
                    heatcol = heat_color(stdval,mode)
                    outgrid[x][y]['Col']=heatcol

        canw = pw*w+20
        canh = ph*h+80
        
        self.infobox = Toplevel(self)
        self.infobox.rowconfigure(1, weight=1)
        self.infobox.columnconfigure(1, weight=1)
        self.infobox.title("Heat Map plot")
        self.infobox.ysc = Scrollbar(self.infobox, orient=VERTICAL)
        self.infobox.ysc.grid(row=1,column=2,sticky=N+S)
        self.infobox.xsc = Scrollbar(self.infobox, orient=HORIZONTAL)
        self.infobox.xsc.grid(row=2,column=1,sticky=E+W)
        self.infobox.plot = Canvas(self.infobox,bd=0,width=canw,height=canh+10,bg="#ffffff", scrollregion=(0, 0, canw, canh+10),
                                   yscrollcommand=self.infobox.ysc.set, xscrollcommand=self.infobox.xsc.set)
        self.infobox.plot.grid(row=1,column=1,sticky=N+E+S+W)
        self.infobox.ysc["command"]=self.infobox.plot.yview
        self.infobox.xsc["command"]=self.infobox.plot.xview

        legmin = 50
        legmax = canw-50

        if legmin < legmax-10:
            for i in range(legmin,legmax):
                std = float(i-legmin)/(legmax-legmin)
                legcol=heat_color(std,mode)
                self.infobox.plot.create_rectangle(i, 10, i+1, 30, fill=legcol,width=0)
            min_f = "{0:.3f}".format(minval)
            max_f = "{0:.3f}".format(maxval)
            self.infobox.plot.create_text(legmin,50,text=min_f)
            self.infobox.plot.create_text(legmax,50,text=max_f)
            self.infobox.plot.create_rectangle(legmin,10,legmax,30,width=1)
        else:
            verb('ERROR: Plotting region too small for legend')

        for x in range(0,w):
            for y in range(0,h):               
                self.infobox.plot.create_rectangle(pw*x+10, canh-ph*y, pw*(x+1)+10, canh-ph*(y+1), fill=outgrid[x][y]['Col'],width=0)
        self.infobox.plot.create_rectangle(10, canh, pw*w+10, canh-ph*h,width=1)
        self.infobox.plot.update()
        mainloop()
            

    def display_dendro(self, dendrogram, clusters):
        anz_c = len(clusters)

        c_elements = {}
        for i in range(0,anz_c):
            cluster_num = str(i + 1)
            cluster_col = heat_color(float(i)/anz_c,'red')
            cluster_cont = flatten(clusters[i])
            for c in cluster_cont:
                c_elements[c] = {}
                c_elements[c]['Cluster']=cluster_num
                c_elements[c]['Color']=cluster_col
        
        self.infobox = Toplevel(self)
        self.infobox.rowconfigure(1, weight=1)
        self.infobox.columnconfigure(1, weight=1)
        self.infobox.title("Dendrogram")
        self.infobox.ysc = Scrollbar(self.infobox, orient=VERTICAL)
        self.infobox.ysc.grid(row=1,column=2,sticky=W+N+S)
        self.infobox.info = Text(self.infobox,width=120,height=30,bg="#ffffcc",wrap=WORD, yscrollcommand=self.infobox.ysc.set,font=("Courier",8),takefocus=0)
        self.infobox.info.grid(row=1,column=1,sticky=N+E+S+W)
        self.infobox.ysc["command"]=self.infobox.info.yview

        self.infobox.info.insert(END,'Legend:\n-----------\n')


        for i in range(0,anz_c):
            cluster_num = str(i + 1)
            cluster_col = heat_color(float(i)/anz_c,'red')
            self.infobox.info.tag_config(cluster_num,background=cluster_col)
            self.infobox.info.insert(END,'Cluster #'+cluster_num+'\n',cluster_num)

        self.infobox.info.insert(END,'-----------\n\n')

        initline = 5+anz_c
        
        
        self.infobox.info.insert(END,dendrogram)
        self.infobox.info.see(END)


        dendroline = dendrogram.split('\n')
        for i in range(len(dendroline)):
            c1 = dendroline[i].find(' -')
            c2 = dendroline[i].find('+ ')
            if dendroline[i][:c1] in c_elements.keys():
                tag = c_elements[dendroline[i][:c1]]['Cluster']
                pos = str(i+initline)+'.0'
                endpos = str(i+initline)+'.'+str(c2)
                self.infobox.info.tag_add(tag,pos,endpos)
         


    def display_text(self, content = "No text to display"):
        self.infobox = Toplevel(self)
        self.infobox.rowconfigure(1, weight=1)
        self.infobox.columnconfigure(1, weight=1)
        self.infobox.title("Debugging-Information")
        self.infobox.ysc = Scrollbar(self.infobox, orient=VERTICAL)
        self.infobox.ysc.grid(row=1,column=2,sticky=W+N+S)
        self.infobox.info = Text(self.infobox,width=100,height=30,bg="#ffffcc",wrap=WORD, yscrollcommand=self.infobox.ysc.set,font=("Arial",9),takefocus=0)
        self.infobox.info.grid(row=1,column=1,sticky=N+E+S+W)
        self.infobox.ysc["command"]=self.infobox.info.yview

        self.infobox.info.tag_config('errormsg',background="#ffcccc", font = ("Arial",9, "bold"))
        self.infobox.info.tag_config('functcall',foreground="#0000cc", font = ("Arial",9, "bold"))
        
        self.infobox.info.insert(END,content)
        self.infobox.info.see(END)

        start = '1.1'
        while not start == '':
            a = self.infobox.info.search('ERROR',start,END)
            if not a == '':
                start_z = str(a) + ' linestart'
                end_z = str(a) + ' lineend'
                self.infobox.info.tag_add('errormsg',start_z,end_z)
                start = end_z
            else:
                start = a
        start = '1.1'
        while not start == '':
            a = self.infobox.info.search('Function',start,END)
            if not a == '':
                start_z = str(a) + ' linestart'
                end_z = str(a) + ' lineend'
                self.infobox.info.tag_add('functcall',start_z,end_z)
                start = end_z
            else:
                start = a
                
    def show_path(self,event=0):
        global settings
        out = str(settings['Path_Log'])
        out = out + ('\n\nVisited pages:'+str(settings['Page_History']))
        self.display_text(out)

    def show_verb(self,event=0):
        global settings
        out = str(settings['Verb_Log'])
        self.display_text(out)

    def show_storage(self,event=0):
        global storage
        out = str(baum_schreiben(storage,trunc=50))
        self.display_text(out)

    def show_settings(self,event=0):
        tmp_dic = {}
        for c in settings.keys():
            if not c in ['Verb_Log','Path_Log']:
                tmp_dic[c] = settings[c]
        out = str(baum_schreiben(tmp_dic,trunc=50))
        self.display_text(out)

    def show_parameters(self):
        global dta_pos
        global prog_pos
        global settings
        out='Current location within program and data:\n----------------------------\n'
        out = out + 'Dta_Pos: '+str(dta_pos)
        out = out + '\nProg_Pos: '+prog_pos
        out = out + '\nPage: '+str(settings['Curr_Page'])
        out = out + '\n\n\nCurrent Subtree:\n----------------------------\n'
        out = out + str(baum_schreiben(curr(),trunc=50))
        self.display_text(out)

    def reset_coding(self):
        global dta_pos
        global prog_pos
        dta_pos=['-','-','-','-']
        prog_pos = settings['First_Page']
        ##Send Log-File
        fname = '..\\'+settings['Coder'] + str(time.time()) + '.txt'
        out = open(fname,'w')
        out.write('Automatisch generiertes Logfile.\n')
        out.write(time.ctime()+'\n----------------------------\n')
        out.write(str(settings['Verb_Log']))
        out.close()
        self.set_window()

    def locate(self,l1,l2,l3):
        log('Calling Function: Locate with '+l1+' and '+l2)
        self.f_location.angabe.delete('1.0', END)
        lev = 0
        if dta_pos[0] == '-':
            lev = 0
        elif dta_pos[2] == '-':
            lev = 1
        elif dta_pos[4] == '-':
            lev = 2
        if lev > 0:
            self.f_location.angabe.insert(INSERT, self.namegetter('Location',l1)+'\n','fett')
            self.f_location.angabe.insert(INSERT, self.namegetter('Location',l2)+' ')
            self.f_location.angabe.insert(INSERT, storage[dta_pos[0]][dta_pos[1]]['#TN'])
        if lev > 1:
            self.f_location.angabe.insert(INSERT, ' '+self.namegetter('Location',l3)+' ')
            self.f_location.angabe.insert(INSERT, storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]]['#TN'])
       
    def show_review(self,level,rm=1,height=3):
        #Show the list of previously coded elements on a level below the current one
        global prog_pos
        log('Calling Function: Show_Review')
        try:
            self.f_review.A_Liste.destroy()
            self.f_review.scroll_A_Liste.destroy()
            self.f_review.b_remove.grid_remove()            
        except:
            verb('No List to remove')
        
        self.f_review.A_Liste = Listbox(self.f_review, selectmode = BROWSE, height=height, width=80)
        self.f_review.A_Liste.grid(row=0,rowspan=3, column=0, sticky=W+E)
        self.f_review.scroll_A_Liste = Scrollbar(self.f_review, orient=VERTICAL, command=self.f_review.A_Liste.yview)
        self.f_review.scroll_A_Liste.grid(row=0,rowspan=3, column=1, sticky=W+N+S)
        self.f_review.A_Liste["yscrollcommand"] = self.f_review.scroll_A_Liste.set
        self.f_review.b_remove = Button(self.f_review, text = "Remove", width=6, command=CMD(self.remove_item,level,height), takefocus = 0)
        self.f_review.b_remove.grid(row=0,column=2,sticky=N+E)
        self.f_review.b_edit = Button(self.f_review, text = "Edit", width=6, command=CMD(self.edit_item,level), takefocus = 0)
        self.f_review.b_edit.grid(row=1,column=2,sticky=N+E)

        curr_tree = curr()

        if type(level) == str:
            level = [level]

        for l in level:
            if l in curr_tree.keys():
                for item in sorted(curr_tree[l].keys()):
                    lab = '<'+item+'>: '+curr_tree[l][item]['#TN']
                    self.f_review.A_Liste.insert(END,lab)
        if rm == 0:
            self.f_review.b_remove.grid_remove()
        else:
            self.f_review.b_remove.grid()

    def hide_review(self): #Hide the review-list
        global prog_pos
        log('Calling Function: Hide Review')
        try:
            self.f_review.A_Liste.destroy()
            self.f_review.scroll_A_Liste.destroy()
            self.f_review.b_remove.grid_remove()
            self.f_review.b_edit.grid_remove()
        except:
            verb('No List to remove')
        self.f_review["height"]=20

    def remove_item(self,level,height):
        #Remove an item from the review-List. Depending on level and requirements this function might need adaption.
        global prog_pos
        log('Calling Function: Remove_Item')
        listsel = self.f_review.A_Liste.curselection()
        if len(listsel) == 0:
            self.message("Invalid-Selection01")
        else:
            select = self.f_review.A_Liste.get(listsel[0])
            verb(str(select))
            c1 = 1
            c2 = select.find('>')
            code = select[c1:c2]
            verb('Code to remove: '+code)

            if type(level) == str:
                level = [level]

            for l in level:
                if dta_pos[0] == '-':
                    if l in storage.keys():
                        if code in storage[l].keys():
                            del storage[l][code]
                            verb('Removing Element: '+code+ 'on Level: '+l)
                            if len(storage[l].keys()) == 0:
                                del storage[l]
                elif dta_pos[2] == '-':
                    if l in storage[dta_pos[0]][dta_pos[1]].keys():
                        if code in storage[dta_pos[0]][dta_pos[1]][l].keys():
                            del storage[dta_pos[0]][dta_pos[1]][l][code]
                            verb('Removing Element: '+code+ 'on Level: '+l)
                            if len(storage[dta_pos[0]][dta_pos[1]][l].keys()) == 0:
                                del storage[dta_pos[0]][dta_pos[1]][l]
                elif dta_pos[4] == '-':
                    if l in storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]].keys():
                        if code in storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][l].keys():
                            del storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][l][code]
                            verb('Removing Element: '+code+ 'on Level: '+l)
                            if len(storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][l].keys()) == 0:
                                del storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][l]
           
            self.hide_review()
            self.show_review(level,1,height)
            storage['Remove_item'] = storage['Remove_item'] + 1


    def edit_item(self,level):
        #Remove an item from the review-List.
        ##Depending on level and requirements this function might need adaption.
        ## Deprecated in NOGROD
        pass

    def level_up(self):
        log('Calling Function: Level Up')
        global dta_pos 
        laenge = len(dta_pos)
        minstrich = 0
        for i in range(0,laenge):
            if dta_pos[i] == '-' and minstrich == 0:
                minstrich = i
        if minstrich == 0:
            minstrich = laenge        
        for i in range(minstrich-2,laenge):
            dta_pos[i] = '-'

    def level_down(self,variable,ebene): ##Change the level of analysis by one. If successful, the function returns 1
        log('Calling Function: Level Down')
        global dta_pos
        accept = 1
        idvar = self.store_var(variable,store=0)
        tstamp = time.time()
        if type(idvar) == tuple:
            ident = str(idvar[1])
            if ident == '[]' or ident == '':
                accept = 0
            tname = str(idvar[0])
        elif type(idvar) == str or type(idvar) == unicode:
            ident = bereinigen(idvar)
            tname = bereinigen(idvar)
            if ident == '':
                accept = 0

        if accept == 1:
            if '-' in dta_pos:
                minstrich = len(dta_pos)
                for i in range(0,len(dta_pos)):
                    if dta_pos[i] == '-' and minstrich == len(dta_pos):
                        minstrich = i
            else:
                minstrich = len(dta_pos)
                dta_pos.append('-')
                dta_pos.append('-')

            if minstrich == 0:
                if ebene in storage.keys():
                    while ident in storage[ebene].keys():
                        ident = ident + 'x'
                    storage[ebene][ident] = {}
                else:
                    storage[ebene] = {}
                    storage[ebene][ident] = {}
                storage[ebene][ident]['#TN'] = tname
                storage[ebene][ident]['#TS'] = tstamp
            elif minstrich == 2:
                if ebene in storage[dta_pos[0]][dta_pos[1]].keys():
                    while ident in storage[dta_pos[0]][dta_pos[1]][ebene].keys():
                        ident = ident + 'x'
                    storage[dta_pos[0]][dta_pos[1]][ebene][ident] = {}
                else:
                    storage[dta_pos[0]][dta_pos[1]][ebene] = {}
                    storage[dta_pos[0]][dta_pos[1]][ebene][ident] = {}
                if len(dta_pos) < 5:
                    dta_pos = dta_pos + ['-','-']
                storage[dta_pos[0]][dta_pos[1]][ebene][ident]['#TN'] = tname
                storage[dta_pos[0]][dta_pos[1]][ebene][ident]['#TS'] = tstamp
            elif minstrich == 4:
                if ebene in storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]].keys():
                    while ident in storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][ebene].keys():
                        ident = ident + 'x'
                    storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][ebene][ident] = {}
                else:
                    storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][ebene] = {}
                    storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][ebene][ident] = {}
                if len(dta_pos) < 7:
                    dta_pos = dta_pos + ['-','-']
                storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][ebene][ident]['#TN'] = tname
                storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][ebene][ident]['#TS'] = tstamp
            elif minstrich == 6:
                if ebene in storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]].keys():
                    while ident in storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][dta_pos[4]][dta_pos[5]][ebene].keys():
                        ident = ident + 'x'
                    storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][dta_pos[4]][dta_pos[5]][ebene][ident] = {}
                else:
                    storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][dta_pos[4]][dta_pos[5]][ebene] = {}
                    storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][dta_pos[4]][dta_pos[5]][ebene][ident] = {}
                storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][dta_pos[4]][dta_pos[5]][ebene][ident]['#TN'] = tname
                storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][dta_pos[4]][dta_pos[5]][ebene][ident]['#TS'] = tstamp

            dta_pos[minstrich] = ebene
            dta_pos[minstrich+1] = ident          
        return accept
        

    def buttons(self, check=1, abort=0, back=1, pause=0):
        log('--Setting Buttons: Check='+str(check)+'; Abort='+str(abort)+'; Back='+str(back)+'; Break='+str(pause),pos=0)
        if check == 1:
            self.f_bottomline.b_check.grid()
            self.f_bottomline.b_check["state"] = NORMAL
        else:
            self.f_bottomline.b_check.grid_remove()
        if abort == 1:
            self.f_bottomline.b_abort.grid()
        else:
            self.f_bottomline.b_abort.grid_remove()
        if back == 1:
            self.f_bottomline.b_back.grid()
        else:
            self.f_bottomline.b_back.grid_remove()
        if pause == 1:
            self.f_bottomline.b_break.grid()
        else:
            self.f_bottomline.b_break.grid_remove()


    def load_dset(self,fname,header,sep,dname='Data',vname='D_Var',designation='Main_Table'):
        ##Loads a dataset and wanrs if anything is wrong with the data.
        ##This function is required for all Nogrod-Procedures. It takes a filename and returns a complete dataset.
        global storage
        global settings
                
        dset = get_dataset(fname, header, sep,master=self)

        if dset[0] in ['invalid',0]:
            verb('ERROR: Invalid File')
            self.message('File not found!',add=dset[3])
            retval = 0
            verbout('\n',master=self)
            verbout('ERROR: File could not be loaded\n','warning',master=self)
        else:
            verbout('\n'+dset[2]+'\n',master=self)
            if len(dset[3])>2: verbout(dset[3],'warning',master=self)
            verbout('\n',master=self)
            storage[dname] = dset[0]
            storage[vname] = dset[1]

            settings['Datasets'][designation] = {}
            settings['Datasets'][designation]['Data'] = dname
            settings['Datasets'][designation]['Var'] = vname

            retval = dset
        return retval    

    def store_var_all(self,setdef=1): ##Storing all variables on this page
        log('--Store_Var_All:',pos=0)
        global prog_pos
        global dta_pos
        for i in range(len(settings['Curr_Page'])):
            anzeige = settings['Curr_Page'][i]
            if len(anzeige[1]) > 1:
                if anzeige[0] == 'list':
                    self.store_var(anzeige[2],i,setdef)
                else:
                    self.store_var(anzeige[1],i,setdef)
                
    def store_var(self,variabel,pos=-1,setdef=1,store=1): ##Storing values from all question-types.
        log('----Storing Variable: '+variabel,pos=0)
        wert = 'invalid'
        if pos == -1:
            for i in range(len(settings['Curr_Page'])):
                if settings['Curr_Page'][i][1] == variabel:
                    pos = i
                elif len(settings['Curr_Page'][i]) == 3:
                    if settings['Curr_Page'][i][2] == variabel:
                        pos = i
        if pos == -1:
            verb('ERROR: Variable "'+variabel+'" not found on this page')
        else:
            element = settings['Curr_Page'][pos]
            if element[0] == 'dd':
                wert = (settings['Input'][pos].get(),self.codegetter(variabel,settings['Input'][pos].get()))
                verb('    Stored Variable:'+variabel+': '+str(wert),1)                       

            elif element[0] in ['txt','file']:
                if pos == 0:
                    wert = bereinigen(self.f_questions.txt1.get())
                elif pos == 1:
                    wert = bereinigen(self.f_questions.txt2.get())
                elif pos == 2:
                    wert = bereinigen(self.f_questions.txt3.get())                   
                verb('    Stored Variable:'+variabel+': '+str(wert),1)

            elif element[0] == 'txt2':
                if pos == 0:
                    wert = bereinigen(self.f_questions.txt1.get('1.0',END))
                elif pos == 1:
                    wert = bereinigen(self.f_questions.txt2.get('1.0',END))
                elif pos == 2:
                    wert = bereinigen(self.f_questions.txt3.get('1.0',END))
                verb('    Stored Variable:'+variabel+': '+str(wert),1)

            elif element[0] == 'rb':
                wert = (self.namegetter(variabel,settings['Input'][pos].get()),settings['Input'][pos].get())
                verb('    Stored Variable:'+variabel+': '+str(wert),1)
                
            elif element[0] == 'rbopen':
                if settings['Input'][pos].get() == 'Open Answer':## Achtung: Analog zu TXT mit pos auslesen!!
                    if pos == 0:
                        wert = bereinigen(self.t1.get())
                    elif pos == 1:
                        wert = bereinigen(self.t2.get())
                    elif pos == 2:
                        wert = bereinigen(self.t3.get())                   
                    verb('    Stored Variable:'+variabel+': '+str(wert),1)
                    #wert = (wert,'Open Answer')
                else:
                    wert = (self.namegetter(variabel,settings['Input'][pos].get()),settings['Input'][pos].get())
                verb('    Stored Variable:'+variabel+': '+str(wert),1)
                
                                    
            elif element[0] == 'spr':
                wert = {}
                wert[codebook[variabel][3][0]] = (bereinigen(settings['Akt_a'].get()),self.codegetter('Akt_a',settings['Akt_a'].get()))
                wert[codebook[variabel][3][1]] = (bereinigen(settings['Akt_b'].get()),self.codegetter('Akt_b',settings['Akt_b'].get()))
                wert[codebook[variabel][3][2]] = (bereinigen(settings['Akt_c'].get()),self.codegetter('Akt_c',settings['Akt_c'].get()))
                verb('    Stored Variable:'+variabel+': '+str(wert),1)

            elif element[0] in ['rating','sd','cb']:
                wert = {}
                for i in range(len(codebook[variabel][3])):
                    code = codebook[variabel][3][i]
                    wert[code] = settings['Input'][pos][i].get()
                verb('    Stored Variable:'+variabel+': '+str(wert),1)                  
            
            elif element[0] in ['list','listseek']:
                listvar = settings['Curr_Page'][0][2]
                variabel = settings['Curr_Page'][0][1]
                listsel = self.f_questions.Aspliste.curselection()
                Namen = []
                Codes = []
                for selection in listsel:
                    Namen.append(self.f_questions.Aspliste.get(selection))
                    Codes.append(self.codegetter(listvar, self.f_questions.Aspliste.get(selection)))
                wert = (Namen,Codes)
                verb('    Stored Variable:'+variabel+': '+str(wert),1)

            elif element[0] in ['listadd']:
                listvar = settings['Curr_Page'][0][2]
                variabel = settings['Curr_Page'][0][1]
                listsel = range(0,self.f_questions.Itmliste.size()) ###Get all!
                Namen = []
                Codes = []
                for i in listsel:
                    selection = str(i)
                    Namen.append(self.f_questions.Itmliste.get(selection))
                    c = self.codegetter(listvar, self.f_questions.Itmliste.get(selection))
                    if c == '':
                        c = self.f_questions.Itmliste.get(selection)
                    Codes.append(c)
                wert = (Namen,Codes)
                verb('    Stored Variable:'+variabel+': '+str(wert),1)
                
            elif element[0] == 'unit_auswahl':
                selection = self.f_questions.Aspliste.curselection()[0]
                label = self.f_questions.Aspliste.get(selection)
                verb('List Selection:'+str(label))
                for u in settings[variabel].keys():
                    if label == settings[variabel][u]['Label']:
                        verb(baum_schreiben(settings[variabel][u]))
                        wert = u
                verb('    Stored Variable:'+variabel+': '+str(wert),1)

        if store == 1:
            if not dta_pos[3] == '-':
                storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]][variabel] = wert
            elif not dta_pos[1] == '-':
                storage[dta_pos[0]][dta_pos[1]][variabel] = wert
            else:
                storage[variabel] = wert
            verb('    Stored in storage',1)
        if setdef == 1:
            if type(wert) == list:
                def_val[variabel] = get_unique(wert)
                verb('    -Setting Default: '+str(def_val[variabel]),1)
            else:
                def_val[variabel] = wert
                verb('    -Setting Default: '+str(def_val[variabel]),1)
            verb('    Set as Default',1)
        return wert

    def check_entries(self):
        log('--Checking submitted entries',pos=0)
        #Check all entries for validity.
        all_correct = 1 ##Begin with an assumption of innocence
        for i in range(0,len(settings['Curr_Page'])):
            typ = settings['Curr_Page'][i][0]
            if typ == 'rb':
                if settings['Input'][i].get()=='98':
                    verb('Please choose an option vor variable: "' + settings['Curr_Page'][i][1] + '"')
                    self.message('Invalid-Selection02',add=settings['Curr_Page'][i][1])
                    all_correct=0
            if typ == 'dd':
                if self.codegetter(settings['Curr_Page'][i][1],settings['Input'][i].get()) == '98':
                    verb('Please choose an option vor variable: "' + settings['Curr_Page'][i][1] + '"')
                    self.message('Invalid-Selection02',add=settings['Curr_Page'][i][1])
                    all_correct=0
            if typ == 'file':
                a = ''
                if i == 0:
                    a = self.f_questions.txt1.get()
                elif i == 1:
                    a = self.f_questions.txt2.get()
                elif i == 2:
                    a = self.f_questions.txt3.get()

                if a == '':
                    all_correct = 0
                    self.message('Invalid-Selection04')
                    
            if typ == 'txt':
                if i == 0:
                    try:
                      verb(bereinigen(self.f_questions.txt1.get()))
                    except:
                        all_correct = 0
                        self.message('Invalid-Selection03')
                if i == 1:
                    try:
                      verb(bereinigen(self.f_questions.txt2.get()))
                    except:
                        all_correct = 0
                        self.message('Invalid-Selection03')
                if i == 2:
                    try:
                      verb(bereinigen(self.f_questions.txt3.get()))
                    except:
                        all_correct = 0
                        self.message('Invalid-Selection03')
            if typ == 'txt2':
                if i == 0:
                    try:
                      verb(bereinigen(self.f_questions.txt1.get('1.0',END)))
                    except:
                        all_correct = 0
                        self.message('Invalid-Selection03')
                if i == 1:
                    try:
                      verb(bereinigen(self.f_questions.txt2.get('1.0',END)))
                    except:
                        all_correct = 0
                        self.message('Invalid-Selection03')
                if i == 2:
                    try:
                      verb(bereinigen(self.f_questions.txt3.get('1.0',END)))
                    except:
                        all_correct = 0
                        self.message('Invalid-Selection03')

            if typ == 'spr':
                wert = {}
                a = self.codegetter('Akt_a',settings['Akt_a'].get())
                b = self.codegetter('Akt_b',settings['Akt_b'].get())
                c = self.codegetter('Akt_c',settings['Akt_c'].get())
                if a == 'akt000':
                    if c == 'orga000':
                        if b == 'fkt000':
                            all_correct = 0
                            self.message('Invalid-Selection02')
        return all_correct

    def clean_up_all(self): ##Clean up the whole page (only question-frame).
        log('--Clean Up All:',pos=0)
        for i in range(0,len(settings['Curr_Page'])):
            typ = settings['Curr_Page'][i][0]
            pos = i + 1
            self.clean_up(typ,pos)

    def clean_up(self,typ='all',pos=1): ##Clean up a specific element on the page (only question_frame)
        log('----Clean Up: Element: '+typ+' Position: '+str(pos))
        if typ == 'all':
            typ = settings['Curr_Page'][pos-1][0]
        try:
            if typ == 'dd':
                if pos == 1:
                    self.f_questions.dd1.destroy()
                    self.f_questions.help1.destroy()
                    self.f_questions.Frage1.delete("1.0",END)
                    if settings['Insecure']=='1':
                        self.f_questions.ins1.destroy()
                    settings['Curr_Page'][pos-1] = ['','']
                if pos == 2:
                    self.f_questions.dd2.destroy()
                    self.f_questions.help2.destroy()
                    self.f_questions.Frage2.delete("1.0",END)
                    if settings['Insecure']=='1':
                        self.f_questions.ins2.destroy()
                    settings['Curr_Page'][pos-1] = ['','']
                if pos == 3:
                    self.f_questions.dd3.destroy()
                    self.f_questions.help3.destroy()
                    self.f_questions.Frage3.delete("1.0",END)
                    if settings['Insecure']=='1':
                        self.f_questions.ins3.destroy()
                    settings['Curr_Page'][pos-1] = ['','']
                    
            if typ == 'menu':
                if pos == 1:
                    self.f_questions.mb1.destroy()
                    self.f_questions.help1.destroy()
                    self.f_questions.Frage1.delete("1.0",END)
                    if settings['Insecure']=='1':
                        self.f_questions.ins1.destroy()
                    settings['Curr_Page'][pos-1] = ['','']
                if pos == 2:
                    self.f_questions.mb2.destroy()
                    self.f_questions.help2.destroy()
                    self.f_questions.Frage2.delete("1.0",END)
                    if settings['Insecure']=='1':
                        self.f_questions.ins2.destroy()
                    settings['Curr_Page'][pos-1] = ['','']
                if pos == 3:
                    self.f_questions.mb3.destroy()
                    self.f_questions.help3.destroy()
                    self.f_questions.Frage3.delete("1.0",END)
                    if settings['Insecure']=='1':
                        self.f_questions.ins3.destroy()
                    settings['Curr_Page'][pos-1] = ['','']
                    
            elif typ in ['txt','txt2','file']:
                if pos == 1:
                    self.f_questions.txt1.destroy()
                    self.f_questions.help1.destroy()
                    self.f_questions.Frage1.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.getselect1.destroy()
                    except:
                        verb('No Button')
                if pos == 2:
                    self.f_questions.txt2.destroy()
                    self.f_questions.help2.destroy()
                    self.f_questions.Frage2.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.getselect2.destroy()
                    except:
                        verb('No Button')
                if pos == 3:
                    self.f_questions.txt3.destroy()
                    self.f_questions.help3.destroy()
                    self.f_questions.Frage3.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.getselect3.destroy()
                    except:
                        verb('No Button')
            elif typ == 'list':
                self.f_questions.Aspliste.destroy() 
                self.f_questions.scroll_AspListe.destroy()
                self.f_questions.h_Aspliste.destroy()
                self.f_questions.Frage1.delete("1.0",END)    
                for i in settings['Curr_Page']:
                    if i[0] == 'list':
                        i[0] = ''
                        i[1] = ''
            elif typ == 'listseek':
                self.f_questions.Aspliste.destroy() 
                self.f_questions.scroll_AspListe.destroy()
                self.f_questions.h_Aspliste.destroy()
                self.f_questions.seektext.destroy()
                self.f_questions.Frage1.delete("1.0",END)    
                for i in settings['Curr_Page']:
                    if i[0] == 'listseek':
                        i[0] = ''
                        i[1] = ''
            elif typ == 'listadd':
                self.f_questions.Aspliste.destroy()
                self.f_questions.scroll_AspListe.destroy()
                self.f_questions.h_Aspliste.destroy()
                self.f_questions.seektext.destroy()
                self.f_questions.adb.destroy()
                self.f_questions.rb.destroy()
                self.f_questions.adb2.destroy()
                self.f_questions.rb2.destroy()
                self.f_questions.Itmliste.destroy()
                self.f_questions.scroll_ItmListe.destroy()
                self.f_questions.Frage1.delete("1.0",END)    
                for i in settings['Curr_Page']:
                    if i[0] == 'listadd':
                        i[0] = ''
                        i[1] = ''
            elif typ == 'unit_auswahl':
                self.f_questions.Aspliste.destroy()
                self.f_questions.scroll_AspListe.destroy()
                self.f_questions.h_Aspliste.destroy()
                self.f_questions.Frage1.delete("1.0",END)
                self.f_questions.fk_hinzu.destroy()
                self.f_questions.fk_weg.destroy()
                self.f_questions.fk_markieren.destroy()
                for i in settings['Curr_Page']:
                    if i[0] == 'list':
                        i[0] = ''
                        i[1] = ''
                        
            elif typ in ['rb','rbopen','sd','rating','cb']:
                if pos == 1:
                    self.f_questions.rblist1.destroy()
                    self.f_questions.help1.destroy()
                    self.f_questions.Frage1.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.ins1.destroy()
                    except:
                        verb('No Insec Button')
                if pos == 2:
                    self.f_questions.rblist2.destroy()
                    self.f_questions.help2.destroy()
                    self.f_questions.Frage2.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.ins2.destroy()
                    except:
                        verb('No Insec Button')
                if pos == 3:
                    self.f_questions.rblist3.destroy()
                    self.f_questions.help3.destroy()
                    self.f_questions.Frage3.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.ins3.destroy()
                    except:
                        verb('No Insec Button')
            elif typ == 'bt':
                if pos == 1:
                    self.f_questions.help1.destroy()
                    self.f_questions.Frage1.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.bu1_1.destroy()
                        self.f_questions.bu1_2.destroy()
                        self.f_questions.bu1_3.destroy()
                        self.f_questions.bu1_4.destroy()
                    except:
                        verb('Less than 4 Buttons')
                if pos == 2:
                    self.f_questions.help2.destroy()
                    self.f_questions.Frage2.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.bu2_1.destroy()
                        self.f_questions.bu2_2.destroy()
                        self.f_questions.bu2_3.destroy()
                        self.f_questions.bu2_4.destroy()
                    except:
                        verb('Less than 4 Buttons')
                if pos == 3:
                    self.f_questions.help3.destroy()
                    self.f_questions.Frage3.delete("1.0",END)
                    settings['Curr_Page'][pos-1] = ['','']
                    try:
                        self.f_questions.bu3_1.destroy()
                        self.f_questions.bu3_2.destroy()
                        self.f_questions.bu3_3.destroy()
                        self.f_questions.bu3_4.destroy()
                    except:
                        verb('Less than 4 Buttons')
            elif typ == 'spr':
                self.f_questions.help_spr.destroy()
                self.f_questions.sprecher.destroy()
                for i in settings['Curr_Page']:
                    if i[0] == 'spr':
                        i[0] = ''
                        i[1] = ''
            elif typ == 'val':
                self.f_questions.value.destroy()
                self.f_questions.help_val.destroy()
                for i in settings['Curr_Page']:
                    if i[0] == 'val':
                        i[0] = ''
                        i[1] = ''
            else:
                if settings['Verbose'] == '1':
                    verb('ERROR: Unknown Element. Cannot clean up')
        except:
            verb('ERROR: Element not found')


    def codegetter(self,variabel,item): #Get the code of a value when only the label is known
        log('Calling Function: Codegetter for Variable: '+variabel+' and Value: '+str(item),pos=0)
        varindex = ''
        for i in range(0,len(codebook[variabel][2])):
            try:
                if bereinigen(codebook[variabel][2][i]) == bereinigen(item):
                    varindex = codebook[variabel][3][i]
            except:
                varindex = 'special character!'

        if varindex == '':
            if settings['Verbose'] == '1':
                verb('--Error: Code not found for:'+str(item))
        else:
            verb('--Code identified: '+str(varindex))           
        return varindex

    def namegetter(self,variabel,item): #Get the label of a value when only the code is known
        log('Calling Function: Namegetter for Variable: '+variabel+' and Value: '+item,pos=0)
        varindex = ''
        for i in range(0,len(codebook[variabel][2])):
            try:
                if codebook[variabel][3][i] == item:
                    varindex = codebook[variabel][2][i]
            except:
                varindex = 'special character!'

        if varindex == '':
            if settings['Verbose'] == '1':
                verb('--Error: Name not found for:'+str(item))
        else:
            verb('--Name identified: '+str(varindex))
        return varindex

    def unit_confirm(self,Einheit,sel_tags=[]):
        log('Calling Function: Unit Confirm')
        global prog_pos
        global dta_pos
        self.f_questions.Frage2.delete('1.0',END)
        tags = []
        if len(sel_tags) == 0:
            for b in settings['Highlight_Buttons']:
                tags.append(b[1])
        else:
            for b in settings['Highlight_Buttons']:
                if b[1] in sel_tags:
                    tags.append(b[1])
        verb('--Storing selection for tags: '+str(tags))
                
        if Einheit in settings.keys():
            verb('--Removing previously highlighted but uncoded items for: '+Einheit)
            for e in settings[Einheit].keys():
                if settings[Einheit][e]['Done']==0:
                    del settings[Einheit][e]
                    verb('----Removed: '+e)
        else:
            settings[Einheit] = {}
            
        rem = []
        try:
            for tag in tags:
                verb('Looking for highlights: '+tag)
                i = 0
                while i in range(0,len(self.Artikel.tag_ranges(tag))):
                    i0 = i/2+1
                    ident = tag + str(i0)
                    while ident in settings[Einheit]:
                        i0 = i0 + 1
                        ident = tag + str(i0)                        
                    settings[Einheit][ident] = {}
                    start_tag = self.Artikel.tag_ranges(tag)[i]
                    end_tag = self.Artikel.tag_ranges(tag)[i+1]
                    start_z = str(start_tag) + ' linestart'
                    end_z = str(end_tag) + ' lineend'
                    verb('Found: '+str(tag)+ '; ' + str(i)+ '; ' + str(i/2)+ '; From: ' + str(start_tag)+ '; To: ' + str(end_tag))
                    verb('Naming it: '+ident)
                    sel_txt = bereinigen(self.Artikel.get(start_tag,end_tag))
                    verb('String: '+sel_txt)
                    labeltext = tag + ': ' + sel_txt[:40]
                    if len(sel_txt) > 40: labeltext = labeltext + '...'
                    select = bereinigen(self.Artikel.get(start_z,end_z))
                    verb('In Paragraph: '+select)
                    settings[Einheit][ident]['Start'] = start_tag
                    settings[Einheit][ident]['End'] = end_tag
                    settings[Einheit][ident]['Fulltext'] = select
                    settings[Einheit][ident]['Wording'] = sel_txt
                    settings[Einheit][ident]['Done'] = 0
                    settings[Einheit][ident]['Label'] = labeltext
                    settings[Einheit][ident]['Typ']=tag
                    
                    rem.append((tag,start_tag,end_tag))
                    i = i + 2
        except:
            verb('Not able to output highlighted text')

        for e in rem:
            self.Artikel.tag_remove(e[0],e[1],e[2])              

        verb(baum_schreiben(settings[Einheit]))
        verb(str(len(settings[Einheit].keys()))+' Units highlighted')

    def unit_select(self,Einheit):
        log('Calling Function: Unit Select')
        global prog_pos
        global dta_pos
        u = self.store_var(Einheit,store=0)
        verb('Selection:'+str(u))
        settings['Fulltext'] = settings[Einheit][u]['Fulltext']
        self.Artikel.tag_add(settings[Einheit][u]['Typ'],settings[Einheit][u]['Start'],settings[Einheit][u]['End'])
        settings['Current_Unit'] = u
        dta_pos = [Einheit,u,'-','-']
        verb(str(dta_pos))
        if not dta_pos[0] in storage.keys():
            storage[dta_pos[0]] = {}
        storage[dta_pos[0]][dta_pos[1]] = {}
        storage[dta_pos[0]][dta_pos[1]]['Fulltext'] = settings[Einheit][u]['Fulltext']
        storage[dta_pos[0]][dta_pos[1]]['#TN'] = settings[Einheit][u]['Wording']
        storage[dta_pos[0]][dta_pos[1]]['Wording'] = settings[Einheit][u]['Wording']
        storage[dta_pos[0]][dta_pos[1]]['#TS'] = time.time()

############################################
##                                        ##
##       Grundfunktionen                  ##
##                                        ##
############################################


    def anzeigen(self): ##Display the text
        log('Calling Function: Anzeigen')
        global def_val
        dn = storage['ID']
       
        settings['Fulltext'] = 'Nogrod 1.0(b)\n\n'
        settings['Fulltext_Art'] = bereinigen(settings['Fulltext'])

        verb(settings['Fulltext'])
        self.ausblenden() ##First try to remove the text to prevent redundancy

        cols = [3,4]       
        if len(settings['Fulltext']) > 5:
            self.yscroller = Scrollbar(self, orient=VERTICAL)
            self.yscroller.grid(row=1, rowspan=6, column=cols[1], sticky=W+N+S)
            self.Artikel = Text(self, width=80, height=30, bg=settings['CSS']['Console']['text']['BG'],
                                fg=settings['CSS']['Console']['text']['FG'],
                                wrap=WORD, yscrollcommand=self.yscroller.set, relief=GROOVE,
                                font = settings['CSS']['Console']['text']['Font'], takefocus = 0)
            self.Artikel.grid(row=1, rowspan=6, column=cols[0], sticky=N+E+S+W)      
            self.yscroller["command"] = self.Artikel.yview
            self.Artikel.insert(INSERT, settings['Fulltext'])
            settings['Text_Aktiv'] = 1

            self.quick = Menubutton(self,text="Quick Tools", relief=RAISED)
            self.quick.grid(row=0,column=cols[0],sticky=N+E+W)
            self.quick.menu = Menu(self.quick, tearoff=0)
            self.quick["menu"] = self.quick.menu

            self.quick.menu.add_command(label="Variable overview",command=self.show_variables)
            self.quick.menu.add_command(label="Descriptives of one variable",command=self.show_descriptives)
            self.quick.menu.add_command(label="Write this output to a textfile",command=self.export_output)

            for t in settings['CSS']['Console'].keys():
                styleset = settings['CSS']['Console'][t]
                self.Artikel.tag_config(t,foreground=styleset['FG'],background=styleset['BG'], font = styleset['Font'])
            
        

##            self.etui = Frame(self, borderwidth=2, bg=farbton_text)
##            self.etui.grid(row=0, column=cols[0], columnspan=2,sticky=S+W+E)
##            self.Artikel.tag_config('worddetect', underline=1, font = (settings['Font'], settings['Fontsize'], "bold"))
##            
##            for i in range(0,len(settings['Highlight_Buttons'])):
##                self.Artikel.tag_config(settings['Highlight_Buttons'][i][1], background=settings['Highlight_Buttons'][i][2])
##                self.etui.leuchtstift = Button(self.etui, text = settings['Highlight_Buttons'][i][0], width=5, command=CMD(self.mark,settings['Highlight_Buttons'][i][1]), background=settings['Highlight_Buttons'][i][2], takefocus = 0)
##                self.etui.leuchtstift.grid(row=0, column=i+3, sticky=W)

##            if 'Hotwords' in settings.keys():
##                self.etui.wd = Button(self.etui, text = "Highl.", width=4, command=CMD(self.mark,'worddetect'), background='#ffffff', takefocus = 0)
##                self.etui.wd.grid(row=0, column=0, sticky=W)
##                if 'Auto_Highlight' in settings.keys():
##                    if settings['Auto_Highlight'] in [1,'1']:
##                        self.mark('worddetect')
##            self.etui.tippex = Button(self.etui, text = "", width=4, command=CMD(self.mark,'blank'), background='#ffffff', takefocus = 0)
##            self.etui.tippex.grid(row=0, column=1, sticky=W)
##            self.etui.columnconfigure(1,minsize=50)
                
           
        else:
            fehlermeld = '\n\nArticle: "' + storage['ID'] + '"\nFolder: "' + settings['Text_Folder'] + '"'
            #self.message("Runtime-Error04",add=fehlermeld)
            settings['Text_Aktiv'] = 0
            verb('ERROR: No Article: '+fehlermeld)

    def verbout(self,zeile,style='text'):
        if settings['Text_Aktiv'] == 1:
            self.Artikel.insert(END, zeile,style)
            self.Artikel.update()
            self.Artikel.see(END)
        else:
            try:
                #print(zeile,end="")
                CMD(print,zeile,end="")
            except Exception as f:
                print(zeile),
        return zeile

    def clean_all_tags(self,sel_tag=[]):
        global settings
        log('Calling Function: Clean All Tags')
        tags = []
        if len(sel_tag) == 0:
            for b in settings['Highlight_Buttons']:
                tags.append(b[1])
        else:
            for b in settings['Highlight_Buttons']:
                if b[1] in sel_tag:
                    tags.append(b[1])
        verb('--Only removing tags: '+str(tags))

        for tag_id in tags:
            try:
                self.Artikel.tag_remove(tag_id,1.0,END)
            except:
                verb('--Tag removal impossible for: '+tag_id)

    def mark(self,tag_id,ls=[]): ##Highlighting the selected text
        log('--Marking as: '+tag_id,pos=0)
        if tag_id == 'blank':
            for tag_desc in settings['Highlight_Buttons']:
                tag_id = tag_desc[1]
                try:
                    self.Artikel.tag_remove(tag_id,SEL_FIRST,SEL_LAST)
                except:
                    if not settings['Verbose'] == '0':
                        verb('----No selection. Highlighting impossible')
        elif tag_id == 'worddetect':
            try:
                verb(str(self.Artikel.tag_ranges('worddetect')[1]))
                self.Artikel.tag_remove('worddetect',1.0,END)
                verb('tags wieder entfernt')
                self.etui.wd['relief']=RAISED
            except:
                wlist=[]
                if ls == []:
                    if settings['Country'] in settings['Hotwords'].keys():                    
                        verb('Suchen nach Worten aus '+str(settings['Country']))
                        verb('Hotwords: '+str(settings['Hotwords'][settings['Country']]))
                        wlist = settings['Hotwords'][settings['Country']]
                    else:
                        verb('ERROR: No Keywords for this country')
                else:
                    wlist = ls
                
                for wort in wlist:
                    verb(wort)
                    start = '1.1'
                    while not start == '':
                        a = self.Artikel.search(wort,start,END)
                        verb(str(a)+':')
                        if not a == '':
                            anfang = self.Artikel.search(' ',a,backwards=TRUE)
                            ende = self.Artikel.search(' ',a,forwards=TRUE)
                            start = ende
                            if math.floor(float(anfang)) < math.floor(float(a)):
                                anfang = str(int(math.floor(float(a))))+".0"
                            self.Artikel.tag_add('worddetect',anfang,ende)
                            if math.floor(float(ende)) > math.floor(float(a)):
                                ende = str(1+int(math.floor(float(a))))+".0"
                            self.Artikel.tag_add('worddetect',anfang,ende)
                        else:
                            start = ''
                            anfang = ''
                            ende = ''
                        verb('>'+str(a)+str(anfang)+str(ende))
                self.etui.wd['relief']=SUNKEN
        else:
            try:
                self.Artikel.tag_add(tag_id,SEL_FIRST,SEL_LAST)
            except:
                verb('No selection. Highlighting impossible')

    def ausblenden(self): #Remove text from window
        log('--(Ausblenden)',pos=0)
        try:
            self.yscroller.destroy()
            self.Artikel.destroy()
            settings['Text_Aktiv'] = 0
            self.quick.destroy() ##etui weg
        except:
            verb('--Unable to remove text')

    def intronase(self): ##Cut loops from the page history. Necessary for the correct application of the back-function
        log('--Intronase')
        current = settings['Page_History'][len(settings['Page_History'])-1]
        i = len(settings['Page_History'])-1
        while i > 0:
            i = i -1
            if settings['Page_History'][i] == current:
                verb('----Double in: '+str(settings['Page_History']))
                while len(settings['Page_History']) > i:
                    settings['Page_History'].pop(i)
                settings['Page_History'].append(current)
        verb('----New History: '+str(settings['Page_History']))
 
    def hilfe_zu(self,htext,event=0): ##Display the help-text for a variable in a new window
        global settings
        log('Calling Function: Hilfe zu')
        storage['Helptexts'] = storage['Helptexts'] + 1
        htext = htext.replace("#","\n")
        if settings['Python_Version']==3:
            messagebox.showinfo("Help", htext)
        else:
            tkMessageBox.showinfo("Help", htext)


    def message(self,m_id,m_type=1,var='Err_Msg',add=''):
        log('Calling Function: Message with message: '+m_id+' in Codebook-Variable: '+var)
        title = m_id
        text = self.namegetter(var,m_id)+add
        r = 0

        text = text.replace("#","\n")

        if settings['Python_Version']==3:
            if m_type == 1:
                messagebox.showwarning(title,text)
            elif m_type == 2:
                messagebox.showinfo(title,text)
            elif m_type == 3:
                r = messagebox.askokcancel(title,text)            
        else:
            if m_type == 1:
                tkMessageBox.showwarning(title,text)
            elif m_type == 2:
                tkMessageBox.showinfo(title,text)
            elif m_type == 3:
                r = tkMessageBox.askokcancel(title,text)
        return r

    def insecure(self,variabel,event=0): ##Report an insecurity
        log('Calling Function: Insecure')
        exp_file = open('..\insec.txt','a')
        zeile = settings['Coder'] + '\t' + storage['ID'] + '\t' + variabel + '\t' + str(dta_pos) + '\t' + str(time.ctime()) + '\t' + bereinigen(settings['Fulltext']) + '\n'
        exp_file.write(zeile)
        exp_file.close()           
        self.message("Info01")
        
    def pause(self): ##Making a break (disable the question-frame and wat for the button to be pressed again)
        log('Calling Function: Pause')
        storage['Breaks'] = storage['Breaks'] + 1
        if settings['Layout'] == 'Lefty':
            c1 = 1
        else:
            c1 = 5

        if settings['Break'] == 0:
            settings['Break'] = time.time()
            self.f_bottomline.b_break["text"] = 'End Break'
            self.pausentext = Text(self, width=80, height=3, wrap=WORD, relief=FLAT, font = (settings['Font'], "9"), bg='#ffaaaa')
            self.pausentext.grid(row=4, column=c1, sticky=N+E+S+W)
        else:
            zeit = time.time() - settings['Break']
            settings['Break'] = 0
            settings['Break_Time'] = settings['Break_Time'] + zeit
            self.f_bottomline.b_break["text"] = 'Break'
            self.pausentext.destroy()

    def cini_schreiben(self):
        log('Calling Function: CINI schreiben')
        c_file = open(settings['Settings'],'w')
        c_file.write('##Coder Information:\n##This ini-File is formatted as the codebook:\n##3 Lines (Question, Information, Help) before any values may be defined.\n\n\n')
        for variable in ['Coder-Settings','Default-Values']:
            c_file.write('[')
            c_file.write(variable)
            c_file.write(']\nFrage: ')
            c_file.write(cini[variable][0])
            c_file.write('Anweisung: ')
            c_file.write(cini[variable][1])
            c_file.write('Hilfe: ')
            c_file.write(cini[variable][4])
            if variable == 'Default-Values':
                #verb('Default Values: '+baum_schreiben(def_val,trunc=40))
                for dvar in sorted(def_val.keys()):
                    c_file.write(dvar)
                    c_file.write(':')
                    dv = str(def_val[dvar])
                    dv = dv.replace("#","")
                    c_file.write(dv)
                    c_file.write('\n')                
            else:                
                for i in range(0,len(cini[variable][2])):
                    c_file.write(cini[variable][3][i])
                    c_file.write(':')
                    c_file.write(cini[variable][2][i])
                    c_file.write('\n')
            c_file.write('\n\n')
        c_file.write('\n\n\n')
        c_file.close()


    def load_cset(self,csettings):
        outcodes = []
        for c in csettings[2]:
            if len(c) > 0:
                if c[0] in ['[','{','(']:
                    try:
                        code=eval(c)
                    except:
                        code = c
                elif c[0] in ['0','1','2','3','4','5','6','7','8','9','-']:
                    try:
                        code = float(c)
                        if code == int(code):
                            code = int(code)
                    except:
                        code = c
                else:
                    code = c
            else:
                code = 0
            outcodes.append(code)
        csettings[2] = outcodes
        return csettings
       

############################################
##                                        ##
##       I/O Functions                    ##
##                                        ##
############################################

def verbout(zeile,style='text',master=''):
    if not master == 'silent':
        try:
            master.Artikel.insert(END, zeile,style)
            master.Artikel.update()
            master.Artikel.see(END)
        except:
            try: #Python 3
                print(zeile,end="")
            except:
                print(zeile),
    return zeile


def artikelholen(ID): ##Get a text from the folder specified
    log('        Artikelholen',pos=0)
    at = ""
    if ID[-4] == '.':
        art_filen = settings['Text_Folder'] + ID
    else:
        art_filen = settings['Text_Folder'] + ID + '.txt'
    if settings['Verbose'] == '1':
        verb('Loading: ' + str(art_filen))
    try:
        art_file = open(art_filen, 'r')
        F_list = art_file.readlines()
        if settings['Verbose'] == '1':
            verb(str(F_list))
        art_file.close()
        textmine(F_list)
        F_listneu = []
        for zeile in F_list: ### Replace unusual characters for a correct display of the text
            zeileneu = bereinigen(zeile,lb=1,uml=1)
            F_listneu.append(zeileneu)                    
        for zeile in F_listneu:
            at = at + zeile
        verb(art_filen)
    except IOError:
        at = "-"

    except:
        ##print("EXCEPTIONAL ERROR WHEN LOADING ARTICLE: "+ID)
        pass
    return at       

def textmine(linelist): ##A list of all lines within the text are submitted to this function. May be used to change default values.
    global def_val
    global settings

    ## Functions for analyzing texts loaded into the tool.
           


def get_codebook(filename): ##Load the codebook from a given file. Returns a codebook-Dictionary for use within the tool.
    log('Calling Function: Get Codebook from: '+filename)
    #Codebook-Enries have the form:
    #
    #[Name of the variable]
    #Question
    #Coder Information
    #Helptext
    #Code1:Item1
    #Code2:Item2
    #...
        
    cb = {}
    cb_file = open(filename, 'r')
    F_list = cb_file.readlines()
    cb_file.close()
    i = 0
    while i < (len(F_list)-1):
        if F_list[i][0] == '[':
            varname = F_list[i][1:-2]
            varfrage = F_list[i+1]
            varanw = F_list[i+2]
            varhilfe = F_list[i+3]

            if varfrage[:6] == 'Frage:':
                varfrage = varfrage[7:]
            if varanw[:10] == 'Anweisung:':
                varanw = varanw[11:]
            if varhilfe[:6] == 'Hilfe:':
                varhilfe = varhilfe[7:]

            i = i + 4
            optionen = []
            codes = []
            while not F_list[i] == '\n':
                cod_zeile = F_list[i]
                cut = cod_zeile.find(':')
                if cut == -1:
                    opt = cod_zeile
                    cod = cod_zeile
                else:
                    opt = cod_zeile[cut+1:-1]
                    cod = cod_zeile[:cut]
##                while '#' in opt:
##                    c = opt.find('#')
##                    opt = opt[:c]+'\n'+opt[c+1:]
                optionen.append(opt)
                codes.append(cod)
                i = i + 1

            optout = str(optionen)
            if len(optout)>30: optout = optout[:30]+'... ('+str(len(optout))+'characters)'
            verb(varname+optout)

            if varname in cb.keys():
                verb('ERROR: Variable "'+varname+'" already defined. Overwriting previous variable')

            cb[varname] = []
            cb[varname].append(varfrage) #codebook['variabel'][0] is the question as string (including final linebreak)
            cb[varname].append(varanw)   #codebook['variabel'][1] is the coder information
            cb[varname].append(optionen) #codebook['variabel'][2] is a list containing all labels
            cb[varname].append(codes)    #codebook['variabel'][3] is a list containing all codes (parallel to [2]
            cb[varname].append(varhilfe)    #codebook['variabel'][4] is the helptext
        i = i + 1
    
    verb('Codebook loaded successfully')
    return cb

def add_varlist(vname,labels,codes=[],excludes=[],retain=0):
    global codebook

    if codes == []:
        codes = list(labels)
    if not len(labels) == len(codes):
        verb('ERROR: Labels and Codes are different in length. Using Labels as codes')
        codes = list(labels)
    remi = []
    for e in excludes:
        for i in range(len(labels)):
            if labels[i] == e or codes[i] == e:
                remi.append(i)
    remi = sorted(remi, reverse = True)
    for i in remi:
        labels.pop(i)
        codes.pop(i)

    if vname in codebook.keys():
        if retain == 0:
            codebook[vname][2] = []
            codebook[vname][3] = []
        else:
            codebook[vname][2] = codebook[vname][2][:retain]
            codebook[vname][3] = codebook[vname][3][:retain]
        for i in range(len(labels)):
            codebook[vname][2].append(labels[i])
            codebook[vname][3].append(codes[i])
    else:
        verb('ERROR: Variable not in codebook: "'+vname+'"')


def define_styleset(styleset="Default"):
    global settings

    settings['CSS'] = {'Console':{}}

    if available('Fontsize'):
        fs = int(settings['Fontsize'])

    if styleset == "Hacker":
        settings['CSS']['Console']['text']={'BG':"#000000",'FG':"#20ff20",'Font':('fixedsys', fs)}
        settings['CSS']['Console']['warning']={'BG':"#40cc40",'FG':"#000000",'Font':('fixedsys', fs)}
        settings['CSS']['Console']['progress']={'BG':"#000000",'FG':"#00ff00",'Font':('fixedsys', fs)}
        settings['CSS']['Console']['table']={'BG':"#000000",'FG':"#ffff00",'Font':('fixedsys', fs)}
        settings['CSS']['Console']['title']={'BG':"#000000",'FG':"#ffffff",'Font':('fixedsys', fs,'underline')}
    elif styleset == "Tardis":
        settings['CSS']['Console']['text']={'BG':"#003b6f",'FG':"#ffffff",'Font':('Verdana', fs)}
        settings['CSS']['Console']['warning']={'BG':"#99b0c5",'FG':"#000000",'Font':('Verdana', fs,'bold')}
        settings['CSS']['Console']['progress']={'BG':"#ffffff",'FG':"#003b6f",'Font':('Courier', fs,'bold')}
        settings['CSS']['Console']['table']={'BG':"#ffffff",'FG':"#003b6f",'Font':('Courier', fs-1,'bold')}
        settings['CSS']['Console']['title']={'BG':"#003b6f",'FG':"#ffffff",'Font':('Verdana', fs+2,'bold')}
    elif styleset == "Simple":
        settings['CSS']['Console']['text']={'BG':"#ffffff",'FG':"#000000",'Font':('Courier', fs)}
        settings['CSS']['Console']['warning']={'BG':"#ffffff",'FG':"#ff0000",'Font':('Courier', fs)}
        settings['CSS']['Console']['progress']={'BG':"#ffffff",'FG':"#000000",'Font':('Courier', fs)}
        settings['CSS']['Console']['table']={'BG':"#ffffff",'FG':"#000000",'Font':('Courier', fs-1)}
        settings['CSS']['Console']['title']={'BG':"#ffffff",'FG':"#000000",'Font':('Courier', fs,'bold underline')}      
    elif styleset == "Debug":
        settings['CSS']['Console']['text']={'BG':"#ffffff",'FG':"#000000",'Font':('Arial', fs)}
        settings['CSS']['Console']['warning']={'BG':"#ffcccc",'FG':"#000000",'Font':('Arial', fs,'bold')}
        settings['CSS']['Console']['progress']={'BG':"#00ffff",'FG':"#005000",'Font':('Courier', fs,'bold')}
        settings['CSS']['Console']['table']={'BG':"#ffff80",'FG':"#000000",'Font':('Courier', fs-1,'bold')}
        settings['CSS']['Console']['title']={'BG':"#cccccc",'FG':"#000050",'Font':('Arial', fs+2,'bold underline')}
    else:
        settings['CSS']['Console']['text']={'BG':"#ffffff",'FG':"#003b6f",'Font':('Arial', fs)}
        settings['CSS']['Console']['warning']={'BG':"#ffcccc",'FG':"#000000",'Font':('Arial', fs,'bold')}
        settings['CSS']['Console']['progress']={'BG':"#ffffff",'FG':"#003b6f",'Font':('Courier', fs)}
        settings['CSS']['Console']['table']={'BG':"#ffffff",'FG':"#003b6f",'Font':('Courier', fs-1,'bold')}
        settings['CSS']['Console']['title']={'BG':"#ffffff",'FG':"#003b6f",'Font':('Arial', fs+2,'bold underline')}
    


def log(name,pos=1):
    try:
    ## Making Log-Entries for called functions and events.
        global settings
        global prog_pos
        global dta_pos
        if settings['Verbose'] == '2':
            if pos == 1:
                settings['Verb_Log']=settings['Verb_Log']+name+'. Prog_Pos='+prog_pos+'; DTA-Pos='+ str(dta_pos) + '\n'
            else:
                settings['Verb_Log']=settings['Verb_Log']+name+'\n'
        if pos == 1:
            settings['Path_Log']=settings['Path_Log']+name+'. Prog_Pos='+prog_pos+'; DTA-Pos='+ str(dta_pos) + '\n'
        else:
            settings['Path_Log']=settings['Path_Log']+name+'\n'
    except:
        pass ##Logging not possible. Probably missing settings or prog_pos because nogrod was opened as a module.

def verb(name,stage=2,nl=1):
    ## Making Verbose-Entries. 
    global settings
    if int(settings['Verbose']) >= int(stage):
        settings['Verb_Log']=settings['Verb_Log']+name
        if nl==1:
            settings['Verb_Log']=settings['Verb_Log']+'\n'

def display_dset(dset, sep='blank'):
    ## Uses display_table to display a dataset with row numbers

    if type(dset) == dict:
        data = dset
        varlist = sorted(data.keys())
    else:
        data = dset[0]
        varlist = dset[1]
        
    rownum = range(1,len(data[varlist[0]])+1)

    table = {}
    for v in varlist:
        table[v] = {}
        for i in range(len(data[varlist[0]])):
            table[v][rownum[i]] = data[v][i]

    return display_table(table,sep,varlist)


def display_table(table, sep="blank", cols_pre=[], rows_pre=[]):
    ## Each Cell has the format: table[x][y]

    maxx = 0
    maxy = 0
    cols = []
    rows = []

    for x in sorted(table.keys()):
        if len(x) > maxx: maxx = len(x)
        cols.append(x)
        for y in table[x].keys():
            if len(str(y)) > maxy: maxy = len(str(y))
            rows.append(y)
            if type(table[x][y]) == str:
                if len(table[x][y]) > maxx: maxx = len(table[x][y])
    rows = get_unique(rows)

    if len(cols_pre) > 0:
        ncols = []
        for c in cols_pre:
            if c in cols:
                ncols.append(c)
        cols = ncols

    if len(rows_pre) > 0:
        nrows = []
        for r in rows_pre:
            if r in rows:
                nrows.append(r)
        rows = nrows

    if sep in ["blank",' ']:
        rowf = "{:>"+str(maxy+1)+"}"
        colf = "{:>"+str(maxx+1)+"}"

        tvis = '\n'+rowf.format('')
        for x in cols:
            tvis = tvis + colf.format(x)
            
        for y in rows:
            tvis = tvis+'\n'+rowf.format(y)
            for x in cols:
                if y in table[x].keys():
                    tvis = tvis + colf.format(table[x][y])
                else:
                    tvis = tvis + colf.format('')
    elif sep in ['tab','\t']:
        tvis = '\n'
        for x in cols:
            tvis = tvis + '\t' + str(x)
        for y in rows:
            tvis = tvis + '\n'+str(y)
            for x in cols:
                if y in table[x].keys():
                    tvis = tvis + '\t' + str(table[x][y])
                else:
                    tvis = tvis + '\t'

    tvis = tvis + '\n'
    return tvis

def curr():
    log('Calling Function: Curr')
    global dta_pos
    global storage
    if dta_pos[0] == '-':
        curr_tree = storage
    elif dta_pos[2] == '-':
        curr_tree = storage[dta_pos[0]][dta_pos[1]]
    elif dta_pos[4] == '-':
        curr_tree = storage[dta_pos[0]][dta_pos[1]][dta_pos[2]][dta_pos[3]]
    #verb(str(curr_tree))
    return curr_tree

def critical_abort():
    Anzeige.destroy()

def available(key):
    ##Checks whether a setting is available in the settings-directory.
    ##Available means: The key is present and if the value is a string this string is non-empty
    global settings
    log('Checking availability of setting: '+key)
    if key in settings.keys():
        if type(settings[key]) == str:
            if len(settings[key]) > 0:
                avail = True
            else:
                avail = False
        else:
            avail = True
    else:
        avail = False

    if avail:
        verb('--Setting found: '+str(settings[key]))
    else:
        verb('--Not available')
    return avail

def get_todo(fname):
    log('Calling Function: Get TODO')
    try:
        todo_file = open(fname,'r')
        art_name = todo_file.readline()
        art_name = art_name[:len(art_name)-1]
        verb('Next Item To do: '+ art_name)
        todo_file.close()
    except:
        verb('Unable to get next Item to do in: '+fname)
        art_name = ''
    if art_name == '':
        message('Runtime-Error02')
    return art_name

def baum_schreiben(tdic, einr = 0, inherit = '', spc = '    ', trunc=0, lists=False):
    if type(tdic)==dict:
        if inherit == '': inherit = '{\n'
        for k in sorted(tdic.keys()):
            if type(tdic[k]) == dict:
                inherit = inherit + einr*spc + str(k) + ': ' + baum_schreiben(tdic[k],einr+1,inherit = '{\n',trunc=trunc)
            elif type(tdic[k]) == list and lists:
                inherit = inherit + einr*spc + str(k) + ': ' + baum_schreiben(tdic[k],einr+1,inherit = '[\n',trunc=trunc)
            else:
                value = tdic[k]
                if type(value)==str:
                    value = "'"+value+"'"
                elif type(value) in [int,float]:
                    value = str(value)
                elif type(value) == list:
                    value = str(value)
                else:
                    value = str(value)
                if len(value) > trunc and trunc > 0:
                    tail = int(trunc/2)
                    value = value[:tail] + '...'+value[-tail:] + ' ('+str(len(value))+' characters)'
                    
                inherit = inherit + einr*spc + str(k) + ': '+ value + '\n'

        inherit = inherit + einr*spc + '}\n'
    elif type(tdic)==list:
        if inherit == '': inherit = '[\n'
        for e in tdic:
            if type(e) == dict:
                inherit = inherit + einr*spc + baum_schreiben(e,einr+1,inherit = spc+'{\n',trunc=trunc)
            elif type(e) == list and lists:
                inherit = inherit + einr*spc + baum_schreiben(e,einr+1,inherit = spc+'[\n',trunc=trunc)
            else:
                value = e
                if type(value)==str:
                    value = "'"+value+"'"
                elif type(value) in [int,float]:
                    value = str(value)
                elif type(value) == list:
                    value = str(value)
                else:
                    value = str(value)
                if len(value) > trunc and trunc > 0:
                    tail = int(trunc/2)
                    value = value[:tail] + '...'+value[-tail:] + ' ('+str(len(value))+' characters)'
                    
                inherit = inherit + einr*spc + value + ',\n'

        inherit = inherit + einr*spc + ']\n'
            
    return inherit

def tempout(d, fname = 'dict_temp.txt'):
    outt = baum_schreiben(d)
    of = open(fname,'w')
    of.write(outt)
    of.close()

def log_settings(out_vars,outfile,append=1):
    global codebook

    if append == 1:
        outf = open(outfile,'a')
    else:
        outf = open(outfile,'w')

    outf.write('\n----------------\nLogging Information from Nogrod.\nFunction: '+str(storage['Methode'])+'\n'+'----------------------\n\n')
    outf.write('User defined settings: \n')
    for v in out_vars:
        outf.write('Variable "'+v+'":\n')
        outf.write(' -Question: '+codebook[v][0]+'     '+codebook[v][1])
        outf.write(' -Answer: '+str(storage[v])+'\n\n')

    outf.close()


##############################################
#
# I / O Functions without reference to Nogrod
#
##############################################


def get_dataset(fname, header=1, sep='\t',master=''):
    data = 0

    if sep == 'xlsx' or ".xlsx" in fname:
        if settings["Excel"]==1:
            verb("Loading Excel-File: "+fname)
            data = get_xlsx(fname, header,master=master)
        else:
            data = 0
        if data == 0:
            verbout("\n",master=master)
            verbout("Unable to load this file as Excel-File.","warning",master=master)
            verbout("\nPackage 'openpyxl' is not available","warning",master=master)
            verbout("\n",master=master)
        elif type(data)==str:
            verbout("\n",master=master)
            verbout("Unable to load this file as Excel-File.\n","warning",master=master)
            verbout(data,"warning",master=master)
            verbout("\n\nTrying to load it as a text file with tabstopps.",master=master)
            data = 0
            sep = '\t' ## Set to this default to probably save the day.
                    
    if data == 0:
        try:
            data = get_data(fname,header,sep,master=master)
        except Exception as f:
            data = [{},[],'Could not get data.','ERROR: Could not get data']
            print(f)
            
        if data[0] == 0:
            data[0] = 'invalid'
            description = ''
    return data

def write_dataset(data,filename,header=1,sep='\t'):
    t = write_data(data[0],data[1],filename,header,sep)
    return t

    
def get_varnames(filename,header=1,sep='\t'):
    log('Calling Function: Get Varnames. Separator: '+sep+'; Header: '+str(header))
    try:
        inp_file = open(filename, 'r')
        varline = inp_file.readline()[:-1]
        inp_file.close()
        vlist = []
    except:
        vlist = 0
    if not vlist == 0:
        varnames = varline.split(sep)
        for varname in varnames:
            if len(varname) == 0: varname = '_'
            if varname[0] in ['"',"'"]: varname = varname[1:]
            if varname[-1] in ['"',"'"]: varname = varname[:-1]
            if len(varname) == 0: varname = '_'
            nr = 1
            vlab = varname
            while varname in vlist:
                varname = vlab + "{0:02}".format(nr)
                nr = nr + 1
            vlist.append(varname)       
        if header == '0':
            nv = []
            for i in range(len(vlist)): nv.append('VAR'+"{0:02}".format(i))
            vlist = nv        
    return vlist

def get_xlsx(filename, header=1, sheet=0, master=''):
    errmsg = ""
    try:
        wb = load_workbook(filename=filename, data_only=True)
        sheets = wb.sheetnames
    except:
        errmsg+="The file '"+filename+"' could not be loaded as Excel file.\n"
        outdata = errmsg
        verb(errmsg)
        sheets = []

    if len(sheets)>0:

        if type(master)==str: ## If this function was called without open Nogrod window
            sheet = sheets[0]
        elif len(sheets)>1:
            question = "Multiple Sheets detected in this workbook. Which one do you want to open?\nNames of all sheets: '"+"'; '".join(sheets)+"'"
            sheet = ''
            while not sheet in sheets and not sheet == None:
                sheet = simpledialog.askstring("Select Sheet",question,parent=master,initialvalue=sheets[0])
            if sheet==None:
                sheet = sheets[0]
        else:
            sheet = sheets[0]

        sheet_temp = wb[sheet]
        a = sheet_temp.dimensions
        
        rmin = sheet_temp.min_row
        rmax = sheet_temp.max_row
        cmin = sheet_temp.min_column
        cmax = sheet_temp.max_column

        data = {}
        varlist = []

        vcount = 1
        for c in sheet_temp.iter_cols(min_row=rmin,max_row=rmax,
                                  min_col=cmin, max_col=cmax):
            column = []
            for v in c:
                val = v.value
                if val == None:
                    val = ''
                column.append(str(val))

            if header in [1,'1']:
                vname = column[0]
                column = column[1:]
                if vname == '':
                    vname = "Var_{0:02}".format(vcount)
                if vname in varlist:
                    vlab = vname
                    vnum = 1
                    vname = "{0}_{1:02}".format(vlab,vnum)
                    while vname in varlist:
                        vnum+=1
                        vname = "{0}_{1:02}".format(vlab,vnum)
            else:
                vname = "Var_{0:02}".format(vcount)

            vcount +=1
            data[vname] = column
            varlist.append(vname)
     
        summary = "Read Excel Worksheet.\n"+str(len(data.keys()))+" Variables\n"+str(len(column))+" Cases."
        errmsg=""
        outdata = [data,varlist,summary,errmsg]

    return outdata


def get_data(filename, header=1, sep='\t', varlist = [],verbose=2,master=''):
    global settings
    errmsg = ''
        
    try:
        inp_file = open(filename, 'r',encoding="latin-1")
        dtalines = inp_file.readlines()
        inp_file.close()
        data_dic = {}
    except:
        data_dic = 0
        errmsg = 'The file "'+filename+'" is not valid or does not exist'

    if data_dic == 0:
        summary = 'Unable to load data.'
    else:
        if varlist == []: varlist = get_varnames(filename,header,sep) #Get variable names
        first_line = 0
        if header in [1,'1']: first_line = 1 #Remove header
        for var in varlist: data_dic[var] = [] #Set up data dic

        for i in range(first_line,len(dtalines)):
            dtaline = dtalines[i]
            if len(dtaline) < 2:
                errmsg = errmsg + 'ERROR: line '+str(i)+' is empty. Ignoring Line\n'
            else:
                if dtaline[-1] == '\n':
                    dtaline = dtaline[:-1]
                dta = dtaline.split(sep)
                if len(dta) == len(varlist):
                    for k in range(len(dta)):
                        value = dta[k]
                        if value == ' ': value = ''
                        if len(value) > 1:
                            if value[0] in ['"',"'"] and value[-1] in ['"',"'"]: value = value[1:-1]
                        data_dic[varlist[k]].append(value)
                elif len(dta) > len(varlist):
                    errmsg = errmsg + 'ERROR: line '+str(i)+' has more values than variables (N='+str(len(dta))+'). Ignoring line\n'
                else:
                    errmsg = errmsg + 'ERROR: line '+str(i)+' has less values than variables (N='+str(len(dta))+'). Ignoring line\n'

        ncases = len(data_dic[varlist[0]])
        nvars = len(varlist)
        nerr = len(errmsg.split('\n'))-1
        summary = 'Data loaded.\n'+str(ncases)+' Cases\n'+str(nvars)+' Variables\n'+str(nerr)+' Invalid lines'

        if len(errmsg)>2: verb('Errors while loading file: "'+filename+'"\n'+errmsg,verbose)
        verb('\n'+summary,verbose)
        
    return [data_dic, varlist, summary, errmsg]



def write_xlsx(data,varlist,filename,header=1):
    ## Gets a well-formed data dictionary and a varlist from the write_data function.
    ## All the function does is to write an Excel-Sheet with an oblong table.
    wb = Workbook()
    ws = wb.active ## Current worksheet to fill

    if header == 1:
        ws.append(varlist)

    for i in range(len(data[varlist[0]])):
        row = []
        for v in varlist:
            row.append(data[v][i])
        ws.append(row)
    
    wb.save(filename)
    


def write_data(data,init_varlist,filename,header=1,sep='\t',verbose=2):
    errmsg = ''
    varlist = []
    varlen = []
    for v in init_varlist: #Check for completeness before starting
        if v in data.keys():
            if not v in varlist:
                varlist.append(v)
            else:
                lnum = 1
                lab = v
                vlab = v+"{0:02}".format(lnum)
                while vlab in varlist:
                    lnum +=1
                    vlab = v+"{0:02}".format(lnum)
                varlist.append(vlab)
                data[vlab] = data[v]
                errmsg = errmsg + 'Warning: Variable "'+v+'" twice in output. Renamed to '+vlab+'\n'
            varlen.append(len(data[v]))
        else:
            errmsg = errmsg + 'ERROR: Variable "'+v+'" not in data. Skipping variable\n'

    if not min(varlen) == max(varlen):
        errmsg = errmsg + 'ERROR: Not all variables have the equal amount of cases:\n'
        errmsg = errmsg + ' - Using the first variable as standard: '+str(varlen[0])+' cases.\n'
        remvar = []
        for v in varlist:
            if not len(data[v]) == varlen[0]:
                errmsg = errmsg + ' - Variable "'+v+'" deviant ('+str(len(data[v]))+' cases). Skipping variable\n'
                remvar.append(v)
        for v in remvar:
            varlist.remove(v)

    if sep=="xlsx":
        if not filename[:-5]==".xlsx":
            filename+=".xlsx"
            filename = filename.replace(".txt","")
            
    try:
        exp_file = open(filename,'w')
    except:
        varlist = []
        errmsg = errmsg + 'ERROR: File '+filename+' could not be created. Invalid filename.\n'

    if len(varlist) > 0:
        if sep == 'xlsx':
            exp_file.close()
            t = write_xlsx(data,varlist,filename,header)

        else:
            if header == 1:
                for k in range(len(varlist)):
                    exp_file.write(varlist[k])
                    if k < len(varlist)-1:
                        exp_file.write(sep)
                    else:
                        exp_file.write('\n')

            for i in range(len(data[varlist[0]])):
                for k in range(len(varlist)):
                    outval = data[varlist[k]][i]
                    if type(outval) == float:
                        if outval == int(outval):
                            outval = int(outval)
                    try:
                        exp_file.write(str(outval))
                    except:
                        exp_file.write('UNICODE_ERROR')
                    if k < len(varlist)-1:
                        exp_file.write(sep)
                    else:
                        exp_file.write('\n')

            exp_file.close()
        outtext = '\nFile "'+filename+'" successfully created.\n'
        outtext = outtext + str(len(data[varlist[0]])) + ' Cases\n'+str(len(varlist))+' Variables\n'
    else:
        outtext = '\nFile "'+filename+'" could not be created.\n'

    if len(errmsg)>2:
        verb('Errors while writing data to: "'+filename+'"\n'+errmsg,verbose)
    verb('\n'+outtext,verbose)
    
    return [outtext,errmsg,'']

def dim(dset): ## Alternate spelling
    return data_dim(dset)

def dset_dim(dset): ## Alternate spelling
    return data_dim(dset)

def data_dim(dset):
    if type(dset) in [list,tuple]:
        dset = dset[0]
    nvar = len(dset.keys())
    ncas = len(dset[list(dset.keys())[0]])
    return [nvar,ncas]

def get_directory(path='',ext=[]):
    ##List the content of a directory. Creates a list with complete paths of all files in a given directory with given extensions
    ##path: Path of the directory. If left empty, the directory is the current working directory
    ##ext: Extension as string or list of extensions as strings.
    
    if path == '': path = os.getcwd()
    path = path.replace('\\','/')
    
    dirlist = os.listdir(path)
    if type(ext) == str: ext = [ext]

    if ext == []:
        truelist = dirlist
    else:
        truelist = []
        for d in dirlist:
            for e in ext:
                if d[-len(e):] == e:
                    truelist.append(d)

    if not path[-1] == '/': path = path + '/'

    outlist = []
    for d in truelist:
        outlist.append(path+d)

    return outlist


def get_unique(liste):
    td = {}
    for element in liste:
        td[element] = 0
    return sorted(td.keys())

def get_sep(s):
    if s == '1':
        sep = '\t'
    elif s == '2':
        sep = ';'
    elif s == '3':
        sep = ','
    elif s == '4':
        sep = '";"'
    elif s == '5':
        sep = "xlsx"
    else:
        sep = '\t'
        verb('ERROR: Invalid choice for separator')
    return sep


############################################
##                                        ##
##       Basic Calculations               ##
##                                        ##
############################################


def sort_table(table,variables,mode=False):
    while len(variables) > 0:
        casevec = range(0,len(table[variables[0]]))
        sortvar = variables.pop()
        vartype = 'Num'
        sortcol = []
        for val in table[sortvar]:
            try:
                a = float(val)
            except:
                if not val == '':
                    vartype = 'Str'
            sortcol.append(val)
            
        if vartype == 'Num':
            for k in range(len(sortcol)):
                try:
                    sortcol[k] = float(sortcol[k])
                except:
                    sortcol[k] = 0               
        casevec = list(list(zip(*sorted(zip(sortcol, casevec),reverse=mode)))[1])
        outdic = {}
        for v in table.keys():
            outdic[v] = []
        for i in casevec:
            for v in table.keys():
                outdic[v].append(table[v][i])
        table = outdic

    return outdic


def sort_dataset(dset,variables,descend=False,master=''):
    if type(variables) == str: variables = [variables]
    verbout('\nSorting data by variables '+str(variables)+'\n',master=master)
    out_data = sort_table(dset[0],variables,descend)
    out_vars = dset[1]
    return [out_data,out_vars]


def dummy(caslist,keylist,mode='dicho',min_case=0,min_anz=0,master=''):
    log('Calling Function: Dummy')
    verb('First 10 cases: '+str(caslist[:10]))
    verb('First 10 keys: '+str(keylist[:10]))
    verb('Mode: '+str(mode))
    verb('Creating variables:')
    cat = get_unique(keylist)
    cases = get_unique(caslist)
    verbout('\nCalculating Dummy Table for ' +str(len(cases)) +' Groups and ' + str(len(cat)) + ' Values',master=master)    
    outdata = {}
    dummies = {}
    for case in cases:
        dummies[case] = {}
    outdata['#Group'] = []
    for c in cat:
        outdata[c] = []
        for case in cases:
            dummies[case][c]=0

    for i in range(len(caslist)):
        dummies[caslist[i]][keylist[i]] = dummies[caslist[i]][keylist[i]] + 1

    step = int(len(dummies.keys())/40)
    panz = 1
    if step<1:
        panz = int(40.0/len(dummies.keys()))
        step = 1
    verbout('\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    i = 0

    for case in cases:
        i = i + 1
        if i%step == 0:verbout('.'*panz,'progress',master=master)
        outdata['#Group'].append(case)
        outstring = case + '> '
        for c in cat:
            wert = dummies[case][c]
            if mode == 'dicho':
                if wert > 0:
                    wert = 1
                else:
                    wert = 0
            elif mode == 'log':
                wert = math.log(wert+1)
            outdata[c].append(wert)
            outstring = outstring + c + ': ' + str(wert) + '; '
        verb(outstring)
    verbout('\n','progress',master=master)

    if min_anz == 0 and min_case == 0:
        verb('No reduction of cases')
    else:
        outdata = desparse(outdata,min_case,min_anz,master=master)

    return outdata

def dummytab(data, cases, keyvar, mode='dicho', min_case=0, min_anz=0,master=''):
    ##Create a dummy table from a dataset with group variables and one nominal variable
    ##data: Data dictionary (no variables)
    ##cases: List of group variables
    ##keyvar: Nominal variable to be used for dummies
    ##mode: Method to create dummies: 'dicho', 'anz', 'log'
    ##min_case/min_anz: Minimum number of counts per group/category
    ##master: Window in which to print the output using verbout.

    log('Calling Function: Dummy for multiple case variables')
    verb('Mode: '+str(mode))
    verb('Creating variables:')
    keylist = data[keyvar]
    cat = get_unique(keylist)

    if type(cases)==str:cases=[cases]

    if not cases == []:
        dummies = {}
        for i in range(len(data[cases[0]])):
            case = ''
            for v in cases:
                case = case + data[v][i]+'-'
            case = case[:-1]
            dummies[case]={}
            for c in cat:
                dummies[case][c]=0
        caslist = dummies.keys()
    else:
        dummies = {}
        for i in range(len(data[cases[0]])):
            dummies[i] = {}
            for c in cat:
                dummies[i][c] = 0

    caslist = dummies.keys()

    for i in range(len(data[keyvar])):
        if cases == []:
            case = i
        else:
            case = ''
            for v in cases:
                case = case + data[v][i]+'-'
            case = case[:-1]
        dummies[case][data[keyvar][i]] = dummies[case][data[keyvar][i]] + 1
        dummies[case]['#CASEVAL']={}
        for v in cases:
            dummies[case]['#CASEVAL'][v] = data[v][i] 

    verbout('\n\nDummy table prepared.\n\n',master=master)

    outdata = {}       
    for c in cases:
        outdata[c] = []            
    for c in cat:
        outdata[c] = []

    step = int(len(dummies.keys())/40)        
    if step<1: step = 1
    verbout('Reshaping to output format: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    i = 0
    
    for case in sorted(dummies.keys()):
        i = i + 1
        for c in cases:
            outdata[c].append(dummies[case]['#CASEVAL'][c])
        for c in cat:
            wert = dummies[case][c]
            if mode == 'dicho':
                if wert > 0:
                    wert = 1
                else:
                    wert = 0
            elif mode == 'log':
                wert = math.log(wert+1)
            outdata[c].append(wert)
        if i%step == 0:verbout('.','progress',master=master)
    verbout('\n','progress',master=master)

    if min_anz == 0 and min_case == 0:
        verb('No reduction of cases')
    else:
        outdata = desparse(outdata,min_case,min_anz,cases,master=master)


    outvars = cases
    for v in sorted(outdata.keys()):
        if not v in cases:
            outdata[keyvar+'_'+v] = outdata[v]
            outvars.append(keyvar+'_'+v)  

    return [outdata,outvars]


def dummy_reshape(data,varlist,rvars,rtype='count',master=''):
    verbout('\nReshaping Table.',master=master)
    nvars = []
    out_vars = []
    for v in varlist:
        if not v in rvars:
            nvars.append(v)
            out_vars.append(v)

    out_data = {}
    for v in out_vars:
        out_data[v] = []

    vname = 'NVar'
    nname = 'NValue'
    vlab = vname
    nlab = nname
    vnum = 1
    while vlab in nvars:
        vlab = vname + "{:02}".format(vnum)
        nlab = nname + "{:02}".format(vnum)
        vnum = vnum + 1

    out_vars.append(vlab)
    out_data[vlab] = []
                    
    if rtype == 'count':
        for i in range(len(data[rvars[0]])):
            for v in rvars:
                try:
                    counter = int(data[v][i])
                except:
                    counter = 0
                for k in range(counter):
                    for nv in nvars:
                        out_data[nv].append(data[nv][i])
                    out_data[vlab].append(v)

    elif rtype == 'value':
        out_vars.append(nlab)
        out_data[nlab]=[]
        for i in range(len(data[rvars[0]])):
            for v in rvars:
                val = data[v][i]
                for nv in nvars:
                    out_data[nv].append(data[nv][i])
                out_data[vlab].append(v)
                out_data[nlab].append(val)

    return [out_data,out_vars]

def cd_proof(var,master=''):
    ##Test whether a variable is a dummy.
    #Creates a new table in storage: 'Dummy_Data'. This table holds correct dummy variables
    
    check = 1
    liste = storage['Data'][var]
    if not 'Dummy_Data' in storage.keys():
        storage['Dummy_Data']={}
    storage['Dummy_Data'][var] = []
    try:
        for e in liste:
            if not e in ['',' ']:
                a = float(e)
        numeric = 1
    except:
        numeric = 0

    anz0 = 0
    anz1 = 0
    if numeric == 1:
        outs = var+': Numeric Variable: '
        for e in liste:
            if e in ['',' ']:
                anz0 = anz0+1
                storage['Dummy_Data'][var].append(0)
            elif float(e) == 0:
                anz0 = anz0+1
                storage['Dummy_Data'][var].append(0)
            else:
                anz1 = anz1+1
                storage['Dummy_Data'][var].append(1)
    else:
        outs = var+': Non-Numeric Variable: '
        for e in liste:
            if e in ['',' ']:
                anz0 = anz0+1
                storage['Dummy_Data'][var].append(0)
            else:
                anz1 = anz1+1
                storage['Dummy_Data'][var].append(1)

    if anz0 > 0 and anz1 > 0:
        check = 1
        outs = outs + str(anz0) + ' / '+str(anz1)+'\n'
        verbout(outs,master=master)
    else:
        verbout(var+': Not suitable as Dummy Variable. Invariant.\n',master=master)
        check = 0
        
    return check


def check_dummytable(dataset):
    outdic = {}
    outvar = []

    for v in dataset[1]:
        outlist = []
        anz1 = 0
        anz0 = 0
        values = dataset[0][v]
        uval = get_unique(values)
        vtypes = stat_type(values)
        method = ''
        if 0 in uval and 1 in uval: method = 'dummy'
        elif '0' in uval and '1' in uval: method = 'dummy'
        elif vtypes['Type_String'] > 0 and vtypes['Type_Missing'] > 0: method = 'stringmis'
        elif vtypes['Type_Int'] > 0 and vtypes['Type_Missing'] > 0: method = 'nmis'
        elif vtypes['Type_Float'] > 0 and vtypes['Type_Missing'] > 0: method = 'nmis'

        if method == 'dummy':
            for e in values:
                try:
                    e = int(e)
                except:
                    e = ''

                if not e in [1,0]: e = ''
                outlist.append(e)
        elif method == 'stringmis':
            for e in values:
                if e in ['',' ']:
                    outlist.append(0)
                else:
                    outlist.append(1)
        elif method == 'nmis':
            for e in values:
                try:
                    e = float(e)
                    outlist.append(1)
                except:
                    outlist.append(0)
        if len(outlist) > 0:
            outdic[v] = outlist
            outvar.append(v)

    return [outdic,outvar]


def desparse(data,minc,mina,remvar=['#Group'],master=''):
    verbout('\nRemoving sparse rows (< '+str(minc)+' Values) and columns (< '+str(mina)+' Groups)\n',master=master)
    varlist = list(data.keys())
    for v in remvar:
        varlist.remove(v)
    remcol = []
    remrow = []
    colanz = {}
    for v in varlist:
        colanz[v] = sum(data[v])
        if colanz[v] < mina: remcol.append(v)
    rowanz = {}
    for i in range(len(data[remvar[0]])):
        rowanz[i] = 0
        for v in varlist:
            rowanz[i] = rowanz[i] + data[v][i]
        if rowanz[i] < minc: remrow = [i] + remrow
    verb('Removing Columns due to small numbers: '+str(remcol))
    verb('Removing Rows due to small numbers: '+str(remrow))

    if type(remvar)==str:
        remvar = [remvar]
    loops = 0

    while len(remcol)>0 or len(remrow)>0:
        loops = loops + 1
        verb('Looping...')
        for r in remrow:
            for v in remvar + varlist:
                data[v].pop(r)
        for c in remcol:
            del data[c]
        varlist = list(data.keys())
        for v in remvar:
            varlist.remove(v)
        remcol = []
        remrow = []
        colanz = {}
        for v in varlist:
            colanz[v] = sum(data[v])
            if colanz[v] < mina: remcol.append(v)
        rowanz = {}
        for i in range(len(data[remvar[0]])):
            rowanz[i] = 0
            for v in varlist:
                rowanz[i] = rowanz[i] + data[v][i]
            if rowanz[i] < minc: remrow = [i] + remrow
        verb('Removing Columns due to small numbers: '+str(remcol))
        verb('Removing Rows due to small numbers: '+str(remrow))

    verbout('\nFinished in '+str(loops)+' Iterations.',master=master)
    return data


def delete_missing(data,varlist,n_allowed=0): #Removing cases with missings. n_allowed specifies how many variables are allowed to be missing before excluding.
    outdata = {}
    outvars = data.keys()
    for v in outvars:
        outdata[v] = []

    for i in range(len(data[varlist[0]])):
        count = 0
        for v in varlist:
            if data[v][i] in ['',' ','.']:
                count = count + 1

        if count <= n_allowed:
            for v in outvars:
                outdata[v].append(data[v][i])

    return outdata          
            

def inspect_variable(values):
    ## Takes list of values and transforms them to a dictionary of descriptives
    ## Desc for str: {'Val': [], 'Weight': [], 'Min': 'Mar 13', 'Max': 'Mar 20', 'Sum': 0, 'N': 0, 'Err_Val': 19, 'Err_Weight': 0, 'N_Total': 19, 'Err_Both': 0}
    ## Desc for num: {'Val': [0.0, 0.0, 0.0, 0.0, 1.0, 0.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 1.0], 'Weight': [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1], 'Min': 0.0, 'Vari': 0.14035087719298242, 'Sum': 3.0, 'M': 0.15789473684210525, 'N': 19, 'Err_Val': 0, 'Max': 1.0, 'Err_Weight': 0, 'N_Total': 19, 'Err_Both': 0, 'SD': 0.37463432463267754}
    descriptives = stat_desc(values)
    types = stat_type(values)
    for t in types.keys():
        descriptives[t] = types[t]

    descriptives['Uniques'] = get_unique(values)

    if len(descriptives['Uniques']) < 20:
        descriptives['Freq_Table'] = stat_frequencies(values)
    else:
        descriptives['Freq_Table'] = {'\nMore than 20 unique values. Frequency table omitted\n':{'':''}}
    
    return(descriptives)

def stat_type(values):
    log('Calling Function: Variable types')
    outdic = {}
    outdic['Type_Int'] = 0
    outdic['Type_Float'] = 0
    outdic['Type_String'] = 0
    outdic['Type_Missing'] = 0

    for e in values:
        if e in ['',' ']:
            outdic['Type_Missing'] = outdic['Type_Missing'] + 1
        elif type(e) == str:
            try:
                e2 = float(e)
                if '.' in e:
                    outdic['Type_Float'] = outdic['Type_Float'] + 1
                else:
                    outdic['Type_Int'] = outdic['Type_Int'] + 1
            except Exception as f:
                outdic['Type_String'] = outdic['Type_String'] + 1
        elif type(e) == int:
            outdic['Type_Int'] = outdic['Type_Int'] + 1
        elif type(e) == float:
            outdic['Type_Float'] = outdic['Type_Float'] + 1

    return outdic


def stat_frequencies(values):
    anz = len(values)
    anz2 = len(values)
    undic = {}

    for u in get_unique(values):
        undic[u] = 0

    for e in values:
        undic[e] = undic[e] + 1
        if e in ['',' ']:
            anz2 = anz2 - 1 
    
    outdic = {'Frequency':{},'Percent':{},'Valid Pct.':{}}

    for e in sorted(undic.keys()):
        #outdic['Value'][e] = e
        outdic['Frequency'][e] = undic[e]

        if anz > 0:
            outdic['Percent'][e] = "{0:0.2%}".format(float(undic[e])/anz)
        else:
            outdic['Percent'][e] = '-'

        if anz2 > 0 and not e in ['',' ']:
            outdic['Valid Pct.'][e] = "{0:0.2%}".format(float(undic[e])/anz2)
        else:
            outdic['Valid Pct.'][e] = '-'
            
    return outdic
    
def stat_desc(values,weight=0,verbose=1):
    ##Calculate complete descriptive statistics of a variable.
    ##If only one value is required, use calculate() instead of stat_desc()
    
    if verbose == 1:log('Calling Function: Descriptive Statistics')
    if type(weight) == list:
        if not len(values) == len(weight):
            weight = 0
            if verbose == 1:verb('ERROR: invalid weighting variable')
        else:
            if verbose == 1:verb('Weighting active')
    else:
        weight = 0
        if verbose == 1:verb('No weighting')
            
    desc = {}
    desc['Val'] = []
    desc['Weight'] = []
    desc['N'] = 0
    desc['Sum'] = 0
    desc['Err_Val'] = 0
    desc['Err_Weight'] = 0
    desc['Err_Both'] = 0
    desc['N_Total']=len(values)
    err = 0
    for i in range(len(values)):
        try:    
            fval = float(values[i])
        except:
            desc['Err_Val'] = desc['Err_Val'] + 1
            fval = 'e'
        if not weight == 0:
            try:
                fweigh = float(weight[i])
            except:
                desc['Err_Weight']=desc['Err_Weight']+1
                fweigh = 'e'
        else:
            fweigh = 1

        if fval == 'e' and fweigh == 'e':
            desc['Err_Both'] = desc['Err_Both']+1
        elif not fval == 'e' and not fweigh == 'e':
            desc['Val'].append(fval)
            desc['Weight'].append(fweigh)
            desc['N'] = desc['N'] + fweigh
            desc['Sum'] = desc['Sum'] + fval*fweigh

    if desc['N'] < 1:
        verb('ERROR: No valid numeric cases for descriptive statistics')
    else:
        desc['M'] = desc['Sum']/desc['N']
        if desc['N'] > 1:
            desc['Vari'] = 0
            for i in range(len(desc['Val'])):
                v = desc['Val'][i]
                w = desc['Weight'][i]
                desc['Vari'] =desc['Vari']+w*(v-desc['M'])**2
            desc['Vari']=desc['Vari']/(desc['N']-1)
            desc['SD'] = desc['Vari']**.5
        else:
            desc['Vari']=0.0
            desc['SD']=0.0

    if len(desc['Val'])>0:
        desc['Min'] = min(desc['Val'])
        desc['Max'] = max(desc['Val'])
        desc['Range'] = desc['Max']-desc['Min']
    elif len(values) > 0:
        desc['Min'] = sorted(values)[0]
        desc['Max'] = sorted(values)[-1]
        desc['Range'] = '-'
    else:
        desc['Min'] = '-'
        desc['Max'] = '-'
        desc['Range'] = '-'

    return desc


def heat_color(sval,mode='bw'): ##Return a color value from a float in range [0;1]
    ##Modes: bw: Black and white (return a grayscale) 0=black
    ##       ibw: Grayscales with 0=white as highest value
    ##       red: Red-Green continuum
    
    ccode = "#ff0000"
    if mode == 'bw':
        v = int(sval * 255)
        if v > 255:
            v = 255
        if v < 0:
            v = 0

        hn = hex(v)[2:]
        if len(hn)==1:
            hn = '0'+hn
        ccode = '#' + hn + hn + hn
    elif mode == 'ibw':
        v = int(sval * 255)
        if v > 255:
            v = 255
        if v < 0:
            v = 0

        v = 255-v
        hn = hex(v)[2:]
        if len(hn)==1:
            hn = '0'+hn
        ccode = '#' + hn + hn + hn
    elif mode == 'red':
        gval = math.sin(sval*3.14159)/2+0.5
        bval = math.sin(sval*4+.9)/2+0.5
        rval = math.sin(sval*3-1)/2+0.5
        rv = int(rval * 255)
        if rv > 255:
            rv = 255
        if rv < 0:
            rv = 0
        gv = int(gval * 255)
        if gv > 255:
            gv = 255
        if gv < 0:
            gv = 0
        bv = int(bval * 200)
        if bv > 255:
            bv = 255
        if bv < 0:
            bv = 0


        rn = hex(rv)[2:]
        if len(rn)==1:
            rn = '0'+rn
        gn = hex(gv)[2:]
        if len(gn)==1:
            gn = '0'+gn
        bn = hex(bv)[2:]
        if len(bn)==1:
            bn = '0'+bn
        ccode = '#' + rn + gn + bn
        #ccode = '#' + bn + bn + bn 

    else:
        ccode = '#000000'
    return ccode


##def tts(invalue,inf,outf,numf='dec',verbose=0,master=''):  ##Transform Timestamp
##    outvalue = ''
##    try:
##        if inf == 'pys':
##            tf = time.strptime(invalue,"%a %b %d %H:%M:%S %Y")
##        elif inf == 'pyn':
##            tf = time.gmtime(float(invalue))
##        elif inf == 'ex':
##            ts = float(invalue)-25569
##            ts = ts *24*3600
##            tf = time.gmtime(ts)
##        elif inf == 'ger':
##            try:
##                tf = time.strptime(invalue,"%d.%m.%Y")
##            except:
##                tf = ' nodate'
##        elif inf == 'eng':
##            try:
##                tf = time.strptime(invalue,"%m/%d/%Y")
##            except:
##                tf = ' nodate'
##        elif inf ==  'gerlong':
##            try:
##                tf = time.strptime(invalue,'%d.%m.%Y %H:%M')
##            except:
##                tf = ' nodate'
##        else:
##            print(invalue,inf,outf)
##            print(time.strptime(invalue,inf))
##            verbose = 1
##            try:
##                tf = time.strptime(invalue,inf)
##            except:
##                tf = ' nodate'
##
##        if verbose == 1:verbout('Transforming Timestamp: "'+str(invalue)+'" ('+str(tf)+') to format '+str(outf))
##
##        if outf == 'pyn':
##            outvalue = time.mktime(tf)
##        elif outf == 'pys':
##            outvalue = time.ctime(time.mktime(tf))
##        elif outf == 'dec_h':
##            h = tf[3]
##            sec = 60*tf[4]+tf[5]
##            outvalue = h+float(sec)/3600
##        elif outf == 'ex':
##            ts = time.mktime(tf)
##            ts = ts / 24 / 3600
##            outvalue = ts + 25569
##        elif outf == 'ex7':
##            ts = time.mktime(tf)
##            ts = ts / 24 / 3600
##            outvalue = ts + 25569 - tf[6]
##        elif outf == 'ex30':
##            ts = time.mktime(tf)
##            ts = ts / 24 / 3600
##            outvalue = ts + 25569 - tf[2] + 2
##        elif outf == 'ger':
##            outvalue = time.strftime("%d.%m.%Y",tf)
##        elif outf == 'eng':
##            outvalue = time.strftime("%m/%d/%Y",tf)
##        elif outf == 'time':
##            outvalue = time.strftime("%H:%M:%S",tf)
##        else:
##            outvalue = time.strftime(outf,tf)
##            
##    except Exception as f:
##        a=str(f)
##        verb('ERROR: '+a)
##        print(a)
##
##    if type(outvalue)==float:
##        if numf == 'ic':
##            outvalue = int(outvalue)
##        elif numf == 'ir':
##            outvalue = round(outvalue)
##    return outvalue

def tts(invalue,inf,outf,numf='dec',verbose=0,master=''):  ##Transform Timestamp for Python 3
    outvalue = ''
    try:
        if inf == 'pys':
            tf = datetime.strptime(invalue,"%a %b %d %H:%M:%S %Y")
        elif inf == 'pyn':
            tf = datetime.utcfromtimestamp(float(invalue))
        elif inf == 'ex':
            ts = float(invalue)-25569
            ts = ts *24*3600
            tf = datetime.utcfromtimestamp(ts)
        elif inf == 'ger':
            try:
                tf = datetime.strptime(invalue,"%d.%m.%Y")
            except:
                tf = ' nodate'
        elif inf == 'eng':
            try:
                tf = datetime.strptime(invalue,"%m/%d/%Y")
            except:
                tf = ' nodate'
        elif inf ==  'gerlong':
            try:
                tf = datetime.strptime(invalue,'%d.%m.%Y %H:%M')
            except:
                tf = ' nodate'
        else:
            try:
                tf = datetime.strptime(invalue,inf)
            except:
                tf = ' nodate'

        #print('Time decoded: ',tf)
        if tf.year == 1900:tf = tf.replace(year=2000)

        if verbose == 1:verbout('Transforming Timestamp: "'+str(invalue)+'" ('+str(tf)+') to format '+str(outf))

        if outf == 'pyn':
            try:
                outvalue = (tf-datetime(1970,1,1,0))/timedelta(seconds=1)
            except:
                outvalue = time.mktime(tf.timetuple())
        elif outf == 'pys':
            outvalue = tf.strftime("%a %b %d %H:%M:%S %Y")
        elif outf == 'dec_h':
            h = tf.hour
            sec = 60*tf.minute+tf.second
            outvalue = h+float(sec)/3600
        elif outf == 'ex':
            try:
                ts = (tf-datetime(1970,1,1,0))/timedelta(seconds=1)
            except:
                ts = time.mktime(tf.timetuple())
            ts = ts / 24 / 3600
            outvalue = ts + 25569
        elif outf == 'ex7':
            ts = (tf-datetime(1970,1,1,0))/timedelta(seconds=1)
            ts = ts / 24 / 3600
            outvalue = ts + 25569 - tf.weekday()
        elif outf == 'ex30':
            ts = (tf-datetime(1970,1,1,0))/timedelta(seconds=1)
            ts = ts / 24 / 3600
            outvalue = ts + 25569 - tf.day + 2
        elif outf == 'ger':
            outvalue = tf.strftime("%d.%m.%Y")
        elif outf == 'eng':
            outvalue = tf.strftime("%m/%d/%Y")
        elif outf == 'time':
            outvalue = tf.strftime("%H:%M:%S")
        else:
            outvalue = tf.strftime(outf)
            
    except Exception as f:
        a=str(f)
        verb('ERROR: '+a)
        #print(a)

    if type(outvalue)==float:
        if numf == 'ic':
            outvalue = int(outvalue)
        elif numf == 'ir':
            outvalue = round(outvalue)
    return outvalue




def binomial_odds(draws, p, critical):
    prob = 0.0
    for i in range(critical-1):
        odd = 0
        if 0 <= i <= draws:
            ntok = 1
            ktok = 1
            for t in range(1, min(i, draws - i) + 1):
                ntok *= draws
                ktok *= t
                draws -= 1
            odd = ntok // ktok
        prob = prob + float(odd)*p**i*(1-p)**(draws-i)
    return 1-prob


def bootstrap_sample(liste,l=0): ##Draw a sample with repetitions
    if l == 0: l=len(liste)
    outlist = []
    for i in range(l):
        rand = random.randint(0,l-1)
        outlist.append(liste[rand])
    return outlist


def calc_regression(x,y): ##Simple OLS regression from two numeric vectors 
    xm = calculate(x, 'mean')
    ym = calculate(y, 'mean')
    n = len(x)
    sx = 0.0
    sy = 0.0
    sxx = 0.0
    sxy = 0.0
    syy = 0.0
    for i in range(len(x)):
        sxy = sxy + (x[i]-xm)*(y[i]-ym)
        sxx = sxx + (x[i]-xm)**2
    b = sxy/sxx
    a = ym - b*xm
    se = 0    
    return [b,a]


def calc_entropy(liste,maxliste=[]): ##Caluclate the entropy of a list (with or without complete list of elements)
    if len(liste)==0:
        entropie = 1.0
    else:
        table = {}
        if maxliste == []:
            maxliste = get_unique(liste)
        for e in maxliste:
            table[e] = 0
        for e in liste:
            if e in table.keys():
                table[e] = table[e] + 1
            else:
                pass
        for e in maxliste:
            table[e] = float(table[e])/len(liste)
        entropie = 0
        for e in get_unique(liste):
            if table[e] > 0:
                try:
                    summand = table[e] * math.log(table[e],len(maxliste))
                    entropie = entropie - summand
                except:
                    entropie = entropie                
    return entropie


def calc_chisquare_dummy(l1,l2,directed=0,cramersv=0): ##Calculate chi-square of two dummy variables
    o11 = 0
    o12 = 0
    o21 = 0
    o22 = 0
    mj1 = 0
    mj2 = 0
    m1j = 0
    m2j = 0
    ntot = 0

    dlist1 = []
    dlist2 = []

    if len(l1)==len(l2):
        for i in range(len(l1)):
            try:
                a = int(l1[i])
                b = int(l2[i])
                if a in [0,1] and b in [0,1]:
                    dlist1.append(a)
                    dlist2.append(b)
            except:
                a=''
                b=''

        for i in range(len(dlist1)):
            ntot = ntot + 1
            if dlist1[i] == 1 and dlist2[i] == 1:
                o11 = o11 + 1
                m1j = m1j + 1
                mj1 = mj1 + 1
            if dlist1[i] == 1 and dlist2[i] == 0:
                o12 = o12 + 1
                m1j = m1j + 1
                mj2 = mj2 + 1
            if dlist1[i] == 0 and dlist2[i] == 1:
                o21 = o21 + 1
                m2j = m2j + 1
                mj1 = mj1 + 1
            if dlist1[i] == 0 and dlist2[i] == 0:
                o22 = o22 + 1
                m2j = m2j + 1
                mj2 = mj2 + 1

        e11 = float(m1j*mj1)/ntot
        e12 = float(m1j*mj2)/ntot
        e21 = float(m2j*mj1)/ntot
        e22 = float(m2j*mj2)/ntot

        if min(e11,e12,e21,e22) == 0:
            chisq = ''
        else:
            chisq = (o11-e11)**2/e11 +(o12-e12)**2/e12 +(o21-e21)**2/e21 +(o22-e22)**2/e22
            v = (chisq/ntot)**0.5
            tp = -1
            if o11 + o22 > o12 + o21: tp = 1      
            if cramersv == 1:
                chisq = v
            if directed == 1:
                chisq = chisq * tp
    else:
        verb('ERROR: Lists not of equal length')
        chisq = 'error'

    return chisq


def calc_correlation(l1,l2):  ## Calculate pearson Correlation
    ml1 = calculate(l1,'mean')
    ml2 = calculate(l2,'mean')
    sd1 = calculate(l1,'sd')
    sd2 = calculate(l2,'sd')
    cov = 0.0
    anz = 0
    miss = 0
    for i in range(len(l1)):
        try:
            cov = cov + (l1[i]-ml1)*(l2[i]-ml2)
            anz = anz + 1
        except:
            miss = miss + 1
    if anz>0:
        cov = cov/anz
    else:
        cov = 0.0

    if '' in (sd1,sd2):
        pcor = 0
    elif min(sd1,sd2)==0:
        pcor = 0
    else:
        pcor = cov/sd1/sd2

    return pcor    


def calculate(liste,method): ##Calculate may different things from a list
    ##If the list is a list of 2-tuples, the first value is the value and the second value is the weight. (zipped lists of values and weights)
    ##Methods:
    ## sum, mean, wsum, wmean: (weighted) sums and means of a list
    ## max, min, range: Maximum, Minimum and Range of a list
    ## freq, first, last: Most frequent, first or last element of a list
    
    out_value = 0
    nvalues = []
    nweight = []
    lenvalues = 0

    for e in liste:
        if type(e) == tuple: ##If the list is a zipped list of values and weights
            try:
                val = float(e[0])
                wei = float(e[1])
                nvalues.append(val)
                nweight.append(wei)
                lenvalues = lenvalues + 1
            except:
                val = 0
        else: ##If the list is a list of numbers or strings
            try:
                val = float(e)
                nvalues.append(val)
                nweight.append(1)
                lenvalues = lenvalues + 1
            except:
                val = 0
    
    if method in ['sum','mean','sd','wsum','wmean','nval']:
        s = 0.0
        anz = 0.0
        for i in range(lenvalues):
            summand = nvalues[i]
            gewicht = nweight[i]
            s = s + summand * gewicht
            anz = anz + gewicht
        summe = s
        if anz > 0:
            mittel = s/anz
            if method == 'sd':
                sdsum = 0.0
                for i in range(len(nvalues)):
                    sdsum = sdsum + ((nvalues[i]-mittel)**2)*nweight[i]
                if anz > 1:
                    sdsum = sdsum/(anz-1)
                    stabw = sdsum**.5
                else:
                    stabw = ''
        else:
            mittel = ''
            stabw = ''

        if method in ['sum','wsum']:
            out_value = summe
        elif method in ['mean','wmean']:
            out_value = mittel
        elif method == 'sd':
            out_value = stabw
        elif method == 'nval':
            out_value = anz
            
    elif method == 'max':
        if len(nvalues) > 0:
            out_value = max(nvalues)
        else:
            out_value = ''

    elif method == 'min':
        if len(nvalues) > 0:
            out_value = min(nvalues)
        else:
            out_value = ''

    elif method == 'range':
        if len(nvalues) > 0:
            out_value = max(nvalues)-min(nvalues)
        else:
            out_value = ''
            
    elif method == 'frequ':
        fdic = {}
        for e in liste:
            if type(e) == tuple:
                val = e[0]
            else:
                val = e
            if not val in fdic.keys():
                fdic[val] = 0
            fdic[val] = fdic[val] +1
        maxe = sorted(fdic.keys())[0]
        maxn = fdic[maxe]
        for e in sorted(fdic.keys()):
            if fdic[e] > maxn:
                maxe = e
                maxn = fdic[e]
        out_value = maxe
    elif method == 'first':
        out_value = liste[0]
        if type(out_value) == tuple:
            out_value = out_value[0]
    elif method == 'last':
        out_value = liste[-1]
        if type(out_value) == tuple:
            out_value = out_value[0]

    return out_value


def calculate_vectors(vectors,method,master=0):
    outlist = []
    fdata = []
    mis = 0
    step = int(len(vectors[0])/40)        
    if step<1: step = 1
    verbout('\n\n',master=master)
    verbout('Calculating with method "'+method+'":\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
            
    for i in range(len(vectors[0])):
        if i%step==0:verbout('.','progress',master=master)
        tup = []
        conc = ''
        for k in range(len(vectors)):
            if method == 'concat':
                conc = conc + vectors[k][i]
            else:
                try:
                    tup.append(float(vectors[k][i]))
                except:
                    verb('ERROR: No number:'+vectors[k][i])
                    mis = mis + 1
        fdata.append(tup)
        if method == 'concat':
            outlist.append(conc)
    verbout('\n','progress',master=master)
    if mis > 0:
        verbout('\nAttention: '+str(mis)+' missing values found\n',master=master)
    if not method == 'concat':
        for i in range(len(fdata)):
            if len(fdata[i]) > 0:
                outlist.append(calculate(fdata[i],method))
            else:
                outlist.append('')
    return outlist


def group_variable(liste,mode='equal',param=2,master=''):
    ##Group a scale to specified groups
    ##modes:
    ##   'equal': Groups of equal size (e.g.: Median Split). param: Integer specifying the number of groups
    ##   'fixed': Groups of equal range (e.g.: Days to Weeks). param: Float spefifying the range
    ##   'tails': Tails of the distribution (-1 and 1) and central bulk (0). param: size of the confidence interval in percent (e.g.: 95)
    verbout('\n\nGrouping variable. Parameters: '+str(mode)+' / '+str(param)+'\n',master=master)
    
    missings = []
    numlist = []
    outlist = []
    for i in range(len(liste)):
        try:
            numlist.append(float(liste[i]))
        except:
            missings.append(i)

    if len(missings)>0:
        verbout('\nAttention: '+str(len(missings))+' missing/string values found.',master=master)

    if mode == 'equal':
        try:
            param = int(param)
            perc = []
            for i in range(param):
                perc.append(1.0/param*(i+1))
        except:
            param = 'invalid: '+str(param)
            perc=[1]
        
        cuts = []
        anz = float(len(numlist))
        for p in perc:
            cut = int(anz*p)
            if cut == anz:
                cut = int(anz-1)
            cuts.append(sorted(numlist)[cut])

        for i in range(len(liste)):
            try:
                val = float(liste[i])
                outval = 1
                for ci in range(len(cuts)):
                    if val > cuts[ci]:
                        outval = ci+2
                outlist.append(outval)
            except:
                outlist.append('')

    if mode == 'fixed':
        step = float(param)           
        for i in range(len(liste)):
            try:
                val = float(liste[i])
                outval = math.floor(val/step)*step
                outlist.append(outval)
            except:
                outlist.append('')

    if mode == 'tails':
        tail = (100-float(param))/200
        perc = [tail,1-tail]
        cuts = []
        anz = float(len(numlist))
        for p in perc:
            cut = int(anz*p)
            if cut == anz:
                cut = int(anz-1)
            cuts.append(sorted(numlist)[cut])

        for i in range(len(liste)):
            try:
                val = float(liste[i])
                outval = 0
                if val < cuts[0]:
                    outval = -1
                elif val > cuts[1]:
                    outval = 1
                outlist.append(outval)
            except:
                outlist.append('')         

    return outlist



def crosstab(l1,l2,mode='dicho'):   ##Creates a cross table of two dummy tables and returns a metric
    #Modes:
    #dicho: Dichotomous cross tabulation: Does co-occur or does not co-occur
    #anz: Count of co-occurrences
    #prob: Probability of co-occurrence
    #cprob: Conditional probability of co-occurrence (Probability of l1 being 1 if l2 is 1)
    #einf: Share of co-occurrence in total N
    #sokal: Sokal distance
    #eukl: Eucilidian distance
    #chi1: Chi Squared
    #chi2: Chi Squared
    
    a = 0
    b = 0
    c = 0
    d = 0
    for i in range(len(l1)):
        if l1[i] == 1 and l2[i] == 1:
            a = a + 1
        elif l1[i] == 0 and l2[i] == 1:
            b = b + 1
        elif l1[i] == 1 and l2[i] == 0:
            c = c + 1
        elif l1[i] == 0 and l2[i] == 0:
            d = d + 1

    if mode == 'dicho':
        if a > 0:
            wert = 1
        else:
            wert = 0
    elif mode == 'anz':
        wert = a
    elif mode == 'prob':
        if (a+b)>0 and (a+c)>0:
            wert = float(a*a)/(a+b)/(a+c)
        else:
            wert = 0
    elif mode == 'cprob':
        if (a+b)>0:
            wert = float(a)/(a+b)
        else:
            wert = 0
    elif mode == 'einf':
        wert = float(a+d)/(a+b+c+d)
    elif mode == 'inv_prob':
        if (a+b)>0 and (a+c)>0:
            p = float(a*a)/(a+b)/(a+c)
        else:
            p = 0
        if p == 0:
            p = 0.5/(a+b+c+d)
        wert = 0-math.log(p)
    elif mode == 'sokal':
        wert = (float(b+c)/(2*a+b+c))**.5
    elif mode == 'eukl':
        wert = float(b+c)**0.5

    elif mode == 'chi1':
        wert = calc_chisquare_dummy(l1,l2,0,0)
    elif mode == 'chi2':
        wert = calc_chisquare_dummy(l1,l2,1,1) 

    else:
        verb('ERROR: Not a valid method')
        wert = 'INVALID'

    return wert


##########################
##
## Text Analysis
##
##########################


def getcaps(zeile):
    log('Calling Function: Getcaps')
    #This function removes all words from a string which do not begin with a capital letter.
    outstring=''
    seg=''
    for i in range(0,len(zeile)):
        if zeile[i] in [' ','\t',',',';','.',':','"',"'"]:
            if len(seg) > 1:
                if ord(seg[0]) in range(65,91):
                    outstring=outstring+seg+' '
            seg = ''
        else:
            seg = seg + zeile[i]
    if len(seg) > 1:
        if ord(seg[0]) in range(65,91):
            outstring=outstring+seg
    verb('Only words with capital letters: '+outstring)
    return outstring


def make_fname(fname,ext='txt',suffix=''):
    contents = fname.split('.')
    if len(contents)>1:
        if len(contents[-1]) in [2,3,4]:
            ext = contents.pop(-1)

    filename = ''
    for c in contents:
        filename = filename + c + '.'
    filename = filename[:-1] + suffix + '.' + ext
    
    return filename


def bereinigen(uml_string, lc=0,lb=0,uml=0,encod='latin-1'):
    #This function removes any special character from a string.
    replace = {}
    replace[126] = {1:'~',0:'tilde'}
    replace[128] = {1:'E',0:'euro'}
    replace[132] = {1:'"',0:'"'}
    replace[133] = {1:'...',0:'...'}
    replace[138] = {1:'S',0:'S'}
    replace[139] = {1:'<',0:'"'}
    replace[140] = {1:'Oe',0:'Oe'}
    replace[142] = {1:'Z',0:'Z'}
    replace[145] = {1:'"',0:'"'}
    replace[146] = {1:'"',0:'"'}
    replace[147] = {1:'"',0:'"'}
    replace[148] = {1:'"',0:'"'}
    replace[149] = {1:'-',0:'-'}
    replace[150] = {1:'-',0:'-'}
    replace[151] = {1:'-',0:'-'}
    replace[152] = {1:'-',0:'-'}
    replace[153] = {1:'(tm)',0:'(tm)'}
    replace[154] = {1:'s',0:'s'}
    replace[155] = {1:'>',0:'"'}
    replace[156] = {1:'oe',0:'oe'}
    replace[158] = {1:'z',0:'z'}
    replace[159] = {1:'Y',0:'Y'}
    replace[160] = {1:'.',0:' '}
    replace[162] = {1:'.',0:''}
    replace[163] = {1:'lbs.',0:'lbs.'}
    replace[165] = {1:'Y',0:'yen'}
    replace[167] = {1:'§',0:'par'}
    replace[169] = {1:'(c)',0:'(c)'}
    replace[171] = {1:'"',0:'"'}
    replace[173] = {1:'-',0:'-'}
    replace[175] = {1:'-',0:'-'}
    replace[176] = {1:'°',0:''}
    replace[177] = {1:'+/-',0:'+/-'}
    replace[186] = {1:'°',0:''}
    replace[187] = {1:'"',0:'"'}
    replace[188] = {1:'1/4',0:'1/4'}
    replace[189] = {1:'1/2',0:'1/2'}
    replace[190] = {1:'3/4',0:'3/4'}
    replace[191] = {1:'?',0:'?'}
    replace[192] = {1:'A',0:'A'}
    replace[193] = {1:'A',0:'A'}
    replace[194] = {1:'Â',0:'A'}
    replace[195] = {1:'Ã',0:'A'}
    replace[196] = {1:'Ä',0:'Ae'}
    replace[197] = {1:'A',0:'A'}
    replace[198] = {1:'Ae',0:'Ae'}
    replace[199] = {1:'Ç',0:'C'}
    replace[200] = {1:'È',0:'E'}
    replace[201] = {1:'É',0:'E'}
    replace[202] = {1:'Ê',0:'E'}
    replace[203] = {1:'Ë',0:'E'}
    replace[204] = {1:'I',0:'I'}
    replace[205] = {1:'I',0:'I'}
    replace[206] = {1:'I',0:'I'}
    replace[207] = {1:'Ï',0:'I'}
    replace[208] = {1:'D',0:'D'}
    replace[209] = {1:'Ñ',0:'N'}
    replace[210] = {1:'Ò',0:'O'}
    replace[211] = {1:'Ó',0:'O'}
    replace[212] = {1:'Ô',0:'O'}
    replace[213] = {1:'O',0:'O'}
    replace[214] = {1:'Ö',0:'Oe'}
    replace[216] = {1:'Oe',0:'Oe'}
    replace[217] = {1:'Ù',0:'U'}
    replace[218] = {1:'Ú',0:'U'}
    replace[219] = {1:'Û',0:'U'}
    replace[220] = {1:'Ü',0:'Ue'}
    replace[221] = {1:'Y',0:'Y'}
    replace[222] = {1:'th',0:'th'}
    replace[223] = {1:'ß',0:'ss'}
    replace[224] = {1:'à',0:'a'}
    replace[225] = {1:'á',0:'a'}
    replace[226] = {1:'â',0:'a'}
    replace[227] = {1:'a',0:'a'}
    replace[228] = {1:'ä',0:'ae'}
    replace[229] = {1:'a',0:'a'}
    replace[230] = {1:'æ',0:'ae'}
    replace[231] = {1:'ç',0:'c'}
    replace[232] = {1:'è',0:'e'}
    replace[233] = {1:'é',0:'e'}
    replace[234] = {1:'ê',0:'e'}
    replace[235] = {1:'ë',0:'e'}
    replace[236] = {1:'i',0:'i'}
    replace[237] = {1:'i',0:'i'}
    replace[238] = {1:'î',0:'i'}
    replace[239] = {1:'ï',0:'i'}
    replace[240] = {1:'dh',0:'dh'}
    replace[241] = {1:'ñ',0:'n'}
    replace[242] = {1:'o',0:'o'}
    replace[243] = {1:'ó',0:'o'}
    replace[244] = {1:'ô',0:'o'}
    replace[245] = {1:'õ',0:'o'}
    replace[246] = {1:'ö',0:'oe'}
    replace[247] = {1:'%',0:'%'}
    replace[248] = {1:'oe',0:'oe'}
    replace[249] = {1:'ù',0:'u'}
    replace[250] = {1:'ú',0:'u'}
    replace[251] = {1:'û',0:'u'}
    replace[252] = {1:'ü',0:'ue'}
    replace[253] = {1:'y',0:'y'}
    replace[254] = {1:'th',0:'th'}
    replace[913] = {1:u'Α',0:'A'}
    replace[914] = {1:u'Β',0:'B'}
    replace[915] = {1:u'Γ',0:'G'}
    replace[916] = {1:u'Δ',0:'D'}
    replace[917] = {1:u'Ε',0:'E'}
    replace[918] = {1:u'Ζ',0:'Z'}
    replace[919] = {1:u'Η',0:'H'}
    replace[920] = {1:u'Θ',0:'Th'}
    replace[921] = {1:u'Ι',0:'I'}
    replace[922] = {1:u'Κ',0:'K'}
    replace[923] = {1:u'Λ',0:'L'}
    replace[924] = {1:u'Μ',0:'M'}
    replace[925] = {1:u'Ν',0:'N'}
    replace[926] = {1:u'Ξ',0:'X'}
    replace[927] = {1:u'Ο',0:'O'}
    replace[928] = {1:u'Π',0:'p'}
    replace[929] = {1:u'Ρ',0:'R'}
    replace[931] = {1:u'Σ',0:'S'}
    replace[932] = {1:u'Τ',0:'T'}
    replace[933] = {1:u'Υ',0:'Y'}
    replace[934] = {1:u'Φ',0:'F'}
    replace[935] = {1:u'Χ',0:'Ch'}
    replace[936] = {1:u'Ψ',0:'Ps'}
    replace[937] = {1:u'Ω',0:'W'}
    replace[940] = {1:u'ά',0:'a'}
    replace[941] = {1:u'έ',0:'e'}
    replace[942] = {1:u'ή',0:'h'}
    replace[943] = {1:u'ί',0:'i'}
    replace[945] = {1:u'α',0:'a'}
    replace[946] = {1:u'β',0:'b'}
    replace[947] = {1:u'γ',0:'g'}
    replace[948] = {1:u'δ',0:'d'}
    replace[949] = {1:u'ε',0:'e'}
    replace[950] = {1:u'ζ',0:'z'}
    replace[951] = {1:u'η',0:'h'}
    replace[952] = {1:u'θ',0:'th'}
    replace[953] = {1:u'ι',0:'i'}
    replace[954] = {1:u'κ',0:'k'}
    replace[955] = {1:u'λ',0:'l'}
    replace[956] = {1:u'μ',0:'m'}
    replace[957] = {1:u'ν',0:'n'}
    replace[958] = {1:u'ξ',0:'x'}
    replace[959] = {1:u'ο',0:'o'}
    replace[960] = {1:u'π',0:'p'}
    replace[961] = {1:u'ρ',0:'r'}
    replace[962] = {1:u'ς',0:'s'}
    replace[963] = {1:u'σ',0:'s'}
    replace[964] = {1:u'τ',0:'t'}
    replace[965] = {1:u'υ',0:'y'}
    replace[966] = {1:u'φ',0:'f'}
    replace[967] = {1:u'χ',0:'ch'}
    replace[968] = {1:u'ψ',0:'ps'}
    replace[969] = {1:u'ω',0:'w'}
    replace[970] = {1:u'ϊ',0:'i'}
    replace[972] = {1:u'ό',0:'o'}
    replace[973] = {1:u'ύ',0:'u'}
    replace[974] = {1:u'ώ',0:'w'}
    replace[977] = {1:u'ϑ',0:'th'}
    replace[978] = {1:u'ϒ',0:'y'}
    replace[982] = {1:u'ϖ',0:'p'}
    replace[1040] = {1:u'А',0:'A'}
    replace[1072] = {1:u'а',0:'a'}
    replace[1041] = {1:u'Б',0:'B'}
    replace[1073] = {1:u'б',0:'b'}
    replace[1042] = {1:u'В',0:'V'}
    replace[1074] = {1:u'в',0:'v'}
    replace[1043] = {1:u'Г',0:'G'}
    replace[1075] = {1:u'г',0:'g'}
    replace[1044] = {1:u'Д',0:'D'}
    replace[1076] = {1:u'д',0:'d'}
    replace[1045] = {1:u'Е',0:'E'}
    replace[1077] = {1:u'е',0:'e'}
    replace[1046] = {1:u'Ж',0:'Zh'}
    replace[1078] = {1:u'ж',0:'zh'}
    replace[1047] = {1:u'З',0:'Z'}
    replace[1079] = {1:u'з',0:'z'}
    replace[1048] = {1:u'И',0:'I'}
    replace[1080] = {1:u'и',0:'i'}
    replace[1049] = {1:u'Й',0:'Y'}
    replace[1081] = {1:u'й',0:'y'}
    replace[1050] = {1:u'К',0:'K'}
    replace[1082] = {1:u'к',0:'k'}
    replace[1051] = {1:u'Л',0:'L'}
    replace[1083] = {1:u'л',0:'l'}
    replace[1052] = {1:u'М',0:'M'}
    replace[1084] = {1:u'м',0:'m'}
    replace[1053] = {1:u'Н',0:'N'}
    replace[1085] = {1:u'н',0:'n'}
    replace[1054] = {1:u'О',0:'O'}
    replace[1086] = {1:u'о',0:'o'}
    replace[1055] = {1:u'П',0:'P'}
    replace[1087] = {1:u'п',0:'p'}
    replace[1056] = {1:u'Р',0:'R'}
    replace[1088] = {1:u'р',0:'r'}
    replace[1057] = {1:u'С',0:'S'}
    replace[1089] = {1:u'с',0:'s'}
    replace[1058] = {1:u'Т',0:'T'}
    replace[1090] = {1:u'т',0:'t'}
    replace[1059] = {1:u'У',0:'U'}
    replace[1091] = {1:u'у',0:'u'}
    replace[1060] = {1:u'Ф',0:'F'}
    replace[1092] = {1:u'ф',0:'f'}
    replace[1061] = {1:u'Х',0:'H'}
    replace[1093] = {1:u'х',0:'h'}
    replace[1062] = {1:u'Ц',0:'Ts'}
    replace[1094] = {1:u'ц',0:'Ts'}
    replace[1063] = {1:u'Ч',0:'Ch'}
    replace[1095] = {1:u'ч',0:'ch'}
    replace[1064] = {1:u'Ш',0:'Sh'}
    replace[1096] = {1:u'ш',0:'sh'}
    replace[1065] = {1:u'Щ',0:'Sht'}
    replace[1097] = {1:u'щ',0:'sht'}
    replace[1066] = {1:u'Ъ',0:'A'}
    replace[1098] = {1:u'ъ',0:'a'}
    replace[1068] = {1:u'Ь',0:"Y"}
    replace[1100] = {1:u'ь',0:"y"}
    replace[1070] = {1:u'Ю',0:'Yu'}
    replace[1102] = {1:u'ю',0:'yu'}
    replace[1071] = {1:u'Я',0:'Ya'}
    replace[1103] = {1:u'я',0:'ya'}  
    replace[260] = {1:u'Ą',0:'A'}
    replace[261] = {1:u'ą',0:'a'}
    replace[262] = {1:u'Ć',0:'C'}
    replace[263] = {1:u'ć',0:'c'}
    replace[280] = {1:u'Ę',0:'E'}
    replace[281] = {1:u'ę',0:'e'}
    replace[321] = {1:u'Ł',0:'L'}
    replace[322] = {1:u'ł',0:'l'}
    replace[323] = {1:u'Ń',0:'N'}
    replace[324] = {1:u'ń',0:'n'}
    replace[346] = {1:u'Ś',0:'S'}
    replace[347] = {1:u'ś',0:'s'}
    replace[377] = {1:u'Ź',0:'Z'}
    replace[378] = {1:u'ź',0:'z'}
    replace[379] = {1:u'Ż',0:'Z'}
    replace[380] = {1:u'ż',0:'z'}
    replace[8211] = {1:u'–',0:'-'}
    replace[8220] = {1:u'"',0:'"'}
    replace[8221] = {1:u'"',0:'"'}


    if encod == 'utf-8':
        uml_string = uml_string.decode('utf-8')
        
    ausgabe = ''
    for i in uml_string:
        if i in ['\n','\r']:
            if lb==0:
                ausgabe = ausgabe + ' / '
            else:
                ausgabe = ausgabe + '\n'
        elif i == '\t':
            if uml == 1:
                ausgabe = ausgabe + '\t'
            else:
                ausgabe = ausgabe + '<TAB>'
        elif ord(i) < 125:
            ausgabe = ausgabe + i
        else:  ##Replacement of common special characters.
            if ord(i) in replace.keys():
                ausgabe = ausgabe + replace[ord(i)][uml]
            else:
                ausgabe = ausgabe + '<ORD:' + str(ord(i)) +'>'

        if lc == 1:
            ausgabe = ausgabe.lower()
    return str(ausgabe)


def stem(token,lang='eng',verbose=0):
    numbers = ['1','2','3','4','5','6','7','8','9','0']
    for n in numbers:
        if n in token:
            token = '_number_'
            lang = 'none'
                
    if lang == 'eng':
        if len(token) > 7:
            if token[-3:] == 'ing':
                token = token[:-3]
        elif len(token) > 5:
            if token[-4:] == 'sses':
                token = token[:-4] + 'ss'  
            elif token[-3:] in ['ied','ies']:
                token = token[-3:] + 'i'
            elif token[-1] == 's':
                token = token[:-1]
        elif len(token) > 4:
            if token[-1] == 's':
                token = token[:-1]

    elif lang=='de':

        suffix = {}
        replace = {}

        suffix[4] = ['sten']
        suffix[3] = ['ern']
        suffix[2] = ['em','er','en','es']
        suffix[1] = ['e']

        replace[3] = {'rst':'r'}
        replace[2] = {'bs':'b','ds':'d','fs':'f','gs':'g','hs':'h','ks':'k',
                      'ls':'l','ms':'m','ns':'n','rs':'r','ts':'t'}
        

        length = len(token)
        maxsuff = length-3
        maxrep = length-3
        if maxsuff > 4: maxsuff = 4
        if maxrep > 3: maxrep = 3
        minsuff = 0
        minrep = 1
        tc = 0
        for rp in range(maxrep,minrep,-1):
            if token[-rp:] in replace[rp].keys() and tc == 0:
                token = token[:-rp] + replace[rp][token[-rp:]]
                tc = 1
                
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1

        suffix = {}
        suffix[4] = ['heit','isch','lich','keit']
        suffix[3] = ['end','ung']
        suffix[2] = ['ig','ik']

        length = len(token)
        maxsuff = length-3      
        if maxsuff > 4: maxsuff = 4  
        minsuff = 1
        tc = 0
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1

    elif lang=='fr':
        suffix = {}
        suffix[9] = ['issements']
        suffix[8] = ['issement','issaient','issantes']
        suffix[7] = ['atrices','iraient','issante','issants','issions','eraient','assions']
        suffix[6] = ['atrice', 'ateurs', 'ations' ,'logies', 'usions', 'utions', 'ements',
                     'amment', 'emment','irions','issais','issait','issant','issent',
                     'issiez','issons','erions','assent','assiez']
        suffix[5] = ['ances', 'iques', 'ismes', 'ables', 'istes' , 'ateur', 'ation',
                     'logie', 'usion' ,'ution', 'ences', 'ences', 'ement','euses', 'ments',
                     'irais','irait','irent','iriez','irons','iront','isses','issez','èrent',
                     'erais','erait',u'eriez',u'erons',u'eront','aient','antes','asses']
        suffix[4] = ['ance', 'ique', 'isme', 'able', 'iste', 'ence', 'ence', 'ités',
                     u'ités', 'ives', 'eaux', 'euse', 'ment','îmes',u'îmes','îtes',u'îtes',
                     'irai','iras','irez','isse','ions','erai','eras','erez',u'âmes','âmes',
                     u'âtes','âtes','ante','ants','asse']
        suffix[3] = ['eux', u'ité', 'ive', 'ifs', 'aux','ité','ies','ira',
                     u'ées','ées','era','iez','ais','ait','ant','oir']
        suffix[2] = [u'és', 'ai','és','ît',u'ît','ie','ir','is','it',
                     u'ée','ée',u'és','és','er','ez',u'ât','ât','ai',u'as']
        suffix[1] = ['a',u'é','é','e','i']
        replace = {}
        replace[6] = {'logies':'log'}
        replace[5] = {'ament':'ant','ement':'ent','logie':'log','ences':'ent'}
        replace[4] = {'eaux':'eau','ence':'ent'}

        length = len(token)
        maxsuff = length-3
        maxrep = length-3
        if maxsuff > 9: maxsuff = 9
        if maxrep > 6: maxrep = 6
        minsuff = 0
        minrep = 3
        tc = 0
        for rp in range(maxrep,minrep,-1):
            if token[-rp:] in replace[rp].keys() and tc == 0:
                token = token[:-rp] + replace[rp][token[-rp:]]
                tc = 1
                
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1

    elif lang=='it':
        suffix = {}
        suffix[6] = ['gliela', 'gliele','glieli','glielo','gliene']
        suffix[5] = []
        suffix[4] = ['sene','mela','mele','meli','melo','mene','tela','tele','teli','telo','tene','cela',
                     'cele','celi','celo','cene','vela','vele','veli','velo','vene','ando','endo']
        suffix[3] = ['gli']
        suffix[2] = ['ci','la','le','li','lo','mi','ne','si','ti','vi']

        verbs = {}
        verbs['avere'] = ['ho', 'hai', 'ha', 'abbiamo', 'avete','hanno', 'abbia', 'abbiate', 'abbiano',
                          'avrò', 'avrai', 'avrà', 'avremo', 'avrete', 'avranno', 'avrei',
                          'avresti', 'avrebbe', 'avremmo', 'avreste', 'avrebbero', 'avevo', 'avevi',
                          'aveva', 'avevamo', 'avevate', 'avevano', 'ebbi', 'avesti','ebbe', 'avemmo',
                          'aveste', 'ebbero', 'avessi', 'avesse', 'avessimo', 'avessero', 'avendo', 'avuto',
                          'avuta', 'avuti', 'avute']
        
        verbs['essere'] = ['sono', 'sei', 'è', 'siamo', 'siete', 'sia', 'siate', 'siano', 'sarò', 'sarai',
                          'sarà', 'saremo', 'sarete', 'saranno', 'sarei', 'saresti',
                          'sarebbe', 'saremmo', 'sareste', 'sarebbero', 'ero', 'eri', 'era',
                          'eravamo', 'eravate', 'erano', 'fui', 'fosti', 'f', 'fummo', 'foste', 'furono',
                          'fossi', 'fosse', 'fossimo', 'fossero', 'essend']

        verbs['fare'] = ['faccio', 'fai', 'fa', 'facciamo', 'fate', 'fanno', 'fatto', 'facevo', 'faceviò', 'faceva',
                         'facevamo', 'facevate', 'facevano', 'feci', 'facesti', 'fece',
                         'facemmo', 'faceste', 'fecero', 'farò', 'farai', 'farà', 'faremo', 'farete',
                         'eravate', 'erano','fui', 'fosti', 'f', 'fummo', 'foste',
                         'faranno', 'faccia', 'facciamo', 'facciate', 'facciano', 'facessi',
                         'facesse','facessimo','faceste','facessero','farei','faresti','farebbe',
                         'faremmo','fareste','farebbero','facciano']

        verbs['stare'] = ['sto', 'stai', 'sta', 'stiamo', 'stanno', 'stia', 'stiate', 'stiano', 'starò',
                          'starai', 'starà', 'staremo', 'starete', 'staranno',
                          'starei', 'staresti', 'starebbe', 'staremmo', 'stareste', 'starebbero',
                          'stavo', 'stavi', 'stava', 'stavamo', 'stavate', 'stavano', 'stetti',
                          'stesti', 'stette', 'stemmo', 'steste', 'stettero', 'stessi', 'stesse',
                          'stessimo', 'stessero', 'stando']    
        

        length = len(token)
        maxsuff = length-3      
        if maxsuff > 6: maxsuff = 6  
        minsuff = 1
        tc = 0
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1
             
        for v in verbs.keys():
            for vf in verbs[v]:
                if token == vf:
                    token = v
                    tc = 1

        if len(token) > 5:
            if token[-2:] in ['ar','er','ir']: token = token[:-2]+'e'

        suffix = {}
        replace = {}
        suffix[8] = ['erebbero', 'irebbero']
        suffix[7] = []
        suffix[6] = ['atrice', 'atrici', 'azione', 'azioni', 'uzione', 'uzioni', 'usione', 'usioni',
                     'amento', 'amenti', 'imento', 'imenti', 'amente','assero','assimo','eranno',
                     'erebbe','eremmo','ereste','eresti','essero','iranno','irebbe','iremmo',
                     'ireste','iresti','iscano','iscono','issero']
        suffix[5] = ['abile', 'abili', 'ibile', 'ibili', 'mente' 'atore', 'atori', 'logia', 'logie',
                     'arono', 'avamo', 'avano', 'avate', 'eremo', 'erete', 'erono', 'evamo', 'evano',
                     'evate', 'iremo', 'irete', 'irono', 'ivamo', 'ivano', 'ivate']
        suffix[4] = ['anza', 'anze', 'iche', 'ichi', 'ismo', 'ismi', 'ista', 'iste', 'isti',
                     'ante', 'anti', 'enza', 'enze', u'istà', 'istà', u'istè','istè', u'istì','istì',
                     'ammo', 'ando', 'asse', 'assi', 'emmo', 'enda', 'ende', 'endi',
                     'endo', 'erai', 'erei', 'yamo', 'iamo', 'immo', 'irai', 'irei',
                     'isca', 'isce', 'isci', 'isco']
        suffix[3] = ['ico', 'ici', 'ica', 'ice', 'oso', 'osi', 'osa', 'ose', u'ità','ità','ivi','iva', 'ive',
                     'ano', 'are', 'ata', 'ate', 'ati', 'ato', 'ava', 'avi', 'avo', u'erà','erà', 'ere', u'erò','erò',
                     'ete', 'eva', 'evi', 'evo', u'irà', 'irà','ire', u'irò','irò','ita', 'ite', 'iti', 'ito', 'iva',
                     'ivi', 'ivo', 'ono', 'uta', 'ute', 'uti', 'uto']
        suffix[2] = ['ar', 'ir']

        replace[6] = {'uzione':'u','uzioni':'u','usione':'u','usioni':'u'}
        replace[5] = {'logia':'log','logie':'log'}
        replace[4] = {'enza':'te','enze':'te'}


        length = len(token)
        maxsuff = length-3
        maxrep = length-3
        if maxsuff > 8: maxsuff = 8
        if maxrep > 6: maxrep = 6
        minsuff = 1
        minrep = 3
        for rp in range(maxrep,minrep,-1):
            if token[-rp:] in replace[rp].keys() and tc == 0:
                token = token[:-rp] + replace[rp][token[-rp:]]
                tc = 1
                
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1
                
    elif lang=='nl':      
        suffix = {}
        replace = {}

        suffix[5] = ['heden']
        suffix[4] = []
        suffix[3] = ['ene']
        suffix[2] = ['en','se']
        suffix[1] = ['s','e']

        length = len(token)
        maxsuff = length-3
        if maxsuff > 5: maxsuff = 5
        minsuff = 0
        tc = 0
                
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1

        suffix = {}
        replace = {}

        suffix[4] = ['lijk','baar']
        suffix[3] = ['bar']
        suffix[2] = ['ig']
        suffix[1] = []

        replace[5] = {'igend':'ig', 'iging':'ig'}
        
        length = len(token)
        maxsuff = length-3
        maxrep = length-3
        if maxsuff > 4: maxsuff = 4
        if maxrep > 5: maxrep = 5
        minsuff = 0
        minrep = 4
        tc = 0
        for rp in range(maxrep,minrep,-1):
            if token[-rp:] in replace[rp].keys() and tc == 0:
                token = token[:-rp] + replace[rp][token[-rp:]]
                tc = 1
                
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1

    elif lang=='pl':      

        suffix = {}
        suffix[6] = ['cyjnym','cyjnej','cyjnym','cznych','cznego']
        suffix[5] = ['ckich','cyjna','cyjny','cznym','cznej']
        suffix[4] = ['owie','cyjn','czna','ckie']
        suffix[3] = ['cki','owi','cja','cji','ach','ych','ego']
        suffix[2] = ['ci','ch','ym','ej']
        suffix[1] = ['a','i','y','u','o','t']
        
        length = len(token)
        maxsuff = length-3
        if maxsuff > 6: maxsuff = 6
        minsuff = 0
        tc = 0
                
        for su in range(maxsuff,minsuff,-1):
            if token[-su:] in suffix[su] and tc == 0:
                token = token[:-su]
                tc = 1

    else:
        pass ##Just return the token.

    return token

##outdset = create_corpus(data,path,ids,nvar,encod,master=self)

def create_corpus(dset,path,idvar,nvar='Fulltext',encod='latin-1',master=''):
    if type(dset) == list:
        data = dset[0]
        varlist = dset[1]
    else:
        data = dset
        varlist = sorted(data.keys())

    step = int(len(data[idvar])/40)
    panz = 1
    if step<1:
        panz = int(40.0/len(data[idvar]))
        step = 1

    verbout('\n\nReading '+str(len(data[idvar]))+' files:\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
        
    data[nvar] = []
    i = 0
    for f in data[idvar]:
        i+=1
        if i%step==0:verbout('.'*panz,'progress',master)
        if len(f) > 4:
            if not f[-4] == '.':
                f = f + '.txt'
        try:
            inf = open(path+f,'r')
            lines = inf.readlines()
            inf.close()
            fulltext = bereinigen(" ".join(lines),encod=encod)
            data[nvar].append(fulltext)
        except:
            data[nvar].append('')
            verb('\nFile not found: '+path+f)
    verbout('\n',master=master)
    
    verbout('\n\nRemoving entries without text...',master=master)
    outdata = {}
    varlist.append(nvar)
    
    removed = 0
    for v in varlist:
        outdata[v] = []
        
    for i in range(len(data[idvar])):
        if len(data[nvar][i]) > 2:
            for v in varlist:
                outdata[v].append(data[v][i])
        else:
            removed = removed + 1
    verbout('\nRemoved '+str(removed)+' Entries from the corpus',master = master)
    return [outdata,varlist]
                

def create_ngrams(tokens,nlen=2,universe=[]):
    nglist = []
    if not universe == []:
        udic = {}
        for u in universe:
            udic[u] = u
    for nglen in range(1,nlen+1):
        for i in range(len(tokens)-nglen+1):
            n = " ".join(tokens[i:i+nglen])
            if len(universe)>0:
                try:
                    nglist.append(udic[n])
                except:
                    n = ''
            else:
                nglist.append(n)

    return nglist

def lemmatize(zeile,lang='none'):
    ## Takes a line of text (one string variable) and returns a list of
    ## stemmed tokens in the goven language.
    
    fixedex = {':-)':'smile',
               ':)':'smile',
               ':D':'smile',
               ':-D':'smile',
               ';-)':'irony',
               ';)':'irony',
               ':(':'sadsmile',
               ':-(':'sadsmile',
               '!':'exclam',
               '?':'question',
               '-.-':'japsmile',
               '^.^':'japsmile',
               '^^':'japsmile'}
    for exp in fixedex.keys():
        zeile = zeile.replace(exp,' _'+fixedex[exp]+'_ ')

    for Sonderzeichen in ['.',',',';',':','<','>','&','\n','\r','\t',
                          '"',"'",'/','-','$','(',')','|']:
        zeile = zeile.replace(Sonderzeichen,' ')

    zeile = zeile.lower()
    worte = zeile.split(' ')
    while '' in worte:
        worte.remove('')

    tokens = []
    if lang == 'none':
        tokens = worte
    else:
        for w in worte:
            tokens.append(stem(w,lang))
        
    return tokens


def context(line, expression, span=10, case=0):
    ## At the moment only works if re is loaded. Probably make it {base} some time in the future
    outlist = []

    if case in [0,'0']:
        pattern = re.compile(expression,re.I)
    else:
        pattern = re.compile(expression)
        
    obj = ''
    end = 0
    while not obj == None:
        obj = pattern.search(line, pos=end)
        if not obj == None:
            end = obj.span()[1]
            start = obj.span()[0]
            #print(str([r]),context(text,obj.span()[0],obj.span()[1]))
    
            substr = line[start:end]

            if start >= span:
                leading = line[(start-span):start]
            else:
                over = "-" * (span-start)
                leading = over+ line[:start]

            if end < len(line)-span:
                tailing = line[end:(end+span)]
            else:
                tailing = line[end:]

            outstr = leading + '**' + substr + '**' + tailing

            outlist.append((leading,substr,tailing))
    return outlist




def get_words(fname, lang='en', laenge=2,encod='latin-1'):
    ## Takes a filename and returns a list of lemmatized words
    outlist = []
    infile = open(fname,'r')
    lines = infile.readlines()
    infile.close()
    for z in lines:
        wordlist = lemmatize(z,lang,encod=encod)
        wordlist = create_ngrams(wordlist,laenge)
        outlist = outlist + wordlist
    return get_unique(outlist)   


def get_univ(flist, lang, ngram, sparsity=[.01,.99], encod='latin-1',master=''):

    step = int(len(flist)/40)
    panz = 1
    if step<1:
        panz = int(40.0/len(flist))
        step = 1

    verbout('\n\nReading '+str(len(flist))+' files:\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    textlist = []
    i = 0
    for f in flist:
        i += 1
        if i%step==0:verbout('.','progress',master=master)
        tmp = open(f,'r')
        tl = tmp.readlines()
        tmp.close()
        document = " ".join(tl)
        for fixedex in ['\n','\t','\r']:
            document = document.replace(fixedex,' ')
        textlist.append(document)
    verbout('\n',master=master)

    step = int(len(textlist)/40)
    panz = 1
    if step<1:
        panz = int(40.0/len(textlist))
        step = 1
    verbout('\nReading texts:\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    ngramdic = {}
    for i in range(len(textlist)):
        if i%step == 0:verbout('.','progress',master=master)
        lemmata = lemmatize(textlist[i],lang)
        ngrams = create_ngrams(lemmata,ngram)
        for n in ngrams:
            try:
                ngramdic[n] +=1
            except:
                ngramdic[n] = 1
    verbout('\n',master=master)

    outdic = {}
    minanz = float(len(textlist)*sparsity[0])
    maxanz = float(len(textlist)*sparsity[1])
    verbout('\nRemoving ngrams appearing in less than '+str(minanz)+' or more than '+str(maxanz)+' texts',master=master)
    for t in ngramdic.keys():
        if ngramdic[t]>minanz and ngramdic[t]<maxanz:
            outdic[t] = float(ngramdic[t])/len(textlist)
    verbout('\nRemaining ngrams: '+str(len(outdic.keys()))+' of '+str(len(ngramdic.keys())),master=master)

    return outdic


def generate_tdm(textlist,idlist=[],lang='none',ngrams=2,sparsity=[.01,.99],universe=[],weight="dicho", master=''):
    verbout('\nGenerating TDM for '+str(len(textlist))+' documents.',master=master)    
    if universe == []:
        verbout('\nNo universe of ngrams loaded. Using all words.',master=master)
    if idlist == []:
        idlist = list(range(len(textlist)))
    tokenlist = []
    tokendic = {}
    for text in textlist:
        tokens = lemmatize(text,lang)
        nglist = create_ngrams(tokens,ngrams,universe)
        for t in nglist:
            try:
                tokendic[t] += 1
            except:
                tokendic[t] = 1
        tokenlist.append(nglist)
    verbout("\nList of ngrams for all texts prepared. N = "+str(len(tokendic.keys())),master=master)
    
    tlist = ['res_Document']
    if type(sparsity)==list:
        minanz = float(len(textlist)*sparsity[0])
        maxanz = float(len(textlist)*sparsity[1])
        verbout('\n\nRemoving ngrams appearing in less than '+str(minanz)+' or more than '+str(maxanz)+' texts',master=master)
        for t in sorted(tokendic.keys()):
            if tokendic[t]>minanz and tokendic[t]<maxanz:
                tlist.append(t)
        verbout('\nRemaining ngrams: '+str(len(tlist)-1),master=master)
    else:
        tlist +=sorted(list(tokendic.keys()))
    
    tdm = {}
    for t in tlist:
        tdm[t] = []
        
    step = int(len(textlist)/40)
    panz = 1
    if step<1:
        panz = int(40.0/len(textlist))
        step = 1
    verbout('\n\nReading texts:\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    for i in range(len(textlist)):
        if i%step == 0:verbout('.','progress',master=master)
        tdm['res_Document'].append(idlist[i])
        if weight=="tf":
            tw = {}
            for t in tlist[1:]:
                tw[t] = 0
            for t in tokenlist[i]:
                try:
                    tw[t]+=1
                except:
                    pass
            for t in tlist[1:]:
                tdm[t].append(tw[t])

        if weight=="dicho":
            for t in tlist[1:]: ##Do not do it for 'res_Document'
                if t in tokenlist[i]:
                    tdm[t].append(1)
                else:
                    tdm[t].append(0)
    verbout('\n\nTDM complete\n',master=master)
    return tdm


################## SVM

def svm_scores(tdm_in,rvec,features=[],cases=[],verbose=0):
    tdm = {}
    for t in tdm_in:
        if not t == 'res_Document': tdm[t] = tdm_in[t]
    
    if features == []: features = list(tdm.keys())
    if cases == []: cases = range(len(tdm[features[0]]))
    scorelist = []
    for i in cases:
        score = rvec['Intercept']
        for f in features:
            try:
                score = score + tdm[f][i]*rvec[f]
            except:
                score = score
        scorelist.append(score)
    return scorelist

def svm_prf(scores,true):
    tp = 0
    tn = 0
    fp = 0
    fn = 0
    for i in range(len(scores)):
        if scores[i] < 0 and true[i] == 0: tn+=1
        if scores[i] < 0 and true[i] == 1: fn+=1
        if scores[i] >= 0 and true[i] == 0: fp+=1
        if scores[i] >= 0 and true[i] == 1: tp+=1

    try:
        precision = tp/(tp+fp)
    except:
        precision = 0.0
    try:
        recall = tp/(tp+fn)
    except:
        recall = 0.0
    if precision+recall > 0:
        f = 2*precision*recall/(precision+recall)
    else:
        f = 0.0

    return [precision,recall,f]

def svm_prf_curve(tdm, rvec, classvec):
    scorelist = svm_scores(tdm,rvec)
    intercepts = []
    nsteps = 50
    step = (max(scorelist)-min(scorelist))/(nsteps-1)
    ms = min(scorelist)
    for s in range(nsteps):
        intercepts.append(rvec['Intercept']-(ms+s*step))

    outdic = {}
    outvar = ['Intercept','Precision','Recall','F_Score']
    for v in outvar:
        outdic[v] = []

    for i in intercepts:
        rvec['Intercept'] = i
        scorelist = svm_scores(tdm,rvec)
        prf = svm_prf(scorelist,classvec)
        outdic['Intercept'].append(i)
        outdic['Precision'].append(prf[0])
        outdic['Recall'].append(prf[1])
        outdic['F_Score'].append(prf[2])
    return [outdic,outvar]



def train_svm(tdm, classvec,master=''):
    ##Trains an SVM using a Term-Document Matrix (document name omitted) and a class vector of equal length

    verbout("\nTraining SVM on "+str(len(classvec))+" cases using "+str(len(tdm.keys()))+" features",master=master)
    terms = list(tdm.keys())
    stats = {}
    for t in terms:
        stats[t] = stat_desc(tdm[t])

    cases = range(len(classvec))

    verbout("\nFeature statistics prepared.",master=master)

    ##Step 1: Create Resulting Vector between Class 1 and 0

    clist = [[],[]]
    for i in cases:
        if classvec[i] == 1:
            clist[1].append(i)
        else: ## Categorizes 1 versus all other
            clist[0].append(i)

    rvec  ={'Intercept':0.0}

    for t in terms:
        rvec[t] = 0

    step = int(len(clist[1])*len(clist[0])/40)
    panz = 1
    if step<1:
        panz = 2 ##approximation which will never be used
        step = 1
    verbout('\n\nSeeking discriminating vector:\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    
    npairs = 0
    for c1 in clist[1]:
        for c0 in clist[0]:
            npairs+=1
            if npairs%step == 0: verbout('.','progress',master)
            for t in terms:
                try:
                    rvec[t] += tdm[t][c1] - tdm[t][c0]
                except:
                    verbout('Fehler',c1,c0,t,master)
    ranking = []
    for term in terms:
        rvec[term] = rvec[term]/npairs
        ranking.append((abs(rvec[term]/stats[term]['SD']),term))

    ranking = sorted(ranking, reverse=True)

    verbout("\n\nFound Vector between 0 and 1. Highest values (first 20): ",master=master)
    for e in ranking[:20]:
        verbout("\n   Feature '"+e[1]+"': "+"{0:.2f}".format(e[0]),'table',master=master)

    cuti =int(len(ranking)/50)
##    for i in range(len(ranking)):
##        if ranking[i][0] < .1:
##            cuti = i
##            break
    if cuti > 0:
        verbout('\n\nPruning list to include only top 2%',master=master)
        ranking = ranking[:cuti]
    verbout('\nRemaining possible discriminating features: '+str(len(ranking)),master=master)

    ##Step 3: Train SVM


    verbout("\n\nSeeking optimal hyperplane by iteratively including features...",master=master)

    use_features = [] ## Replace by iterative loop later.
    highest_f = 0
    count_noninvar = 0
    for nfeats in range(0,len(ranking)):
        use_features.append(ranking[nfeats][1])
        #verbout('\n'+str(nfeats+1)+' - Including "'+ranking[nfeats][1]+'": ',master=master)
        scorelist = svm_scores(tdm,rvec,use_features)
        intercepts = []
        nsteps = 50
        step = (max(scorelist)-min(scorelist))/(nsteps-1)
        ms = min(scorelist)
        for s in range(nsteps):
            intercepts.append(-1*(ms+s*step))

        maxf = 0
        maxint = 0
        for i in intercepts:
            rvec['Intercept'] = i
            scorelist = svm_scores(tdm,rvec,use_features)
            #print(min(scorelist),max(scorelist))
            prf = svm_prf(scorelist,classvec)
            f = prf[2]
            #print(i,f)
            if f > maxf:
                maxf = f
                maxint = i

        #verbout(" : "+str([maxint,maxf]))
        if maxf > highest_f:
            df = maxf-highest_f
            verbout('\n   Added: "'+ranking[nfeats][1]+'\t\tdF='+"{0:.5f}".format(df)+" (Feature "+str(nfeats)+"/"+str(len(ranking))+") ",'table',master=master)
            highest_f = maxf
            optinum = nfeats+1
            optrvec = {'Intercept':maxint}
            for t in use_features:
                optrvec[t] = rvec[t]
            count_noninvar = 0
        else:
            count_noninvar +=1
            verbout('.','table',master=master)
            use_features = use_features[:-1] ##Remove latest addition
            #if count_noninvar > 20: break

    verbout('\n\nBest number of features: '+str(len(use_features)),master=master)
    verbout('\nFirst 20:\n',master=master)
    verbout(str(use_features[:20]),'table',master=master)

    return[optrvec,maxf]



#### Ngram Shingling

def naive_tokenizer(line, num=0):
    line = line.lower()
    allowed = [' ','a','b','c','d','e','f','g','h','i','j','k','l',
               'm','n','o','p','q','r','s','t','u','v','w','x','y','z']
    if num == 1:
        allowed = allowed + ['1','2','3','4','5','6','7','8','9','0']
    outline = ''
    for c in line:
        if c in allowed:
            outline = outline + c

    outline = outline.split(' ')
    while '' in outline:
        outline.remove('')

    return outline


def nghash(ngram,ns=5):
    l = len(ngram)
    step = int(l/ns)
    if step < 1: step =1
    cid = list(range(0,l,step))
    while len(cid)<ns+1:
        cid.append(l-1)
    if len(cid)>ns+1:
        cid=cid[:ns+1]
    
    summe=len(ngram)*10**(ns+2)

    for i in range(ns+1):
        val = (ord(ngram[cid[i]])-97)*10**i
        summe=summe+val

    return str(summe)

def shinglehash(line,tid='none',prev={},ng=5):
    tokens = naive_tokenizer(line)
    shingles = []
    outdic = prev
    for i in range(len(tokens)-ng-1):
        ngl = tokens[i:i+ng]
        ngs = ''
        wording = ''
        ngs = ''.join(ngl)
        wording = ' '.join(ngl)
        sh = nghash(ngs)
        result = [wording,tid]
        try:
            outdic[sh].append(result)
        except:
            outdic[sh] = [result]

    return outdic


def hash_texts(tid,fulltext,mode=1,ngl=5,master=''):
    all_hashes = {}

    step = int(len(tid)/40)
    panz = 1
    if step<1:
        panz = 2 ##approximation which will never be used
        step = 1
    verbout('\n\nHashing Texts:\n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    for i in range(len(tid)):
        all_hashes = shinglehash(fulltext[i],tid[i],all_hashes,ngl)
        if i%step==0: verbout(".","progress",master=master)

    verbout("\nHashes prepared. Removing hashes only occurring once or in >10% of texts...",master=master)
    maxtexts = int(0.1*len(tid))

    if mode == 1:
        remkey = []
        for k in all_hashes.keys():
            if len(all_hashes[k])==1 or len(all_hashes[k])>maxtexts:
                remkey.append(k)
        verbout("\nRemoving "+str(len(remkey))+" hashes",master=master)
        for r in remkey:
            del all_hashes[r]

    verbout("\nReturning "+str(len(all_hashes.keys()))+" relevant hashes",master=master)

    return all_hashes

def duplicate_shingling(tdata,idvar,tvar,ngl=5,master=''):
    sh = hash_texts(tdata[0][idvar],tdata[0][tvar],1,ngl,master)

    verbout("\n\nPreparing Overlap Matrix..",master=master)
    textmatrix = {}
    texts = tdata[0][idvar]

    for t1 in texts:
        textmatrix[t1] = {}
        for t2 in tdata[0][idvar]:
            textmatrix[t1][t2] = 0

    maxover = 0
    for dupli in sh.keys():
        aff_texts = {}
        for e in sh[dupli]:
            aff_texts[e[1]] = 1

        at = aff_texts.keys()

        if len(at) > 1:
            for t1 in at:
                for t2 in at:
                    if not t1 == t2:
                        textmatrix[t1][t2] += 1
                        if textmatrix[t1][t2]>maxover:
                            maxover = textmatrix[t1][t2]
    verbout("\nDone. Maximal overlap between two texts: "+str(maxover)+" hashes.",master=master)

    distr = {}
    for i in range(maxover+1):
        distr[i] = 0
    for t1 in texts:
        for t2 in texts:
            if not t1==t2:
                distr[textmatrix[t1][t2]] +=1

    displim = 30
    if maxover<displim:displim=maxover

    dtab = {'N_Overlap':{},'N_Pairs':{}}
    rn=[]

    for i in range(displim):
        dtab['N_Overlap'][str(i)]=i
        dtab['N_Pairs'][str(i)]=distr[i]/2
        rn.append(str(i))

    verbout("\nDistribution of overlapping hashes up (max 30 overlaps):\n",master=master)
    verbout(display_table(dtab,rows_pre=rn,cols_pre=['N_Overlap','N_Pairs']),'table',master=master)

    return textmatrix
    #return ndupli





#########################################
##                                     ##
##    Individual Analysis Functions    ##
##                                     ##
#########################################



##########################
##
## Aggregate
##
##########################

        

def aggregate(dset,key,var,method,weight=0,master=''):
    ##Aggregates a dataset with a specific method.
    ##dset: Dataset or data dictionary
    ##key: Grouping variables (list of variable names)
    ##var: Variables to be aggregated (list of variable names)
    ##method: Method for calculation (see calculate())
    ##weight:Weighting factor: vector containing the weights for each line in the dataset

    if type(dset)==dict:
        data = dset
    else:
        data = dset[0]
    if type(key) == str:key=[key]
    
    verbout('\n\nAggragating dataset.\nAssessing data for aggregation:\n',master=master)
    if weight == 0:
        tw = []
        for e in data[key[0]]:
            tw.append(1.0)
        weight = tw
    out_data = {}
    out_var = []
    for v in key:
        out_data[v] = []
        out_var.append(v)
    out_data['Number_of_Cases'] = []
    out_var.append('Number_of_Cases')
    out_data['WNumber_of_Cases'] = []
    out_var.append('WNumber_of_Cases')
    if method in ['all','wall']:
        for v in var:
            for prefix in ['M_','SD_','N_']:
                out_data[prefix+v]=[]
                out_var.append(prefix+v)
    else:  
        for v in var:
            out_data[v] = []
            out_var.append(v)

    step = int(len(data[key[0]])/40)        
    if step<1: step = 1
    verbout('\nIdentifying groups of '+str(len(data[key[0]]))+' cases: \n',master=master)
    verbout('0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    agg_dic = {}

    for i in range(len(data[key[0]])):
        if i % step == 0:
            verbout('.','progress',master=master)
        agg_key = []
        for v in key:
            agg_key.append(data[v][i])
        agg_key = str(agg_key)
        agg_dic[agg_key] = {}
    for a in agg_dic.keys():
        for v in var+key:
            agg_dic[a][v] = []
    verbout('\n','progress',master=master)
    verbout('\nFound '+str(len(agg_dic.keys()))+' Groups.',master=master)

    verbout('\n\n',master=master)
    verbout('Assigning '+str(len(data[key[0]]))+' cases to respective groups: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    stepstone = int(len(data[key[0]]) / 10)+1
    for i in range(len(data[key[0]])):
        if i % step == 0:
            verbout('.','progress',master=master)
        agg_key = []
        for v in key:
            agg_key.append(data[v][i])
        agg_key = str(agg_key)
        for v in var+key:
            agg_dic[agg_key][v].append((data[v][i],weight[i]))

    verbout('\n','progress',master=master)
    verbout('\n',master=master)

    if method == 'broad':
        out_var = []
        for k in key:
            out_var.append(k)
        ml = 0
        for k in agg_dic.keys():
            if len(agg_dic[k][var[0]]) > ml:
                ml = len(agg_dic[k][var[0]])
        verbout('\nHighest number of cases per group: '+str(ml),master=master)

        for num in range(0,ml):
            for v in var:
                out_var.append(v + "{0:02}".format(num+1))
                out_data[v + "{0:02}".format(num+1)] = []

        verbout('\nVariables in new dataset:\n',master=master)
        verbout(str(out_var)+'\n','table',master=master)
        
        for a in sorted(agg_dic.keys()):
            for v in key:
                out_data[v].append(agg_dic[a][v][0][0])
            length = len(agg_dic[a][key[0]])
            for i in range(0,length):
                for v in var:
                    vname = v + "{0:02}".format(i+1)
                    out_data[vname].append(agg_dic[a][v][i][0])
            for i in range(length,ml):
                for v in var:
                    vname = v + "{0:02}".format(i+1)
                    out_data[vname].append('')

    elif method in ['all','wall']:
        step = int(len(agg_dic.keys())/40)          
        if step<1: step = 1
        verbout('\n',master=master)
        verbout('Aggregating to '+str(len(agg_dic.keys()))+' groups with function "'+str(method)+'": \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
        i = 0

        for a in sorted(agg_dic.keys()):
            i = i + 1
            if i % step == 0:
                verbout('.','progress',master=master)
            for v in key:
                out_data[v].append(agg_dic[a][v][0][0])
            out_data['Number_of_Cases'].append(len(agg_dic[a][key[0]]))
            wn = 0
            for tup in agg_dic[a][key[0]]:
                wn = wn + tup[1]
            out_data['WNumber_of_Cases'].append(wn)
            for v in var:
                m = calculate(agg_dic[a][v],'mean')
                sd = calculate(agg_dic[a][v],'sd')
                n = calculate(agg_dic[a][v],'nval')
                out_data['M_'+v].append(m)
                out_data['SD_'+v].append(sd)
                out_data['N_'+v].append(n)  

    else:
        step = int(len(agg_dic.keys())/40)          
        if step<1: step = 1
        verbout('\n',master=master)
        verbout('Aggregating to '+str(len(agg_dic.keys()))+' groups with function "'+str(method)+'": \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
        i = 0

        for a in sorted(agg_dic.keys()):
            i = i + 1
            if i % step == 0:
                verbout('.','progress',master=master)
            for v in key:
                out_data[v].append(agg_dic[a][v][0][0])
            out_data['Number_of_Cases'].append(len(agg_dic[a][key[0]]))
            wn = 0
            for tup in agg_dic[a][key[0]]:
                wn = wn + tup[1]
            out_data['WNumber_of_Cases'].append(wn)
            for v in var:
                wert = calculate(agg_dic[a][v],method)
                out_data[v].append(wert)

    verbout('\n','progress',master=master)

    return [out_data,out_var]


def agg_entropy(data,group,mvar,mode='1',master=''):
    ##Aggregates data by calculating the entropy within each group.
    ##data: Data dictionary (not dataset)
    ##group: Group variables (List of variable names)
    ##mvar: Variables to be aggregated (List of variable names)
    ##mode: 1: All groups have the same population of values. 0: The population of values may differ between groups.
    
    out_dic = {}
    out_var = []
    keydic = {}
    groupdic = {}
    val_dic = {}

    verbout('\n\nCalculating entropy within groups.',master=master)

    for g in group:
        out_var.append(g)
    for m in mvar:
        keydic[m] = {}
        val_dic[m] = {}
        keydic[m]['Vname'] = 'Entropy_'+m
        keydic[m]['Nname'] = 'N_Elements_'+m
        out_var.append(keydic[m]['Nname'])
        out_var.append(keydic[m]['Vname'])
    for v in out_var:
        out_dic[v] = []

    edic = {}
    for i in range(len(data[group[0]])):
        g = ''
        vl = {}
        for v in group:
            vl[v] = data[v][i]
            g = g+data[v][i]+'/'
            groupdic[g] = vl
        if not g in edic.keys():
            edic[g] = {}
            for v in mvar:
                edic[g][v] = []
        for v in mvar:
            e = data[v][i]
            if not e in ['',' ']:
                edic[g][v].append(e)
                val_dic[v][e] = 1
    
    verbout('\n\nTotal number of groups: '+str(len(edic.keys())),master=master)
    for v in mvar:
        verbout('\nTotal number of values vor variable "'+v+'": '+str(len(val_dic[v].keys())),master=master)

    verbout('\n\nCalculating and writing entropies.\n',master=master)
    for g in sorted(edic.keys()):
        valid = 0
        for v in mvar: #Check group validity
            if len(edic[g][v]) > 1:
                valid = 1

        if valid == 1:
            for gvar in group:
                out_dic[gvar].append(groupdic[g][gvar])
            for v in mvar:
                nvar = keydic[v]['Nname']
                vvar = keydic[v]['Vname']
                out_dic[nvar].append(len(edic[g][v]))
                if len(edic[g][v]) > 1:
                    if mode == '1':
                        out_dic[vvar].append(calc_entropy(edic[g][v],val_dic[v].keys()))
                    else:
                        out_dic[vvar].append(calc_entropy(edic[g][v]))
                else:
                    out_dic[vvar].append('')

    return [out_dic,out_var]



def analyze_entropy(data,group,varlist,opt={'comp':0,'boot':0},report = '',master=''):
    verbout('\n\nAnalysis of Entropy',master=master)
    outreport = report

    bootstrap_steps = 1000
    comparison_steps = 200

    groupvar = data[group]
    groups = get_unique(groupvar)
    verbout('\nNumber of Groups: '+str(len(groups)),master=master)
    verbout('\nNumber of cases: '+str(len(groupvar)),master=master)
    comparison = 0


    if opt['comp'] == 1:
        comparison = 1
        ratiolist = {}
        compare = {}
        compare['mean'] = {}
        compare['group'] = {}
        for v in varlist:
            ratiolist[v] = []
            compare['group'][v] = {}
            for g in groups:
                compare['group'][v][g] = []
        outreport = outreport + '\n\nResults for random group assignment (N='+str(comparison_steps)+')\n'
        verbout('\n\n',master=master)
        verbout('Calculating results for random group distribution: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
        step = comparison_steps/40

        for randi in range(comparison_steps):           
            groupvar = data[group]
            random.shuffle(groupvar)
            groups = get_unique(groupvar)
            if randi%step == 0:
                verbout('.','progress',master=master)
            
            group_id = {}
            for g in groups:
                group_id[g] = []
            for i in range(len(groupvar)):
                group_id[groupvar[i]].append(i)

            for v in varlist:
                cats = get_unique(data[v])
                ent_overall = calc_entropy(data[v])
                group_ents = []
                for g in groups:
                    catlist = []
                    for i in group_id[g]:
                        catlist.append(data[v][i])
                    ent_group = calc_entropy(catlist,cats)
                    compare['group'][v][g].append(ent_group)
                    group_ents.append(ent_group)

                within_entropy = sum(group_ents)/len(group_ents)
                ratio_entropy = within_entropy / ent_overall
                ratiolist[v].append(ratio_entropy)

        verbout('\n','progress',master=master)
        verbout('\n',master=master)

        for v in varlist:
            mean_ratio = sum(ratiolist[v])/len(ratiolist[v])
            sd_ratio = calculate(ratiolist[v],'sd')
            compare['mean'][v] = [mean_ratio,sd_ratio]
            outreport = outreport + verbout('\nMean ratio vor variable "'+v+'": '+"{0:.3f}".format(mean_ratio)+ '('+ "{0:.3f}".format(sd_ratio) +')',master=master)
            for g in groups:
                mean_ent = sum(compare['group'][v][g])/len(compare['group'][v][g])
                sd_ent = calculate(compare['group'][v][g],'sd')
                compare['group'][v][g] = [mean_ent,sd_ent]
        outreport = outreport + '\n'

    var_report = ''

    group_id = {}
    for g in groups:
        group_id[g] = []
    for i in range(len(groupvar)):
        group_id[groupvar[i]].append(i)

    for g in groups:
        outreport = outreport + verbout('\nCases in Group "'+g+'": '+str(len(group_id[g])),master=master)
        

    for v in varlist:
        outreport = outreport + verbout('\n\nAnalysis for variable: '+v,master=master)
        cats = get_unique(data[v])
        outreport = outreport + verbout('\nValues: '+str(cats),master=master)
        if opt['boot'] == 1:
            verbout('\n\n',master=master)
            verbout('Bootstrapping ('+str(bootstrap_steps)+'): \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
            step = bootstrap_steps/40
            ov_entlist = []
            for bs in range(bootstrap_steps):
                if bs%step==0:verbout('.','progress',master=master)
                bs_sample = bootstrap_sample(data[v])
                ent = calc_entropy(bs_sample)
                ov_entlist.append(ent)
            verbout('\n','progress',master=master)
            ent_overall = sum(ov_entlist)/len(ov_entlist)
            ent_sd = calculate(ov_entlist,'sd')
            overall_ent_lab = "{0:.3f}".format(ent_overall)+ '('+ "{0:.3f}".format(ent_sd) +')'
        else:
            ent_overall = calc_entropy(data[v])
            overall_ent_lab = "{0:.3f}".format(ent_overall)
        outreport = outreport + verbout('\nOverall Entropy: '+overall_ent_lab+'\n',master=master)

        group_ents = []

        for g in groups:
            outreport = outreport + verbout('\n   Analysis for group: '+g,'table',master=master)
            catlist = []
            for i in group_id[g]:
                catlist.append(data[v][i])
                
            if opt['boot']==1:
                gr_entlist = []
                for bs in range(bootstrap_steps):
                    bs_sample = bootstrap_sample(catlist)
                    ent = calc_entropy(bs_sample)
                    gr_entlist.append(ent)
                ent_group = sum(gr_entlist)/len(gr_entlist)
                ent_sd = calculate(gr_entlist,'sd')
                ent_group_lab = "{0:.3f}".format(ent_group)+ '('+ "{0:.3f}".format(ent_sd) +')'
                group_ents = group_ents + gr_entlist
            else:
                ent_group = calc_entropy(catlist,cats)
                ent_group_lab = "{0:.3f}".format(ent_group)
                group_ents.append(ent_group)

            if comparison == 1:
                if ent_group > compare['group'][v][g][0] + 2.57 * compare['group'][v][g][1]:
                    ent_group_lab = ent_group_lab + '** (higher)'
                elif ent_group > compare['group'][v][g][0] + 1.96 * compare['group'][v][g][1]:
                    ent_group_lab = ent_group_lab + '* (higher)'
                elif ent_group < compare['group'][v][g][0] - 2.57 * compare['group'][v][g][1]:
                    ent_group_lab = ent_group_lab + '** (lower)'
                elif ent_group < compare['group'][v][g][0] - 1.96 * compare['group'][v][g][1]:
                    ent_group_lab = ent_group_lab + '* (lower)'
            
            outreport = outreport + verbout('\n   Group Entropy: '+ent_group_lab+'\n','table',master=master)


        if opt['boot']==1:
            within_entropy = sum(group_ents)/len(group_ents)
            within_sd = calculate(group_ents,'sd')
            ratiolist = []
            for e in ov_entlist:
                try:
                    ratiolist.append(within_entropy/e)
                except:
                    ratiolist.append(0)
            ratio_entropy = sum(ratiolist)/len(ratiolist)
            ratio_sd = calculate(ratiolist,'sd')
            within_lab = "{0:.3f}".format(within_entropy)+ '('+ "{0:.3f}".format(within_sd) +')'
            ratio_lab = "{0:.3f}".format(ratio_entropy)+ '('+ "{0:.3f}".format(ratio_sd) +')'
        else:
            within_entropy = sum(group_ents)/len(group_ents)
            ratio_entropy = within_entropy / ent_overall              
            within_lab = "{0:.3f}".format(within_entropy)
            ratio_lab = "{0:.3f}".format(ratio_entropy)

        if opt['comp'] == 1:
            if ratio_entropy > compare['mean'][v][0] + 2.57*compare['mean'][v][1]:
                ratio_lab = ratio_lab + '** (higher)'
            elif ratio_entropy > compare['mean'][v][0] + 1.96*compare['mean'][v][1]:
                ratio_lab = ratio_lab + '* (higher)'
            elif ratio_entropy < compare['mean'][v][0] - 2.57*compare['mean'][v][1]:
                ratio_lab = ratio_lab + '** (lower)'
            elif ratio_entropy < compare['mean'][v][0] - 1.96*compare['mean'][v][1]:
                ratio_lab = ratio_lab + '* (lower)'
            else:
                ratio_lab = ratio_lab + ' (n.s.)'
      

        outreport = outreport + verbout('\nMean within Entropy: '+str(within_lab),master=master)
        outreport = outreport + verbout('\nRatio of within and without Entropy: '+str(ratio_lab),master=master)

        var_report = var_report + '\n\nVariable "'+v+'": \n - Overall Entropy: '+str(overall_ent_lab)
        var_report = var_report + '\n - Mean Entropy within Groups: '+str(within_lab)
        var_report = var_report + '\n - Entropy Ratio: '+str(ratio_lab)
        if opt['comp'] == 1:
            var_report = var_report + '\n - Expected Ratio by chance: '+"{0:.3f}".format(compare['mean'][v][0])+ '('+ "{0:.3f}".format(compare['mean'][v][1]) +')'

    outreport = outreport + verbout('\n\nExplanation:\nRatios higher than expected by chance indicate that the distribution of elements within groups are evenly distributed.\nRatios lover than expected indicate uneven distributions within groups.',master=master)
        

    outreport = outreport + verbout('\n\nSummary for variables: \n-----------------\n',master=master)
    outreport = outreport + verbout(var_report,'table',master=master)

    return outreport
    



##########################
##
## Sequence analysis
##
##########################

def collapse_repetitions(sequences,minlen,maxlen):
    patterndic = {}
    for rlen in range(minlen,maxlen+1):
        for seq in sequences.keys():
            s = sequences[seq]
            i = 0
            while i < len(s):
                pat = s[i:i+rlen]
                k = i + rlen
                rep = 1
                anzrep = 1
                while k<len(s)-rlen and rep == 1:
                    rep = 0
                    if s[k:k+rlen] == pat:
                        k = k+rlen
                        rep = 1
                        anzrep=anzrep+1
                    elif anzrep>1:
                        k2 = k
                        for j in range(1,rlen):
                            if s[k:k+j] == pat[:j]:
                                k2 = k+j
                        k = k2
                if k > i+rlen:
                    niceprint = '('
                    for symb in sorted(pat):
                        niceprint = niceprint + symb+','
                    niceprint = niceprint[:-1]+')*'
                    s = s[:i]+[niceprint]+s[k:]
                    if not niceprint in patterndic.keys():
                        patterndic[niceprint] = []
                    patterndic[niceprint].append(anzrep)
                i = i+1
            sequences[seq] = s

    report = 'Repetitious patterns found: '
    
    for p in sorted(patterndic.keys()):
        if len(patterndic[p]) > 1:
            report = report + '\n'+str(p)+' Occurrences: '+str(len(patterndic[p]))+'; Mean length: {0:3.2f}'.format(float(sum(patterndic[p]))/len(patterndic[p]))

    return sequences, report

def calc_adios(data,symb,gvar='',tvar='',mini=0,rep=[1,3],length=[2,7],eta=.9,subst=1,master=''):
    if tvar == '':
        data['notime'] = range(len(data[symb]))
        tvar = 'notime'
    else:
        data[tvar] = transform_float(data[tvar])  
    if gvar == '':
        data['nogroup'] = []
        for i in range(len(data[symb])):
            data['nogroup'].append(1)
        gvar = 'nogroup'

    outstring = ''
    sequences = {}
    outstring = outstring + 'Pattern recognition.\nSettings:\nSymbol Variable: '+str(symb)+'\n'
    outstring = outstring + 'Group Variable: '+str(gvar)+'\n'
    outstring = outstring + 'Time Variable: '+str(tvar)+'\n'
    outstring = outstring + 'Collapsing repetitious sequences with lengths in interval: '+str(rep)+'\n'
    outstring = outstring + 'Minimal length of sequence to be counted: '+str(mini)+'\n'
    outstring = outstring + 'Extracting sequences with lengths in interval: '+str(length)+'\n'
    outstring = outstring + 'Eta-Coefficient (discrimination): '+str(eta)

    ### Sort dataset

    data = sort_table(data,[gvar,tvar])
    verbout('\nTable sorted by Group and Time',master=master)

    outstring = outstring + '\nTotal cases in datafile: '+str(len(data[gvar]))
    data = delete_missing(data,[gvar,tvar],0)
    outstring = outstring + '\nValid cases in datafile: '+str(len(data[gvar]))+'\n-------------------------------\n\n'
    
    ###Create sequences to operate with

    verbout('\nExtracting sequences...',master=master)

    currsymb = 0
    anzsymb = 0
    currgroup = 0
    groupstats = {}
    currsent = []
    stat_len = []
    stat_glitch = 0
    for i in range(len(data[symb])):
        if currsymb == 0:
            currgroup = data[gvar][i]
            currsymb = data[symb][i]
            anzsymb = 1
        elif currgroup == data[gvar][i]:
            if currsymb == data[symb][i]:
                anzsymb = anzsymb + 1
            else:
                if anzsymb >= mini:
                    for k in range(anzsymb):
                        currsent.append(currsymb)
                else:
                    stat_glitch = stat_glitch + 1
                currsymb = data[symb][i]
                anzsymb = 1
        else:
            if anzsymb >= mini:
                for k in range(anzsymb):
                    currsent.append(currsymb)
            else:
                stat_glitch = stat_glitch + 1
            sequences[currgroup] = currsent
            stat_len.append(len(currsent))
            groupstats[currgroup]={}
            groupstats[currgroup]['NSymb']=len(currsent)
            currsent = []
            currgroup = data[gvar][i]
            currsymb = data[symb][i]
            anzsymb = 1

    if anzsymb >= mini:
        for k in range(anzsymb):
            currsent.append(currsymb)
    else:
        stat_glitch = stat_glitch + 1
    sequences[currgroup] = currsent
    stat_len.append(len(currsent))
    groupstats[currgroup]={}
    groupstats[currgroup]['NSymb']=len(currsent)

    if len(sequences.keys())>0:      
        lstat = stat_desc(stat_len)
        verbout('\n\nNumber of sequences: '+str(len(sequences.keys()))+'\nMean length: {0:5.2f}'.format(lstat['M']),master=master)
        verbout('\nStandard deviation: {0:5.2f}'.format(lstat['SD'])+'\nMinimum length: '+str(int(lstat['Min'])),master=master)
        verbout('\nMaximum length: '+str(int(lstat['Max']))+'\nNumber of glitches total: '+str(stat_glitch),master=master)
    else:
        verbout('\n\nNo valid sequences found.\n','warning',master=master)

    ### Collapsing repetitious patterns

    if rep[1] > 0:
        minrep = rep[0]
        maxrep = rep[1]
        verbout('\n\nCollapsing repetitious patterns...',master=master)

        outsent, rep_report = collapse_repetitions(sequences,minrep,maxrep)
        verbout(rep_report,master=master)

        stat_len = []
        for s in sequences.keys():
            stat_len.append(len(sequences[s]))
            groupstats[s]['NNgram']=len(sequences[s])
            
        if len(sequences.keys())>0:
            lstat = stat_desc(stat_len)
            verbout('\n\nNumber of sequences: '+str(len(sequences.keys()))+'\nMean length: {0:5.2f}'.format(lstat['M']),master=master)
            verbout('\nStandard deviation: {0:5.2f}'.format(lstat['SD'])+'\nMinimum length: '+str(int(lstat['Min'])),master=master)
            verbout('\nMaximum length: '+str(int(lstat['Max']))+'\nNumber of glitches total: '+str(stat_glitch),master=master)
        else:
            verbout('\n\nNo valid sequences found.\n','warning',master=master)
        outstring = outstring + '\n' + str(rep_report)+'\n'


    ### Looking for Paths

    verbout('\n\nGetting Nodes and Paths inside parameters...'+str(length),master=master)

    minlen=length[0]
    maxlen=length[1]

    nodes = {}
    paths = {}
    pathlists = {}
    n_paths = 0

    for s in sequences.keys():
        for c in sequences[s]:
            nodes[c] = 0

    for s in sequences.keys():
        for c in sequences[s]:
            nodes[c] = nodes[c] + 1

    for length in range(1,maxlen+1):
        for s in sequences.keys():
            for i in range(0,len(sequences[s])-length+1):
                paths[str(sequences[s][i:i+length])] = 0
                pathlists[str(sequences[s][i:i+length])] = sequences[s][i:i+length]
        for s in sequences.keys():
            for i in range(0,len(sequences[s])-length+1):
                paths[str(sequences[s][i:i+length])] = paths[str(sequences[s][i:i+length])]+1
                n_paths = n_paths+1

    verbout('\nFound '+str(len(paths.keys()))+' individual paths. Npaths = '+str(n_paths),master=master)
    outstring = outstring + '\n\nExtracting repetitious sequences from '+str(n_paths)+' possible sequences in dataset:\n'

    verbout('\n\nSearching for common patterns with Eta: '+str(eta),master=master)

    patterns = {}

    for sent in sequences.keys():
        s = sequences[sent]
        rpatterns = {}
        lpatterns = {}
        for i in range(len(s)):
            prev_Pr = 0.1
            if i+maxlen+1<=len(s):
                maxj = i+maxlen+1
            else:
                maxj = len(s)
            for j in range(i+1,maxj):
                seq = s[i:j]
                if len(seq) == 1:
                    Pr = float(paths[str(seq)])/n_paths
                else:
                    Pr = float(paths[str(seq[:-1])])/paths[str(seq)]
                if Pr/prev_Pr<eta:
                    if len(seq)>minlen:
                        rpatterns[str(seq[:-1])]={'Sequence':seq[:-1],'Pr':prev_Pr,'Dr':Pr/prev_Pr}
                prev_Pr = Pr

        for i in range(len(s),0,-1):
            prev_Pl = 0.1
            if i-maxlen-1>=0:
                minj = i-maxlen-1
            else:
                minj = 0
            for j in range(i-1,minj,-1):
                seq = s[j:i]
                if len(seq) == 1:
                    Pl = float(paths[str(seq)])/n_paths
                else:
                    Pl = float(paths[str(seq)])/paths[str(seq[:-1])]
                if Pl/prev_Pl<eta:
                    if len(seq)>minlen:
                        lpatterns[str(seq[:-1])]={'Sequence':seq[:-1],'Pl':prev_Pl,'Dl':Pl/prev_Pl}
                prev_Pl = Pl

        for r in rpatterns.keys():
            if r in lpatterns.keys():
                patterns[r]={}
                patterns[r]['Sequence'] = rpatterns[r]['Sequence']
                patterns[r]['Pr'] = rpatterns[r]['Pr']
                patterns[r]['Dr'] = rpatterns[r]['Dr']
                patterns[r]['Pl'] = lpatterns[r]['Pl']
                patterns[r]['Dl'] = lpatterns[r]['Dl']
                patterns[r]['AvD'] = (rpatterns[r]['Dr'] + lpatterns[r]['Dl'])/2
                patterns[r]['N_raw'] = paths[r]

    ranking = []            

    for s in patterns:
        ranking.append((patterns[s]['AvD'],s))
    rangliste = sorted(ranking)

    verbout('\nIntermediate result: Found '+str(len(rangliste))+' patterns',master=master)
    outstring = outstring + 'Number of patterns possibly subsuming smaller patterns: '+str(len(rangliste))

    select = []

    for r in rangliste:
        select.append(r[1])


    if subst in [1,2]:
        verbout('\n\nLooking for subsuming patterns...',master=master)

        patlist = []
        patmax = 0
        for s in sorted(patterns.keys()):
            patlist.append(patterns[s]['Sequence'])
            if len(patterns[s]['Sequence']) > patmax:
                patmax=len(patterns[s]['Sequence'])
        
        clean_patlist = []
        for l in range(patmax,minlen-1,-1):
            for p in patlist:
                if len(p) == l:
                    present = 0
                    for cp in clean_patlist:
                        if str(p)[1:-1] in str(cp)[1:-1] and present == 0:
                            verbout('\nFound subsuming pattern: '+str(p)+' / '+str(cp),master=master)
                            outstring = outstring + '\nFound subsuming pattern: '+str(p)+' / '+str(cp)
                            present = 1
                            if subst == 1:
                                del patterns[str(p)]
                            elif subst == 2:
                                del patterns[str(cp)]
                    if present == 0:
                        clean_patlist.append(p)
    else:
        verbout('\n\nSkipped looking for subsuming patterns because of user input.',master=master)


    verbout('\nRemaining: '+str(len(patterns.keys()))+' patterns:',master=master)
    outstring = outstring + '\n\nRemaining: '+str(len(patterns.keys()))+' patterns:'

    npat = {}
    symbols = {}
    nr = 1
    plabels = []
    for p in sorted(patterns.keys()):
        plab = 'P_'+"{0:03}".format(nr)
        patterns[p]['Label'] = plab
        plabels.append(plab)
        npat[plab]=patterns[p]
        verbout('\n'+plab+'\t:'+ p,master=master)
        outstring = outstring + '\n'+plab+'\t:'+ p
        for s in patterns[p]['Sequence']:
            symbols[s] = 0
        nr = nr + 1

    ### Formatting output

    outstring = outstring + '\n\nIndividual sequences in dataset after replacing patterns:\n'

    occurrences = {}

    for s in sorted(sequences.keys()):
        ind_seq = str(sequences[s])
        outstring = outstring + str(s) + ': '+ ind_seq +'\n'
        occ = {}
        for p in sorted(patterns.keys()):
            currlab = patterns[p]['Label']
            p_seq = p[1:-1]
            start=0
            while ind_seq.find(p_seq,start)>0:
                if currlab in occ.keys():
                    occ[currlab] = occ[currlab]+1
                else:
                    occ[currlab] = 1
                start = ind_seq.find(p_seq,start)+1
        occurrences[s] = occ
        verbout('\nPattern occurrence for group: '+str(s) + ':\n'+baum_schreiben(occ),master=master)


    outstring = outstring + '\n\nSequences by Group:\n--------------\nGroup\tLength\tNgrams'
    for p in plabels:
        outstring = outstring + '\t'+p
    outstring = outstring + '\n'

    for s in sorted(sequences.keys()):
        outstring = outstring + str(s)
        outstring = outstring + '\t' + str(groupstats[s]['NSymb'])
        outstring = outstring + '\t'+ str(len(sequences[s]))
        for p in plabels:
            if p in occurrences[s].keys():
                outstring = outstring + '\t' + str(occurrences[s][p])
            else:
                outstring = outstring + '\t0'
        outstring = outstring + '\n'
    outstring = outstring + '\n\n'

    return outstring 
       

def find_tpats(data,tvar,gvar,varlist,p_level=0.05,long_events=1,master=''):
    timedic = {}
    begindic = {}
    enddic = {}
    posdic = {}
    tbegindic = {}
    tenddic = {}
    timepoints = {}
    occurrence = {}
    timespan = {}

    for g in get_unique(data[gvar]):
        timepoints[g] = []
    
    for v in varlist:
        occurrence[v] = 0
    for i in range(len(data[tvar])):
        try:
            t = float(data[tvar][i])                    
            g = data[gvar][i]
            timedic[(t,g)]={}
            begindic[(t,g)]={}
            enddic[(t,g)]={}
            tbegindic[(t,g)]={}
            tenddic[(t,g)]={}
            posdic[(t,g)]={}

            timepoints[g].append(t)

            for v in varlist:
                if data[v][i] == '1':
                    timedic[(t,g)][v] = 1
                    begindic[(t,g)][v] = 1
                    enddic[(t,g)][v] = 1
                    tbegindic[(t,g)][v] = t
                    tenddic[(t,g)][v] = t
                    posdic[(t,g)][v] = [t]
                    occurrence[v] = occurrence[v] + 1
                else:
                    timedic[(t,g)][v] = 0
                    begindic[(t,g)][v] = 0
                    enddic[(t,g)][v] = 0
        except:
            t = 0

    timespan['SUMMED_UP'] = 0

    gremove = []

    for g in timepoints.keys():
        if len(timepoints[g]) > 2:
            timepoints[g] = get_unique(timepoints[g])
            timespan[g] = max(timepoints[g])-min(timepoints[g])
            timespan['SUMMED_UP'] = timespan['SUMMED_UP'] + timespan[g]
        else:
            gremove.append(g)


    for g in gremove:
        del timepoints[g]
        verbout('\nRemoving sequence group '+str(g)+' because it had less than two elements',master=master)
        

    verbout('\n\nFound '+str(len(timepoints.keys()))+' different groups.\nTimespans:',master=master)
    for g in timepoints.keys():
        verbout('\n  '+str(g)+': '+str(timespan[g]),master=master)

    if long_events == 1:
        for g in timepoints.keys():
            for v in varlist:
                for i in range(len(timepoints[g])):
                    t = timepoints[g][i]
                    if timedic[(t,g)][v] == 1:
                        if i == 0:
                            begindic[(t,g)][v] == 1
                        elif i == len(timepoints[g])-1:
                            enddic[(t,g)][v] == 1
                        else:
                            tp = timepoints[g][i-1]
                            tn = timepoints[g][i+1]
                            if timedic[(tp,g)][v] == 1:
                                begindic[(t,g)][v] = 0
                                occurrence[v] = occurrence[v] -1 
                            if timedic[(tn,g)][v] == 1:
                                enddic[(t,g)][v] = 0

    tpattern = {}
    all_pattern_found = 0

    while not all_pattern_found == 1:
        verbout('\n\nIdentifying patterns...',master=master)

        pairs = []
        for v1 in varlist:
            for v2 in varlist:
                tpname = 'Tpat_'+v1+'_'+v2
                if not tpname in tpattern.keys():
                    pairs.append((v1,v2))

        for pair in pairs:
            event_a = pair[0]
            event_b = pair[1]
            tpname = 'Tpat_'+event_a+'_'+event_b
            tpattern[tpname] = {}
            tpattern[tpname]['A'] = event_a
            tpattern[tpname]['B'] = event_b
            if event_a in tpattern.keys():
                list_a = tpattern[event_a]['List']
                tree_a = tpattern[event_a]['Tree']
            else:
                list_a = [event_a]
                tree_a = event_a
            if event_b in tpattern.keys():
                list_b = tpattern[event_b]['List']
                tree_b = tpattern[event_b]['Tree']
            else:
                list_b = [event_b]
                tree_b = event_b

            length_a = len(list_a)
            length_b = len(list_b)
            
            tpattern[tpname]['List']=list_a + list_b
            tpattern[tpname]['Tree']=[tree_a,tree_b]

            tpattern[tpname]['Start']=[]
            tpattern[tpname]['End']=[]
            tpattern[tpname]['Dist']=[]
            tpattern[tpname]['Length']=[]                
            tpattern[tpname]['Anz']=0
            tpattern[tpname]['Valid']=0
            tpattern[tpname]['Positions']=[]
            tpattern[tpname]['p']=1.0
            
            for g in timepoints.keys():
                i = 0
                t = timepoints[g][i]
                while not i == -1:
                    t = timepoints[g][i]
                    if enddic[(t,g)][event_a] == 1:
                        tstart = t
                        i = i + 1
                        if i > len(timepoints[g])-1:
                            i = -1
                        i2 = i
                        while not i2 == -1:
                            t = timepoints[g][i2]
                            if begindic[(t,g)][event_b] == 1:
                                tend = t
                                truend = tenddic[(t,g)][event_b]
                                i = i2
                                while not timepoints[g][i] <= tstart:
                                    i = i - 1
                                    t = timepoints[g][i]
                                    if enddic[(t,g)][event_a] == 1:
                                        tstart = t
                                        trustart = tbegindic[(t,g)][event_a]
                                        i2 = -2
                                        tpattern[tpname]['Start'].append((trustart,g))
                                        tpattern[tpname]['End'].append((truend,g))
                                        tpattern[tpname]['Dist'].append(tend-tstart)
                                        tpattern[tpname]['Length'].append(truend-trustart+1)
                                        tpattern[tpname]['Anz']=tpattern[tpname]['Anz']+1
                                        tpattern[tpname]['Positions'].append(posdic[(trustart,g)][event_a]+posdic[(tend,g)][event_b])
                            i2 = i2+1
                            if i2 > len(timepoints[g])-1:
                                i2 = -1
                                
                    i = i + 1
                    if i > len(timepoints[g])-1 or i == 0:
                        i = -1

        verbout('\n\nAssessing p-values:\n',master=master)


        for pair in pairs:
            event_a = pair[0]
            event_b = pair[1]
            tpat = 'Tpat_'+event_a+'_'+event_b
            #verbout('\nCalculating p-value for: '+str(tpattern[tpat]['List'])+'...',master=master)
            occurrence[tpat] = tpattern[tpat]['Anz']
            #verbout(str(tpattern[tpat]['Anz']),master=master)
            if occurrence[tpat] > 0:
                d0_list = sorted(get_unique(tpattern[tpat]['Dist']))
                d1_list = sorted(get_unique(tpattern[tpat]['Dist']), reverse=True)

                dist_tupel = []
                ##Fast Patterns:
                for d in d1_list:
                    dist_tupel.append((0,d))

                sp_temp = []

                ##Slow Patterns:
                for d0 in d0_list:
                    for d1 in d1_list:
                        if d0<d1:
                            d = d1-d0
                            sp_temp.append((d,(d0,d1)))

                for sp in sorted(sp_temp,reverse=True):
                    dist_tupel.append(sp[1])
                
                p = 1
                while not p < p_level and len(dist_tupel)>0:
                    d0 = dist_tupel[0][0]
                    d1 = dist_tupel[0][1]
                    d = d1-d0
                    
                    occ = 0
                    rd = []
                    for k in range(len(tpattern[tpat]['Dist'])):
                        distocc = tpattern[tpat]['Dist'][k]
                        if distocc <= d1 and distocc >=d0:
                            occ = occ + 1
                        else:
                            rd.append(k)

                    if occ > 2:                      
                        prob_b = float(occurrence[tpattern[tpat]['B']])/timespan['SUMMED_UP']
                        invprob_b = 1-(1-prob_b)**d

                        prob_a = float(occurrence[tpattern[tpat]['A']])/timespan['SUMMED_UP']
                        invprob_a = 1-(1-prob_a)**d

                        p_b = binomial_odds(occurrence[tpattern[tpat]['A']],invprob_b,occ)
                        p_a = binomial_odds(occurrence[tpattern[tpat]['B']],invprob_a,occ)
                        p = max([p_b,p_a]) ##Beidseitiger Test auf Signifikanz
                    else:
                        p = 1.0
                    
                    tpattern[tpat]['p'] = p
                    if p < p_level:
                        tpattern[tpat]['Valid'] = 1
                        tpattern[tpat]['d'] = d
                        tpattern[tpat]['d1'] = d1
                        tpattern[tpat]['d0'] = d0
                        occurrence[tpat] = occ
                        tpattern[tpat]['Anz'] = occ
                        verbout('.',master=master)
                        ns = []
                        ne = []
                        nd = []
                        nl = []
                        np = []
                        for k in range(len(tpattern[tpat]['Dist'])):
                            if not k in rd:
                                ns.append(tpattern[tpat]['Start'][k])
                                ne.append(tpattern[tpat]['End'][k])
                                nd.append(tpattern[tpat]['Dist'][k])
                                nl.append(tpattern[tpat]['Length'][k])
                                np.append(tpattern[tpat]['Positions'][k])
                        
                        tpattern[tpat]['Start'] = ns
                        tpattern[tpat]['End'] = ne
                        tpattern[tpat]['Dist'] = nd
                        tpattern[tpat]['Length'] = nl
                        tpattern[tpat]['Positions'] = np 
                    else:
                        dist_tupel.pop(0)
                        occurrence[tpat] = 0

        for tpat in tpattern.keys():
            if tpattern[tpat]['Valid']==0:
                tpattern[tpat]['Start'] = []
                tpattern[tpat]['End'] = []
                tpattern[tpat]['Dist'] = []
                tpattern[tpat]['Length'] = []

        newpats = []
        for pair in pairs:
            event_a = pair[0]
            event_b = pair[1]
            tpname = 'Tpat_'+event_a+'_'+event_b
            if tpattern[tpname]['p'] < p_level:
                newpats.append(tpname)

        verbout('\nRemoving redundant patterns\n',master=master)

        redundic = {}
        for tpat in tpattern.keys():
            redundic[str(tpattern[tpat]['List'])] = []
        for tpat in tpattern.keys():
            redundic[str(tpattern[tpat]['List'])].append((tpattern[tpat]['p'],tpat))

        bestlist = []


        for r in sorted(redundic.keys()):
            bestlist.append(sorted(redundic[r])[0][1])

        np2 = []
        for n in newpats:
            if n in bestlist:
                np2.append(n)

        newpats = np2

        for tpat in newpats:
            verbout('\nFound T-Pattern: '+str(tpattern[tpat]['Tree']),master=master)
            verbout(' p='+"{:5.4f}".format(tpattern[tpat]['p']),master=master)
            verbout('; N='+str(tpattern[tpat]['Anz']),master=master)
            verbout('; d0='+str(tpattern[tpat]['d0']),master=master)
            verbout('; d1='+str(tpattern[tpat]['d1']),master=master)
            for g in timepoints.keys():
                for t in timepoints[g]:
                    timedic[(t,g)][tpat] = 0
                    begindic[(t,g)][tpat] = 0
                    enddic[(t,g)][tpat] = 0

            for i in range(len(tpattern[tpat]['Dist'])):
                if tpattern[tpat]['Dist'][i] <= tpattern[tpat]['d1'] and tpattern[tpat]['Dist'][i] >= tpattern[tpat]['d0']:
                    g = tpattern[tpat]['Start'][i][1]
                    start = tpattern[tpat]['Start'][i][0]
                    end = tpattern[tpat]['End'][i][0]
                    for t in timepoints[g]:
                        if t >= start and t <= end:
                            timedic[(t,g)][tpat] = 1
                    begindic[(start,g)][tpat] = 1
                    tenddic[(start,g)][tpat]=end
                    enddic[(end,g)][tpat] = 1
                    tbegindic[(end,g)][tpat] = start
                    posdic[(start,g)][tpat] = tpattern[tpat]['Positions'][i]

        if len(newpats) == 0:
            all_pattern_found = 1
        else:
            varlist = varlist + newpats

        remlist = []
        for tpat in tpattern.keys():
            if not tpat in varlist:
                remlist.append(tpat)

        for r in remlist:
            del tpattern[r]

    tmp = open('tpattern_report.txt','w')

    outdic = {}

    for tpat in sorted(tpattern.keys()):
        if tpattern[tpat]['Valid']==1:
            tmp.write('\n'+tpat+'\n\n')
            tmp.write(str(tpattern[tpat]['List'])+'\n')
            tmp.write(str(tpattern[tpat]['Tree'])+'\n\n')
            tmp.write(' p='+"{:5.4f}".format(tpattern[tpat]['p']))
            tmp.write('; N='+str(tpattern[tpat]['Anz']))
            tmp.write('; d0='+str(tpattern[tpat]['d0']))
            tmp.write('; d1='+str(tpattern[tpat]['d1'])+'\n\n')
            for i in range(len(tpattern[tpat]['Dist'])):
                tmp.write(str(tpattern[tpat]['Start'][i])+' - Length: '+str(tpattern[tpat]['Length'][i])+' Pos: ')
                positionlist = ''
                for k in range(len(tpattern[tpat]['Positions'][i])):
                    positionlist = positionlist + str(tpattern[tpat]['List'][k])+': '+str(tpattern[tpat]['Positions'][i][k])+' -> '
                tmp.write(positionlist[:-4] + '\n')

            tmp.write('\n-----------------\n\n')
            outdic[tpat] = tpattern[tpat]
    tmp.close()

    return outdic


def find_sequence(data,svar,tvar='tmp_Time',gvar='tmp_Group',slen=4,somit=1,mode=1,master=''):
    if tvar == 'tmp_Time':
        data['tmp_Time'] = range(0,len(data[svar]))
    if gvar == 'tmp_Group':
        data['tmp_Group']=[]
        for i in data[svar]:
            data['tmp_Group'].append(1)

    seq_dic = {}
    for i in range(0,len(data[svar])):
        if not data[gvar][i] in seq_dic.keys():
            seq_dic[data[gvar][i]] = {}
        seq_dic[data[gvar][i]][data[tvar][i]] = {'S':data[svar][i]}

    if type(mode)==str:
        try:
            mode = int(mode)
        except:
            mode = 1

    seq_vars = []
    seq_collection = []
    slen = str(slen)
    somit = str(somit)

    for p in seq_dic.keys():
        times = sorted(seq_dic[p].keys())
        for t in range(0,len(times)):
            t0 = times[t]
            if len(times) > t+1 and slen in ['2','3','4']:
                t1 = times[t+1]
                seq_dic[p][t0]['Seq_2'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t1]['S']
                seq_vars.append('Seq_2')
                seq_collection.append(seq_dic[p][t0]['Seq_2'])
                
            if len(times) > t+2 and slen in ['2','3','4']:
                t2 = times[t+2]
                if slen in ['3','4']:
                    seq_dic[p][t0]['Seq_3'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t1]['S'] + '; ' + seq_dic[p][t2]['S']
                    seq_vars.append('Seq_3')
                    seq_collection.append(seq_dic[p][t0]['Seq_3'])
                if somit == '1':
                    seq_dic[p][t0]['Seq_2a'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t2]['S']
                    seq_vars.append('Seq_2a')
                    seq_collection.append(seq_dic[p][t0]['Seq_2a'])

            if len(times) > t+3 and slen in ['3','4']:
                t3 = times[t+3]
                if slen == '4':
                    seq_dic[p][t0]['Seq_4'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t1]['S'] + '; ' + seq_dic[p][t2]['S'] + '; ' + seq_dic[p][t3]['S']
                    seq_vars.append('Seq_4')
                    seq_collection.append(seq_dic[p][t0]['Seq_4'])
                if somit == '1':
                    seq_dic[p][t0]['Seq_3a'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t2]['S'] + '; ' + seq_dic[p][t3]['S']
                    seq_vars.append('Seq_3a')
                    seq_collection.append(seq_dic[p][t0]['Seq_3a'])
                    
                    seq_dic[p][t0]['Seq_3b'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t1]['S'] + '; ' + seq_dic[p][t3]['S']
                    seq_vars.append('Seq_3b')
                    seq_collection.append(seq_dic[p][t0]['Seq_3b'])

            if len(times) > t+4 and slen in ['4'] and somit == '1':
                t4 = times[t+4]
                seq_dic[p][t0]['Seq_4a'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t2]['S'] + '; ' + seq_dic[p][t3]['S'] + '; ' + seq_dic[p][t4]['S']
                seq_vars.append('Seq_4a')
                seq_collection.append(seq_dic[p][t0]['Seq_4a'])
                seq_dic[p][t0]['Seq_4b'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t1]['S'] + '; ' + seq_dic[p][t3]['S'] + '; ' + seq_dic[p][t4]['S']
                seq_vars.append('Seq_4b')
                seq_collection.append(seq_dic[p][t0]['Seq_4b'])
                seq_dic[p][t0]['Seq_4c'] = seq_dic[p][t0]['S'] + '; ' + seq_dic[p][t1]['S'] + '; ' + seq_dic[p][t2]['S'] + '; ' + seq_dic[p][t4]['S']
                seq_vars.append('Seq_4c')
                seq_collection.append(seq_dic[p][t0]['Seq_4c'])

                         

    freq = {}
    for f in seq_collection:
        freq[f] = 0
    for f in seq_collection:
        freq[f] = freq[f] + 1

    zlist = []
    for f in freq.keys():
        zlist.append((freq[f],f))

    verbout('\n\nMost frequent sequences:\n',master=master)

    for val in sorted(zlist,reverse=True)[:5]:
        verbout(val[1]+' (' + str(val[0]) + ')\n','table',master=master)


    out_data = {}
    seq_vars = get_unique(seq_vars)

    if mode == 1:
        verbout('\n\nExporting in long format:\n',master=master)
        out_data[gvar] = []
        out_data[tvar] = []
        out_data[svar] = []
        out_data['Type'] = []
        out_data['Sequence'] = []
        outvars = [gvar,tvar,svar,'Type','Sequence']
        for p in sorted(seq_dic.keys()):
            for t in sorted(seq_dic[p].keys()):
                for v in seq_vars:
                    if v in seq_dic[p][t].keys():
                        out_data[gvar].append(p)
                        out_data[tvar].append(t)
                        out_data[svar].append(seq_dic[p][t]['S'])
                        out_data['Sequence'].append(seq_dic[p][t][v])
                        out_data['Type'].append(v)

    if mode == 2:
        verbout('\n\nExporting in broad format:\n',master=master)
        out_data[gvar] = []
        out_data[tvar] = []
        out_data[svar] = []
        for v in seq_vars:
            out_data[v] = []
        outvars = [gvar,tvar,svar]+seq_vars
        for p in sorted(seq_dic.keys()):
            for t in sorted(seq_dic[p].keys()):
                out_data[gvar].append(p)
                out_data[tvar].append(t)
                out_data[svar].append(seq_dic[p][t]['S'])                    
                for v in seq_vars:
                    if v in seq_dic[p][t].keys():
                        out_data[v].append(seq_dic[p][t][v])
                    else:
                        out_data[v].append('')

    return [out_data,outvars]


##########################
##
## Time Series
##
##########################


def recentpeak(date,values,threshold):
    rp = -1
    for i in range(len(date)):
        if values[i] > threshold:
            rp = date[i]

    return rp


def flatten_curve(tvar,svar,window=10,master=''):
    ## Flatten a time series
    ## tvar: Timestamps
    ## svar: Timeseries
    ## window: Window to be considered

    verbout('\n\nFlattening curve\n',master=master)    
    glide= {}
    missings = 0
    step = float(window)/2
    tv = []
    sv = []
    for i in range(len(tvar)):
        try:
           t = float(tvar[i])
           s = float(svar[i])
           glide[t] = []
           tv.append(t)
           sv.append(s)
        except:
            missings = missings + 1

    for i in range(len(tv)):
        wmin = tv[i]-step
        wmax = tv[i]+step
        for ts in glide.keys():
            if ts >= wmin and ts <=wmax:
                glide[ts].append(sv[i])

    outvar1 = []
    outvar2 = []
    for i in range(len(tv)):
        o1 = calculate(glide[tv[i]],'mean')
        o2 = sv[i]-o1
        outvar1.append(o1) ##Flattened curve
        outvar2.append(o2) ##Residuals

    return (outvar1,outvar2)
        

def find_peaks(svar,tvar=0,pdir='1',pthres=95,master=''):
    ##Find peaks in a timeseries
    ##tvar: vector containing the timestamps. If no vector is offered, correct sorting is assumed
    ##svar: vector containing the timeseries
    ##pdir: direction of the peaks to be identified: '1': Both, '2': Positive, '3': Negative
    ##pthres: Confidence interval which has to be left to be counted as peak
    ##Note: The function does not sort the time series by timestamp but leaves it in the order it is submitted. It can be re-attached to data tables. Invalid or missing cases are left missing.

    verbout('\n\nIdentifying peaks in time series\n',master=master)
    series = {}
    missings = 0
    outlist = []

    if tvar == 0:
        tvar = []
        for i in range(len(svar)):tvar.append(i)
        
    possible = group_variable(svar,'tails',pthres,master=master)

    for i in range(len(tvar)):
        outlist.append('')
        try:
            a = float(tvar[i])
            b = float(svar[i])
            series[a]={}
            series[a]['Value']=b
            series[a]['Position']=i
            series[a]['Peak']=0
            series[a]['Tail']=possible[i]
        except:
            missings = missings + 1

    if missings > 0:
        verbout('\nAttention: '+str(missings)+' missing/string values found.',master=master)

    sk = sorted(series.keys())
    for i in range(2,len(sk)):
        v1 = series[sk[i-2]]['Value']
        v2 = series[sk[i-1]]['Value']
        v3 = series[sk[i]]['Value']
        if series[sk[i-1]]['Tail'] == -1 and v2<=v1 and v2<=v3:
            if pdir in ['1','3']:
                series[sk[i-1]]['Peak'] = -1
        elif series[sk[i-1]]['Tail'] == 1 and v2>=v1 and v2>=v3:
            if pdir in ['1','2']:
                series[sk[i-1]]['Peak'] = 1

    for e in series.keys():
        outlist[series[e]['Position']]=series[e]['Peak']

    return outlist


def mpdetection(seq,pat,minlen=5,maxlen=30,master=''):
    ## Find a pattern in longiudinal data (correctly sorted!)
    ## seq: dictionary with subsequent keys, starting with 0, containing all sequences to be searched 
    ## pat: dictionary with identical keys containing the pattern to look for in each sequence
    ## minlen: minimal length of patterns (if it is lower than the smallest pattern, minimal length is overwritten with shortest pattern length.
    ## maxlen: maximal length of patterns
    
    patlen = []
    for p in pat.keys():
        patlen.append(len(pat[p]))
    if minlen < min(patlen):
        minlen = min(patlen)

    patn = {}
    for i in pat.keys():
        patn[i] = {}
        cpat = pat[i]
        patlen = float(len(cpat))
        for tlen in range(minlen,maxlen+1):
            patn[i][tlen] = []
            for t in range(tlen-1):
                conv = (patlen-1)/(tlen-1)*t
                lower = int(conv)
                rel = conv%1
                wert = (1-rel)*cpat[lower]+rel*cpat[lower+1]
                patn[i][tlen].append(wert)
            patn[i][tlen].append(cpat[-1])

    verbout('\n\n',master=master)
    verbout('Looking for parallel pattern in sequences: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    step = int(len(seq[0])/40)
    if step<1:
        step = 1

    outlist = [] ##Correlations
    altlist = [] ##Altitude algebraic
    alt2list = [] ##Altitude geometric
    lenlist = [] ##Length of optimal match
    
    for i in range(len(seq[0])):

        if i%step == 0:verbout('.','progress',master=master)
        if i > len(seq[0])-minlen-1:
            outlist.append(0)
            altlist.append(0)
            alt2list.append(0)
            lenlist.append(0)
        else:
            if i > len(seq[0])-maxlen-1:
                maxlen = len(seq[0])-i-1

            maxcor = -2
            alt = 0
            for k in range(minlen,maxlen+1):
                addcor = 0
                addalt = 0
                mulalt = 1
                for pnr in pat.keys():
                    s1 = patn[pnr][k]
                    s2 = seq[pnr][i:i+k]
                    cor = calc_correlation(s1,s2)
                    addcor = addcor + cor
                    addalt = addalt + calculate(s2,'range')
                    mulalt = mulalt*calculate(s2,'range')
                if addcor > maxcor:
                    maxcor = addcor
                    alt = addalt
                    alt2 = mulalt
                    length = k

            maxcor = maxcor/len(pat.keys())
            alt = alt/len(pat.keys())
            alt2 = alt2**(1.00/len(pat.keys()))
            
            outlist.append(maxcor)
            altlist.append(alt)
            alt2list.append(alt2)
            lenlist.append(length)
    verbout('\n','progress',master=master)
    
    return [outlist,altlist,alt2list,lenlist]



def create_window(data,gvar,nvar,units=7,pos='1',master=''):
    ##
    ## Create a gliding window
    ## data: data table or dataset
    ## gvar: Variable containing Timestamps
    ## nvar: Variable to be created
    ## units: Number of units comprising one window.
    ## pos: Position of the denominator of the window. 1: First, 2: Middle, 3: Last Element of the window.
    ##

    verbout('\n\nGenerating gliding window.\n',master=master)

    if type(data) in [tuple,list]:
        varlist = data[1]
        data = data[0]
    else:
        varlist = sorted(data.keys())
    
    window = {}
    unit_range = []
    if pos == '1':
        start = 1-units
    elif pos == '2':
        start = -0.5*float(units)+0.5
    elif pos == '3':
        start = 0
    for u in range(units):
        unit_range.append(start)
        start = start + 1

    verbout('\nPreparing windows...',master=master)
        
    for i in range(len(data[gvar])):
        try:
            ts = float(data[gvar][i])
        except:
            ts = ''
        if not ts == '':
            for p in unit_range:
                ts2 = int(ts) + p
                window[ts2]=[]

    verbout('Done.\nDefining windows...',master=master)

    for i in range(len(data[gvar])):
        try:
            ts = float(data[gvar][i])
        except:
            ts = ''
        if not ts == '':
            for p in unit_range:
                ts2 = int(ts) + p
                window[ts2].append(i)

    verbout('Done.\nExtending data...',master=master)

    outdic = {nvar:[]}
    for v in data.keys():
        outdic[v] = []

    for group in sorted(window.keys()):
        for i in window[group]:               
            for v in data.keys():
                outdic[v].append(data[v][i])
            outdic[nvar].append(group)

    varlist.append(nvar)

    return [outdic,varlist]


def detect_gaps(data,tvar,nvar,length,gvar='',sorting=1,master=''):
    dims = data_dim(data) ## Returns [vars,cases]
    
    if gvar in ['','res_nogroup']:
        data[0]['#Group'] = [1]*dims[1]
    else:
        data[0]['#Group'] = data[0][gvar]

    if sorting == 1:
        data = sort_dataset(data,['#Group',tvar],master='silent')
        data[0]['#Sorting'] = list(range(dims[1]))
    else:
        data[0]['#Sorting'] = list(range(dims[1]))
        data = sort_dataset(data,['#Group',tvar],master='silent')

    gi = 1
    data[0][nvar] = [gi]
    data[1].append(nvar)

    for i in range(1,dims[1]):
        try:
            t1 = float(data[0][tvar][i-1])
            t2 = float(data[0][tvar][i])
            dt = t2-t1
        except:
            verbout("\nWarning: Discontinuous time variable. Missing value on line "+str(i),'warning',master=master)
            dt = 0
        if dt > length or not data[0]['#Group'][i-1]==data[0]['#Group'][i]:
            gi +=1

        data[0][nvar].append(gi)

    data = sort_dataset(data,['#Sorting'],master='silent')
    
    return data
    
#detect_gaps([data,varlist],tvar,nvar,length,gvar='',sorting=1,master=self)


def normalize_ts(dset,tvar,gvar,length,addvars=[],method='retain',relts='0',master=''):

    groups = get_unique(dset[0][gvar])
    print(groups)

    if addvars == []:
        addvars = dset[1]
        addvars.remove(tvar)
        addvars.remove(gvar)


    outdata = {}
    outvars = [gvar,tvar]+addvars
    for v in outvars:
        outdata[v] = []

    verbout('\n\nProcessing time series..',master = master)
                
    for g in groups:
        verbout('\nGroup: "'+str(g)+'"',master=master)
        sdata = {}
        for v in [gvar,tvar]+addvars:
            sdata[v] = []
        for i in range(len(dset[0][tvar])):
            if g == dset[0][gvar][i]:  
                for v in [gvar,tvar]+addvars:
                    sdata[v].append(dset[0][v][i])

        correction = 0
        if relts in [1,'1']:
            correction = min(sdata[tvar])

        print(g,dim(sdata))

        point = sdata[tvar][0]
        start = sdata[tvar][0]-0.5*length
        end = sdata[tvar][0]+0.5*length
        idx = 0

        while not point > sdata[tvar][-1]:
            #print([point,start,end])
            context = {}
            context['prev'] = idx
            context['while'] = []
            context['while_ts'] = []
            scani = idx
            while sdata[tvar][scani] < end and scani < (len(sdata[tvar])-1):
                if sdata[tvar][scani] > start:
                    context['while'].append(scani)
                    context['while_ts'].append(sdata[tvar][scani])
                scani += 1
            if scani < len(sdata[tvar]):
                context['next'] = scani
            else:
                context['next'] = scani-1
            idx = context['next']-1

            context['range'] = sdata[tvar][context['next']]-sdata[tvar][context['prev']]
            mindist = context['range']
            context['proximal'] = idx
            for scani in range(context['prev'],context['next']+1):
                if abs(sdata[tvar][scani]-point)<mindist:
                    mindist = abs(sdata[tvar][scani]-point)
                    context['proximal'] = scani

            #print(baum_schreiben(context))

            if method == 'mode':
                #print(context['while'])
                if len(context['while']) == 0:
                    for v in addvars:
                        outdata[v].append(sdata[v][context['prev']])
                elif len(context['while']) == 1:
                    for v in addvars:
                        outdata[v].append(sdata[v][context['while'][0]])
                else:
                    for v in addvars:
                        prox = sdata[v][context['proximal']]
                        dist = {}
                        #print(context['while'])
                        for w in context['while']:
                            try:
                                dist[sdata[v][w]] += 1
                            except:
                                dist[sdata[v][w]] = 1
                        #print(dist)
                        ranking = []
                        for e in dist.keys():
                            ranking.append((dist[e],e))

                        ranking = sorted(ranking,reverse=True)
                        #print(ranking)

                        if len(ranking) == 1:
                            take_value = ranking[0][1]
                        else:
                            if ranking[0][0] > ranking[1][0]:
                                take_value = ranking[0][1]
                            else:
                                take_value = prox ## Unsaubere Lösung: Wenn Prox nur einmal und andere Werte häufig vorkommen. Aber im Moment egal.

                        #print(ranking,take_value)

                        outdata[v].append(take_value)
            elif method == 'proxy':
                for v in addvars:
                    outdata[v].append(sdata[v][context['proximal']])

            elif method == 'recent':
                if len(context['while']) == 0:
                    for v in addvars:
                        outdata[v].append(sdata[v][context['prev']])
                else:
                    for v in addvars:
                        take_value = sdata[v][context['prev']]
                        for w in context['while']:
                            if sdata[tvar][w] < point:
                                take_value = sdata[v][w]
                        outdata[v].append(take_value)
                        

            elif method == 'interpol':
                if sdata[tvar][context['proximal']] > point:
                    upper = context['proximal']
                    lower = upper-1
                else:
                    lower = context['proximal']
                    upper = lower + 1

                if lower == -1:
                    upper+=1
                    lower+=1
                elif upper == len(sdata[tvar]):
                    upper-=1
                    lower-=1 

                for v in addvars:
                    upperval = ''
                    lowerval = ''
                    reach = 0
                    while upperval == '' and (upper+reach) < len(sdata[tvar]):
                        try:
                            uppertime = sdata[tvar][upper+reach] ## Set time first in case an error occurs here
                            upperval = float(sdata[v][upper+reach])
                        except:
                            reach+=1

                    reach = 0
                    while lowerval == '' and (lower-reach) > -1:
                        try:
                            lowertime = sdata[tvar][lower-reach] ## Set time first in case an error occurs here
                            lowerval = float(sdata[v][lower-reach])
                        except:
                            reach+=1

                    if type(upperval) == float and type(lowerval) == float:
                        #print(upperval,lowerval,uppertime,lowertime)
                        b = (upperval-lowerval)/(uppertime-lowertime)

                        value = lowerval + (point-lowertime)*b

                        #print(value,point,'\n')                        
                        
                    else:
                        value = ''

                    outdata[v].append(value)

                            

            outdata[tvar].append(point-correction)
            outdata[gvar].append(g)
                        

            point+=length
            start+=length
            end+=length

    verbout('\n',master=master)

    return [outdata,outvars]

        
    
#Call: normalize_ts([data,varlist],tvar,gvar,length,addvars,method,master=self)        




def focus_timeseries(data,issvar,actvar,windir,winlen,master=''):
    verbout('\nSetting up data',master=master)
    ddic = {}
    for d in data['res_Day']:
        if not d == '':
            ddic[d] = {}


    mindat = sorted(ddic.keys())[0]
    maxdat = sorted(ddic.keys())[-1]

    for d in range(mindat,maxdat+1):
        ddic[d] = {'Date':tts(d,'ex','ger'),
                   'Volume':[],
                   'Weight':[],
                   'Issue':[],
                   'Actor':[]}

    mindatn = tts(mindat,'ex','ger')
    maxdatn = tts(maxdat,'ex','ger')
    verbout('\nCreated data for '+str(len(ddic.keys()))+' dates, ranging from '+mindatn+' to '+maxdatn+'.',master=master)


    for i in range(len(data['res_Textcount'])):
        d = data['res_Day'][i]
        if not d == '':  
            ddic[d]['Volume'].append(data['res_Textcount'][i])
            ddic[d]['Weight'].append(data['res_Weighting'][i])
            for v in issvar:
                try:
                    n = int(data[v][i])
                except:
                    n = 0
                for k in range(n):
                    ddic[d]['Issue'].append(v)
            for v in actvar:
                try:
                    n = int(data[v][i])
                except:
                    n = 0
                for k in range(n):
                    ddic[d]['Actor'].append(v)

    for d in ddic.keys():
        ddic[d]['Volume'] = calculate(list(zip(ddic[d]['Volume'],ddic[d]['Weight'])),'sum')
   
    for d in sorted(ddic.keys()):
        window = []
        if windir == 'retro':
            for wl in range(int(winlen)):
                window.append(d+wl)
        elif windir == 'prosp':
            for wl in range(int(winlen)):
                window.append(d-wl)

        issues = []
        actors = []
        volumes = []
        for d2 in window:
            try:
                issues = issues + ddic[d2]['Issue']
                actors = actors + ddic[d2]['Actor']
                volumes.append(ddic[d2]['Volume'])
            except:
                verb('No date: '+str(d2))

        ddic[d]['Iss_Focus'] = 1-calc_entropy(issues,issvar)
        ddic[d]['Act_Focus'] = 1-calc_entropy(actors,actvar)
        ddic[d]['Win_Volume'] = calculate(volumes,'mean')

    verbout('\nCalculated entropies and volume. Writing to table',master=master)


    outdata = {}
    outvars = ['Day_Int','Date','Volume','Issue_Focus','Actor_Focus']
    for v in outvars:
        outdata[v] = []
    for d in sorted(ddic.keys()):
        outdata['Day_Int'].append(d)
        outdata['Date'].append(ddic[d]['Date'])
        outdata['Volume'].append(ddic[d]['Win_Volume'])
        outdata['Issue_Focus'].append(ddic[d]['Iss_Focus'])
        outdata['Actor_Focus'].append(ddic[d]['Act_Focus'])
                

    return (outdata,outvars)



def event_agreement(l1,l2,lag,meas):

    minn = 0
    maxn = min(len(l1),len(l2))

    if lag<0:
        minn = minn - lag
    elif lag>0:
        maxn = maxn - lag

    crosstab = {'a':0,'b':0,'c':0,'d':0}
    comp = 0
    errors = 0
    
    for i in range(minn,maxn):
        e1 = l1[i]
        e2 = l2[i+lag]
        
        if e1 == 1 and e2 == 1:
            crosstab['a'] +=1
            comp+=1
        elif e1 == 1 and e2 == 0:
            crosstab['b'] +=1
            comp+=1
        elif e1 == 0 and e2 == 1:
            crosstab['c'] +=1
            comp+=1
        elif e1 == 0 and e2 == 0:
            crosstab['d'] +=1
            comp+=1
        else:
            errors +=1

    if comp > 1:
        if meas == 'percent':
            agreement = float(crosstab['a']+crosstab['b'])/comp
        elif meas == 'sokal':
            agreement = float(crosstab['a']*2)/(2*crosstab['a']+crosstab['b']+crosstab['c'])
        elif meas == 'cohen':
            p1 = float(2*crosstab['a']+crosstab['b']+crosstab['c'])/comp
            p0 = 1-p1
            pc = p1**2+p0**2
            a = float(crosstab['a']+crosstab['b'])/comp
            verb(pc,a)
            if pc < 1:
                agreement = (a-pc)/(1-pc)
            else:
                agreement = 1.0
            verb(agreement)

    return agreement

def event_agree_overall(ad):
    agreelist = []
    vlist = list(ad.keys())
    for i in range(len(vlist)-1):
        for k in range(i,len(vlist)):
            agreelist.append(ad[vlist[i]][vlist[k]][0])
    stats = stat_desc(agreelist)

    return [stats['M'],stats['SD']]



def synch_move(ad,v,l):
    verb('Moving vector '+v+' by '+str(l)+' Frames')
    vectors = sorted(list(ad.keys()))

    for v1 in vectors:
        for v2 in vectors:
            shift = 0
            if v1 == v: shift = shift + l
            if v2 == v: shift = shift - l

            if not shift == 0:
                print('shifting',v1,v2,shift)
                values = dict(ad[v1][v2])
                for lag in ad[v1][v2].keys():
                    try:
                        ad[v1][v2][lag] = values[lag+shift]

                    except:
                        ad[v1][v2][lag] = ''
    return ad



def synch_wiggle(ad,master=''):
    vectors = sorted(list(ad.keys()))

    wiggled = 1
    iteration = 0

    while wiggled == 1:
        iteration +=1
        verbout('\nWiggling iteration: '+str(iteration),master=master)
        wiggled = 0

        for v in vectors:
            ma = 0.0
            for v2 in vectors:
                try:
                    ma = ma + ad[v][v2][0]
                except:
                    pass

            ma = ma/(len(vectors)-1)
            #print('Initial agreement for '+v+': '+str(ma))

            pf = sorted(list(ad[v][v].keys()))
            ranking = []
            for l in pf:
                ma = 0.0
                invalid = 0
                for v2 in vectors:
                    if not v==v2:
                        try:
                            ma += ad[v][v2][l]
                        except:
                            invalid = 1
                if invalid == 0:
                    ranking.append((ma,l))

            #print(ranking)
            #print(max(ranking))

            if not max(ranking)[1] == 0:
                ad = synch_move(ad,v,max(ranking)[1])
                wiggled = 1
                ad[v]['res_wiggled'] += max(ranking)[1]

    return ad



def synch_events(data,varlist,frame,meas,master=''):
    #event_corr(storage['Dummy_Data'],cvars,frame,meas,master=self)
    twoframe = 2*frame

    agree_dic = {}
    for v in varlist:
        agree_dic[v] = {'res_wiggled':0}    
        for v2 in varlist:
            agree_dic[v][v2] = {}
            for l in range(-twoframe,twoframe+1):
                agree_dic[v][v2][l] = ''
                
    pairs = []

    for i in range(len(varlist)-1):
        for k in range(i+1,len(varlist)):
            pairs.append((varlist[i],varlist[k]))

    for p in pairs:
        maxa = -100.0
        optlag = 0
        for lag in range(-frame,frame+1):
            a = event_agreement(data[p[0]],data[p[1]],lag,meas)
            if a > maxa:
                maxa = a
                optlag = lag

            agree_dic[p[0]][p[1]][lag] = a
            agree_dic[p[1]][p[0]][-lag] = a

    #print(baum_schreiben(agree_dic))
    overall = event_agree_overall(agree_dic)
    verbout('\nOverall Agreement: '+str(round(overall[0],3))+' (SD='+str(round(overall[1],3))+')',master=master)

    agree_dic = synch_wiggle(agree_dic,master=master)

    #print(baum_schreiben(agree_dic))
    verbout('\nWiggling Complete.',master=master)
    overall = event_agree_overall(agree_dic)
    verbout('\n\nOverall Agreement: '+str(round(overall[0],3))+' (SD='+str(round(overall[1],3))+')',master=master)

    verbout('\n\nReport of frame shifts to get to optimal solution:','table',master=master)

    minshift = 0
    maxshift = 0
    for v in varlist:
        verbout('\n - '+v+' shifted by: '+str(agree_dic[v]['res_wiggled']),'table',master=master)
        if agree_dic[v]['res_wiggled'] < minshift: minshift = agree_dic[v]['res_wiggled']
        if agree_dic[v]['res_wiggled'] > maxshift: maxshift = agree_dic[v]['res_wiggled']
    verbout('\n',master=master)

    outdic = {}
    for v in varlist:
        outdic[v] = []

    ldata = data_dim(data)[1]
    #print(ldata)

    for i in range(-maxshift,ldata-minshift):
        for v in varlist:
            icorr = i - agree_dic[v]['res_wiggled']
            if icorr >= 0 and icorr < ldata:
                outdic[v].append(data[v][icorr])
            else:
                outdic[v].append('')

    return [outdic,varlist]

        
##########################
##
## Merge Tables
##
##########################


def create_keydic(data,keyvars,varlist=[],master=''):
    ##Creates a key dictionary from data and a list of key variables.
    ##Function required to shape the keydic used for merge_data()
    
    verbout('\nKey Variables in Key Table:\n',master=master)
    if type(keyvars)==str:keyvars = [keywars] ##Catching single variables
    if len(keyvars) > 0:
        for v in keyvars:
            verbout(v+'\n',master=master)

    if varlist == []: varlist = sorted(data.keys())
    for v in keyvars:
        while v in varlist:
            varlist.remove(v)

    #d = storage['Schl']
    schldic = {}
    dupdic = {}
    duplicates = []
    
    step = int(len(data[keyvars[0]])/40)
    if step<1: step = 1
    verbout('Getting informations from key table: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    for i in range(len(data[keyvars[0]])):
        if i % step == 0:
            verbout('.','progress',master=master)
        key = []
        for v in keyvars:
            key.append(data[v][i])
        keylab = str(key)

        schldic[keylab] = {}
        dupdic[keylab]=0
        for v in varlist:
            schldic[keylab][v] = data[v][i]

    verbout('\n','progress',master=master)
    verbout('\n\n',master=master)

    verbout('Checking for duplicates: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    for i in range(len(data[keyvars[0]])):
        if i % step == 0:
            verbout('.','progress',master=master)
        key = []
        for v in keyvars:
            key.append(data[v][i])
        keylab = str(key)
        dupdic[keylab]=dupdic[keylab]+1
        if dupdic[keylab] > 1:
            duplicates.append(keylab)
    verbout('\n','progress',master=master)
    verbout('\n',master=master)
    
    
    if len(duplicates) > 0:
        verbout('\n',master=master)
        verbout('Warning!!\nThere were duplicates in the keytable:\n','warning',master=master)
        for d in duplicates[:20]:
            verbout(str(d)+'\n','warning',master=master)
        if len(duplicates)>20:
            verbout('...List trunctuated. More than 20 duplicates (N=)\n','warning',master=master)
    verbout('\n',master=master)

    return schldic


def merge_data(data,keyvars,schldic,vadd=[],vretain=[],master=''):
    ##Merges the contents of a keydic to a dataset
    ##data: Data dictionary of main table
    ##keyvars: Key variables in main table
    ##schldic: key dictionary as produced by create_keydic()
    ##vadd: List of variables to be added from schldic
    ##vretain: List of variables to be retained from main table
    ##master: Master TK object to verbout in.

    if vadd == []:
        vadd = schldic[list(schldic.keys())[0]].keys()

    if vretain == []:
        vretain = sorted(data.keys())


    tmp = {}
    for v in vadd:
        tmp[v] = []

    verbout('\n',master=master)
    success = 0
    fail = 0

    step = int(len(data[keyvars[0]])/40)
    if step<1: step = 1
    verbout('Linking cases: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    for i in range(len(data[keyvars[0]])):
        key = []
        for v in keyvars:
            key.append(data[v][i])
        keylab = str(key)
        for v in vadd:
            try:
                tmp[v].append(schldic[keylab][v])
                success = success + 1
            except:
                tmp[v].append('')
                fail = fail + 1
        if i%step == 0:
            verbout('.','progress',master=master)
    verbout('\n','progress',master=master)

    success = int(success/len(vadd))
    fail = int(fail/len(vadd))

    verbout('\nDatasets linked.\n'+str(success)+' successfully merged cases\n'+str(fail)+' failed to find corresponding cases in keytable',master=master)
    verbout('\n\nCreating new variables in dataset:\n',master=master)

    vout = vretain

    for v in vadd:
        vlab = v
        idnr = 0
        while vlab in data.keys():
            idnr = idnr + 1
            vlab = v + str(idnr)
        data[vlab] = tmp[v]
        vout.append(vlab)
        verbout(vlab+'\n',master=master)

    return [data,vout]

def merge_files(filelist,varlist=[],master=''):
    if varlist == []:
        varlist = get_varnames(filelist[0])
        for fname in filelist[1:]:
            nv = get_varnames(fname)
            for v in nv:
                if not v in varlist: varlist.append(v)
    verbout('\nList of variables in all files: \n',master=master)
    verbout(str(varlist),'table',master=master)
    verbout('\n\nAdding files:',master=master)
    
    out_data = {}
    for v in varlist:
        out_data[v] = []
    for fname in filelist:
        dta = get_data(fname)[0]
        dl = data_dim(dta)[1]
        vl = dta.keys()
        verbout('\n- '+fname+': ',master=master)
        verbout(str(len(vl))+' Variables and '+str(dl)+' Cases',master=master)
        for i in range(dl):
            for v in varlist:
                if v in vl:
                    out_data[v].append(dta[v][i])
                else:
                    out_data[v].append('')

    return [out_data,varlist]


def merge_elong(d1,d2,v1,v2,varlist,master=''):
    verbout('Merging the datasets.\n\n',master=master)         

    outdata = {}
    outvar = []
    
    for v in v1:
        if v in varlist:
            outvar.append(v)
    for v in v2:
        if v in varlist and not v in outvar:
            outvar.append(v)

    verbout('\nVariables in final dataset:\n  '+str(outvar),'table',master=master)

    for v in outvar:
        outdata[v] = []

    for i in range(len(d1[v1[0]])):
        for v in outvar:
            if v in v1:
                outdata[v].append(d1[v][i])
            else:
                outdata[v].append('')
    for i in range(len(d2[v2[0]])):
        for v in outvar:
            if v in v2:
                outdata[v].append(d2[v][i])
            else:
                outdata[v].append('')

    return [outdata,outvar]



##########################
##
## Co-Occurrence Analysis
##
##########################

def co_occurrence(data,dmode='dummy',method='anz',margin=1,prefix='',cases='',keyvar='',min_case=0,min_anz=0,dummylist=[],master=''):
    #Do Co-Occurrence Analysis of data. The data may either be a group and a nominal variable or a dummy table.
    #data: Dataset, containing either dummies or a group and a nominal variable
    #dmode: specifier whether it is a dummy ('dummy') or a nominal ('nominal') dataset
    #method: Method to calculate the co-occurrence.
    #prefix: Prefix for dummy variables (only for 'nominal')
    #cases: Name of the case variable (only for 'nominal')
    #keyvar: Name of the group variable (only for 'nominal')
    #min_case / min_anz: Integer specifying the mininmal count of cases/groups to be counted in the dummy table
    #dummylist: List of dummy variables in the data (only for 'dummy')
    #master: Class of the open Nogrod window

    if dmode=='nominal':
        d_mat = dummy(data[cases],data[keyvar],'dicho',min_case,min_anz,master=master)
    elif dmode=='dummy':
        if dummylist==[]:
            d = check_dummytable([data,sorted(data.keys())])
            data = d[0]
            dummylist = d[1]    
        d_mat = {}
        d_mat['#Group']=[]
        for dvar in dummylist:
            d_mat[dvar]=[]
        for i in range(len(data[dummylist[0]])):
            d_mat['#Group'].append(i)
            for dvar in dummylist:
                d_mat[dvar].append(data[dvar][i])

    verbout('\nCalculating Co-Occurrence scores within matrix',master=master)

    out_mat = {}
    out_mat[''] = []
    varlist = []
    for v in sorted(d_mat.keys()):
        if not v == '#Group':
            out_mat[prefix+v] = []
            varlist.append(v)

    for v1 in varlist:
        out_mat[''].append(prefix+v1)
        for v2 in varlist:
            ntup = crosstab(d_mat[v1],d_mat[v2],method)
            out_mat[prefix+v2].append(ntup)

    if margin in ['1',1]:
        fullcol = []
        for i in d_mat[v1]:
            fullcol.append(1)
        for v2 in varlist:
            ntup = crosstab(d_mat[v2],fullcol,method)
            out_mat[prefix+v2].append(ntup)
        out_mat[''].append('Total')

    outvarlist = ['']
    for v in varlist:
        outvarlist.append(prefix+v)

    return [out_mat,outvarlist]







##########################
##
## Visone
##
##########################


def create_visone(subj,obj,rel=[],method='all',min_anz=0,obj_is_subj=1,master=''):
    ##Creates an adjacency matrix and link and node files as input for Visone
    ##subj: Vector of subjects
    ##obj: Vector of objects
    ##rel: Vector of relations (may be empty list, if no relation is necessary)
    ##method: Method for calculating links: 'anz', 'dicho','entf','mean','sum','typ','all'
    ##min_anz: Minimum number of relations to be counted as a link
    ##The output is three datasets in a dictionary: Adjacency matrix (data and variables), Nodes table (data and variables), Link table (data and variables, if method='all'
    
    log('Calling Function: Create Visone')
    verbout('\n\nCreating adjacency matrix\n',master=master)
    global storage

    s_names = []
    o_names = []
    storage['Attributes'] = {'id':[],'anz':[],'typ':[]}

    if obj_is_subj == 1:
        n_dic = {}
        for n in get_unique(subj + obj): n_dic[n] = 0
        for n in subj + obj: n_dic[n] = n_dic[n] + 1
        if '' in n_dic.keys(): del n_dic['']
        if ' ' in n_dic.keys(): del n_dic[' ']
        for n in sorted(n_dic.keys()):
            if n_dic[n] >= min_anz:
                s_names.append(n)
                o_names.append(n)
                storage['Attributes']['id'].append(n)
                storage['Attributes']['anz'].append(n_dic[n])
                storage['Attributes']['typ'].append('Case')
    else:
        s_dic = {}
        o_dic = {}
        for n in get_unique(subj): s_dic[n] = 0
        for n in get_unique(obj): o_dic[n] = 0
        for n in subj: s_dic[n] = s_dic[n] + 1
        for n in obj: o_dic[n] = o_dic[n] + 1
        if '' in s_dic.keys(): del s_dic['']
        if ' ' in s_dic.keys(): del s_dic[' ']
        if '' in o_dic.keys(): del o_dic['']
        if ' ' in o_dic.keys(): del o_dic[' ']
        for n in sorted(s_dic.keys()):
            if s_dic[n] >= min_anz:
                s_names.append(n)
                storage['Attributes']['id'].append(n)
                storage['Attributes']['anz'].append(s_dic[n])
                storage['Attributes']['typ'].append('Subj')
        for n in sorted(o_dic.keys()):
            if o_dic[n] >= min_anz:
                o_names.append(n)
                storage['Attributes']['id'].append(n)
                storage['Attributes']['anz'].append(o_dic[n])
                storage['Attributes']['typ'].append('Obj')      

    anzdic = {'_Case':[]}
    valdic = {'_Case':[]}
    verbout('\n'+str(len(s_names))+' possible subjects, '+str(len(o_names))+ ' possible objects ',master=master)
    verb(str(o_names))
    verb(str(s_names))
    v_list = []
    coord = {}
    i = 0
    for n in sorted(o_names):
        anzdic[n] = []
        valdic[n] = []
        v_list.append(n)
    for n in sorted(s_names):
        for k in v_list:
            anzdic[k].append(0)
            valdic[k].append(0)        
        anzdic['_Case'].append(n)
        valdic['_Case'].append(n)
        coord[n] = i
        i = i + 1
    verb(baum_schreiben(anzdic))
    verb(str(coord))

    verbout('and '+str(len(rel))+'relations.\n\n',master=master)
    if rel == []:
        for e in subj: rel.append(1)
        
    numeric = 1
    for e in rel:
        try:
            e = float(e)
        except:
            if not e in ['',' ']:
                numeric = 0

    step = int(len(subj)/40)
    panz = 1
    if step<1:
        panz = int(40.0/len(subj))
        step = 1
    verbout('Analyzing relations: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    i = 0
    
    for i in range(len(subj)):
        if i % step == 0:verbout('.'*panz,'progress',master=master)
        if subj[i] in s_names and obj[i] in o_names:
            if method in ['anz','dicho','entf']:
                anzdic[obj[i]][coord[subj[i]]] = anzdic[obj[i]][coord[subj[i]]] + 1
            elif method in ['mean','sum'] or (method == 'all' and numeric ==1):
                try:
                    valdic[obj[i]][coord[subj[i]]] = valdic[obj[i]][coord[subj[i]]] + float(rel[i])
                    anzdic[obj[i]][coord[subj[i]]] = anzdic[obj[i]][coord[subj[i]]] + 1
                except:
                    verb('ERROR: No float number '+str(rel[i]))
            elif method in ['typ','all']:
                valdic[obj[i]][coord[subj[i]]] = rel[i]
                anzdic[obj[i]][coord[subj[i]]] = anzdic[obj[i]][coord[subj[i]]] + 1
    
    verbout('\n','progress',master=master)
    verbout('\n\nAssignining scores.',master=master)

    linknr = 0
    linkdic = {'csv value':[],'Count':[],'Dicho':[],'Sum':[],'Mean':[]}

    outdic = {}
    outdic['_Case']=anzdic['_Case']
    if method == 'anz':
        outdic = anzdic
    elif method == 'dicho':
        for k in v_list:
            outdic[k] = []
            for i in range(len(anzdic[k])):
                if anzdic[k][i] > 0:
                    outdic[k].append(1)
                else:
                    outdic[k].append(0)
    elif method == 'entf':
        for k in v_list:
            outdic[k] = []
            for i in range(len(anzdic[k])):
                outdic[k].append(int(9.0/(anzdic[k][i]+1)))
    elif method == 'mean':
        for k in v_list:
            outdic[k] = []
            for i in range(len(anzdic[k])):
                if anzdic[k][i] > 0:
                    outdic[k].append(int(100.0*valdic[k][i]/anzdic[k][i]))
                else:
                    outdic[k].append(0)
    elif method in ['sum','typ']:
        outdic = valdic
        
    elif method == 'all':
        for k in v_list:
            outdic[k] = []
            for i in range(len(anzdic[k])):
                linknr = linknr + 1
                linklab = "L_{0:05}".format(linknr)
                count = anzdic[k][i]
                if count > 0:
                    dicho = 1
                else:
                    dicho = 0
                summe = valdic[k][i]

                if type(summe) in [int,float] and count>0:
                    mittel = int(summe*100)/count
                else:
                    mittel = ''

                if count > 0:                                   
                    linkdic['csv value'].append(linklab)
                    linkdic['Count'].append(count)
                    linkdic['Dicho'].append(dicho)
                    linkdic['Sum'].append(summe)
                    linkdic['Mean'].append(mittel)
                    outdic[k].append(linklab)
                else:
                    outdic[k].append(0)
                
    verbout('\n\nAdjacency matrix created.\n',master=master)

    nanz = {}
    for node in (subj+obj):
        nanz[node] = 0
    for node in (subj+obj):
        nanz[node] = nanz[node]+1

    nodedic = {'id':[],'Count':[]}
    for node in sorted(nanz.keys()):
        nodedic['id'].append(node)
        nodedic['Count'].append(nanz[node])

    outdata = {}
    outdata['Adjacency'] = [outdic,['_Case']+v_list]
    outdata['Nodes'] = [nodedic,['id','Count']]
    outdata['Links'] = [linkdic,['csv value','Count','Dicho','Sum','Mean']]

    return outdata



##########################
##
## Cluster Analysis
##
##########################


def write_ssa(data,filename,rownames='',bildname=""):
    log('Calling Function: Write SSA')
    matexp = open(filename,'w')

    matexp.write('library(smacof)\n\n')
    matexp.write('diffm <- matrix(c(\n')
    zeile = 0
    liste = list(data.keys())
    liste.remove('')
    laenge = len(liste)
    
    titel_bez = 'SSA-Visualization of Distance Matrix'

    titel = 'd1 <- list("'
    for i in range(len(data[rownames])):
        element = data[rownames][i]
        zeile = zeile + 1
        titel = titel+element
        if zeile < laenge:
            titel = titel + '","'
        else:
            titel = titel + '")'

        for element2 in data[rownames]:
            matexp.write(str(data[element2][i]))
            if ((zeile == laenge)&(element==element2)):
                matexp.write(')')
            matexp.write(',')
        if zeile < laenge:
            matexp.write('\n')

    matexp.write('nrow=')
    matexp.write(str(laenge))
    matexp.write(',ncol=')
    matexp.write(str(laenge))
    matexp.write(')\n\n')
    matexp.write(titel)
    matexp.write('\nrownames(diffm) <- d1\n')
    matexp.write('colnames(diffm) <- d1\n')
    matexp.write('darst = smacofSym(diffm, ndim=2)\n')
    matexp.write('plot(darst, main="')
    matexp.write(titel_bez)
    matexp.write('")\n\n')

    if len(bildname) > 1:       
        matexp.write('png(filename="')
        matexp.write(bildname)
        matexp.write('", height=1100, width=1600, pointsize=22, bg="white")\n')
        matexp.write('plot(darst, main="')
        matexp.write(titel_bez)
        matexp.write('", xlab="", ylab="")\n')
        matexp.write('dev.off()\n')
    matexp.close()


def find_cluster(dset,varlist,outfile,add_outputs={'ssa':0,'dendro':0,'hist':0,'dist':0,'vector':0},row_std=0,table_std=0,master=''):
    verbout('\n\nClustering '+str(len(varlist))+' Elements.\n',master=master)
    if type(dset) == dict:
        data = dset
        dvar = sorted(data.keys())
    else:
        data = dset[0]
        dvar = dset[1]

    out_dendro = outfile[:-4]+'_Dendrogram.R'
    out_ssa = outfile[:-4]+'_SSA.R'
    out_log = outfile[:-4]+'_Logfile.dat'
    out_hist = outfile[:-4]+'_History.dat'
    out_loading = outfile[:-4]+'_Loadings.dat'
    var_detail = lcopy(varlist)
    out_vector = outfile[:-4]+'_Vectors.dat'

    progress = {}
    cluster_dict = {}
    maxdist_dict = {}
    anz_element = {}
    solutions = []
    listwise = {}

    clus_dic = transform_float(data,varlist)
    for v in varlist:
        listwise[v] = v
        cluster_dict[v] = 1
        anz_element[v] = sum(clus_dic[v])
    if table_std == 1:
        verbout('\nStandardizing Table...',master=master)
        clus_dic = standardize_table(clus_dic)
        verbout('Done.\n',master=master)
    if row_std == 1:
        verbout('\nStandardizing Rows...',master=master)
        clus_dic = standardize_rows(clus_dic)
        verbout('Done.\n',master=master)

    verbout('\nNormalizing vectors...',master=master)
    for v in clus_dic.keys():
        clus_dic[v] = univec(clus_dic[v])
    verbout('Done.\n\n',master=master)
              
    p_i = 1
    cluster_nr = 1

    cl_file = open(out_log,'w')

    dist_dic = {}

    anz = 0
    anz_todo = len(varlist)**2
    step = int(anz_todo/40)
    if step<1: step = 1
    verbout('Calculating Distances: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    
    for i in varlist:
        dist_dic[i] = {}
        for k in varlist:
            dist_dic[i][k] = distance(clus_dic[i],clus_dic[k])
            anz = anz + 1
            if  anz%step == 0:verbout('.','progress',master=master)
    verbout('\n','progress',master=master)

    verbout('\nClustering.\n',master=master)
    while len(varlist) > 1:
        verbout('.','progress',master=master)
        mindist = 10000
        for i in varlist:
            for k in varlist:
                if not i == k:
                    if dist_dic[i][k] < mindist:
                        k1 = i
                        k2 = k
                        mindist = dist_dic[i][k]
        cl_file.write(str(mindist)+'\t'+k1+'\t'+k2+'\n')
        progress[p_i] = {}
        progress[p_i]['K1'] = k1
        progress[p_i]['K2'] = k2
        progress[p_i]['Dist'] = mindist

        varlist.remove(k1)
        varlist.remove(k2)

        cluster = [listwise[k1],listwise[k2]]          
        clus_dic[str(cluster)] = gravity(clus_dic,cluster)
        listwise[str(cluster)] = cluster
        varlist.append(str(cluster))
        dvar.append(str(cluster))
        data[str(cluster)] = clus_dic[str(cluster)]

        dist_dic[str(cluster)] = {}
        for v in varlist:
            dist_dic[v][str(cluster)] = distance(clus_dic[v],clus_dic[str(cluster)])
            dist_dic[str(cluster)][v] = dist_dic[v][str(cluster)]
        
        progress[p_i]['Result'] = str(cluster)
        progress[p_i]['Maxdist'] = maxdist(clus_dic,cluster)
        progress[p_i]['Components'] = flatten(cluster)
        if progress[p_i]['Maxdist'] > 0:
            progress[p_i]['Quota'] = progress[p_i]['Dist']/progress[p_i]['Maxdist']
        else:
            progress[p_i]['Quota'] = 0
        cluster_dict[str(cluster)] = progress[p_i]['Quota']
        maxdist_dict[str(cluster)] = progress[p_i]['Maxdist']

        if cluster_dict[str(cluster)] < 1:
            if cluster_dict[k1] > 1 and len(flatten(listwise[k1])) > 3:
                verbout(cluster_nr,master=master)
                cluster_nr = cluster_nr + 1
                solutions.append(k1)
                if cluster_dict[k2] < 1:
                    sister = 0
                    for s in solutions:
                        if s in k2:
                            sister = 1
                    if sister == 1:
                        solutions.append(k2)
            if cluster_dict[k2] > 1 and len(flatten(listwise[k2])) > 3:
                verbout(cluster_nr,master=master)
                cluster_nr = cluster_nr + 1
                solutions.append(k2)
                if cluster_dict[k1] < 1:
                    sister = 0
                    for s in solutions:
                        if s in k1:
                            sister = 1
                    if sister == 1:
                        solutions.append(k1)
           
                
        p_i = p_i + 1

    cl_file.write(str(flatten(varlist[0]))+'\n')
    cl_file.close()
    verbout('\n','progress',master=master)
    
    verbout('\n\nDetailed Log-File completed: '+out_log,master=master)

    full_dendro = printDendrogram(varlist[0])

    real_clusters = []
    #real_cluster_elements = []
    real_nr = 1
    verbout('\n\nExtracting real clusters (no redundancy)\n',master=master)
    for e in solutions:
        invalid = 0
        for f in solutions:
            if f in e and len(f) < len(e):
                invalid = 1
        if invalid == 0:
            verbout('\nCluster'+str(real_nr)+': '+str(e)+'\n'+printDendrogram(e)+'\n\n',master=master)
            real_clusters.append(e)
            #real_cluster_elements = real_cluster_elements + flatten(e)
            real_nr = real_nr + 1

    try:
        master.display_dendro(full_dendro,real_clusters)
    except:
        verbout(full_dendro+'\n\n',master=master)

    super_clusters = []
    for e in solutions:
        if not e in real_clusters:
            container = []
            for i in range(0,len(real_clusters)):
                if real_clusters[i] in e and len(real_clusters[i]) < len(e):
                    container.append(i+1)
            if len(container) > 1:
                super_clusters.append(str(container)+ ' (Radius: '+str(maxdist_dict[str(e)])+')')
                verbout('\nSupercluster subsuming: '+str(container) + ' (Radius: '+str(maxdist_dict[str(e)])+')',master=master)


    verbout('\n\nOutput of plain text report:\n',master=master)
    ssa_elemente = []
    try:
        tmode = 'Performed cluster analysis of count data\n\nInput-File:'+storage['Cluster_Input']
    except:
        tmode = 'Performed cluster analysis of count data\n\n'
    tmode = tmode + '\nParameters:\nRow Standardization: '+str(row_std)
    tmode = tmode + '\nTable Standardization: '+str(table_std)

    if add_outputs['ssa'] == 1: tmode = tmode+'\nOutput for SSA-Script: '+out_ssa
    if add_outputs['dendro'] == 1: tmode = tmode+'\nOutput for Dendrogram-Script: '+out_dendro
    if add_outputs['hist'] == 1: tmode = tmode+'\nOutput for History: '+out_hist
    if add_outputs['dist'] == 1: tmode = tmode+'\nOutput for Proximity Scores: '+out_loading
    if add_outputs['vector'] == 1: tmode = tmode+'\nOutput for Vectors: '+out_vector
    tmode = tmode + '\n\n-------------------------------\n'
    
    report_elements = []
    
    txt_file = open(outfile,'w')
    txt_file.write(tmode)
    txt_file.write('Number of Clusters: '+str(len(real_clusters))+'\n\n')
    for i in range(0,len(real_clusters)):
        md = maxdist_dict[real_clusters[i]]
        txt_file.write('Cluster #'+str(i+1)+' (radius='+str(md)+')'+'\n-------------\n')
        elist = flatten(listwise[real_clusters[i]])
        report_elements.append(elist)
        for e in elist:
            ssa_elemente.append(e)
            txt_file.write(e)
            txt_file.write(' (N=' + str(anz_element[e]) +')\n')
        txt_file.write('\n\n')

    tmp_report = open('tmp_report_cluster.txt','a')
    try:
        tmp_report.write(storage['Cluster_Input']+'\t'+ str(len(real_clusters))+'\t'+str(report_elements)+'\n')
    except:
        pass
    tmp_report.close()
    

    txt_file.write('\nSuperclusters subsuming clusters:\n----------------------------------\n')
    for i in range(0,len(super_clusters)):
        txt_file.write('Supercluster #'+str(i+1)+' subsumes: '+super_clusters[i]+'\n')

    txt_file.write('\n\n')

    txt_file.write('\nCorrelations:\n-------------\n\n\t')
    for ci1 in range(0,len(real_clusters)):
        tn1 = 'Cluster' + str(ci1+1)
        txt_file.write(tn1+'\t')

    for ci1 in range(0,len(real_clusters)):
        c1 = real_clusters[ci1]
        tn1 = 'Cluster' + str(ci1+1)
        txt_file.write('\n'+tn1+'\t')
        for ci2 in range(0,len(real_clusters)):
            c2 = real_clusters[ci2]
            tn2 = 'Cluster' + str(ci2+1)
            dist = distance(univec(clus_dic[c1]),univec(clus_dic[c2]))
            load = math.cos(2*math.asin(dist/2))
            txt_file.write("{:1.3f}".format(load)+'\t')
    txt_file.write('\n\n')
    txt_file.close()
    verbout('\nFile "'+outfile+'" created.',master=master)

    if add_outputs['ssa'] == 1:
        verbout('\n\nCreating SSA-Script (R-Script)',master=master)
        ssa_elemente = ssa_elemente + real_clusters
        distmat = distmatrix(clus_dic,ssa_elemente)
        write_ssa_cluster(distmat,ssa_elemente,out_ssa,titel_bez='Proximity of Elements')
        verbout('\nFile "'+out_ssa+'" created.',master=master)

    if add_outputs['dendro'] == 1:
        verbout('\n\nOutput of Dendrogram (R-Script)...',master=master)
        write_dendrogram_R(progress,flatten(varlist[0]),out_dendro,'Dendrogram')
        verbout('\nFile "'+out_dendro+'" created.',master=master)
        
    if add_outputs['hist'] == 1:
        verbout('\n\nOutput of detailed history: ',master=master)
        out_prog_v = ['Result','Components','Dist','Maxdist','Quota']
        out_prog = tabulate(progress,out_prog_v)
        t = write_data(out_prog,out_prog_v,out_hist)
        verbout(t[0],master=master)
        if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=master)

    if add_outputs['dist'] == 1:
        verbout('\nCalculating Proximity scores for '+str(len(var_detail))+' elements',master=master)
        v_out = ['Element']
        v_out = v_out + real_clusters
        out_dic = {}
        for v in v_out:
            out_dic[v] = []
        for v in var_detail:
            out_dic['Element'].append(v)
            for rc in real_clusters:
                dist = distance(univec(clus_dic[v]),univec(clus_dic[rc]))
                load = math.cos(2*math.asin(dist/2))
                out_dic[rc].append(load)
        for i in range(len(data[dvar[0]])):
            out_dic['Element'].append(data[dvar[0]][i])
            for rc in real_clusters:
                out_dic[rc].append(clus_dic[rc][i])
        t = write_data(out_dic,v_out,out_loading)
        verbout(t[0],master=master)
        if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=master)

    if add_outputs['vector'] == 1:
        verbout('\nOutput of all vectors: ',master=master)
        for v in dvar:
            if not v in clus_dic.keys():
                clus_dic[v] = data[v]
        t = write_data(clus_dic,dvar,out_vector)
        verbout(t[0],master=master)
        if len(t[1])>2:verbout('\n'+t[1]+'\n','warning',master=master)



def write_ssa_cluster(distmat,ssa_elemente,out_ssa,titel_bez='Proximity of Elements'):
    elements = sorted(distmat.keys())
    disttable = {'':elements}
    for e1 in elements:
        disttable[e1] = []
        for e2 in elements:
            disttable[e1].append(distmat[e1][e2])       
    write_ssa(disttable,out_ssa,bildname=titel_bez)


def lcopy(liste):
    outlist = []
    for e in liste:
        outlist.append(e)
    return outlist

def univec(vector): ## Transform vector to Unit vector (length 1)
    outlist = []
    laenge = 0
    for e in vector:
        laenge = laenge + float(e)**2            
    laenge = laenge**.5
        
    if laenge == 0:
        laenge = 1

    for e in vector:
       outlist.append(float(e)/laenge)
    return outlist    

def transform_float(data,variablen=[]): ##Either submit a dataset and get the dataset back transformed or submit individual columns and get columns back
    if type(data) == dict:
        if variablen == []:
            variablen == data.keys()
        outdic = {}
        for v in variablen:
            outdic[v] = []
        for i in range(0,len(data[variablen[0]])):           
            for v in variablen:
                try:
                    f = float(data[v][i])
                except:
                    f = 0.0
                outdic[v].append(f)
        transformed = outdic

    elif type(data) == list:
        outlist = []
        for e in data:
            try:
                outlist.append(float(e))
            except:
                outlist.append('')
        transformed = outlist
    return transformed

def standardize_rows(data):
    outdic = {}
    variablen = sorted(data.keys())
    for v in variablen:
        outdic[v] = []
    for i in range(0,len(data[variablen[0]])):
        row = []
        for v in variablen:
            row.append(float(data[v][i]))
        quot = max(row)
        for v in variablen:
            outdic[v].append(float(data[v][i])/quot)
    return outdic

def standardize_table(data):
    cols = sorted(data.keys())
    ncols = len(cols)
    nrows = len(data[cols[0]])
    margins_x = []
    margins_y = []
    for i in range(0,ncols):
        margins_x.append(0.0)
    for i in range(0,nrows):
        margins_y.append(0.0)
    total = 0.0

    for x in range(0,len(cols)):
        v = cols[x]
        for y in range(0,nrows):
            margins_x[x] = margins_x[x] + data[v][y]
            margins_y[y] = margins_y[y] + data[v][y]
            total = total + data[v][y]

    outdic = {}
    for v in cols:
        outdic[v] = []

    for x in range(0,len(cols)):
        v = cols[x]
        for y in range(0,nrows):
            exp = float(margins_x[x] + margins_y[y]) / total
            dev = (float(data[v][y])/exp)**.5
            outdic[v].append(dev)
    return outdic
    
def distance(list1,list2,uni=0):
    dist = 0
    l1 = []
    l2 = []
    for i in range(len(list1)):
        try:
            a1 = float(list1[i])
            a2 = float(list2[i])
            l1.append(a1)
            l2.append(a2)
        except:
            pass ##One of the vectors contains string variables
    
    if uni == 1:
        l1 = univec(l1)
        l2 = univec(l2)
    if len(l1) == len(l2):
        for i in range(0,len(l1)):
            dist = dist + (l1[i]-l2[i])**2
        dist = dist**0.5  
    else:
        dist = -1

    return dist


def distmatrix(data,liste):
    out_dic = {}
    anz_dic = {}
    n = len(data.keys())
    for a in liste:
        out_dic[a] = {}
        anz_dic[a] = 0
        for b in liste:
            out_dic[a][b] = 0

    for e1 in liste:
        for e2 in liste:
            dist = distance(univec(data[e1]),univec(data[e2]))
            out_dic[e1][e2] = dist
            
    return out_dic


def flatten(nestlist):
    a = str(nestlist)
    b = ''
    for c in a:
        if not c in ['[',']']:
            b = b + c
    b = '['+b+']'
    outlist = eval(b)   
    return outlist

def gravity(data, cluster):
    outlist = []
    vliste = flatten(cluster)
    for i in range(0, len(data[vliste[0]])):
        mean = 0
        for var in vliste:
            mean = mean + data[var][i]/len(vliste)
        outlist.append(mean)
    
    outlist = univec(outlist) #Damit auch die Cluster immer Einheitsvektoren bleiben    
    return outlist


def maxdist(data,nestlist):
    cluster = data[str(nestlist)]
    liste = flatten(nestlist)

    max_distance = 0
    for var in liste:
        dist = distance(cluster,data[var])
        if dist > max_distance:
            max_distance = dist
    return max_distance


def get_pos(liste,element): ##Help function for dendrogram: Finding the position of a leaf
    ret = -1
    for i in range(0,len(liste)):
        if liste[i] == element:
            ret = i
    return ret

def get_node(prog,result): ##Help function for dendrogram: Finding the position of a node
    ret = -1
    for i in prog.keys():
        if prog[i]['Result'] == result:
            ret = i
    return ret


def get_cluster(data,frames,mode='loading'): ##Herausfinden, welches Item zu welchem Cluster gehört
    outlist = []
    if mode == 'dist': 
        for i in range(0,len(data[frames[0]])):
            dist_list = []
            for f in frames:
                dist_list.append(data[f][i])
            mindist = min(dist_list)
            minpos = -1
            for k in range(0,len(dist_list)):
                if dist_list[k] == mindist:
                    minpos = k + 1

            quota = sum(dist_list)/len(dist_list)/mindist

            outcluster = '0'
            if quota > 1.5:
                outcluster = 'F' + str(minpos)
            elif quota > 1.1:
                outcluster = 'F' + str(minpos)+ '.1'
            outlist.append(outcluster)
    elif mode == 'loading':
        for i in range(0,len(data[frames[0]])):
            load_list = []
            for f in frames:
                load_list.append(data[f][i])
            maxload = max(load_list)
            maxpos = -1
            for k in range(0,len(load_list)):
                if load_list[k] == maxload:
                    maxpos = k + 1

            outcluster = 'None'
            if maxload > 0.8:
                outcluster = 'F' + str(maxpos)
            elif maxload > 0.6:
                outcluster = 'F' + str(maxpos)+ '.1'

            outcluster = outcluster + '\t' + str(maxload)
            outlist.append(outcluster) 

    return outlist           
        
def tabulate(prog,variablen): ##Make table from a dendrogram
    outdic = {}
    for v in variablen:
        outdic[v] = []

    for k in prog.keys():
        for v in variablen:
            outdic[v].append(prog[k][v])

    return outdic

def write_dendrogram_R(prog,order,fname,plottitel='Clusteranalyse'):
    out = 'a <- list()\na$merge <- matrix(c('
    merge = ''
    i = 1
    while i in prog.keys():
        k1 = prog[i]['K1']
        k2 = prog[i]['K2']
        k1_pos = get_pos(order,k1)+1
        k2_pos = get_pos(order,k2)+1

        if k1_pos == 0:
            k1_pos = get_node(prog,k1)
            merge = merge + str(k1_pos)
        else:
            merge = merge + '-' + str(k1_pos)
        merge = merge + ','

        if k2_pos == 0:
            k2_pos = get_node(prog,k2)
            merge = merge + str(k2_pos)
        else:
            merge = merge + '-' + str(k2_pos)
        i = i + 1
        if i < len(order):
            merge = merge + ',\n     '
    out = out + merge
    out = out + '), nc=2, byrow=TRUE )\n'
    out = out + 'a$height <- 1:' + str(i-1) + '\n'
    out = out + 'a$order <- 1:' + str(i) + '\n'
    out = out + 'a$labels <- c("' + str(order[0]) + '"'
    for i in range(1,len(order)):
        out = out + ',"' + str(order[i]) + '"'
    out = out + ')\n'
    out = out + 'class(a) <- "hclust"\nplot(a,cex=0.6,main="'
    out = out + plottitel
    out = out + '")\n\n'

    exp_file = open(fname,'w')
    exp_file.write(out)
    exp_file.close()

def printDendrogram(T, sep=2):
    global outdend
    outdend = ''
    
    """Print dendrogram of a binary tree.  Each tree node is represented by a length-2 tuple."""
	
    def isPair(T):
        return type(T) == tuple and len(T) == 2
    
    def maxHeight(T):
        if isPair(T):
            h = max(maxHeight(T[0]), maxHeight(T[1]))
        else:
            h = len(str(T))
        return h + sep
        
    activeLevels = {}

    def traverse(T, h, isFirst):
        global outdend
        if isPair(T):
            traverse(T[0], h-sep, 1)
            s = [' ']*(h-sep)
            s.append('|')
        else:
            s = list(str(T))
            s.append(' ')

        while len(s) < h:
            s.append('-')
        
        if (isFirst >= 0):
            s.append('+')
            if isFirst:
                activeLevels[h] = 1
            else:
                del activeLevels[h]
        
        A = list(activeLevels)
        A.sort()
        for L in A:
            if len(s) < L:
                while len(s) < L:
                    s.append(' ')
                s.append('|')

        outdend = outdend + '\n'+''.join(s)
        
        if isPair(T):
            traverse(T[1], h-sep, 0)

    fullc = str(T)
    fullc = fullc.replace('[','(')
    fullc = fullc.replace(']',')')
    T = eval(fullc)
    
    traverse(T, maxHeight(T), -1)
    return outdend


######################################  K-Means


def multi_kmeans(data, dvar, varlist, direction, numrange, iterations, icenters, gv,master=''):
    verbout('\n\nAdjusting solution for multigroup analysis.\n\nGroups:',master=master)
    detailed_report = ''
    groups = get_unique(data[gv])
    
    groupmembers = {}
    groupdata = {}
    for g in groups:
        groupmembers[g] = []
        groupdata[g] = {}
        for v in dvar:
            groupdata[g][v] = []
            
    for i in range(len(data[gv])):
        groupmembers[data[gv][i]].append(i)
        for v in dvar:
            groupdata[data[gv][i]][v].append(data[v][i])

    ogroups = []
    gweights = []
    for g in groups:
        if len(groupmembers[g])>2:
            verbout('\n - "'+str(g)+'": '+str(len(groupmembers[g]))+' Cases',master=master)
            gweights.append(len(groupmembers[g]))
            ogroups.append(g)

    ncenters = {}
    membership = {}
    for num in numrange:
        detailed_report = detailed_report + '\n------------------------\nReport for '+str(num)+' Clusters\n------------------------\n'
        verbout('\n\n\n------------------------\nMultigroup for '+str(num)+' Clusters\n------------------------\n',master=master)
        ncenters[num] = {}
        ncenters[num]['#Pooled'] = icenters[num]
        detailed_report = detailed_report + '\nInitial cluster solution: \n'+baum_schreiben(icenters[num])
        membership[num] = {}

        movement = 1

        while movement > 0:   
            for g in ogroups:
                verbout('\n\nRealigning Group '+str(g)+':',master=master)
                solution = kmeans(groupdata[g],varlist,direction,num,iterations=1,precenter=ncenters[num]['#Pooled'],master=master)
                ncenters[num][g] = {}
                for c in solution.keys():
                    ncenters[num][g][c] = solution[c]['Center']
                
            ncenters[num]['#New'] = {}
            for n in range(num):
                ncenters[num]['#New'][n] = []
                for d in range(len(ncenters[num]['#Pooled'][n])):
                    clist = []
                    for g in ogroups:
                        clist.append(ncenters[num][g][n][d])
                    m = calculate(clist,'mean',gweights)
                    sd = calculate(clist,'sd',gweights)
                    
                    ncenters[num]['#New'][n].append(m)

            movement = 0
            for n in range(num):
                dist = distance(ncenters[num]['#New'][n], ncenters[num]['#Pooled'][n])
                verbout('\n - Moving cluster center '+str(n)+' by '+str(dist),master=master)
                movement = movement + dist
                ncenters[num]['#Pooled'][n] = ncenters[num]['#New'][n]

        verbout('\n\nFound optimal solution for overall cluster centers.\n\n',master=master)
        verbout(baum_schreiben(ncenters[num]['#Pooled']),'table',master=master)
        verbout('\n',master=master)

        detailed_report = detailed_report + '\n\nPooled cluster center after alignment: \n'+baum_schreiben(ncenters[num]['#Pooled'])    

        for g in ogroups:
            membership[num][g] = {}
            verbout('\n\nReasserting center for Group: ' +str(g),master=master)
            solution = kmeans(groupdata[g],varlist,direction,num,iterations=100,precenter=ncenters[num]['#Pooled'],master=master)
            for c in solution.keys():
                membership[num][g][c] = solution[c]['Members']
                ncenters[num][g][c] = solution[c]['Center']
            detailed_report = detailed_report + '\n\nCluster centers for group: '+str(g)+'\n'
            detailed_report = detailed_report + baum_schreiben(ncenters[num][g])

        membership[num]['#Pooled'] = {}
        verbout('\n\nReasserting center for overall solution',master=master)
        solution = kmeans(data,varlist,direction,num,iterations=100,precenter=ncenters[num]['#Pooled'],master=master)
        for c in solution.keys():
            membership[num]['#Pooled'][c] = solution[c]['Members']
            ncenters[num]['#Pooled'][c] = solution[c]['Center']


    return [detailed_report,membership,ogroups]
            
                    

def kmeans(dset,varlist,direction=1,num=4,iterations = 100, precenter = {},master=''):
    if type(dset) in [list,tuple]:
        data = dset[0]
    else:
        data = dset
    entity = {}
    dimensions = {}
    center = {}
    ncases = len(data[varlist[0]])
    varlist = sorted(varlist)
    caslist = range(ncases)
    ranges = []


    if not '#Original_i' in data.keys():
        data['#Original_i'] = caslist

    if direction in [1,'1']:
        for v in varlist:
            dimensions[v] = []
        for i in caslist:
            ori = data['#Original_i'][i]
            entity[ori] = []
            for v in varlist:
                entity[ori].append(data[v][i])
                dimensions[v].append(data[v][i])
    elif direction in [2,'2']:
        for i in caslist:
            dimensions[i] = []
        for v in varlist:
            entity[v] = []
            for i in caslist:
                entity[v].append(data[v][i])
                dimensions[i].append(data[v][i])

    for d in dimensions.keys():
        dimensions[d] = stat_desc(dimensions[d])
        ranges.append(dimensions[d]['Range'])

    ndim = len(dimensions.keys())
    nent = len(entity.keys())

    if iterations > 1:
        verbout('\nClustering '+str(nent)+' entities in '+str(ndim)+'-dimensional space\n',master=master)

    if precenter == {} or not len(precenter.keys()) == num:
        verbout('\nSeeding random centers',master=master)
        mindist = 0
        while mindist == 0:
            mindist = 1
            for i in range(num):
                center[i] = []
                for d in sorted(dimensions.keys()):
                    low = dimensions[d]['Min']
                    hig = dimensions[d]['Max']
                    ran = dimensions[d]['Range']
                    coord = random.random()*ran+low
                    center[i].append(coord)
            for i in range(num):
                for k in range(num):
                    if not i == k:
                        if distance(center[i],center[k]) == 0:
                            mindist = 0
                            
    else:
        for c in precenter.keys():
            center[c] = precenter[c]


    rsum = 0
    for r in ranges:
        rsum = rsum + r**2
    maxrange = rsum**.5
    cluster = {}
    for c in center.keys():
        cluster[c] = {'Center':center[c]}


    ##Iterative k-means clustering
    change = 1
    iteration = 0
    eta = 0.0001
    if iterations > 1:
        verbout('\niterating',master=master)
    while change == 1 and iteration < iterations:
        iteration = iteration + 1
        if iterations > 1:
            verbout('.',master=master)

        for c in center.keys():
            cluster[c]['Members'] = []
        undefined = 0
        for e in entity.keys():
            mindist = maxrange
            membership = -1
            for c in sorted(center.keys()):
                dist = distance(entity[e],center[c])
                if dist < mindist:
                    mindist = dist
                    membership = c
                elif dist == mindist:
                    membership = -1
            if membership == -1:
                verb('Undefined case: '+str(e))
                undefined = undefined + 1
            else:
                cluster[membership]['Members'].append(e) ##if there is an error on this line, no cluster center could be found

        deltalist = []
        for c in center.keys():
            newcenter = []
            for i in range(ndim):
                dcoord = []
                for m in cluster[c]['Members']:
                    dcoord.append(entity[m][i])
                if len(dcoord) > 0:
                    mcoord = calculate(dcoord, 'mean') #stat_desc(dcoord)['M']
                else:
                    mcoord = 0
                newcenter.append(mcoord)
            delta = distance(cluster[c]['Center'],newcenter)
            deltalist.append(delta)
            cluster[c]['Center'] = newcenter
            center[c] = newcenter

        movement = calculate(deltalist,'mean') #stat_desc(deltalist)['M']
        if iterations == 1:
            verbout(' Mean center movement: '+str(movement),master=master)
        if movement < eta:
            change = 0
        
##            for c in sorted(cluster.keys()):
##                verbout('Cluster#'+str(c)+': '+str(len(cluster[c]['Members']))+' Members\n',master=master)

    if iterations > 1:
        if change == 0:
            verbout('\n\nCluster analysis concluded in '+str(iteration)+' iterations\n',master=master)
        else:
            verbout('\n\nCluster analysis NOT concluded in '+str(iteration)+' iterations\nFinal movement of centers (eta) was '+str(eta),master=master)            
    if undefined > 0:
        verbout('\n - '+str(undefined)+' Cases do not definitely belong to a cluster.\n',master=master)

##        alldist = []
##        for e1 in entity.keys():
##            for e2 in entity.keys():
##                if not e1==e2:
##                    alldist.append(distance(entity[e1],entity[e2]))
##
##        verbout('\n\nMean distance of entities: '+str(stat_desc(alldist)['M']),master=master)

    cendist = []        
    for c1 in cluster.keys():
        for c2 in cluster.keys():
            if not c1==c2:
                cendist.append(distance(cluster[c1]['Center'],cluster[c2]['Center']))

    if iterations > 1:
        verbout('\n\nMean distance of cluster centers: '+str(stat_desc(cendist)['M']),master=master)
        cbdists = []
        for c in sorted(cluster.keys()):
            verbout('\nCluster#'+str(c)+': '+str(len(cluster[c]['Members']))+' Members; ',master=master)
            bindist = []
            mb = '-'
            for e in cluster[c]['Members']:
                bindist.append(distance(cluster[c]['Center'],entity[e]))
                mb = stat_desc(bindist)['M']
                cbdists.append(mb)
            verbout('Mean distance: '+str(mb),master=master)

        verbout('\n\nMean distance from cluster center: '+str(stat_desc(cbdists)['M']),master=master)       
    
    return cluster
        

def create_cluster_table(data, varlist, std, mg,master=''):
    global storage

    cluster_table = {}
    cluster_table['#Original_i'] = []
    cluster_varlist = ['#Original_i']
    ncases = len(data[varlist[0]])

    if type(mg) == str:
        cluster_table[mg] = []
        cluster_varlist.append(mg)

    for i in range(ncases):
        cluster_table['#Original_i'].append(i)
        if type(mg) == str:
            cluster_table[mg].append(data[mg][i])

    verbout('\n\nPreparing numeric Clustering table:\nVariable overview:',master=master)
        
    for v in sorted(varlist):
        column = []
        valid = 0
        for i in range(ncases):
            try:
                column.append(float(data[v][i]))
                valid = valid + 1
            except:
                column.append('')
        verbout('\n - "'+v+'": '+str(valid)+' valid values',master=master)
        cluster_table[v] = column
        cluster_varlist.append(v)

    incompletes = []
    for i in range(ncases):
        valid = 1
        for v in varlist:
            if cluster_table[v][i] == '':
                valid = 0

        if valid == 0: incompletes.append(i)

    for remi in sorted(incompletes, reverse=True):
        for v in cluster_table.keys():
            cluster_table[v].pop(remi)

    ncases = len(cluster_table[varlist[0]])

    if std == 'none':
        pass
    elif std == 'rstand':
        verbout('\nStandardizing rows',master=master)
        for i in range(ncases):
            tmprow = []
            for v in varlist:
                tmprow.append(cluster_table[v][i])
            stat = stat_desc(tmprow)
            m = stat['M']
            sd = stat['SD']
            if sd > 0:
                for v in varlist:
                    cluster_table[v][i] = (cluster_table[v][i]-m)/sd
            else:
                for v in varlist:
                    cluster_table[v][i] = 0
    elif std == 'rnorm':
        verbout('\nNormalizing rows',master=master)
        for i in range(ncases):
            tmprow = []
            for v in varlist:
                tmprow.append(cluster_table[v][i])
            stat = stat_desc(tmprow)
            low = stat['Min']
            hig = stat['Max']
            ran = hig-low
            if ran > 0:
                for v in varlist:
                    cluster_table[v][i] = (cluster_table[v][i]-low)/ran
            else:
                for v in varlist:
                    cluster_table[v][i] = 0   
    elif std == 'cstand':
        verbout('\nStandardizing columns',master=master)
        for v in varlist:
            tmpcol = cluster_table[v]
            stat = stat_desc(tmpcol)
            m = stat['M']
            sd = stat['SD']
            if sd > 0:
                for i in range(ncases):
                    cluster_table[v][i] = (cluster_table[v][i]-m)/sd
            else:
                for i in range(len(tmpcol)):
                    cluster_table[v][i] = 0               
    elif std == 'cnorm':
        verbout('\nNormalizing columns',master=master)
        for v in varlist:
            tmpcol = cluster_table[v]
            stat = stat_desc(tmpcol)
            low = stat['Min']
            hig = stat['Max']
            ran = hig-low
            if hig-low > 0:
                for i in range(ncases):
                    cluster_table[v][i] = (cluster_table[v][i]-low)/ran
            else:
                for i in range(len(tmpcol)):
                    cluster_table[v][i] = 0
    elif std == 'tstand':
        verbout('\nStandardizing table',master=master)
        tcoll = []
        for v in varlist:
            tcoll = tcoll + cluster_table[v]
        stat = stat_desc(tcoll)
        for v in varlist:
            if stat['SD']>0:
                for i in range(ncases):
                    cluster_table[v][i] = (cluster_table[v][i]-stat['M'])/stat['SD']
            else:
                for i in range(ncases):
                    cluster_table[v][i] = 0
                verbout('\n',master=master)
                verbout('Warning: The overall variance of your table was zero. All values are set to 0\n','warning',master=master)
    elif std == 'tnorm':
        verbout('\nNormalizing table',master=master)
        tcoll = []
        for v in varlist:
            tcoll = tcoll + cluster_table[v]
        stat = stat_desc(tcoll)
        ran = stat['Max']-stat['Min']
        for v in varlist:
            if ran>0:
                for i in range(ncases):
                    cluster_table[v][i] = (cluster_table[v][i]-stat['Min'])/ran
            else:
                for i in range(ncases):
                    cluster_table[v][i] = 0
                verbout('\n',master=master)
                verbout('Warning: The overall range of your table was zero. All values are set to 0\n','warning',master=master)
                
    else:
        verbout('\nError: unknown command\n','warning',master=master)

    storage['CData'] = cluster_table
    storage['CVars'] = cluster_varlist
    
    settings['Datasets']['Cluster Table'] = {}
    settings['Datasets']['Cluster Table']['Data'] = 'CData'
    settings['Datasets']['Cluster Table']['Var'] = 'CVars'

    



##########################
##
## Reliability testing
##
##########################


def calc_ic_reliability(dset,uvar,cvar,varlist,unitlist=[],coderlist=[],core_cod='no_core',methods=['PA'],options=[],master=''):
    if type(dset)==dict:
        data = dset
    else:
        data = dset[0]

    if unitlist == []: unitlist = get_unique(data[uvar])
    if coderlist == []: coderlist = get_unique(data[cvar])
        
    cdic = create_coding_dic(data,uvar,cvar,varlist,unitlist,master=master)
    
    return reltest(cdic,varlist,unitlist,coderlist,core_cod,methods,options,master=master)


def create_coding_dic(data,uvar,cvar,varlist,unitlist=[],master=''): ##Reliability testing: Table of coders and units
    outdic = {}
    for u in unitlist:
        outdic[u] = {}

    verbout('\nCreating dictionary of codings: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
    step = int(len(data[uvar])/40)
    panz = 1
    if step<1:
        step=1
        panz = int(40.0/len(pairs))

    for i in range(len(data[uvar])):
        if i%step==0:verbout('.'*panz,'progress',master=master)
        c = data[cvar][i]
        u = data[uvar][i]
        if u in unitlist:
            outdic[u][c] = {}
            for v in varlist:
                value = data[v][i]
                if value in ['',' ','.']: value = ''
                outdic[u][c][v] = value
                
    verbout('\n','progress',master=master)         
    return outdic




def reltest(codings,tvars,units,coders,kerncod='no_core',methods=['PA'],options=[],master=''):
##        print(codings)
##        print(tvars)
##        print(units)
##        print(coders)
##        print(kerncod)
##        print(methods)
##        print(options)
    

    text_table = 'Reliability test output:\n\n'
    
    min_c_count = 0
    if 'min2' in options:
        min_c_count = 1
    if 'min5' in options:
        min_c_count = 4

    if not kerncod == 'no_core':
        if 'PA' in methods:
            methods.append('PA_w_core')
        if 'Kappa' in methods:
            methods.append('Kappa_w_core')
        if 'Kappan' in methods:
            methods.append('Kappan_w_core')
        if 'Pi' in methods:
            methods.append('Pi_w_core')


    verbout('\nCalculating variable distributions...\n',master=master)

    vinfo = {}
    for v in tvars:
        vinfo[v] = {}
        vinfo[v]['Values'] = {}
        vallist = []
        vinfo[v]['Codings'] = 0
        for u in codings.keys():
            for c in codings[u].keys():
                if v in codings[u][c].keys():
                    if 'cm' in options or not codings[u][c][v] == '':
                        if not codings[u][c][v] in vinfo[v]['Values'].keys():
                            vinfo[v]['Values'][codings[u][c][v]] = 0
                        vinfo[v]['Values'][codings[u][c][v]] = vinfo[v]['Values'][codings[u][c][v]] + 1
                        vinfo[v]['Codings'] = vinfo[v]['Codings'] + 1
                        vallist.append(codings[u][c][v])

        pc = 0
        modus = ('invalid',0)
        for val in vinfo[v]['Values'].keys():
            pc = pc + (float(vinfo[v]['Values'][val]) / vinfo[v]['Codings'])**2
            if vinfo[v]['Values'][val] > modus[1]: modus = (val,vinfo[v]['Values'][val])

        vinfo[v]['Coded N']=str(len(vallist))
        vinfo[v]['Codes']=str(len(get_unique(vallist)))
        vinfo[v]['Pc'] = pc
        vinfo[v]['P(chance)'] = "{:4.2f}".format(pc)+'%'
        vinfo[v]['Entropy']="{:4.3f}".format(calc_entropy(vallist))
        vinfo[v]['Mode'] = modus[0]

    v_table = {}
    for prop in ['Codes','Coded N','Mode','P(chance)','Entropy']:
        v_table[prop] = {}
        for v in vinfo.keys():
            v_table[prop][v] = vinfo[v][prop]

    verbout(display_table(v_table)+'\n','table',master=master)
    text_table = text_table + '\n\nVariable Properties: \n'+display_table(v_table,sep='tab')+'\n'

    results = {}
    nresults = {}
    for v in tvars:
        results[v] = {}
        nresults[v] = {}

    if 'PA' in methods or 'Pi' in methods or 'Kappa' in methods or 'Kappan' in methods:

        verbout('\nCalculating Pairwise comparisons...\n\n',master=master)
       


        pairs = []
        for i1 in range(len(coders)-1):
            for i2 in range(i1+1,len(coders)):
                pairs.append((coders[i1],coders[i2]))
                
        verbout('Comparing '+str(len(pairs)) + ' Pairings: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
        step = len(pairs)/40
        stepi = 0
        panz = 1
        if step<1:
            step=1
            panz = int(40.0/len(pairs))
        

        for p in pairs:
            stepi = stepi + 1
            if stepi%step == 0:verbout('.'*panz,'progress',master=master)
            results[p] = {}
            for v in tvars:
                results[p][v] = {}
                crosstab = {}
                margins = {}
                for v1 in vinfo[v]['Values'].keys():
                    crosstab[v1] = {}
                    margins[v1] = 0
                    for v2 in vinfo[v]['Values'].keys():
                        crosstab[v1][v2]=0

                c_count = 0
                for u in units:
                    if p[0] in codings[u].keys() and p[1] in codings[u].keys():
                        val1 = codings[u][p[0]][v]
                        val2 = codings[u][p[1]][v]
                        if 'cm' in options or not '' in [val1,val2]:
                            margins[val1] = margins[val1] + 1
                            margins[val2] = margins[val2] + 1
                            if not val1 == '' and not val2 == '':
                                crosstab[val1][val2] = crosstab[val1][val2] +1
                                c_count = c_count + 1
                    
                results[p][v]['Crosstab'] = crosstab

                hit = 0
                for val in vinfo[v]['Values'].keys():
                    hit = hit + crosstab[val][val]
                kappan = ''

                if c_count > min_c_count:
                    pc_kappa = 0
                    for m in margins.keys():
                        pc_kappa = pc_kappa + (float(margins[m])/c_count/2)**2
                    pc_kappan = 1.0/len(margins.keys())
                    pa = float(hit) / c_count

                    if pc_kappa < 1.0:
                        kappa = (pa-pc_kappa)/(1-pc_kappa)
                    elif 'km' in options:
                        kappa = ''
                    else:
                        kappa = 1.0

                    if pc_kappan < 1.0:
                        kappan = (pa-pc_kappan)/(1-pc_kappan)
                    elif 'km' in options:
                        kappan = ''
                    else:
                        kappan = 1.0
                    
                    if vinfo[v]['Pc'] < 1.0:
                        pi = (pa-vinfo[v]['Pc'])/(1-vinfo[v]['Pc'])
                    elif 'km' in options:
                        pi = ''
                    else:
                        pi = 1.0
                else:
                    pa = ''
                    pi = ''
                    kappa = ''
                    kappan = ''

                results[p][v]['PA'] = pa
                results[p][v]['Kappa'] = kappa
                results[p][v]['Pi'] = pi
                results[p][v]['Kappan'] = kappan

            pa = []
            kappa = []
            kappan = []
            pi = []
            for v in tvars:
                if type(results[p][v]['PA']) == float: pa.append(results[p][v]['PA'])
                if type(results[p][v]['Kappa']) == float: kappa.append(results[p][v]['Kappa'])
                if type(results[p][v]['Kappan']) == float: kappan.append(results[p][v]['Kappan'])
                if type(results[p][v]['Pi']) == float: pi.append(results[p][v]['Pi'])

            if len(pa) > 0:
                results[p]['PA'] = "{:3.1f}".format(sum(pa)/len(pa)*100)+'%'
            else:
                results[p]['PA'] = ''

            if len(kappa) > 0:
                results[p]['Kappa'] = "{:4.3f}".format(sum(kappa)/len(kappa))
            else:
                results[p]['Kappa'] = ''

            if len(kappan) > 0:
                results[p]['Kappan'] = "{:4.3f}".format(sum(kappan)/len(kappan))
            else:
                results[p]['Kappan'] = ''

            if len(pi) > 0:
                results[p]['Pi'] = "{:4.3f}".format(sum(pi)/len(pi))
            else:
                results[p]['Pi'] = ''
                            
        verbout('\n','progress',master=master)

        pa_table = {}
        kappa_table = {}
        kappan_table = {}
        pi_table = {}
        pa_total = []
        kappa_total = []
        kappan_total = []
        pi_total = []
        for v in tvars:                   
            pa_table[v] = {}
            kappa_table[v] = {}
            kappan_table[v] = {}
            pi_table[v] = {}
            pa = []
            kappa = []
            kappan = []
            pi = []
            for p in pairs:
                pair = p[0]+' X '+p[1]
                if type(results[p][v]['PA']) == float:
                    pa_table[v][pair] = "{:3.1f}".format(results[p][v]['PA']*100) + '%'
                    pa.append(results[p][v]['PA'])
                    pa_total.append(results[p][v]['PA'])
                else:
                    pa_table[v][pair] = ''

                if type(results[p][v]['Kappa']) == float:
                    kappa_table[v][pair] = "{:4.3f}".format(results[p][v]['Kappa'])
                    kappa.append(results[p][v]['Kappa'])
                    kappa_total.append(results[p][v]['Kappa'])
                else:
                    kappa_table[v][pair] = ''

                if type(results[p][v]['Kappan']) == float:
                    kappan_table[v][pair] = "{:4.3f}".format(results[p][v]['Kappan'])
                    kappan.append(results[p][v]['Kappan'])
                    kappan_total.append(results[p][v]['Kappan'])
                else:
                    kappan_table[v][pair] = ''

                if type(results[p][v]['Pi']) == float:
                    pi_table[v][pair] = "{:4.3f}".format(results[p][v]['Pi'])
                    pi.append(results[p][v]['Pi'])
                    pi_total.append(results[p][v]['Pi'])
                else:
                    pi_table[v][pair] = ''

            if len(pa) > 0:
                results[v]['PA'] = "{:3.1f}".format(sum(pa)/len(pa)*100)+'%'
                nresults[v]['PA'] = sum(pa)/len(pa)*100
            else:
                results[v]['PA'] = ''
                nresults[v]['PA'] = ''

            if len(kappa) > 0:
                results[v]['Kappa'] = "{:4.3f}".format(sum(kappa)/len(kappa))
                nresults[v]['Kappa'] = sum(kappa)/len(kappa)
            else:
                results[v]['Kappa'] = ''
                nresults[v]['Kappa'] = ''

            if len(kappan) > 0:
                results[v]['Kappan'] = "{:4.3f}".format(sum(kappan)/len(kappan))
                nresults[v]['Kappan'] = sum(kappan)/len(kappan)
            else:
                results[v]['Kappan'] = ''
                nresults[v]['Kappan'] = ''

            if len(pi) > 0:
                results[v]['Pi'] = "{:4.3f}".format(sum(pi)/len(pi))
                nresults[v]['Pi'] = sum(pi)/len(pi)
            else:
                results[v]['Pi'] = ''
                nresults[v]['Pi'] = ''


        if 'PA' in methods:
            if len(pa_total) > 0:
                results['PA'] = "{:3.1f}".format(sum(pa_total)/len(pa_total)*100)+'%'
            else:
                results['PA'] = 'No valid comparisons'
            text_table = text_table + 'Percent Agreement (Holsti) between coders: '+results['PA']+display_table(pa_table,sep='tab')     

        if 'Kappa' in methods:
            if len(kappa_total) > 0:
                results['Kappa'] = "{:4.3f}".format(sum(kappa_total)/len(kappa_total))
            else:
                results['Kappa'] = 'No valid comparisons'
            text_table = text_table + "\n\nCohen's Kappa between coders: "+results['Kappa']+display_table(kappa_table,sep='tab')

        if 'Kappan' in methods:
            if len(kappan_total) > 0:
                results['Kappan'] = "{:4.3f}".format(sum(kappan_total)/len(kappan_total))
            else:
                results['Kappan'] = 'No valid comparisons'
            text_table = text_table + "\n\nBrennan & Prediger's Kappa between coders: "+results['Kappan']+display_table(kappan_table,sep='tab')

        if 'Pi' in methods:
            if len(pi_total) > 0:
                results['Pi'] = "{:4.3f}".format(sum(pi_total)/len(pi_total))
            else:
                results['Pi'] = 'No valid comparisons'
            text_table = text_table + "\n\nScott's Pi between coders: "+results['Pi']+display_table(pi_table,sep='tab')

        verbout('\nSummary of pairwise comparisons by coder pair:\n',master=master)

        pcomp = []
        for m in methods:
            if m in ['PA','Kappa','Kappan','Pi']:
                pcomp.append(m)

        total_pairs = {}
        for coeff in pcomp:
            total_pairs[coeff] = {}
            for p in pairs:
                pair = p[0]+' X '+p[1]
                total_pairs[coeff][pair] = results[p][coeff]

        verbout(display_table(total_pairs)+'\n','table',master=master)
        text_table = text_table + '\n\n\n\nSum of agreement between coders:'+display_table(total_pairs,sep='tab')


    if 'PA_w_core' in methods or 'Pi_w_core' in methods or 'Kappa_w_core' in methods or 'Kappan_w_core' in methods:
        verbout('\nCalculating comparisons with core coder: '+kerncod,master=master)

        pairs = []
        for c in coders:
            if not c == kerncod:
                pairs.append((kerncod,c))

        verbout('\n\n',master=master)

        verbout('Comparing '+str(len(pairs)) + ' Pairings with core coder "'+kerncod+'": \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)
        step = len(pairs)/40
        if step <1:
            step = 1
            panz = int(40.0/len(pairs))
        stepi = 0

       
        for p in pairs:
            stepi = stepi + 1
            if stepi%step == 0:
                verbout("."*panz,'progress',master=master)
            results[p] = {}
            for v in tvars:
                results[p][v] = {}
                crosstab = {}
                margins = {}
                for v1 in vinfo[v]['Values'].keys():
                    crosstab[v1] = {}
                    margins[v1] = 0
                    for v2 in vinfo[v]['Values'].keys():
                        crosstab[v1][v2]=0

                c_count = 0
                for u in units:
                    if p[0] in codings[u].keys() and p[1] in codings[u].keys():
                        val1 = codings[u][p[0]][v]
                        val2 = codings[u][p[1]][v]
                        if 'cm' in options or not '' in [val1,val2]:
                            margins[val1] = margins[val1] + 1
                            margins[val2] = margins[val2] + 1
                            if not val1 == '' and not val2 == '':
                                crosstab[val1][val2] = crosstab[val1][val2] +1
                                c_count = c_count + 1
                results[p][v]['Crosstab'] = crosstab

                hit = 0
                for val in vinfo[v]['Values'].keys():
                    hit = hit + crosstab[val][val]

                if c_count > min_c_count:
                    pc_kappa = 0
                    for m in margins.keys():
                        pc_kappa = pc_kappa + (float(margins[m])/c_count/2)**2

                    pc_kappan = 1.0/len(margins.keys())
                        
                    pa = float(hit) / c_count

                    if pc_kappa < 1.0:
                        kappa = (pa-pc_kappa)/(1-pc_kappa)
                    elif 'km' in options:
                        kappa = ''
                    else:
                        kappa = 1.0

                    if pc_kappan < 1.0:
                        kappan = (pa-pc_kappan)/(1-pc_kappan)
                    elif 'km' in options:
                        kappan = ''
                    else:
                        kappan = 1.0
                    
                    if vinfo[v]['Pc'] < 1.0:
                        pi = (pa-vinfo[v]['Pc'])/(1-vinfo[v]['Pc'])
                    elif 'km' in options:
                        pi = ''
                    else:
                        pi = 1.0
                else:
                    pa = ''
                    pi = ''
                    kappa = ''
                    kappan = ''

                results[p][v]['PA_w_core'] = pa
                results[p][v]['Kappa_w_core'] = kappa
                results[p][v]['Kappan_w_core'] = kappan
                results[p][v]['Pi_w_core'] = pi

            pa = []
            kappa = []
            kappan = []
            pi = []
            for v in tvars:
                if type(results[p][v]['PA_w_core']) == float: pa.append(results[p][v]['PA_w_core'])
                if type(results[p][v]['Kappa_w_core']) == float: kappa.append(results[p][v]['Kappa_w_core'])
                if type(results[p][v]['Kappan_w_core']) == float: kappan.append(results[p][v]['Kappan_w_core'])
                if type(results[p][v]['Pi_w_core']) == float: pi.append(results[p][v]['Pi_w_core'])

            if len(pa) > 0:
                results[p]['PA_w_core'] = "{:3.1f}".format(sum(pa)/len(pa)*100)+'%'
            else:
                results[p]['PA_w_core'] = ''

            if len(kappa) > 0:
                results[p]['Kappa_w_core'] = "{:4.3f}".format(sum(kappa)/len(kappa))
            else:
                results[p]['Kappa_w_core'] = ''

            if len(kappan) > 0:
                results[p]['Kappan_w_core'] = "{:4.3f}".format(sum(kappan)/len(kappan))
            else:
                results[p]['Kappan_w_core'] = ''

            if len(pi) > 0:
                results[p]['Pi_w_core'] = "{:4.3f}".format(sum(pi)/len(pi))
            else:
                results[p]['Pi_w_core'] = ''

        verbout('\n','progress',master=master)

        pa_table = {}
        kappa_table = {}
        kappan_table = {}
        pi_table = {}
        pa_total = []
        kappa_total = []
        kappan_total = []
        pi_total = []
        for v in tvars:
            pa_table[v] = {}
            kappa_table[v] = {}
            kappan_table[v] = {}
            pi_table[v] = {}
            pa = []
            kappan = []
            kappa = []
            pi = []
            for p in pairs:
                pair = p[0]+' X '+p[1]
                if type(results[p][v]['PA_w_core']) == float:
                    pa_table[v][pair] = "{:3.1f}".format(results[p][v]['PA_w_core']*100) + '%'
                    pa.append(results[p][v]['PA_w_core'])
                    pa_total.append(results[p][v]['PA_w_core'])
                else:
                    pa_table[v][pair] = ''

                if type(results[p][v]['Kappa_w_core']) == float:
                    kappa_table[v][pair] = "{:4.3f}".format(results[p][v]['Kappa_w_core'])
                    kappa.append(results[p][v]['Kappa_w_core'])
                    kappa_total.append(results[p][v]['Kappa_w_core'])
                else:
                    kappa_table[v][pair] = ''

                if type(results[p][v]['Kappan_w_core']) == float:
                    kappan_table[v][pair] = "{:4.3f}".format(results[p][v]['Kappan_w_core'])
                    kappan.append(results[p][v]['Kappan_w_core'])
                    kappan_total.append(results[p][v]['Kappan_w_core'])
                else:
                    kappan_table[v][pair] = ''

                if type(results[p][v]['Pi_w_core']) == float:
                    pi_table[v][pair] = "{:4.3f}".format(results[p][v]['Pi_w_core'])
                    pi.append(results[p][v]['Pi_w_core'])
                    pi_total.append(results[p][v]['Pi_w_core'])
                else:
                    pi_table[v][pair] = ''

            if len(pa) > 0:
                results[v]['PA_w_core'] = "{:3.1f}".format(sum(pa)/len(pa)*100)+'%'
                nresults[v]['PA_w_core'] = sum(pa)/len(pa)*100
            else:
                results[v]['PA_w_core'] = ''
                nresults[v]['PA_w_core'] = ''

            if len(kappa) > 0:
                results[v]['Kappa_w_core'] = "{:4.3f}".format(sum(kappa)/len(kappa))
                nresults[v]['Kappa_w_core'] = sum(kappa)/len(kappa)
            else:
                results[v]['Kappa_w_core'] = ''
                nresults[v]['Kappa_w_core'] = ''

            if len(kappan) > 0:
                results[v]['Kappan_w_core'] = "{:4.3f}".format(sum(kappan)/len(kappan))
                nresults[v]['Kappan_w_core'] = sum(kappan)/len(kappan)
            else:
                results[v]['Kappan_w_core'] = ''
                nresults[v]['Kappan_w_core'] = ''

            if len(pi) > 0:
                results[v]['Pi_w_core'] = "{:4.3f}".format(sum(pi)/len(pi))
                nresults[v]['Pi_w_core'] = sum(pi)/len(pi)
            else:
                results[v]['Pi_w_core'] = ''
                nresults[v]['Pi_w_core'] = ''


        if 'PA_w_core' in methods:
            if len(pa_total) > 0:
                results['PA_w_core'] = "{:3.1f}".format(sum(pa_total)/len(pa_total)*100)+'%'
            else:
                results['PA_w_core'] = 'No valid comparisons'                    
            text_table = text_table + 'Percent Agreement (Holsti) with core coder: '+results['PA_w_core']+display_table(pa_table,sep='tab')     

        if 'Kappa_w_core' in methods:
            if len(kappa_total) > 0:
                results['Kappa_w_core'] = "{:4.3f}".format(sum(kappa_total)/len(kappa_total))
            else:
                results['Kappa_w_core'] = 'No valid comparisons'
            text_table = text_table + "\n\nCohen's Kappa with core coder: "+results['Kappa_w_core']+display_table(kappa_table,sep='tab')

        if 'Kappan_w_core' in methods:
            if len(kappan_total) > 0:
                results['Kappan_w_core'] = "{:4.3f}".format(sum(kappan_total)/len(kappan_total))
            else:
                results['Kappan_w_core'] = 'No valid comparisons'
            text_table = text_table + "\n\nBrennan and Prediger's Kappa with core coder: "+results['Kappan_w_core']+display_table(kappan_table,sep='tab')

        if 'Pi_w_core' in methods:
            if len(pi_total) > 0:
                results['Pi_w_core'] = "{:4.3f}".format(sum(pi_total)/len(pi_total))
            else:
                results['Pi_w_core'] = 'No valid comparisons'
            text_table = text_table + "\n\nScott's Pi with core coder: "+results['Pi_w_core']+display_table(pi_table,sep='tab')

        verbout('\nSummary of comparisons with core coder by coder pair:\n',master=master)

        pcomp = []
        for m in methods:
            if m in ['PA_w_core','Kappa_w_core','Kappan_w_core','Pi_w_core']:
                pcomp.append(m)

        total_pairs = {}
        for coeff in pcomp:
            total_pairs[coeff] = {}
            for p in pairs:
                pair = p[0]+' X '+p[1]
                total_pairs[coeff][pair] = results[p][coeff]

        verbout(display_table(total_pairs)+'\n','table',master=master)
        text_table = text_table + '\n\n\n\nSum of agreement with core coder:'+display_table(total_pairs,sep='tab')

    if 'PRF' in methods:
        verbout('\nCalculating comparisons of dichotomous variables with core coder: '+kerncod+'\n',master=master)

        pairs = []
        for c in coders:
            if not c == kerncod:
                pairs.append((kerncod,c))

        ntvars = []
        for v in tvars:
            if vinfo[v]['Values'].keys() in [['0','1'],['1','0']]:
                ntvars.append(v)
            else:
                verbout('ERROR: Variable '+v+' is not dichotomous. Variable excluded. Values: '+str(sorted(vinfo[v]['Values'].keys()))+'\n','warning',master=master)
       
        for p in pairs: ##Core Coder is always the first coder
            results[p] = {}
            for v in ntvars:
                results[p][v] = {}

                c_count = 0
                tp=0
                fp=0
                tn=0
                fn=0       
                for u in units:
                    if p[0] in codings[u].keys() and p[1] in codings[u].keys():
                        val1 = codings[u][p[0]][v]
                        val2 = codings[u][p[1]][v]

                        if val1 == '1' and val2 == '1':
                            tp = tp + 1
                        elif val1 == '0' and val2 == '1':
                            fp = fp + 1
                        elif val1 == '0' and val2 == '0':
                            tn = tn + 1
                        elif val1 == '1' and val2 == '0':
                            fn = fn + 1

                if tp == 0:
                    precision = 0
                    recall = 0
                else:
                    precision = float(tp)/(tp+fp)
                    recall = float(tp)/(tp+fn)

                if precision + recall > 0:
                    fvalue = 2*precision*recall/(precision+recall)
                else:
                    fvalue = ''

                if tp+fp+tn+fn <= min_c_count:
                    precision = ''
                    recall = ''
                    fvalue = ''


                results[p][v]['Precision'] = precision
                results[p][v]['Recall'] = recall
                results[p][v]['F_Measure'] = fvalue

            precision = []
            recall = []
            fvalue = []
            pi = []
            for v in ntvars:
                if type(results[p][v]['Precision']) == float: precision.append(results[p][v]['Precision'])
                if type(results[p][v]['Recall']) == float: recall.append(results[p][v]['Recall'])
                if type(results[p][v]['F_Measure']) == float: fvalue.append(results[p][v]['F_Measure'])


            if len(precision) > 0:
                results[p]['Precision'] = "{:4.3f}".format(sum(precision)/len(precision))
            else:
                results[p]['Precision'] = ''

            if len(recall) > 0:
                results[p]['Recall'] = "{:4.3f}".format(sum(recall)/len(recall))
            else:
                results[p]['Recall'] = ''

            if len(fvalue) > 0:
                results[p]['F_Measure'] = "{:4.3f}".format(sum(fvalue)/len(fvalue))
            else:
                results[p]['F_Measure'] = ''

        precision_table = {}
        recall_table = {}
        fvalue_table = {}
        precision_total = []
        recall_total = []
        fvalue_total = []
        
        for v in ntvars:
            precision_table[v] = {}
            recall_table[v] = {}
            fvalue_table[v] = {}
            precision = []
            recall = []
            fvalue = []
            for p in pairs:
                pair = p[0]+' X '+p[1]

                if type(results[p][v]['Precision']) == float:
                    precision_table[v][pair] = "{:4.3f}".format(results[p][v]['Precision'])
                    precision.append(results[p][v]['Precision'])
                    precision_total.append(results[p][v]['Precision'])
                else:
                    precision_table[v][pair] = ''

                if type(results[p][v]['Recall']) == float:
                    recall_table[v][pair] = "{:4.3f}".format(results[p][v]['Recall'])
                    recall.append(results[p][v]['Recall'])
                    recall_total.append(results[p][v]['Recall'])
                else:
                    recall_table[v][pair] = ''

                if type(results[p][v]['F_Measure']) == float:
                    fvalue_table[v][pair] = "{:4.3f}".format(results[p][v]['F_Measure'])
                    fvalue.append(results[p][v]['F_Measure'])
                    fvalue_total.append(results[p][v]['F_Measure'])
                else:
                    fvalue_table[v][pair] = ''

            if len(precision) > 0:
                results[v]['Precision'] = "{:4.3f}".format(sum(precision)/len(precision))
                nresults[v]['Precision'] = sum(precision)/len(precision)
            else:
                results[v]['Precision'] = ''
                nresults[v]['Precision'] = ''

            if len(recall) > 0:
                results[v]['Recall'] = "{:4.3f}".format(sum(recall)/len(recall))
                nresults[v]['Recall'] = sum(recall)/len(recall)
            else:
                results[v]['Recall'] = ''
                nresults[v]['Recall'] = ''

            if len(fvalue) > 0:
                results[v]['F_Measure'] = "{:4.3f}".format(sum(fvalue)/len(fvalue))
                nresults[v]['F_Measure'] = sum(fvalue)/len(fvalue)
            else:
                results[v]['F_Measure'] = ''
                nresults[v]['F_Measure'] = ''


        if len(precision_total) > 0:
            results['Precision'] = "{:4.3f}".format(sum(precision_total)/len(precision_total))
        else:
            results['Precision'] = 'No valid comparisons'
        text_table = text_table + "\n\nPrecision, compared to core coder: "+results['Precision']+display_table(precision_table,sep='tab')

        if len(recall_total) > 0:
            results['Recall'] = "{:4.3f}".format(sum(recall_total)/len(recall_total))
        else:
            results['Recall'] = 'No valid comparisons'
        text_table = text_table + "\n\nRecall, compared to core coder: "+results['Recall']+display_table(recall_table,sep='tab')

        if len(fvalue_total) > 0:
            results['F_Measure'] = "{:4.3f}".format(sum(fvalue_total)/len(fvalue_total))
        else:
            results['F_Measure'] = 'No valid comparisons'
        text_table = text_table + "\n\nF-Measure, compared to core coder: "+results['F_Measure']+display_table(fvalue_table,sep='tab')

        verbout('\n\nSummary of Precision, Recall, and F-Measure comparisons with core coder:\n',master=master)

        pcomp = ['Precision','Recall','F_Measure']

        total_pairs = {}
        for coeff in pcomp:
            total_pairs[coeff] = {}
            for p in pairs:
                pair = p[0]+' X '+p[1]
                total_pairs[coeff][pair] = results[p][coeff]

        verbout(display_table(total_pairs)+'\n','table',master=master)
        text_table = text_table + '\n\n\n\nSum of PRF with core coder:'+display_table(total_pairs,sep='tab')


    if 'Lotus' in methods or 'SLotus' in methods:
        verbout('\nCalculating Lotus...',master=master)
        results['Lotus_c']={}
        results['SLotus_c']={}
    
        uinfo = {}
        for u in units:
            uinfo[u] = {}
            uinfo[u]['Modi']={}
            if kerncod == 'no_core':
                for v in tvars:
                    values = {}
                    for c in codings[u].keys():
                        if not codings[u][c][v] in values.keys():
                            values[codings[u][c][v]] = 0
                        values[codings[u][c][v]] = values[codings[u][c][v]] + 1
                    if '' in values.keys():
                        del values['']

                    mod = ['invalid',0]
                    for val in values.keys():
                        if values[val] > mod[1]:
                            mod = [val,values[val]]
                        elif values[val] == mod[1]:
                            mod = ['invalid',0]
                    uinfo[u]['Modi'][v] = mod

            else:
                for v in tvars:
                    try:
                        uinfo[u]['Modi'][v] = [codings[u][kerncod][v],1]
                    except:
                        uinfo[u]['Modi'][v] = ['invalid',0]

        total_coders = {}
        total_coders['Lotus_c'] = {}
        total_coders['SLotus_c'] = {}
        lotus_total = []
        slotus_total = []
        for c in coders:
            lotuslist = []
            slotuslist = []
            results['Lotus_c'][c] = {}
            results['SLotus_c'][c] = {}
            for v in tvars:
                cats = 1.0/len(vinfo[v]['Values'].keys())
                hit = 0
                c_count = 0
                for u in units:
                    if c in codings[u].keys():
                        if 'cm' in options or not codings[u][c][v] == '':
                            if codings[u][c][v] == uinfo[u]['Modi'][v][0]:
                                hit = hit + 1                           
                            c_count = c_count + 1

                if c_count > min_c_count:
                    lotus = float(hit)/c_count
                    results['Lotus_c'][c][v] = "{:4.3f}".format(lotus)
                    lotuslist.append(lotus)
                    if cats < 1:
                        slotus = (lotus-cats)/(1-cats)
                        results['SLotus_c'][c][v] = "{:4.3f}".format(slotus)
                        slotuslist.append(slotus)
                    else:
                        results['SLotus_c'][c][v] = ""
                else:
                    results['Lotus_c'][c][v] = ''
                    results['SLotus_c'][c][v] = ''

            if len(lotuslist) > 0:
                total_coders['Lotus_c'][c] = "{:4.3f}".format(sum(lotuslist)/len(lotuslist))
                lotus_total.append(sum(lotuslist)/len(lotuslist))
            else:
                total_coders['Lotus_c'][c] = ''

            if len(slotuslist) > 0:
                total_coders['SLotus_c'][c] = "{:4.3f}".format(sum(slotuslist)/len(slotuslist))
                slotus_total.append(sum(slotuslist)/len(slotuslist))
            else:
                total_coders['SLotus_c'][c] = ''

        if len(lotus_total) > 0:
            results['Lotus'] = "{:4.3f}".format(sum(lotus_total)/len(lotus_total))

        else:
            results['Lotus'] = 'No valid lotus for any coder'
           
        if len(slotus_total) > 0:
            results['SLotus'] = "{:4.3f}".format(sum(slotus_total)/len(slotus_total))
        else:
            results['SLotus'] = 'No valid standardized lotus for any coder'


        verbout('\nFretwurst Lotus for all Coders: '+results['Lotus'] + ' ('+results['SLotus']+')\n',master=master)
        verbout(display_table(total_coders)+'\n','table',master=master)
        text_table = text_table + '\n\n\nFretwurst Lotus: '+results['Lotus']+display_table(results['Lotus_c'],sep='tab')
        text_table = text_table + '\n\n\nStandardized Lotus: '+results['SLotus']+display_table(results['SLotus_c'],sep='tab')

        text_table = text_table + '\n\nMean Fretwurst Lotus for all Coders: '+display_table(total_coders,sep='tab')  


        for v in tvars:
            cats = 1.0/len(vinfo[v]['Values'].keys())
            results[v]['Lotus'] = {}
            lotuslist = []
            for c in coders:
                if not results['Lotus_c'][c][v] == '':
                    lotuslist.append(float(results['Lotus_c'][c][v]))
                    
            if len(lotuslist) > 0:
                lotus = float(sum(lotuslist))/len(lotuslist)
                results[v]['Lotus'] = "{:4.3f}".format(lotus)
                nresults[v]['Lotus'] = lotus
                if cats < 1:
                    slotus = (lotus-cats)/(1-cats)
                    results[v]['SLotus'] = "{:4.3f}".format(slotus)
                    nresults[v]['SLotus'] = slotus
                else:
                    results[v]['SLotus'] = ""
                    nresults[v]['SLotus'] = ""
            else:
                results[v]['Lotus'] = ''
                results[v]['SLotus'] = ''
                nresults[v]['Lotus'] = ''
                nresults[v]['SLotus'] = ''


    if 'Alpha Nominal' in methods:
        verbout('\n\nCalculating Krippendorff Alpha for nominal variables...',master=master)
        n_codings = {}
        n_dist = {}
        n_alpha = {}
        n_alpha_total = []
        verbout('\nReshaping nominal codings...',master=master)
        for v in tvars:
            verbout('\n..for variable: "'+v+'"',master=master)
            n_codings[v] = []
            for u in codings.keys():
                for c in codings[u].keys():
                    if c in coders and u in units:
                        if not codings[u][c][v] == '' or 'cm' in options:
                            n_codings[v].append(codings[u][c][v])

            verbout("(N="+str(len(n_codings[v]))+")",master=master)
            if len(n_codings[v]) > 1000:
                n_codings[v] = random.sample(n_codings[v],1000)
                verbout('*',master=master)
                        
            dist = 0.0
            anz = 0
            for i1 in range(len(n_codings[v])):
                for i2 in range(len(n_codings[v])):
                    anz = anz + 1
                    if not n_codings[v][i1] == n_codings[v][i2]:
                        dist = dist + 1
            if anz > 0:
                n_dist[v] = dist/anz
            else:
                n_dist[v] = 0.0

            verbout(' -> mean nominal distance = '+str(n_dist[v]),master=master)

        verbout('\n\nNominal codings recorded, calculating pairwise...',master=master)


        for v in tvars:
            verbout('\n..for variable: "'+v+'"',master=master)
            dist = []
            if kerncod == 'no_core':
                coders1 = coders
            else:
                coders1 = [kerncod]
            for u in units:
                for c1 in coders1:
                    for c2 in coders:
                        if c1 in codings[u].keys() and c2 in codings[u].keys() and not c1==c2:
                            if 'cm' in options or not '' in [codings[u][c1][v],codings[u][c2][v]]:
                                if codings[u][c1][v] == codings[u][c2][v]:
                                    dist.append(0)
                                else:
                                    dist.append(1)

            if len(dist) > 0 and n_dist[v] > 0:
                odist = float(sum(dist))/len(dist)
                n_alpha[v] = 1-odist/n_dist[v]
                verbout(": Dist="+str(sum(dist))+" N="+str(len(dist))+" Rdist="+str(n_dist[v])+" Alpha="+str(n_alpha[v]),master=master)
                results[v]['Alpha Nominal'] = "{:4.3f}".format(n_alpha[v])
                nresults[v]['Alpha Nominal'] = n_alpha[v]
                n_alpha_total.append(n_alpha[v])
            else:
                n_alpha[v] = ''
                results[v]['Alpha Nominal'] = ""
                nresults[v]['Alpha Nominal'] = ""

        if len(n_alpha_total) > 0:
            results['Alpha Nominal'] = "{:4.3f}".format(sum(n_alpha_total)/len(n_alpha_total))
        else:
            results['Alpha Nominal'] = "No valid variables"

        if 'Alpha Nominal' in methods:
            text_table = text_table + "\n\nKrippendorff's Alpha (nominal):"+results['Alpha Nominal']

        #text_table = text_table + "\n\n\nKrippendorff's Alpha (nominal):\n"+baum_schreiben(n_alpha)     
        
    if 'Alpha Ordinal' in methods:
        verbout('\n\n\nCalculating Krippendorff Alpha for ordinal variables...',master=master)

        rangliste = {}

        #Alte n_codings Liste aus den metrischen nehmen ist schlechte Idee. Darum auf jeden Fall eine neue Rangliste erstellen.
        for v in tvars:
            tlist = []
            for u in codings.keys():
                for c in codings[u].keys():
                    if c in coders and u in units:
                        if not codings[u][c][v] == '' or 'cm' in options:
                            tlist.append(codings[u][c][v])
            rangliste[v] = get_unique(tlist)


        n_codings = {}
        o_dist = {}
        o_alpha = {}
        o_rank = {}
        o_nrank = {}
        o_alpha_total = []
        verbout('\n\nReshaping ordinal codings...',master=master)
        for v in tvars:
            verbout('\n..for variable: "'+v+'"',master=master)
            o_rank[v] = {}
            o_nrank[v] = {}
            n_codings[v] = []
            for i in range(0,len(rangliste[v])):
                o_rank[v][rangliste[v][i]] = i+1
                o_nrank[v][i+1] = 0

            for u in codings.keys():
                for c in codings[u].keys():
                    if c in coders and u in units:
                        n = codings[u][c][v]
                        if not n == '' or 'cm' in options:
                            n_codings[v].append(o_rank[v][n])
                            o_nrank[v][o_rank[v][n]] = o_nrank[v][o_rank[v][n]]+1

            if len(n_codings[v]) > 1000:
                n_codings[v] = random.sample(n_codings[v],1000)
                verbout('*',master=master)                

            dist = 0.0
            anz = 0
            for i1 in range(len(n_codings[v])):
                for i2 in range(len(n_codings[v])):
                    anz = anz + 1
                    d_sum = 0
                    c1 = n_codings[v][i1]
                    c2 = n_codings[v][i2]
                    if c1 == c2:
                        d_sum = 0
                    else:
                        if c1 < c2:
                            minc = c1
                            maxc = c2
                        else:
                            minc = c2
                            maxc = c1
                        for g in range(minc,maxc+1):
                            d_sum = d_sum + o_nrank[v][g]
                        d_sum = d_sum-(o_nrank[v][minc]+o_nrank[v][maxc])/2
                    dist = dist + d_sum**2
                    
            if anz > 0:
                o_dist[v] = dist/anz
            else:
                o_dist[v] = 0.0


            verbout(' -> mean ordinal distance = '+str(o_dist[v]),master=master)

        verbout('\n\nOrdinal codings recorded, calculating pairwise...',master=master)

        for v in tvars:
            verbout('\n..for variable: "'+v+'"',master=master)
            dist = []
            if kerncod == 'no_core':
                coders1 = coders
            else:
                coders1 = [kerncod]
            for c1 in coders1:
                for c2 in coders:
                    for u in units:
                        if c1 in codings[u].keys() and c2 in codings[u].keys()  and not c1==c2:
                            d_sum = 0
                            code1 = codings[u][c1][v]
                            code2 = codings[u][c2][v]
                            if not code1 == code2 and code1 in o_rank[v].keys() and code2 in o_rank[v].keys():
                                if code1 < code2:
                                    minc = o_rank[v][code1]
                                    maxc = o_rank[v][code2]
                                else:
                                    minc = o_rank[v][code2]
                                    maxc = o_rank[v][code1]
                                for g in range(minc,maxc+1):
                                    d_sum = d_sum + o_nrank[v][g]
                                d_sum = d_sum -(o_nrank[v][minc]+o_nrank[v][maxc])/2
                                
                            if code1 == '' or code2 == '':
                                if 'cm' in options:
                                    dist.append(d_sum**2)
                            else:
                                dist.append(d_sum**2)
                  
            if len(dist) > 0 and o_dist[v] > 0:
                odist = float(sum(dist))/len(dist)
                o_alpha[v] = 1-odist/o_dist[v]
                results[v]['Alpha Ordinal'] = "{:4.3f}".format(o_alpha[v])
                nresults[v]['Alpha Ordinal'] = o_alpha[v]
                o_alpha_total.append(o_alpha[v])
            else:
                o_alpha[v] = ''
                results[v]['Alpha Ordinal'] = ""
                nresults[v]['Alpha Ordinal'] = ""

        if len(o_alpha_total) > 0:
            results['Alpha Ordinal'] = "{:4.3f}".format(sum(o_alpha_total)/len(o_alpha_total))
        else:
            results['Alpha Ordinal'] = "No valid variables"

        text_table = text_table + "\n\nKrippendorff's Alpha (ordinal):"+results['Alpha Ordinal']

        #text_table = text_table + "\n\n\nKrippendorff's Alpha (ordinal):\n"+baum_schreiben(o_alpha)  

    if 'Alpha Metric' in methods:
        verbout('\n\n\nCalculating Krippendorff Alpha for metric variables...',master=master)               

        n_codings = {}
        m_dist = {}
        m_alpha = {}
        m_alpha_total = []
        verbout('\n\nReshaping metric codings...',master=master)
        for v in tvars:
            verbout('\n..for variable: "'+v+'"',master=master)
            n_codings[v] = []
            for u in codings.keys():
                for c in codings[u].keys():
                    if c in coders and u in units:
                        n = codings[u][c][v]
                        try:
                            n_codings[v].append(float(n))
                        except:
                            n = 0

            if len(n_codings[v]) > 1000:
                n_codings[v] = random.sample(n_codings[v],1000)
                verbout('*',master=master)
                
            dist = []
            for i1 in range(len(n_codings[v])):
                for i2 in range(len(n_codings[v])):
                    dist.append((n_codings[v][i1]-n_codings[v][i2])**2)
                    
            if len(dist) > 0:
                m_dist[v] = float(sum(dist))/len(dist)
            else:
                m_dist[v] = 0.0

            verbout(' -> mean ordinal distance = '+str(m_dist[v]),master=master)

        verbout('\n\nMetric codings recorded, calculating pairwise...',master=master)


        for v in tvars:
            verbout('\n..for variable: "'+v+'"',master=master)
            dist = []
            if kerncod == 'no_core':
                coders1 = coders
            else:
                coders1 = [kerncod]
            for c1 in coders1:
                for c2 in coders:
                    for u in units:
                        if c1 in codings[u].keys() and c2 in codings[u].keys() and not c1==c2:
                            try:
                                val1 = float(codings[u][c1][v])
                                val2 = float(codings[u][c2][v])
                                dist.append((val1-val2)**2)
                            except:
                                catch = 0
                                
            if len(dist) > 0 and m_dist[v] > 0:
                odist = float(sum(dist))/len(dist)
                m_alpha[v] = 1-odist/m_dist[v]
                results[v]['Alpha Metric'] = "{:4.3f}".format(m_alpha[v])
                nresults[v]['Alpha Metric'] = m_alpha[v]
                m_alpha_total.append(m_alpha[v])
            else:
                m_alpha[v] = ''
                results[v]['Alpha Metric'] = ""
                nresults[v]['Alpha Metric'] = ""

        if len(m_alpha_total) > 0:
            results['Alpha Metric'] = "{:4.3f}".format(sum(m_alpha_total)/len(m_alpha_total))
        else:
            results['Alpha Metric'] = "No valid variables"

        text_table = text_table + "\n\nKrippendorff's Alpha (metric):"+results['Alpha Metric']
        
        #text_table = text_table + "\n\n\nKrippendorff's Alpha (metric):\n"+baum_schreiben(m_alpha)  


    pcomp = []
    for m in ['PA','Kappa','Kappan','Pi','Lotus','SLotus','Alpha Nominal','Alpha Ordinal','Alpha Metric','PA_w_core','Kappa_w_core','Kappan_w_core','Pi_w_core']:
        if m in methods:
            pcomp.append(m)
    if 'PRF' in methods:
        pcomp = pcomp + ['Precision','Recall','F_Measure']

    total_vars = {}
    for coeff in pcomp:
        total_vars[coeff] = {}
        for v in tvars:
            if coeff in results[v].keys():
                total_vars[coeff][v] = results[v][coeff]
            else:
                total_vars[coeff][v] = '-'

    verbout('\n\n',master=master)
    verbout('Summary of Results','title',master=master)
    verbout('\n',master=master)
    verbout(display_table(total_vars, cols_pre=pcomp)+'\n','table',master=master)

    text_table = text_table + '\n\n\nSum of agreement for variables:'+display_table(total_vars,sep='tab')

    text_table = text_table + '\n\n\nSummary of coefficients:'
    verbout('\n\n\nSummary of coefficients:\n',master=master)

    shortrun = display_table(total_vars,sep='tab')+'\n'

    numres = {}

    for m in methods:
        if m in results.keys():
            text_table = text_table + '\n'+m+':\t'+results[m]
            shortrun = shortrun + '\n'+m+':\t'+results[m]
            numres[m] = results[m]
            verbout('\n'+m+': '+results[m],'table',master=master)
        elif m == 'PRF':
            for c in ['Precision','Recall','F_Measure']:
                text_table = text_table + '\n'+c+':\t'+results[c]
                shortrun = shortrun + '\n'+c+':\t'+results[c]
                numres[c] = results[c]
                verbout('\n'+c+': '+results[c],'table',master=master)

    verbout('\n','table',master=master)

    legend = '\n\n\nLegend of coefficients:\n'
    legend = legend + '\nCoefficient\tLabel\tDescription\tReferences'
    legend = legend + '\nPA\tPercent Agreement/Holsti Coefficient\tPercent of cases in which coders agree\t[5]'
    legend = legend + '\nAlpha Nominal\tKrippendorff Alpha for nominal variables\tAgreement corrected for chance using all data\t[5][6]'
    legend = legend + '\nAlpha Ordinal\tKrippendorff Alpha for ordinal variables\tAgreement corrected for chance using all data\t[4][5]'
    legend = legend + '\nAlpha Metric\tKrippendorff Alpha for ordinal variables\tAgreement corrected for chance using all data\t[4][5]'
    legend = legend + "\nKappa\tCohen's Kappa\tAgreement corrected for chance using pairwise empirical margins\t[2]"
    legend = legend + "\nKappan\tBrennan & Prediger's Kappa\tAgreement corrected for chance using theoretical margins\t[1]"
    legend = legend + "\nPi\tScott's Pi\tAgreement corrected for chance using empirical margins for all coders\t[7]"
    legend = legend + "\nLotus\tFretwurst Lotus\tAgreement of coders on a common modus per decision\t[3]"
    legend = legend + "\nSLotus\tFretwurst Lotus\tAgreement of coders on a common modus, corrected for number of possible decisions\t[3]"
    legend = legend + "\nPrecision\tPrecision\tShare of correct positive choices (high = few false positives)\t"
    legend = legend + "\nRecall\tRecall\tShare of true positives found (high = few false negatives)\t"
    legend = legend + "\nF_Measure\tF-Measure\tHarmonic mean of Precision and Recall\t"
    legend = legend + "\n\nReferences:"        
    legend = legend + "\n[1] Brennan, R. L., & Prediger, D. J. (1981). Coefficient Kappa: Some Uses, Misuses, and Alternatives. Educational and Psychological Measurement, 41(3), 687-699."
    legend = legend + "\n[2] Cohen, J. (1960). A Coefficient of Agreement for Nominal Scales. Educational and Psychological Measurement, 20(1), 37-46."
    legend = legend + "\n[3] Fretwurst, B. (2015). Reliabilität und Validität von Inhaltsanalysen: Mit Erläuterungen zur Berechnung des Reliabilitätskoeffizienten Lotus mit SPSS. In W. Wirth, K. Sommer, M. Wettstein, & J. Matthes (Eds.), Methoden und Forschungslogik der Kommunikationswissenschaft: Vol. 12. Qualitätskriterien in der Inhaltsanalyse (pp. 177-204). Köln: Herbert von Halem Verlag."
    legend = legend + "\n[4] Krippendorff, K. (1970). Estimating the Reliability, Systematic Error and Random Error of Interval Data. Educational and Psychological Measurement, 30(1), 61-70."
    legend = legend + "\n[5] Krippendorff, K. (2004). Content analysis: An introduction to its methodology. Thousand Oaks, Calif: Sage Publ."
    legend = legend + "\n[6] Krippendorff, K. (2008). Systematic and Random Disagreement and the Reliability of Nominal Data. Communication Methods and Measures, 2(4), 323-338."
    legend = legend + "\n[7] Scott, W. A. (1955). Reliability of Content Analysis: The Case of Nominal Scale Coding. Public Opinion Quarterly, 19(3), 323-325."

    text_table = text_table + legend
                
    return (shortrun,text_table,numres,nresults)













##########################
##
## NCCR
## Add Populusm
## And Linkage Analysis / Matching Content and Survey
##
##########################


def add_populism(data,master=''):

    data['STRAT_Blame'] = []
    data['STRAT_Achiev'] = []
    data['STRAT_Sovereign_Pro'] = []
    data['STRAT_Sovereign_Con'] = []
    
    data['Tone_Pos_Grp'] = []
    data['Tone_Pos_Mean'] = []
    data['POP_Blame'] = []
    data['POP_Achiev'] = []
    data['POP_Denouncing'] = []
    data['POP_Virtues'] = []
    data['POP_Exclusion_Elite'] = []
    data['POP_Exclusion_Someone'] = []
    data['POP_Closeness_Self'] = []
    data['POP_Closeness_Someone'] = []
    data['POP_Sovereign_adv'] = []
    data['POP_Sovereign_con'] = []
    data['POP_Monolith'] = []

    data['POPULIST'] = []
    data['POPULIST_Advocative'] = [] #pop_ad
    data['POPULIST_Conflictive'] = [] #pop_con
    data['POPULIST_PeopleCent'] = [] #pop1
    data['POPULIST_AntiElite'] = [] #pop2
    data['POPULIST_Sovereign'] = [] #pop3


    data['APOP_Blame'] = []
    data['APOP_Achiev'] = []
    data['APOP_Denouncing'] = []
    data['APOP_Virtues'] = []
    data['APOP_Closeness_Elite'] = []
    data['APOP_Sovereign_adv'] = []
    data['APOP_Sovereign_con'] = []
    data['APOP_Pluralist'] = []

    data['ANTIPOPULIST'] = [] #apop
    data['APOPULIST_Advocative'] = [] #apop_ad
    data['APOPULIST_Conflictive'] = [] #apop_con
    data['APOPULIST_PeopleCent'] = [] #apop1
    data['APOPULIST_AntiElite'] = [] #apop2
    data['APOPULIST_Sovereign'] = [] #apop3
    
    data['POPSHARE_Advocative'] = []
    data['POPSHARE_Conflictive'] = []
    data['POPSHARE_PeopleCent'] = []
    data['POPSHARE_AntiElite'] = []
    data['POPSHARE_Sovereign'] = []

    data['POPULIST_BIAS'] = []
    data['POPULIST_PC_BIAS'] = []
    data['POPULIST_AE_BIAS'] = []
    data['POPULIST_PS_BIAS'] = []
    
    data['Filter_Pop_possible'] = []


### Populismus im erweiterten Sinn (erweiterte Elite, erweitertes Volk)

    data['POP_Blame_ext'] = []
    data['POP_Achiev_ext'] = []
    data['POP_Denouncing_ext'] = []
    data['POP_Virtues_ext'] = []
    data['POP_Exclusion_Elite_ext'] = []
    data['POP_Exclusion_Someone_ext'] = []
    data['POP_Closeness_Self_ext'] = []
    data['POP_Closeness_Someone_ext'] = []
    data['POP_Sovereign_adv_ext'] = []
    data['POP_Sovereign_con_ext'] = []
    data['POP_Monolith_ext'] = []

    data['POPULIST_ext'] = []
    data['POPULIST_Advocative_ext'] = [] #pop_ad
    data['POPULIST_Conflictive_ext'] = [] #pop_con
    data['POPULIST_PeopleCent_ext'] = [] #pop1
    data['POPULIST_AntiElite_ext'] = [] #pop2
    data['POPULIST_Sovereign_ext'] = [] #pop3


    data['APOP_Blame_ext'] = []
    data['APOP_Achiev_ext'] = []
    data['APOP_Denouncing_ext'] = []
    data['APOP_Virtues_ext'] = []
    data['APOP_Closeness_Elite_ext'] = []
    data['APOP_Sovereign_adv_ext'] = []
    data['APOP_Sovereign_con_ext'] = []
    data['APOP_Pluralist_ext'] = []

    data['ANTIPOPULIST_ext'] = [] #apop
    data['APOPULIST_Advocative_ext'] = [] #apop_ad
    data['APOPULIST_Conflictive_ext'] = [] #apop_con
    data['APOPULIST_PeopleCent_ext'] = [] #apop1
    data['APOPULIST_AntiElite_ext'] = [] #apop2
    data['APOPULIST_Sovereign_ext'] = [] #apop3
    
    data['POPSHARE_Advocative_ext'] = []
    data['POPSHARE_Conflictive_ext'] = []
    data['POPSHARE_PeopleCent_ext'] = []
    data['POPSHARE_AntiElite_ext'] = []
    data['POPSHARE_Sovereign_ext'] = []

    data['POPULIST_BIAS_ext'] = []
    data['POPULIST_PC_BIAS_ext'] = []
    data['POPULIST_AE_BIAS_ext'] = []
    data['POPULIST_PS_BIAS_ext'] = []
    
    data['Filter_Pop_possible_ext'] = []

### End of extended populism

    data['Migrants_Neg'] = []
    data['Migrants_Pos'] = []
    data['Migrants_Mentioned'] = []
    data['LW_Attack'] = []
    data['Spr_Category'] = []
    data['Tgt_Category'] = []
    data['Spr_Orga']=[]
    data['Tgt_Orga']=[]
    data['Spr_Group']=[]
    data['Tgt_Group']=[]
    data['Tgt_Elite'] = []
    data['Tgt_People'] = []
    data['St_On_People'] = []
    data['St_On_People_ext'] = []
    data['St_On_Power'] = []
    data['Filter_Auto'] = []

    all_vars = ['Coder','ID','Level','Unit_ID','Level01','Unit_ID01','#TS','Spr_ID','Auto_Coding',
                'Tgt_ID','Def_Actor','Def_Volk','Def_Elit','Def_ForC','Def_MPer','Def_Othr','Def_OwnP',
                'Embod','Monolith','Distance','Iss_Link','Iss_Link_Pos','Agreement','Att_Pos_good',
                'Att_Pos_char','Att_Pos_comm','Att_Pos_cred','Att_Pos_lead','Att_Pos_cons','Att_Pos_oth',
                'Att_Neg_malev','Att_Neg_crim','Att_Neg_lazy','Att_Neg_stu','Att_Neg_pop','Att_Neg_right',
                'Att_Neg_left','Att_Neg_ext','Att_Neg_raci','Att_Neg_unde','Att_Neg_oth','Att_Impact_thre',
                'Att_Impact_burd','Att_Impact_enri','Att_Impact_aneg','Att_Impact_apos','Att_Impact_abil',
                'Impact_Tgt','Att_People_belo','Att_People_clos','Att_People_know','Att_People_care',
                'Att_People_beha','Att_People_pow','Att_People_dece','Att_Act_every','Att_Act_symb',
                'Att_Act_mist','Att_Act_right','Att_Act_immo','Att_Act_crim','Att_Act_dem','Att_Act_prom',
                'Att_Act_plan','Att_Act_other','Att_Power_gain','Att_Power_lose','Att_Power_have',
                'Privat_Fam','Privat_Pas','Privat_Lei','Privat_Lov','PrivAtt','Namecall','Stereo',
                'Sourcing_1','Sourcing_2','Sourcing_3','Sourcing_4','Sourcing_5','Sourcing_6','Sourcing_7',
                'Sourcing_8','Sourcing_9','Sourcing_99','Rhetoric_abs','Rhetoric_imm','Rhetoric_sarc',
                'Rhetoric_exagg','Rhetoric_war','Rhetoric_emerg','Rhetoric_scand','Rhetoric_quest',
                'Rhetoric_patri','V_Cues_gest','V_Cues_action','V_Cues_smile','V_Cues_sad','V_Cues_anger',
                'Emot_anger','Emot_uneasiness','Emot_happiness','Emot_contentment','Emot_fear','Emot_contempt',
                'Emot_sadness','Emot_regret','Emot_affection','Emot_surprise','Emot_hope','Emot_pride',
                'Emot_trust','STYLE_Facts','STYLE_Sense','STYLE_BlackWhite','STYLE_Sarcasm','STYLE_Drama',
                'STYLE_EmoTone','STYLE_CommMan','STYLE_UsThem','STYLE_Privat','STRAT_ShiftingBlame',
                'STRAT_Closeness','STRAT_Exclusion','STRAT_Virtues','STRAT_Denouncing','STRAT_Sovereignty',
                'STRAT_Monolith']
    
    pos_att = ['Att_Pos_good','Att_Pos_char','Att_Pos_comm','Att_Pos_cred','Att_Pos_lead','Att_Pos_cons','Att_Pos_oth',
               'Att_Impact_enri','Att_Impact_apos','Att_Impact_abil','Att_Act_right']
    neg_att = ['Att_Neg_malev','Att_Neg_crim','Att_Neg_lazy','Att_Neg_stu','Att_Neg_ext','Att_Neg_raci',
               'Att_Neg_unde','Att_Neg_oth','Att_Impact_thre','Att_Impact_burd','Att_Impact_aneg','Att_Act_mist',
               'Att_Act_immo','Att_Act_crim']

    virtue_vars = ['Att_Pos_good','Att_Pos_char','Att_Pos_comm','Att_Pos_cred','Att_Pos_lead',
                   'Att_Pos_cons','Att_Pos_oth']
    denounc_vars = ['Att_Neg_malev','Att_Neg_crim','Att_Neg_lazy','Att_Neg_stu',
                    'Att_Neg_ext','Att_Neg_raci','Att_Neg_unde','Att_Neg_oth']
    achiev_vars = ['Att_Impact_enri','Att_Impact_apos','Att_Impact_abil','Att_Act_right']
    blame_vars = ['Att_Impact_thre','Att_Impact_burd','Att_Impact_aneg','Att_Act_mist','Att_Act_crim','Att_Act_immo']
    close_vars = ['Att_People_belo','Att_People_clos','Att_People_know','Att_People_care',
                'Att_People_beha','Att_People_pow','Att_Act_every']
    excl_vars = ['Att_People_dece']

    power_vars = ['Att_Power_gain','Att_Power_have']
    nopower_vars = ['Att_Power_lose']

    for i in range(len(data['ID'])):
        data['STRAT_Blame'].append('')
        data['STRAT_Achiev'].append('')
        data['STRAT_Sovereign_Pro'].append('')
        data['STRAT_Sovereign_Con'].append('')
    


    step = int(len(data['ID'])/40)
    if step<1: step = 1
    verbout('\nCalculating Additional Variables for each case: \n0%-------25%-------50%-------75%-------100%\n',master=master)

    for i in range(len(data['ID'])):
        if i%step == 0: verbout('.',master=master)
        tonepos = 0
        toneneg = 0
        tonesum = 0
        for v in pos_att:
            if data[v][i] == '1':
                tonepos = tonepos + 1
                tonesum = tonesum + 1
            elif data[v][i] == '-1':
                toneneg = toneneg + 1
                tonesum = tonesum + 1
        for v in neg_att:
            if data[v][i] == '1':
                toneneg = toneneg + 1
                tonesum = tonesum + 1
            elif data[v][i] == '-1':
                tonepos = tonepos + 1
                tonesum = tonesum + 1

        if tonesum > 0:
            tonemean = float(tonepos - toneneg)/tonesum
            data['Tone_Pos_Mean'].append(tonemean)
        else:
            data['Tone_Pos_Mean'].append(0)

        if tonepos > toneneg:
            data['Tone_Pos_Grp'].append(1)
        elif tonepos < toneneg:
            data['Tone_Pos_Grp'].append(-1)
        else:
            data['Tone_Pos_Grp'].append(0)


#### Defining all Strategies irregardless of target



        data['STRAT_Blame'][i] = '0'
        for v in blame_vars:
            if data[v][i] == '1':
                data['STRAT_Blame'][i] = '1'
        for v in achiev_vars:
            if data[v][i] == '-1':
                data['STRAT_Blame'][i] = '1'
            
        data['STRAT_Achiev'][i] = '0'
        for v in blame_vars:
            if data[v][i] == '-1':
                data['STRAT_Achiev'][i] = '1'
        for v in achiev_vars:
            if data[v][i] == '1':
                data['STRAT_Achiev'][i] = '1'


        data['STRAT_Virtues'][i] = '0'
        for v in virtue_vars:
            if data[v][i] == '1':
                data['STRAT_Virtues'][i] = '1'
        for v in denounc_vars:
            if data[v][i] == '-1':
                data['STRAT_Virtues'][i] = '1'
        
        data['STRAT_Denouncing'][i] = '0'
        for v in virtue_vars:
            if data[v][i] == '-1':
                data['STRAT_Denouncing'][i] = '1'
        for v in denounc_vars:
            if data[v][i] == '1':
                data['STRAT_Denouncing'][i] = '1'

        if data['Namecall'][i] == '1':
            data['STRAT_Virtues'][i] = '1'
        elif data['Namecall'][i] == '-1':
            data['STRAT_Denouncing'][i] = '1'

        
        data['STRAT_Closeness'][i] = '0'
        for v in close_vars:
            if data[v][i] == '1':
                data['STRAT_Closeness'][i] = '1'
        for v in excl_vars:
            if data[v][i] == '-1':
                data['STRAT_Closeness'][i] = '1'

        data['STRAT_Exclusion'][i] = '0'
        for v in close_vars:
            if data[v][i] == '-1':
                data['STRAT_Exclusion'][i] = '1'
        for v in excl_vars:
            if data[v][i] == '1':
                data['STRAT_Exclusion'][i] = '1'

        
        data['STRAT_Sovereign_Pro'][i] = '0'
        for v in power_vars:
            if data[v][i] == '2':
                data['STRAT_Sovereign_Pro'][i] = '1'
        for v in nopower_vars:
            if data[v][i] == '-2':
                data['STRAT_Sovereign_Pro'][i] = '1'

        data['STRAT_Sovereign_Con'][i] = '0'
        for v in power_vars:
            if data[v][i] == '-2':
                data['STRAT_Sovereign_Con'][i] = '1'
        for v in nopower_vars:
            if data[v][i] == '2':
                data['STRAT_Sovereign_Con'][i] = '1'


        ##Namentliche Akteure
        ##1: Oberste Regierungsmitglieder (Bundesrat, (Bundes)Präsident/-kanzler, Ministerpräsident, Premierminister, Staatspräsident)
        ##2: Ehem. Oberste Regierungsmitglieder (Bundesrat, (Bundes)Präsident/-kanzler, Premierminister, Staatspräsident)
        ##3: Kabinettmitglieder (Minister, Staatssekretär)
        ##4: Parlamentsmitglieder (National-/Ständerat, Bundestagsmitglied, Senat, Nationalversammlung, Abgeordnetenkammer, House of Commons/Lords, Repräsentantenhaus,
        ##5: Politiker (Ohne wichtiges Amt auf staatlicher/internationaler Ebene)
        ##6: Führung einer Partei (Parteipräsident/-vorsitzender/-sprecher/-obmann/-generalsekretär, Senate Leader, House Leader, Fraktionsführender)
        ##7: Amtsinhaber auf europäischer Ebene (Europarat, Europaparlament, Europäische Kommission)
        ##8: Verstorben od. explizit aus Politik zurückgezogen
        ##9: Unbekannt, anderes


        current_fct1 = ['11001', '11069', '12001', '12028', '13001', '13002', '13003', '13004', '13005', '13006', '13016', '15001',
                        '15009', '16001', '16030', '17006', '17007', '17041', '17098', '18071', '18120', '18149', '19001', '20001',
                        '20046', '20048', '20049', '20050', '21001', '21025', '22001', '22031', '22084', '22086', '22093', '23001',
                        '23002', '24001', '24005']


        current_fct2 = ['11062', '11063', '12023', '13007', '13017', '13057', '13058', '13069', '13070', '13071', '13072', '13073',
                        '13074', '13084', '13085', '13087', '13099', '13101', '13131', '13132', '15011', '15063', '15064', '15065',
                        '15066', '15067', '15068', '15069', '15074', '15100', '15102', '17001', '17002', '17003', '17099', '17100',
                        '17104', '17105', '17107', '17108', '18001', '18063', '18064', '18098', '18099', '18102', '18106', '18115',
                        '18116', '18117', '18126', '18127', '18130', '18131', '18132', '19055', '19056', '19069', '19071', '19103',
                        '20032', '20040', '20060', '21005', '21061', '21067', '21071', '21072', '21075', '21076', '22098', '22099',
                        '22108', '22120', '22123', '22126', '22127', '23056', '23077']


        current_fct3 = ['11005', '11007', '11021', '11022', '11025', '11027', '11030', '11033', '12002', '12003', '12004', '12005',
                        '12006', '12007', '12008', '12009', '12010', '12011', '12012', '12013', '12014', '12015', '12016', '12017',
                        '12039', '12046', '12056', '12057', '12058', '12059', '12062', '12063', '15002', '15004', '15008', '15010',
                        '15017', '15022', '15026', '15029', '15044', '15045', '15047', '15054', '16002', '16004', '16005', '16007',
                        '16008', '16012', '16014', '16016', '16017', '16018', '16020', '16021', '16022', '16040', '16041', '16043',
                        '17009', '17016', '17036', '17038', '17040', '17042', '17043', '17044', '17045', '17046', '17047', '17048',
                        '17049', '17050', '17051', '17052', '17053', '17054', '17055', '17056', '17057', '17058', '17059', '17060',
                        '17061', '17062', '17063', '17064', '17065', '17066', '17067', '17068', '17069', '17070', '17092', '17093',
                        '17113', '18002', '18005', '18006', '18007', '18010', '18011', '18012', '18013', '18016', '18019', '18022',
                        '18023', '18024', '18029', '18031', '18032', '18034', '18035', '18037', '18039', '18042', '18044', '18050',
                        '18055', '18056', '18057', '18086', '18136', '18144', '18147', '18150', '18151', '18152', '18153', '18154',
                        '18155', '18162', '18163', '18167', '19002', '19003', '19004', '19005', '19006', '19007', '19008', '19009',
                        '19010', '19011', '19012', '19013', '19014', '19016', '19017', '19018', '19019', '19111', '20002', '20004',
                        '20005', '20006', '20008', '20009', '20010', '20012', '20016', '20017', '20018', '20020', '20059', '20060',
                        '20061', '21002', '21003', '21004', '21005', '21007', '21008', '21009', '21010', '21011', '21012', '21013',
                        '21014', '21015', '21016', '21017', '21018', '21019', '21020', '21021', '21022', '21023', '21024', '21028',
                        '21029', '21039', '21040', '21041', '21042', '21043', '21044', '21045', '21046', '21047', '21048', '21049',
                        '21050', '21051', '21052', '21053', '21054', '21055', '21056', '21057', '21058', '21059', '21094', '22002',
                        '22003', '22004', '22005', '22006', '22007', '22009', '22012', '22013', '22014', '22015', '22018', '22020',
                        '22021', '22022', '22023', '22029', '22032', '22033', '22034', '22035', '22040', '22041', '22042', '22043',
                        '22044', '22046', '22047', '22048', '22049', '22050', '22051', '22052', '22053', '22054', '22055', '22056',
                        '22057', '22058', '22059', '22060', '22061', '22062', '22063', '22064', '22065', '22066', '22067', '22068',
                        '22089', '22092', '22095', '22104', '23003', '23004', '23005', '23006', '23007', '23008', '23009', '23010',
                        '23011', '23012', '23013', '23014', '23015', '23016', '23018', '23019', '23020', '23021', '23022', '23023',
                        '24002', '24003', '24004', '24007', '24012', '24013']


        current_fct4 = ['11002', '11008', '11011', '11012', '11014', '11015', '11017', '11019', '11020', '11028', '11029', '11031',
                        '11032', '11034', '11038', '11039', '11040', '11041', '11043', '11046', '11048', '11058', '11068', '12006',
                        '12018', '12024', '12026', '12029', '12030', '12033', '12034', '12036', '12037', '12051', '12053', '12060',
                        '13008', '13009', '13010', '13011', '13012', '13013', '13018', '13019', '13020', '13021', '13022', '13023',
                        '13026', '13027', '13031', '13032', '13033', '13038', '13040', '13041', '13042', '13043', '13046', '13052',
                        '13053', '13059', '13061', '13064', '13115', '13116', '13123', '13124', '13125', '13126', '13127', '13128',
                        '13129', '13130', '13134', '13135', '13136', '13137', '13138', '13139', '13140', '13141', '13142', '13143',
                        '13144', '13145', '13146', '13147', '13148', '13149', '13150', '13151', '13152', '13153', '13154', '13155',
                        '13156', '13157', '13158', '13159', '13160', '13161', '13162', '13163', '13164', '13165', '13166', '13167',
                        '13168', '13170', '13171', '13172', '13173', '13174', '13175', '13176', '13177', '13178', '13179', '13180',
                        '13181', '13182', '13183', '13185', '13186', '13187', '13188', '13189', '13190', '13191', '13192', '13193',
                        '13194', '13195', '13196', '13197', '13198', '13199', '13200', '13201', '13202', '13203', '13204', '13205',
                        '13206', '13207', '13208', '13209', '13210', '13211', '13212', '13213', '13214', '13215', '13216', '13217',
                        '13218', '13220', '13221', '13222', '13223', '13224', '13226', '13227', '13228', '13229', '13230', '13231',
                        '13232', '13233', '13234', '13235', '13236', '13237', '13238', '13239', '13240', '13241', '13242', '13243',
                        '13244', '13245', '13246', '13247', '13249', '13250', '13251', '13252', '13253', '13254', '13255', '13256',
                        '13257', '13258', '13259', '13260', '13261', '13262', '13263', '13264', '13265', '13266', '13267', '13268',
                        '13269', '13270', '13273', '13274', '13275', '13277', '13278', '13279', '13280', '13281', '13282', '13284',
                        '13285', '13286', '13287', '13288', '13289', '13290', '13291', '13292', '13293', '13294', '13295', '13296',
                        '13297', '13298', '13299', '13300', '13301', '13302', '13303', '13304', '13306', '13307', '13308', '13309',
                        '13310', '13312', '13313', '13314', '13315', '13316', '13317', '13318', '13319', '13320', '13322', '13323',
                        '13324', '13326', '13327', '13329', '13330', '13332', '13334', '15016', '15019', '15027', '15030', '15033',
                        '15038', '15039', '15041', '15042', '15049', '15050', '15053', '15056', '15059', '15084', '15085', '15107',
                        '15109', '15110', '16006', '16009', '16010', '16011', '16013', '16015', '16019', '16024', '16032', '16034',
                        '16036', '16037', '16038', '17004', '17018', '17026', '17027', '17028', '17029', '17032', '17033', '17034',
                        '17035', '17079', '17080', '17082', '17083', '17084', '17087', '17090', '17091', '17097', '17101', '18033',
                        '18036', '18040', '18071', '18079', '18080', '18106', '18138', '18141', '18142', '18145', '18157', '18161',
                        '18165', '18166', '18168', '19018', '19023', '19024', '19027', '19028', '19029', '19030', '19032', '19036',
                        '19038', '19039', '19041', '19042', '19043', '19044', '19046', '19050', '19057', '19070', '19086', '19089',
                        '20002', '20005', '20008', '20010', '20014', '20016', '20018', '20019', '20020', '20027', '20032', '20035',
                        '20046', '20047', '20058', '20059', '20060', '20062', '20067', '20069', '20072', '20074', '20075', '20077',
                        '21002', '21006', '21007', '21008', '21009', '21010', '21011', '21013', '21014', '21015', '21016', '21017',
                        '21019', '21020', '21021', '21023', '21024', '21026', '21027', '21028', '21029', '21035', '21036', '21037',
                        '21038', '21040', '21041', '21043', '21047', '21048', '21049', '21051', '21052', '21053', '21054', '21055',
                        '21058', '21084', '21094', '22008', '22010', '22011', '22016', '22017', '22019', '22024', '22025', '22026',
                        '22027', '22028', '22030', '22036', '22037', '22038', '22039', '22045', '22070', '22071', '22072', '22073',
                        '22075', '22076', '22077', '22083', '22086', '22090', '22096', '22097', '22098', '22104', '22105', '22106',
                        '22109', '22110', '22115', '22116', '22131', '22132', '22133', '22134', '22135', '22137', '23025', '23026',
                        '23034', '23035', '23036', '23037', '23038', '23039', '23041', '23043', '23044', '23045', '23046', '23051',
                        '23052', '23053', '23054', '23057', '23058', '23060', '23062', '23063', '23064', '23066', '23067', '23071',
                        '23074', '23076', '23078', '23079', '23083', '23084', '23085', '23086', '23087', '23092', '23095', '23096',
                        '24001', '24002', '24003', '24004', '24007', '24008', '24009', '24010', '24011', '24012', '24013', '24014']


        current_fct5 = ['11003', '11004', '11016', '11023', '11026', '11047', '11050', '11051', '11052', '11053', '11054', '11056',
                        '11057', '11061', '11065', '11070', '11071', '11072', '11073', '11074', '11076', '11079', '11081', '11082',
                        '11083', '11089', '11090', '11096', '12020', '12021', '12022', '12032', '12044', '12047', '12052', '12054',
                        '12055', '13025', '13028', '13029', '13034', '13035', '13036', '13037', '13039', '13044', '13045', '13047',
                        '13049', '13050', '13054', '13055', '13056', '13060', '13068', '13075', '13076', '13078', '13079', '13080',
                        '13081', '13088', '13090', '13092', '13094', '13095', '13104', '13105', '13106', '13109', '13110', '13121',
                        '13133', '13169', '13219', '13271', '13272', '13276', '13283', '13305', '13311', '13321', '13325', '13328',
                        '13331', '13333', '15003', '15005', '15006', '15007', '15014', '15018', '15020', '15021', '15023', '15025',
                        '15028', '15031', '15032', '15034', '15035', '15036', '15037', '15043', '15048', '15051', '15052', '15055',
                        '15057', '15058', '15060', '15062', '15070', '15071', '15072', '15078', '15079', '15081', '15082', '15087',
                        '15088', '15089', '15092', '15093', '15095', '15099', '15104', '15106', '15109', '16003', '16023', '16027',
                        '16028', '16033', '16042', '17008', '17017', '17021', '17022', '17023', '17024', '17025', '17030', '17031',
                        '17077', '17078', '17089', '17094', '17102', '17103', '17106', '17114', '18003', '18004', '18008', '18009',
                        '18014', '18015', '18017', '18018', '18020', '18021', '18025', '18026', '18030', '18038', '18041', '18045',
                        '18046', '18047', '18048', '18049', '18051', '18052', '18053', '18054', '18058', '18061', '18062', '18067',
                        '18072', '18073', '18075', '18076', '18077', '18081', '18082', '18083', '18084', '18085', '18088', '18089',
                        '18091', '18092', '18093', '18094', '18095', '18096', '18097', '18100', '18101', '18103', '18105', '18108',
                        '18109', '18110', '18112', '18113', '18114', '18118', '18122', '18125', '18133', '18135', '18139', '18158',
                        '18159', '18160', '18164', '18169', '19015', '19021', '19034', '19048', '19049', '19051', '19054', '19058',
                        '19059', '19062', '19063', '19064', '19065', '19066', '19067', '19072', '19073', '19074', '19075', '19076',
                        '19077', '19078', '19079', '19080', '19081', '19082', '19083', '19084', '19085', '19087', '19088', '19090',
                        '19091', '19092', '19095', '19096', '19098', '19101', '19102', '19104', '19106', '19109', '19112', '19121',
                        '19124', '20003', '20011', '20013', '20015', '20021', '20023', '20033', '20036', '20037', '20038', '20055',
                        '20056', '20063', '20065', '20068', '20070', '20071', '20073', '20076', '21033', '21062', '21063', '21064',
                        '21066', '21068', '21069', '21073', '21078', '21080', '21083', '21085', '21086', '21087', '21088', '21089',
                        '21090', '21092', '21093', '22074', '22079', '22082', '22085', '22101', '22102', '22107', '22111', '22112',
                        '22117', '22119', '22136', '23017', '23024', '23027', '23029', '23030', '23031', '23032', '23040', '23047',
                        '23048', '23049', '23050', '23055', '23059', '23061', '23065', '23068', '23069', '23070', '23072', '23073',
                        '23075', '23080', '23081', '23082', '23088', '23089', '23090', '23091', '23094', '23097', '23098', '23099',
                        '24006']


        current_fct6 = ['11009', '11010', '11013', '11015', '11017', '11018', '11020', '11024', '11031', '11032', '11035', '11036',
                        '11042', '11045', '11049', '11058', '11085', '12018', '12019', '12023', '12024', '12025', '12026', '12027',
                        '12028', '12029', '12030', '12033', '12034', '12035', '12036', '12037', '12038', '12039', '12040', '12041',
                        '12043', '12045', '12048', '12049', '13008', '13009', '13010', '13011', '13012', '13013', '13014', '13015',
                        '13018', '13024', '13048', '13051', '13062', '13063', '13066', '13082', '15013', '15024', '15039', '15040',
                        '15046', '15049', '15057', '15061', '15075', '15076', '15080', '15083', '15105', '15108', '16001', '16011',
                        '16014', '16021', '16024', '16025', '16026', '16029', '16031', '16032', '16035', '16036', '16038', '16039',
                        '17001', '17010', '17012', '17013', '17014', '17018', '17019', '17020', '17071', '17072', '17073', '17074',
                        '17075', '17076', '17080', '17081', '17082', '17083', '17085', '17086', '17087', '17095', '17096', '17097',
                        '18060', '18064', '18065', '18066', '18068', '18069', '18070', '18074', '18078', '18090', '18104', '18111',
                        '19001', '19020', '19022', '19023', '19024', '19025', '19026', '19027', '19028', '19029', '19030', '19032',
                        '19041', '19042', '19043', '19044', '19046', '19070', '20001', '20002', '20022', '20024', '20025', '20026',
                        '20028', '20030', '20031', '20032', '20035', '20039', '20040', '20041', '20042', '20043', '20045', '20046',
                        '20051', '20052', '20053', '20054', '20062', '20064', '21011', '21014', '21015', '21019', '21020', '21025',
                        '21026', '21027', '21028', '21029', '21030', '21031', '21032', '21034', '21035', '21036', '21037', '21038',
                        '21053', '21060', '21084', '21094', '22001', '22031', '22039', '22040', '22051', '22069', '22076', '22080',
                        '22084', '22087', '22088', '22093', '22094', '22130', '23025', '23026', '23028', '23033', '23035', '23038',
                        '23039', '23041', '23042', '23043', '23060', '23083', '23093', '24001', '24003', '24005', '24008', '24009',
                        '24010', '24011', '24012', '24014']

   
        current_fct7 = ['11012', '11037', '11043', '11044', '11075', '11089', '12013', '12023', '12031', '12050', '12051', '12057',
                        '16002', '16026', '17005', '17010', '17011', '17012', '17015', '17037', '17039', '17085', '17088', '18028',
                        '18059', '18134', '18137', '18138', '18140', '18143', '18146', '18148', '18156', '19004', '19031', '19033',
                        '19035', '19037', '19040', '19045', '20001', '20006', '20007', '20020', '20029', '20048', '20049', '20050',
                        '20064', '20066', '20074', '21059', '21079', '22080', '22081', '22088', '22091']


        current_fct8 = ['11002', '11006', '11034', '11055', '11059', '11064', '11066', '11067', '11077', '11078', '11080', '11084',
                        '11086', '11087', '11088', '11091', '11092', '11093', '11094', '11095', '11097', '11098', '11099', '11100',
                        '11101', '13030', '13077', '13083', '13086', '13089', '13091', '13093', '13096', '13097', '13098', '13100',
                        '13102', '13103', '13107', '13108', '13111', '13112', '13113', '13114', '13117', '13118', '13119', '13120',
                        '13122', '13184', '13225', '13248', '15012', '15073', '15077', '15086', '15090', '15091', '15094', '15096',
                        '15097', '15098', '15101', '17109', '17110', '17111', '17112', '18107', '18115', '18116', '18117', '18119',
                        '18121', '18123', '18124', '18126', '18127', '18128', '18129', '18130', '18131', '18132', '19047', '19056',
                        '19060', '19061', '19068', '19094', '19097', '19100', '19105', '19107', '19108', '19110', '19111', '19113',
                        '19114', '19116', '19117', '19118', '19120', '19122', '19123', '19125', '21065', '21070', '21074', '21077',
                        '21081', '21082', '21091', '22100', '22103', '22113', '22114', '22118', '22120', '22121', '22122', '22123',
                        '22124', '22125', '22126', '22127', '22128', '22129']


        current_fct9 = ['12042', '12061', '13067', '15015', '15103', '18027', '18087', '19052', '19053', '20034', '20044', '20057']

        orga_dic = {}
        orga_dic["11001"] ='11801'
        orga_dic["11002"] ='11801'
        orga_dic["11003"] ='11802'
        orga_dic["11004"] ='11802'
        orga_dic["11005"] ='11801'
        orga_dic["11006"] ='11801'
        orga_dic["11007"] ='11802'
        orga_dic["11008"] ='11801'
        orga_dic["11009"] ='11801'
        orga_dic["11010"] ='11802'
        orga_dic["11011"] ='11802'
        orga_dic["11012"] ='11810'
        orga_dic["11013"] ='11808'
        orga_dic["11014"] ='11803'
        orga_dic["11015"] ='11804'
        orga_dic["11016"] ='11810'
        orga_dic["11017"] ='11803'
        orga_dic["11018"] ='11811'
        orga_dic["11019"] ='11806'
        orga_dic["11020"] ='11810'
        orga_dic["11021"] ='11801'
        orga_dic["11022"] ='11802'
        orga_dic["11023"] ='11802'
        orga_dic["11024"] ='11806'
        orga_dic["11025"] ='11801'
        orga_dic["11026"] ='11802'
        orga_dic["11027"] ='11801'
        orga_dic["11028"] ='11801'
        orga_dic["11030"] ='11801'
        orga_dic["11031"] ='11802'
        orga_dic["11032"] ='11801'
        orga_dic["11033"] ='11802'
        orga_dic["11034"] ='11801'
        orga_dic["11035"] ='11801'
        orga_dic["11036"] ='11801'
        orga_dic["11037"] ='11801'
        orga_dic["11038"] ='11801'
        orga_dic["11039"] ='11802'
        orga_dic["11040"] ='11803'
        orga_dic["11041"] ='11803'
        orga_dic["11042"] ='11803'
        orga_dic["11043"] ='11803'
        orga_dic["11044"] ='11803'
        orga_dic["11045"] ='11803'
        orga_dic["11046"] ='11803'
        orga_dic["11047"] ='11804'
        orga_dic["11048"] ='11804'
        orga_dic["11049"] ='11805'
        orga_dic["11050"] ='11807'
        orga_dic["11051"] ='11806'
        orga_dic["11052"] ='11806'
        orga_dic["11053"] ='11806'
        orga_dic["11054"] ='11806'
        orga_dic["11055"] ='11811'
        orga_dic["11056"] ='11811'
        orga_dic["11057"] ='11801'
        orga_dic["11058"] ='11811'
        orga_dic["11059"] ='11809'
        orga_dic["11060"] ='11814'
        orga_dic["11061"] ='11802'
        orga_dic["11062"] ='11801'
        orga_dic["11063"] ='11801'
        orga_dic["11064"] ='11803'
        orga_dic["11065"] ='11803'
        orga_dic["11066"] ='11803'
        orga_dic["11067"] ='11803'
        orga_dic["11068"] ='11801'
        orga_dic["11069"] ='11801'
        orga_dic["11070"] ='11803'
        orga_dic["11071"] ='11803'
        orga_dic["11072"] ='11802'
        orga_dic["11073"] ='11802'
        orga_dic["11074"] ='11804'
        orga_dic["11075"] ='11804'
        orga_dic["11076"] ='11802'
        orga_dic["11077"] ='11801'
        orga_dic["11078"] ='11802'
        orga_dic["11079"] ='11802'
        orga_dic["11080"] ='11803'
        orga_dic["11081"] ='11804'
        orga_dic["11082"] ='11804'
        orga_dic["11083"] ='11804'
        orga_dic["11084"] ='11814'
        orga_dic["11085"] ='11814'
        orga_dic["11086"] ='11801'
        orga_dic["11087"] ='11801'
        orga_dic["11088"] ='11801'
        orga_dic["11089"] ='11802'
        orga_dic["11090"] ='11802'
        orga_dic["11091"] ='11812'
        orga_dic["11092"] ='11812'
        orga_dic["11093"] ='11801'
        orga_dic["11094"] ='11801'
        orga_dic["11095"] ='11802'
        orga_dic["11096"] ='11802'
        orga_dic["11097"] ='11802'
        orga_dic["11098"] ='11803'
        orga_dic["11099"] ='11803'
        orga_dic["11100"] ='11808'
        orga_dic["11101"] ='11808'
        orga_dic["11701"] ='11801'
        orga_dic["11702"] ='11803'
        orga_dic["11703"] ='11802'
        orga_dic["11704"] ='11804'
        orga_dic["11705"] ='11805'
        orga_dic["11706"] ='11806'
        orga_dic["11707"] ='11808'
        orga_dic["11708"] ='11807'
        orga_dic["11709"] ='11809'
        orga_dic["11710"] ='11810'
        orga_dic["11711"] ='11811'
        orga_dic["11712"] ='11812'
        orga_dic["11713"] ='11814'
        orga_dic["11801"] ='11801'
        orga_dic["11802"] ='11802'
        orga_dic["11803"] ='11803'
        orga_dic["11804"] ='11804'
        orga_dic["11805"] ='11805'
        orga_dic["11806"] ='11806'
        orga_dic["11807"] ='11807'
        orga_dic["11808"] ='11808'
        orga_dic["11809"] ='11809'
        orga_dic["11810"] ='11810'
        orga_dic["11811"] ='11811'
        orga_dic["11812"] ='11812'
        orga_dic["11814"] ='11814'
        orga_dic["12003"] ='12805'
        orga_dic["12004"] ='12805'
        orga_dic["12005"] ='12805'
        orga_dic["12006"] ='12805'
        orga_dic["12007"] ='12805'
        orga_dic["12008"] ='12805'
        orga_dic["12009"] ='12808'
        orga_dic["12010"] ='12808'
        orga_dic["12012"] ='12805'
        orga_dic["12013"] ='12808'
        orga_dic["12014"] ='12805'
        orga_dic["12016"] ='12808'
        orga_dic["12018"] ='12801'
        orga_dic["12019"] ='12802'
        orga_dic["12020"] ='12803'
        orga_dic["12021"] ='12803'
        orga_dic["12022"] ='12804'
        orga_dic["12023"] ='12805'
        orga_dic["12024"] ='12806'
        orga_dic["12025"] ='12807'
        orga_dic["12026"] ='12808'
        orga_dic["12027"] ='12809'
        orga_dic["12028"] ='12810'
        orga_dic["12029"] ='12810'
        orga_dic["12030"] ='12811'
        orga_dic["12031"] ='12812'
        orga_dic["12032"] ='12812'
        orga_dic["12033"] ='12813'
        orga_dic["12034"] ='12813'
        orga_dic["12035"] ='12813'
        orga_dic["12036"] ='12817'
        orga_dic["12037"] ='12814'
        orga_dic["12038"] ='12815'
        orga_dic["12039"] ='12815'
        orga_dic["12040"] ='12822'
        orga_dic["12041"] ='12822'
        orga_dic["12042"] ='12822'
        orga_dic["12043"] ='12813'
        orga_dic["12044"] ='12804'
        orga_dic["12045"] ='12819'
        orga_dic["12046"] ='12821'
        orga_dic["12047"] ='12808'
        orga_dic["12048"] ='12899'
        orga_dic["12049"] ='12820'
        orga_dic["12050"] ='12899'
        orga_dic["12051"] ='12899'
        orga_dic["12052"] ='12805'
        orga_dic["12053"] ='12805'
        orga_dic["12054"] ='12805'
        orga_dic["12055"] ='12819'
        orga_dic["12056"] ='12819'
        orga_dic["12057"] ='12810'
        orga_dic["12058"] ='12810'
        orga_dic["12059"] ='12810'
        orga_dic["12060"] ='12810'
        orga_dic["12061"] ='12810'
        orga_dic["12062"] ='12810'
        orga_dic["12063"] ='12851'
        orga_dic["12701"] ='12801'
        orga_dic["12702"] ='12802'
        orga_dic["12703"] ='12803'
        orga_dic["12704"] ='12804'
        orga_dic["12705"] ='12805'
        orga_dic["12706"] ='12806'
        orga_dic["12707"] ='12807'
        orga_dic["12708"] ='12808'
        orga_dic["12709"] ='12809'
        orga_dic["12710"] ='12810'
        orga_dic["12711"] ='12811'
        orga_dic["12712"] ='12812'
        orga_dic["12713"] ='12813'
        orga_dic["12714"] ='12814'
        orga_dic["12715"] ='12815'
        orga_dic["12716"] ='12816'
        orga_dic["12717"] ='12817'
        orga_dic["12718"] ='12818'
        orga_dic["12801"] ='12801'
        orga_dic["12802"] ='12802'
        orga_dic["12803"] ='12803'
        orga_dic["12804"] ='12804'
        orga_dic["12805"] ='12805'
        orga_dic["12806"] ='12806'
        orga_dic["12807"] ='12807'
        orga_dic["12808"] ='12808'
        orga_dic["12809"] ='12809'
        orga_dic["12810"] ='12810'
        orga_dic["12811"] ='12811'
        orga_dic["12812"] ='12812'
        orga_dic["12813"] ='12813'
        orga_dic["12814"] ='12814'
        orga_dic["12815"] ='12815'
        orga_dic["12816"] ='12816'
        orga_dic["12817"] ='12817'
        orga_dic["12818"] ='12818'
        orga_dic["12819"] ='12819'
        orga_dic["12820"] ='12820'
        orga_dic["12821"] ='12821'
        orga_dic["12822"] ='12822'
        orga_dic["12850"] ='12850'
        orga_dic["12851"] ='12851'
        orga_dic["12852"] ='12852'
        orga_dic["12899"] ='12899'
        orga_dic["13001"] ='13802'
        orga_dic["13002"] ='13801'
        orga_dic["13003"] ='13808'
        orga_dic["13004"] ='13804'
        orga_dic["13005"] ='13809'
        orga_dic["13006"] ='13804'
        orga_dic["13007"] ='13809'
        orga_dic["13008"] ='13802'
        orga_dic["13009"] ='13801'
        orga_dic["13010"] ='13804'
        orga_dic["13011"] ='13806'
        orga_dic["13012"] ='13808'
        orga_dic["13013"] ='13809'
        orga_dic["13014"] ='13807'
        orga_dic["13015"] ='13810'
        orga_dic["13016"] ='13809'
        orga_dic["13017"] ='13808'
        orga_dic["13018"] ='13808'
        orga_dic["13019"] ='13808'
        orga_dic["13020"] ='13808'
        orga_dic["13021"] ='13808'
        orga_dic["13022"] ='13808'
        orga_dic["13023"] ='13808'
        orga_dic["13024"] ='13808'
        orga_dic["13025"] ='13808'
        orga_dic["13026"] ='13808'
        orga_dic["13027"] ='13808'
        orga_dic["13028"] ='13808'
        orga_dic["13029"] ='13808'
        orga_dic["13030"] ='13808'
        orga_dic["13031"] ='13809'
        orga_dic["13032"] ='13809'
        orga_dic["13033"] ='13809'
        orga_dic["13034"] ='13809'
        orga_dic["13035"] ='13811'
        orga_dic["13036"] ='13809'
        orga_dic["13037"] ='13809'
        orga_dic["13038"] ='13809'
        orga_dic["13039"] ='13809'
        orga_dic["13040"] ='13809'
        orga_dic["13041"] ='13809'
        orga_dic["13042"] ='13809'
        orga_dic["13043"] ='13809'
        orga_dic["13044"] ='13809'
        orga_dic["13045"] ='13801'
        orga_dic["13046"] ='13802'
        orga_dic["13047"] ='13802'
        orga_dic["13048"] ='13802'
        orga_dic["13049"] ='13803'
        orga_dic["13050"] ='13803'
        orga_dic["13051"] ='13803'
        orga_dic["13052"] ='13804'
        orga_dic["13053"] ='13804'
        orga_dic["13054"] ='13804'
        orga_dic["13055"] ='13804'
        orga_dic["13056"] ='13804'
        orga_dic["13057"] ='13804'
        orga_dic["13058"] ='13805'
        orga_dic["13059"] ='13805'
        orga_dic["13060"] ='13805'
        orga_dic["13061"] ='13805'
        orga_dic["13062"] ='13805'
        orga_dic["13063"] ='13805'
        orga_dic["13064"] ='13809'
        orga_dic["13066"] ='13806'
        orga_dic["13068"] ='13802'
        orga_dic["13069"] ='13804'
        orga_dic["13070"] ='13802'
        orga_dic["13071"] ='13809'
        orga_dic["13072"] ='13809'
        orga_dic["13073"] ='13804'
        orga_dic["13074"] ='13801'
        orga_dic["13075"] ='13813'
        orga_dic["13076"] ='13804'
        orga_dic["13077"] ='13804'
        orga_dic["13078"] ='13809'
        orga_dic["13079"] ='13802'
        orga_dic["13080"] ='13805'
        orga_dic["13081"] ='13805'
        orga_dic["13082"] ='13816'
        orga_dic["13083"] ='13809'
        orga_dic["13084"] ='13808'
        orga_dic["13085"] ='13802'
        orga_dic["13086"] ='13804'
        orga_dic["13087"] ='13802'
        orga_dic["13088"] ='13813'
        orga_dic["13089"] ='13810'
        orga_dic["13090"] ='13802'
        orga_dic["13091"] ='13802'
        orga_dic["13092"] ='13804'
        orga_dic["13093"] ='13804'
        orga_dic["13094"] ='13809'
        orga_dic["13095"] ='13816'
        orga_dic["13096"] ='13802'
        orga_dic["13097"] ='13809'
        orga_dic["13098"] ='13804'
        orga_dic["13099"] ='13809'
        orga_dic["13100"] ='13808'
        orga_dic["13101"] ='13802'
        orga_dic["13102"] ='13804'
        orga_dic["13103"] ='13813'
        orga_dic["13104"] ='13813'
        orga_dic["13105"] ='13812'
        orga_dic["13106"] ='13812'
        orga_dic["13107"] ='13804'
        orga_dic["13108"] ='13808'
        orga_dic["13109"] ='13808'
        orga_dic["13110"] ='13809'
        orga_dic["13111"] ='13808'
        orga_dic["13112"] ='13809'
        orga_dic["13113"] ='13804'
        orga_dic["13114"] ='13802'
        orga_dic["13115"] ='13805'
        orga_dic["13116"] ='13805'
        orga_dic["13117"] ='13813'
        orga_dic["13118"] ='13804'
        orga_dic["13119"] ='13804'
        orga_dic["13120"] ='13808'
        orga_dic["13121"] ='13808'
        orga_dic["13122"] ='13802'
        orga_dic["13123"] ='13804'
        orga_dic["13124"] ='13808'
        orga_dic["13125"] ='13809'
        orga_dic["13126"] ='13808'
        orga_dic["13127"] ='13804'
        orga_dic["13128"] ='13809'
        orga_dic["13129"] ='13808'
        orga_dic["13130"] ='13802'
        orga_dic["13131"] ='13809'
        orga_dic["13132"] ='13808'
        orga_dic["13133"] ='13802'
        orga_dic["13134"] ='13802'
        orga_dic["13135"] ='13809'
        orga_dic["13136"] ='13806'
        orga_dic["13137"] ='13802'
        orga_dic["13138"] ='13808'
        orga_dic["13139"] ='13809'
        orga_dic["13140"] ='13802'
        orga_dic["13141"] ='13802'
        orga_dic["13142"] ='13806'
        orga_dic["13143"] ='13808'
        orga_dic["13144"] ='13808'
        orga_dic["13145"] ='13804'
        orga_dic["13146"] ='13808'
        orga_dic["13147"] ='13808'
        orga_dic["13148"] ='13802'
        orga_dic["13149"] ='13808'
        orga_dic["13150"] ='13802'
        orga_dic["13151"] ='13802'
        orga_dic["13152"] ='13802'
        orga_dic["13153"] ='13809'
        orga_dic["13154"] ='13804'
        orga_dic["13155"] ='13804'
        orga_dic["13156"] ='13806'
        orga_dic["13157"] ='13809'
        orga_dic["13158"] ='13804'
        orga_dic["13159"] ='13808'
        orga_dic["13160"] ='13804'
        orga_dic["13161"] ='13808'
        orga_dic["13162"] ='13804'
        orga_dic["13163"] ='13804'
        orga_dic["13164"] ='13808'
        orga_dic["13165"] ='13804'
        orga_dic["13166"] ='13802'
        orga_dic["13167"] ='13808'
        orga_dic["13168"] ='13802'
        orga_dic["13169"] ='13809'
        orga_dic["13170"] ='13804'
        orga_dic["13171"] ='13808'
        orga_dic["13172"] ='13804'
        orga_dic["13173"] ='13809'
        orga_dic["13174"] ='13809'
        orga_dic["13175"] ='13804'
        orga_dic["13176"] ='13806'
        orga_dic["13177"] ='13806'
        orga_dic["13178"] ='13808'
        orga_dic["13179"] ='13804'
        orga_dic["13180"] ='13808'
        orga_dic["13181"] ='13802'
        orga_dic["13182"] ='13804'
        orga_dic["13183"] ='13808'
        orga_dic["13184"] ='13804'
        orga_dic["13185"] ='13809'
        orga_dic["13186"] ='13809'
        orga_dic["13187"] ='13801'
        orga_dic["13188"] ='13806'
        orga_dic["13189"] ='13808'
        orga_dic["13190"] ='13804'
        orga_dic["13191"] ='13808'
        orga_dic["13192"] ='13808'
        orga_dic["13193"] ='13805'
        orga_dic["13194"] ='13805'
        orga_dic["13195"] ='13802'
        orga_dic["13196"] ='13804'
        orga_dic["13197"] ='13802'
        orga_dic["13198"] ='13805'
        orga_dic["13199"] ='13809'
        orga_dic["13200"] ='13808'
        orga_dic["13201"] ='13809'
        orga_dic["13202"] ='13806'
        orga_dic["13203"] ='13801'
        orga_dic["13204"] ='13802'
        orga_dic["13205"] ='13801'
        orga_dic["13206"] ='13804'
        orga_dic["13207"] ='13802'
        orga_dic["13208"] ='13809'
        orga_dic["13209"] ='13801'
        orga_dic["13210"] ='13809'
        orga_dic["13211"] ='13801'
        orga_dic["13212"] ='13808'
        orga_dic["13213"] ='13809'
        orga_dic["13214"] ='13808'
        orga_dic["13215"] ='13809'
        orga_dic["13216"] ='13804'
        orga_dic["13217"] ='13801'
        orga_dic["13218"] ='13804'
        orga_dic["13219"] ='13805'
        orga_dic["13220"] ='13804'
        orga_dic["13221"] ='13802'
        orga_dic["13222"] ='13808'
        orga_dic["13223"] ='13804'
        orga_dic["13225"] ='13804'
        orga_dic["13226"] ='13803'
        orga_dic["13227"] ='13809'
        orga_dic["13228"] ='13809'
        orga_dic["13229"] ='13808'
        orga_dic["13230"] ='13805'
        orga_dic["13231"] ='13809'
        orga_dic["13232"] ='13808'
        orga_dic["13233"] ='13808'
        orga_dic["13234"] ='13804'
        orga_dic["13235"] ='13806'
        orga_dic["13236"] ='13809'
        orga_dic["13237"] ='13808'
        orga_dic["13238"] ='13808'
        orga_dic["13239"] ='13808'
        orga_dic["13240"] ='13802'
        orga_dic["13241"] ='13809'
        orga_dic["13242"] ='13802'
        orga_dic["13243"] ='13802'
        orga_dic["13244"] ='13801'
        orga_dic["13245"] ='13802'
        orga_dic["13246"] ='13806'
        orga_dic["13247"] ='13809'
        orga_dic["13248"] ='13804'
        orga_dic["13249"] ='13804'
        orga_dic["13250"] ='13809'
        orga_dic["13251"] ='13809'
        orga_dic["13252"] ='13802'
        orga_dic["13254"] ='13806'
        orga_dic["13255"] ='13805'
        orga_dic["13256"] ='13802'
        orga_dic["13257"] ='13808'
        orga_dic["13258"] ='13804'
        orga_dic["13259"] ='13802'
        orga_dic["13260"] ='13808'
        orga_dic["13261"] ='13809'
        orga_dic["13262"] ='13802'
        orga_dic["13263"] ='13808'
        orga_dic["13264"] ='13802'
        orga_dic["13265"] ='13809'
        orga_dic["13266"] ='13804'
        orga_dic["13267"] ='13809'
        orga_dic["13268"] ='13810'
        orga_dic["13269"] ='13809'
        orga_dic["13270"] ='13808'
        orga_dic["13271"] ='13804'
        orga_dic["13272"] ='13808'
        orga_dic["13273"] ='13804'
        orga_dic["13274"] ='13802'
        orga_dic["13275"] ='13809'
        orga_dic["13276"] ='13807'
        orga_dic["13277"] ='13801'
        orga_dic["13278"] ='13810'
        orga_dic["13279"] ='13809'
        orga_dic["13280"] ='13805'
        orga_dic["13281"] ='13802'
        orga_dic["13282"] ='13809'
        orga_dic["13283"] ='13804'
        orga_dic["13284"] ='13802'
        orga_dic["13285"] ='13808'
        orga_dic["13286"] ='13802'
        orga_dic["13287"] ='13802'
        orga_dic["13288"] ='13809'
        orga_dic["13289"] ='13808'
        orga_dic["13290"] ='13809'
        orga_dic["13291"] ='13808'
        orga_dic["13292"] ='13805'
        orga_dic["13293"] ='13809'
        orga_dic["13294"] ='13802'
        orga_dic["13295"] ='13804'
        orga_dic["13296"] ='13802'
        orga_dic["13297"] ='13804'
        orga_dic["13298"] ='13802'
        orga_dic["13299"] ='13809'
        orga_dic["13300"] ='13802'
        orga_dic["13301"] ='13808'
        orga_dic["13302"] ='13809'
        orga_dic["13303"] ='13802'
        orga_dic["13304"] ='13809'
        orga_dic["13305"] ='13808'
        orga_dic["13306"] ='13806'
        orga_dic["13307"] ='13808'
        orga_dic["13308"] ='13809'
        orga_dic["13309"] ='13809'
        orga_dic["13310"] ='13803'
        orga_dic["13311"] ='13805'
        orga_dic["13312"] ='13804'
        orga_dic["13313"] ='13805'
        orga_dic["13314"] ='13809'
        orga_dic["13315"] ='13809'
        orga_dic["13316"] ='13805'
        orga_dic["13317"] ='13808'
        orga_dic["13318"] ='13805'
        orga_dic["13319"] ='13804'
        orga_dic["13320"] ='13819'
        orga_dic["13321"] ='13805'
        orga_dic["13322"] ='13808'
        orga_dic["13323"] ='13809'
        orga_dic["13324"] ='13808'
        orga_dic["13325"] ='13808'
        orga_dic["13326"] ='13806'
        orga_dic["13327"] ='13808'
        orga_dic["13328"] ='13809'
        orga_dic["13329"] ='13809'
        orga_dic["13330"] ='13801'
        orga_dic["13331"] ='13808'
        orga_dic["13332"] ='13808'
        orga_dic["13333"] ='13805'
        orga_dic["13334"] ='13808'
        orga_dic["13701"] ='13801'
        orga_dic["13702"] ='13802'
        orga_dic["13703"] ='13803'
        orga_dic["13704"] ='13804'
        orga_dic["13705"] ='13805'
        orga_dic["13706"] ='13806'
        orga_dic["13707"] ='13807'
        orga_dic["13708"] ='13808'
        orga_dic["13709"] ='13810'
        orga_dic["13710"] ='13809'
        orga_dic["13711"] ='13811'
        orga_dic["13712"] ='13812'
        orga_dic["13713"] ='13813'
        orga_dic["13714"] ='13814'
        orga_dic["13715"] ='13815'
        orga_dic["13716"] ='13816'
        orga_dic["13717"] ='13817'
        orga_dic["13718"] ='13818'
        orga_dic["13719"] ='13819'
        orga_dic["13720"] ='13820'
        orga_dic["13721"] ='13821'
        orga_dic["13722"] ='13822'
        orga_dic["13723"] ='13809'
        orga_dic["13724"] ='13804'
        orga_dic["13725"] ='13802'
        orga_dic["13726"] ='13808'
        orga_dic["13727"] ='13808'
        orga_dic["13801"] ='13801'
        orga_dic["13802"] ='13802'
        orga_dic["13803"] ='13803'
        orga_dic["13804"] ='13804'
        orga_dic["13805"] ='13805'
        orga_dic["13806"] ='13806'
        orga_dic["13807"] ='13807'
        orga_dic["13808"] ='13808'
        orga_dic["13809"] ='13809'
        orga_dic["13810"] ='13810'
        orga_dic["13811"] ='13811'
        orga_dic["13812"] ='13812'
        orga_dic["13813"] ='13813'
        orga_dic["13814"] ='13814'
        orga_dic["13815"] ='13815'
        orga_dic["13816"] ='13816'
        orga_dic["13817"] ='13817'
        orga_dic["13818"] ='13818'
        orga_dic["13819"] ='13819'
        orga_dic["13820"] ='13820'
        orga_dic["13821"] ='13821'
        orga_dic["13822"] ='13822'
        orga_dic["13823"] ='13808'
        orga_dic["15001"] ='15802'
        orga_dic["15002"] ='15801'
        orga_dic["15003"] ='15802'
        orga_dic["15004"] ='15802'
        orga_dic["15005"] ='15801'
        orga_dic["15006"] ='15802'
        orga_dic["15007"] ='15803'
        orga_dic["15008"] ='15802'
        orga_dic["15010"] ='15802'
        orga_dic["15011"] ='15802'
        orga_dic["15012"] ='15807'
        orga_dic["15013"] ='15809'
        orga_dic["15014"] ='15803'
        orga_dic["15015"] ='15801'
        orga_dic["15016"] ='15806'
        orga_dic["15017"] ='15806'
        orga_dic["15018"] ='15805'
        orga_dic["15019"] ='15805'
        orga_dic["15020"] ='15801'
        orga_dic["15021"] ='15803'
        orga_dic["15022"] ='15802'
        orga_dic["15023"] ='15804'
        orga_dic["15024"] ='15805'
        orga_dic["15025"] ='15801'
        orga_dic["15026"] ='15802'
        orga_dic["15027"] ='15804'
        orga_dic["15028"] ='15802'
        orga_dic["15029"] ='15802'
        orga_dic["15030"] ='15802'
        orga_dic["15031"] ='15803'
        orga_dic["15032"] ='15803'
        orga_dic["15033"] ='15802'
        orga_dic["15034"] ='15802'
        orga_dic["15035"] ='15801'
        orga_dic["15036"] ='15803'
        orga_dic["15037"] ='15803'
        orga_dic["15038"] ='15804'
        orga_dic["15039"] ='15805'
        orga_dic["15040"] ='15805'
        orga_dic["15041"] ='15804'
        orga_dic["15042"] ='15804'
        orga_dic["15043"] ='15808'
        orga_dic["15044"] ='15806'
        orga_dic["15045"] ='15806'
        orga_dic["15046"] ='15809'
        orga_dic["15047"] ='15806'
        orga_dic["15048"] ='15806'
        orga_dic["15049"] ='15804'
        orga_dic["15050"] ='15806'
        orga_dic["15051"] ='15806'
        orga_dic["15052"] ='15802'
        orga_dic["15053"] ='15804'
        orga_dic["15054"] ='15802'
        orga_dic["15055"] ='15808'
        orga_dic["15056"] ='15802'
        orga_dic["15057"] ='15803'
        orga_dic["15058"] ='15806'
        orga_dic["15059"] ='15804'
        orga_dic["15060"] ='15806'
        orga_dic["15061"] ='15804'
        orga_dic["15062"] ='15806'
        orga_dic["15063"] ='15806'
        orga_dic["15064"] ='15802'
        orga_dic["15065"] ='15802'
        orga_dic["15066"] ='15806'
        orga_dic["15067"] ='15806'
        orga_dic["15068"] ='15802'
        orga_dic["15069"] ='15806'
        orga_dic["15070"] ='15801'
        orga_dic["15071"] ='15806'
        orga_dic["15072"] ='15806'
        orga_dic["15073"] ='15801'
        orga_dic["15074"] ='15806'
        orga_dic["15075"] ='15808'
        orga_dic["15076"] ='15808'
        orga_dic["15077"] ='15804'
        orga_dic["15078"] ='15806'
        orga_dic["15079"] ='15804'
        orga_dic["15080"] ='15804'
        orga_dic["15081"] ='15802'
        orga_dic["15082"] ='15803'
        orga_dic["15083"] ='15811'
        orga_dic["15084"] ='15090'
        orga_dic["15085"] ='15090'
        orga_dic["15086"] ='15810'
        orga_dic["15087"] ='15803'
        orga_dic["15088"] ='15803'
        orga_dic["15089"] ='15803'
        orga_dic["15090"] ='15090'
        orga_dic["15091"] ='15811'
        orga_dic["15092"] ='15807'
        orga_dic["15093"] ='15806'
        orga_dic["15094"] ='15803'
        orga_dic["15095"] ='15803'
        orga_dic["15096"] ='15802'
        orga_dic["15097"] ='15803'
        orga_dic["15098"] ='15804'
        orga_dic["15099"] ='15804'
        orga_dic["15100"] ='15803'
        orga_dic["15101"] ='15802'
        orga_dic["15102"] ='15802'
        orga_dic["15104"] ='15806'
        orga_dic["15105"] ='15809'
        orga_dic["15106"] ='15806'
        orga_dic["15107"] ='15806'
        orga_dic["15108"] ='15803'
        orga_dic["15109"] ='15802'
        orga_dic["15110"] ='15802'
        orga_dic["15701"] ='15801'
        orga_dic["15702"] ='15802'
        orga_dic["15703"] ='15803'
        orga_dic["15704"] ='15804'
        orga_dic["15705"] ='15805'
        orga_dic["15706"] ='15806'
        orga_dic["15707"] ='15807'
        orga_dic["15708"] ='15808'
        orga_dic["15709"] ='15809'
        orga_dic["15710"] ='15810'
        orga_dic["15711"] ='15811'
        orga_dic["15801"] ='15801'
        orga_dic["15802"] ='15802'
        orga_dic["15803"] ='15803'
        orga_dic["15804"] ='15804'
        orga_dic["15805"] ='15805'
        orga_dic["15806"] ='15806'
        orga_dic["15807"] ='15807'
        orga_dic["15808"] ='15808'
        orga_dic["15809"] ='15809'
        orga_dic["15810"] ='15810'
        orga_dic["15811"] ='15811'
        orga_dic["16001"] ='16809'
        orga_dic["16002"] ='16807'
        orga_dic["16003"] ='16810'
        orga_dic["16004"] ='16809'
        orga_dic["16005"] ='16809'
        orga_dic["16006"] ='16809'
        orga_dic["16007"] ='16809'
        orga_dic["16008"] ='16807'
        orga_dic["16009"] ='16810'
        orga_dic["16010"] ='16810'
        orga_dic["16011"] ='16810'
        orga_dic["16012"] ='16809'
        orga_dic["16013"] ='16809'
        orga_dic["16014"] ='16807'
        orga_dic["16015"] ='16807'
        orga_dic["16016"] ='16807'
        orga_dic["16017"] ='16807'
        orga_dic["16018"] ='16809'
        orga_dic["16019"] ='16810'
        orga_dic["16020"] ='16809'
        orga_dic["16021"] ='16809'
        orga_dic["16022"] ='16809'
        orga_dic["16023"] ='16810'
        orga_dic["16024"] ='16801'
        orga_dic["16025"] ='16802'
        orga_dic["16026"] ='16803'
        orga_dic["16027"] ='16803'
        orga_dic["16028"] ='16803'
        orga_dic["16029"] ='16804'
        orga_dic["16030"] ='16811'
        orga_dic["16031"] ='16805'
        orga_dic["16032"] ='16806'
        orga_dic["16033"] ='16809'
        orga_dic["16034"] ='16812'
        orga_dic["16035"] ='16801'
        orga_dic["16036"] ='16804'
        orga_dic["16037"] ='16806'
        orga_dic["16038"] ='16806'
        orga_dic["16039"] ='16807'
        orga_dic["16040"] ='16807'
        orga_dic["16041"] ='16899'
        orga_dic["16042"] ='16810'
        orga_dic["16043"] ='16811'
        orga_dic["16701"] ='16801'
        orga_dic["16702"] ='16802'
        orga_dic["16703"] ='16803'
        orga_dic["16704"] ='16804'
        orga_dic["16705"] ='16805'
        orga_dic["16706"] ='16806'
        orga_dic["16707"] ='16807'
        orga_dic["16708"] ='16808'
        orga_dic["16709"] ='16809'
        orga_dic["16710"] ='16810'
        orga_dic["16711"] ='16811'
        orga_dic["16712"] ='16812'
        orga_dic["16801"] ='16801'
        orga_dic["16802"] ='16802'
        orga_dic["16803"] ='16803'
        orga_dic["16804"] ='16804'
        orga_dic["16805"] ='16805'
        orga_dic["16806"] ='16806'
        orga_dic["16807"] ='16807'
        orga_dic["16808"] ='16808'
        orga_dic["16809"] ='16809'
        orga_dic["16810"] ='16810'
        orga_dic["16811"] ='16811'
        orga_dic["16812"] ='16812'
        orga_dic["16850"] ='16850'
        orga_dic["16851"] ='16851'
        orga_dic["16852"] ='16852'
        orga_dic["16899"] ='16899'
        orga_dic["17001"] ='17814'
        orga_dic["17002"] ='17814'
        orga_dic["17003"] ='17814'
        orga_dic["17004"] ='17814'
        orga_dic["17005"] ='17814'
        orga_dic["17006"] ='17807'
        orga_dic["17007"] ='17807'
        orga_dic["17008"] ='17807'
        orga_dic["17009"] ='17807'
        orga_dic["17010"] ='17817'
        orga_dic["17011"] ='17817'
        orga_dic["17013"] ='17805'
        orga_dic["17014"] ='17810'
        orga_dic["17015"] ='17809'
        orga_dic["17016"] ='17809'
        orga_dic["17017"] ='17809'
        orga_dic["17018"] ='17899'
        orga_dic["17020"] ='17802'
        orga_dic["17021"] ='17807'
        orga_dic["17022"] ='17814'
        orga_dic["17023"] ='17810'
        orga_dic["17024"] ='17814'
        orga_dic["17025"] ='17814'
        orga_dic["17026"] ='17814'
        orga_dic["17027"] ='17814'
        orga_dic["17028"] ='17814'
        orga_dic["17029"] ='17814'
        orga_dic["17031"] ='17814'
        orga_dic["17032"] ='17811'
        orga_dic["17033"] ='17814'
        orga_dic["17034"] ='17811'
        orga_dic["17035"] ='17814'
        orga_dic["17036"] ='17807'
        orga_dic["17037"] ='17807'
        orga_dic["17038"] ='17806'
        orga_dic["17039"] ='17807'
        orga_dic["17040"] ='17807'
        orga_dic["17041"] ='17807'
        orga_dic["17042"] ='17807'
        orga_dic["17043"] ='17807'
        orga_dic["17044"] ='17807'
        orga_dic["17045"] ='17807'
        orga_dic["17046"] ='17807'
        orga_dic["17047"] ='17807'
        orga_dic["17048"] ='17807'
        orga_dic["17049"] ='17807'
        orga_dic["17050"] ='17807'
        orga_dic["17051"] ='17807'
        orga_dic["17052"] ='17806'
        orga_dic["17053"] ='17807'
        orga_dic["17054"] ='17807'
        orga_dic["17055"] ='17807'
        orga_dic["17056"] ='17807'
        orga_dic["17057"] ='17807'
        orga_dic["17058"] ='17807'
        orga_dic["17059"] ='17807'
        orga_dic["17060"] ='17807'
        orga_dic["17061"] ='17807'
        orga_dic["17062"] ='17807'
        orga_dic["17063"] ='17809'
        orga_dic["17064"] ='17899'
        orga_dic["17065"] ='17807'
        orga_dic["17066"] ='17807'
        orga_dic["17067"] ='17807'
        orga_dic["17068"] ='17806'
        orga_dic["17069"] ='17807'
        orga_dic["17070"] ='17807'
        orga_dic["17072"] ='17802'
        orga_dic["17073"] ='17803'
        orga_dic["17074"] ='17804'
        orga_dic["17075"] ='17806'
        orga_dic["17076"] ='17807'
        orga_dic["17077"] ='17808'
        orga_dic["17078"] ='17809'
        orga_dic["17079"] ='17818'
        orga_dic["17080"] ='17811'
        orga_dic["17082"] ='17813'
        orga_dic["17083"] ='17814'
        orga_dic["17084"] ='17815'
        orga_dic["17085"] ='17816'
        orga_dic["17086"] ='17810'
        orga_dic["17087"] ='17814'
        orga_dic["17088"] ='17817'
        orga_dic["17089"] ='17853'
        orga_dic["17090"] ='17807'
        orga_dic["17091"] ='17813'
        orga_dic["17092"] ='17807'
        orga_dic["17093"] ='17807'
        orga_dic["17095"] ='17820'
        orga_dic["17097"] ='17821'
        orga_dic["17098"] ='17807'
        orga_dic["17100"] ='17807'
        orga_dic["17101"] ='17805'
        orga_dic["17102"] ='17824'
        orga_dic["17103"] ='17825'
        orga_dic["17105"] ='17816'
        orga_dic["17106"] ='17809'
        orga_dic["17107"] ='17819'
        orga_dic["17108"] ='17807'
        orga_dic["17109"] ='17805'
        orga_dic["17110"] ='17899'
        orga_dic["17113"] ='17807'
        orga_dic["17701"] ='17801'
        orga_dic["17702"] ='17802'
        orga_dic["17703"] ='17803'
        orga_dic["17704"] ='17804'
        orga_dic["17705"] ='17805'
        orga_dic["17706"] ='17806'
        orga_dic["17707"] ='17807'
        orga_dic["17708"] ='17808'
        orga_dic["17709"] ='17809'
        orga_dic["17710"] ='17810'
        orga_dic["17711"] ='17811'
        orga_dic["17712"] ='17812'
        orga_dic["17713"] ='17813'
        orga_dic["17714"] ='17814'
        orga_dic["17715"] ='17815'
        orga_dic["17716"] ='17816'
        orga_dic["17717"] ='17817'
        orga_dic["17718"] ='17818'
        orga_dic["17719"] ='17819'
        orga_dic["17720"] ='17820'
        orga_dic["17721"] ='17821'
        orga_dic["17722"] ='17822'
        orga_dic["17723"] ='17823'
        orga_dic["17724"] ='17824'
        orga_dic["17725"] ='17825'
        orga_dic["17726"] ='17853'
        orga_dic["17801"] ='17801'
        orga_dic["17802"] ='17802'
        orga_dic["17803"] ='17803'
        orga_dic["17804"] ='17804'
        orga_dic["17805"] ='17805'
        orga_dic["17806"] ='17806'
        orga_dic["17807"] ='17807'
        orga_dic["17808"] ='17808'
        orga_dic["17809"] ='17809'
        orga_dic["17810"] ='17810'
        orga_dic["17811"] ='17811'
        orga_dic["17812"] ='17812'
        orga_dic["17813"] ='17813'
        orga_dic["17814"] ='17814'
        orga_dic["17815"] ='17815'
        orga_dic["17816"] ='17816'
        orga_dic["17817"] ='17817'
        orga_dic["17818"] ='17818'
        orga_dic["17819"] ='17819'
        orga_dic["17820"] ='17820'
        orga_dic["17821"] ='17821'
        orga_dic["17822"] ='17822'
        orga_dic["17823"] ='17823'
        orga_dic["17824"] ='17824'
        orga_dic["17825"] ='17825'
        orga_dic["17850"] ='17850'
        orga_dic["17851"] ='17851'
        orga_dic["17852"] ='17852'
        orga_dic["17853"] ='17853'
        orga_dic["17899"] ='17899'
        orga_dic["18001"] ='18801'
        orga_dic["18005"] ='18801'
        orga_dic["18006"] ='18801'
        orga_dic["18007"] ='18801'
        orga_dic["18008"] ='18805'
        orga_dic["18010"] ='18809'
        orga_dic["18011"] ='18801'
        orga_dic["18013"] ='18801'
        orga_dic["18015"] ='18801'
        orga_dic["18018"] ='18801'
        orga_dic["18019"] ='18805'
        orga_dic["18020"] ='18805'
        orga_dic["18021"] ='18809'
        orga_dic["18022"] ='18801'
        orga_dic["18023"] ='18805'
        orga_dic["18024"] ='18805'
        orga_dic["18025"] ='18801'
        orga_dic["18026"] ='18801'
        orga_dic["18027"] ='18805'
        orga_dic["18028"] ='18801'
        orga_dic["18029"] ='18809'
        orga_dic["18031"] ='18805'
        orga_dic["18032"] ='18801'
        orga_dic["18033"] ='18805'
        orga_dic["18034"] ='18801'
        orga_dic["18035"] ='18805'
        orga_dic["18036"] ='18801'
        orga_dic["18037"] ='18805'
        orga_dic["18039"] ='18810'
        orga_dic["18040"] ='18801'
        orga_dic["18042"] ='18809'
        orga_dic["18044"] ='18805'
        orga_dic["18045"] ='18801'
        orga_dic["18047"] ='18801'
        orga_dic["18048"] ='18805'
        orga_dic["18049"] ='18801'
        orga_dic["18050"] ='18801'
        orga_dic["18051"] ='18805'
        orga_dic["18052"] ='18809'
        orga_dic["18053"] ='18805'
        orga_dic["18054"] ='18801'
        orga_dic["18055"] ='18801'
        orga_dic["18056"] ='18801'
        orga_dic["18057"] ='18801'
        orga_dic["18058"] ='18805'
        orga_dic["18059"] ='18801'
        orga_dic["18060"] ='18810'
        orga_dic["18061"] ='18805'
        orga_dic["18062"] ='18801'
        orga_dic["18063"] ='18805'
        orga_dic["18064"] ='18802'
        orga_dic["18065"] ='18806'
        orga_dic["18066"] ='18803'
        orga_dic["18067"] ='18804'
        orga_dic["18068"] ='18806'
        orga_dic["18069"] ='18807'
        orga_dic["18070"] ='18808'
        orga_dic["18071"] ='18809'
        orga_dic["18072"] ='18810'
        orga_dic["18073"] ='18899'
        orga_dic["18074"] ='18812'
        orga_dic["18076"] ='18802'
        orga_dic["18077"] ='18801'
        orga_dic["18078"] ='18801'
        orga_dic["18079"] ='18801'
        orga_dic["18080"] ='18801'
        orga_dic["18081"] ='18807'
        orga_dic["18083"] ='18823'
        orga_dic["18086"] ='18809'
        orga_dic["18087"] ='18808'
        orga_dic["18088"] ='18808'
        orga_dic["18089"] ='18805'
        orga_dic["18090"] ='18804'
        orga_dic["18091"] ='18801'
        orga_dic["18103"] ='18825'
        orga_dic["18116"] ='18837'
        orga_dic["18119"] ='18832'
        orga_dic["18121"] ='18835'
        orga_dic["18122"] ='18834'
        orga_dic["18123"] ='18838'
        orga_dic["18126"] ='18837'
        orga_dic["18129"] ='18832'
        orga_dic["18131"] ='18835'
        orga_dic["18133"] ='18801'
        orga_dic["18134"] ='18899'
        orga_dic["18135"] ='18823'
        orga_dic["18136"] ='18809'
        orga_dic["18137"] ='18801'
        orga_dic["18138"] ='18801'
        orga_dic["18139"] ='18823'
        orga_dic["18140"] ='18806'
        orga_dic["18141"] ='18823'
        orga_dic["18142"] ='18808'
        orga_dic["18143"] ='18801'
        orga_dic["18144"] ='18801'
        orga_dic["18145"] ='18808'
        orga_dic["18146"] ='18823'
        orga_dic["18148"] ='18899'
        orga_dic["18149"] ='18801'
        orga_dic["18150"] ='18801'
        orga_dic["18151"] ='18801'
        orga_dic["18153"] ='18801'
        orga_dic["18154"] ='18801'
        orga_dic["18155"] ='18801'
        orga_dic["18156"] ='18801'
        orga_dic["18157"] ='18808'
        orga_dic["18158"] ='18814'
        orga_dic["18159"] ='18801'
        orga_dic["18160"] ='18815'
        orga_dic["18161"] ='18801'
        orga_dic["18162"] ='18809'
        orga_dic["18163"] ='18801'
        orga_dic["18164"] ='18801'
        orga_dic["18165"] ='18806'
        orga_dic["18166"] ='18802'
        orga_dic["18167"] ='18801'
        orga_dic["18169"] ='18835'
        orga_dic["18701"] ='18801'
        orga_dic["18702"] ='18802'
        orga_dic["18703"] ='18803'
        orga_dic["18704"] ='18804'
        orga_dic["18705"] ='18805'
        orga_dic["18706"] ='18806'
        orga_dic["18707"] ='18807'
        orga_dic["18708"] ='18808'
        orga_dic["18709"] ='18809'
        orga_dic["18710"] ='18810'
        orga_dic["18712"] ='18812'
        orga_dic["18713"] ='18815'
        orga_dic["18714"] ='18814'
        orga_dic["18716"] ='18816'
        orga_dic["18717"] ='18817'
        orga_dic["18720"] ='18820'
        orga_dic["18721"] ='18821'
        orga_dic["18722"] ='18822'
        orga_dic["18723"] ='18823'
        orga_dic["18724"] ='18824'
        orga_dic["18725"] ='18825'
        orga_dic["18726"] ='18826'
        orga_dic["18727"] ='18827'
        orga_dic["18728"] ='18828'
        orga_dic["18729"] ='18829'
        orga_dic["18730"] ='18830'
        orga_dic["18731"] ='18831'
        orga_dic["18732"] ='18832'
        orga_dic["18733"] ='18833'
        orga_dic["18734"] ='18834'
        orga_dic["18735"] ='18835'
        orga_dic["18736"] ='18836'
        orga_dic["18737"] ='18837'
        orga_dic["18738"] ='18838'
        orga_dic["18739"] ='18839'
        orga_dic["18740"] ='18840'
        orga_dic["18741"] ='18841'
        orga_dic["18742"] ='18842'
        orga_dic["18801"] ='18801'
        orga_dic["18802"] ='18802'
        orga_dic["18803"] ='18803'
        orga_dic["18804"] ='18804'
        orga_dic["18805"] ='18805'
        orga_dic["18806"] ='18806'
        orga_dic["18807"] ='18807'
        orga_dic["18808"] ='18808'
        orga_dic["18809"] ='18809'
        orga_dic["18810"] ='18810'
        orga_dic["18812"] ='18812'
        orga_dic["18814"] ='18814'
        orga_dic["18815"] ='18815'
        orga_dic["18816"] ='18816'
        orga_dic["18817"] ='18817'
        orga_dic["18820"] ='18820'
        orga_dic["18821"] ='18821'
        orga_dic["18822"] ='18822'
        orga_dic["18823"] ='18823'
        orga_dic["18824"] ='18824'
        orga_dic["18825"] ='18825'
        orga_dic["18826"] ='18826'
        orga_dic["18827"] ='18827'
        orga_dic["18828"] ='18828'
        orga_dic["18829"] ='18829'
        orga_dic["18830"] ='18830'
        orga_dic["18831"] ='18831'
        orga_dic["18832"] ='18832'
        orga_dic["18833"] ='18833'
        orga_dic["18834"] ='18834'
        orga_dic["18835"] ='18835'
        orga_dic["18836"] ='18836'
        orga_dic["18837"] ='18837'
        orga_dic["18838"] ='18838'
        orga_dic["18839"] ='18839'
        orga_dic["18840"] ='18840'
        orga_dic["18841"] ='18841'
        orga_dic["18842"] ='18842'
        orga_dic["18850"] ='18850'
        orga_dic["18851"] ='18851'
        orga_dic["18852"] ='18852'
        orga_dic["18899"] ='18899'
        orga_dic["19001"] ='19801'
        orga_dic["19002"] ='19802'
        orga_dic["19003"] ='19802'
        orga_dic["19004"] ='19802'
        orga_dic["19005"] ='19802'
        orga_dic["19006"] ='19801'
        orga_dic["19007"] ='19801'
        orga_dic["19008"] ='19802'
        orga_dic["19009"] ='19801'
        orga_dic["19010"] ='19802'
        orga_dic["19011"] ='19801'
        orga_dic["19012"] ='19801'
        orga_dic["19013"] ='19801'
        orga_dic["19014"] ='19802'
        orga_dic["19015"] ='19801'
        orga_dic["19016"] ='19801'
        orga_dic["19017"] ='19802'
        orga_dic["19018"] ='19801'
        orga_dic["19019"] ='19802'
        orga_dic["19020"] ='19801'
        orga_dic["19021"] ='19803'
        orga_dic["19022"] ='19802'
        orga_dic["19023"] ='19802'
        orga_dic["19024"] ='19803'
        orga_dic["19025"] ='19805'
        orga_dic["19026"] ='19806'
        orga_dic["19027"] ='19806'
        orga_dic["19028"] ='19805'
        orga_dic["19029"] ='19804'
        orga_dic["19030"] ='19801'
        orga_dic["19031"] ='19801'
        orga_dic["19032"] ='19802'
        orga_dic["19033"] ='19802'
        orga_dic["19034"] ='19806'
        orga_dic["19035"] ='19806'
        orga_dic["19036"] ='19803'
        orga_dic["19037"] ='19803'
        orga_dic["19038"] ='19804'
        orga_dic["19039"] ='19804'
        orga_dic["19040"] ='19805'
        orga_dic["19041"] ='19805'
        orga_dic["19042"] ='19807'
        orga_dic["19043"] ='19807'
        orga_dic["19044"] ='19807'
        orga_dic["19045"] ='19807'
        orga_dic["19046"] ='19803'
        orga_dic["19047"] ='19803'
        orga_dic["19048"] ='19803'
        orga_dic["19049"] ='19803'
        orga_dic["19050"] ='19803'
        orga_dic["19051"] ='19803'
        orga_dic["19052"] ='19803'
        orga_dic["19053"] ='19803'
        orga_dic["19054"] ='19802'
        orga_dic["19055"] ='19802'
        orga_dic["19056"] ='19802'
        orga_dic["19057"] ='19802'
        orga_dic["19058"] ='19802'
        orga_dic["19059"] ='19802'
        orga_dic["19060"] ='19802'
        orga_dic["19061"] ='19801'
        orga_dic["19062"] ='19801'
        orga_dic["19063"] ='19801'
        orga_dic["19064"] ='19801'
        orga_dic["19065"] ='19801'
        orga_dic["19066"] ='19801'
        orga_dic["19067"] ='19801'
        orga_dic["19068"] ='19801'
        orga_dic["19069"] ='19806'
        orga_dic["19070"] ='19806'
        orga_dic["19071"] ='19806'
        orga_dic["19072"] ='19806'
        orga_dic["19073"] ='19806'
        orga_dic["19074"] ='19806'
        orga_dic["19075"] ='19807'
        orga_dic["19076"] ='19807'
        orga_dic["19077"] ='19807'
        orga_dic["19078"] ='19807'
        orga_dic["19079"] ='19807'
        orga_dic["19080"] ='19807'
        orga_dic["19081"] ='19806'
        orga_dic["19082"] ='19806'
        orga_dic["19083"] ='19801'
        orga_dic["19084"] ='19806'
        orga_dic["19085"] ='19806'
        orga_dic["19086"] ='19801'
        orga_dic["19087"] ='19803'
        orga_dic["19088"] ='19801'
        orga_dic["19089"] ='19802'
        orga_dic["19090"] ='19801'
        orga_dic["19091"] ='19802'
        orga_dic["19092"] ='19802'
        orga_dic["19094"] ='19806'
        orga_dic["19095"] ='19802'
        orga_dic["19096"] ='19802'
        orga_dic["19097"] ='19802'
        orga_dic["19098"] ='19806'
        orga_dic["19100"] ='19802'
        orga_dic["19101"] ='19802'
        orga_dic["19102"] ='19802'
        orga_dic["19103"] ='19806'
        orga_dic["19104"] ='19803'
        orga_dic["19105"] ='19806'
        orga_dic["19106"] ='19806'
        orga_dic["19107"] ='19803'
        orga_dic["19108"] ='19806'
        orga_dic["19109"] ='19806'
        orga_dic["19110"] ='19806'
        orga_dic["19111"] ='19808'
        orga_dic["19112"] ='19809'
        orga_dic["19113"] ='19801'
        orga_dic["19114"] ='19809'
        orga_dic["19118"] ='19801'
        orga_dic["19120"] ='19801'
        orga_dic["19121"] ='19810'
        orga_dic["19122"] ='19808'
        orga_dic["19123"] ='19806'
        orga_dic["19124"] ='19802'
        orga_dic["19125"] ='19811'
        orga_dic["19701"] ='19801'
        orga_dic["19702"] ='19802'
        orga_dic["19703"] ='19803'
        orga_dic["19704"] ='19804'
        orga_dic["19705"] ='19805'
        orga_dic["19706"] ='19806'
        orga_dic["19707"] ='19807'
        orga_dic["19708"] ='19808'
        orga_dic["19709"] ='19809'
        orga_dic["19710"] ='19810'
        orga_dic["19711"] ='19811'
        orga_dic["19801"] ='19801'
        orga_dic["19802"] ='19802'
        orga_dic["19803"] ='19803'
        orga_dic["19804"] ='19804'
        orga_dic["19805"] ='19805'
        orga_dic["19806"] ='19806'
        orga_dic["19807"] ='19807'
        orga_dic["19808"] ='19808'
        orga_dic["19809"] ='19809'
        orga_dic["19810"] ='19810'
        orga_dic["19811"] ='19811'
        orga_dic["19850"] ='19850'
        orga_dic["19851"] ='19851'
        orga_dic["19852"] ='19852'
        orga_dic["19899"] ='19899'
        orga_dic["20001"] ='20812'
        orga_dic["20002"] ='20813'
        orga_dic["20003"] ='20812'
        orga_dic["20005"] ='20818'
        orga_dic["20006"] ='20812'
        orga_dic["20008"] ='20812'
        orga_dic["20010"] ='20812'
        orga_dic["20012"] ='20813'
        orga_dic["20013"] ='20812'
        orga_dic["20014"] ='20812'
        orga_dic["20015"] ='20812'
        orga_dic["20016"] ='20813'
        orga_dic["20017"] ='20812'
        orga_dic["20018"] ='20812'
        orga_dic["20019"] ='20812'
        orga_dic["20020"] ='20812'
        orga_dic["20021"] ='20801'
        orga_dic["20022"] ='20802'
        orga_dic["20023"] ='20803'
        orga_dic["20024"] ='20804'
        orga_dic["20025"] ='20805'
        orga_dic["20026"] ='20806'
        orga_dic["20027"] ='20807'
        orga_dic["20028"] ='20807'
        orga_dic["20029"] ='20808'
        orga_dic["20030"] ='20809'
        orga_dic["20031"] ='20810'
        orga_dic["20032"] ='20811'
        orga_dic["20033"] ='20814'
        orga_dic["20034"] ='20814'
        orga_dic["20035"] ='20816'
        orga_dic["20036"] ='20816'
        orga_dic["20037"] ='20816'
        orga_dic["20038"] ='20817'
        orga_dic["20039"] ='20818'
        orga_dic["20040"] ='20899'
        orga_dic["20041"] ='20820'
        orga_dic["20042"] ='20821'
        orga_dic["20043"] ='20822'
        orga_dic["20044"] ='20823'
        orga_dic["20045"] ='20823'
        orga_dic["20046"] ='20812'
        orga_dic["20047"] ='20811'
        orga_dic["20048"] ='20811'
        orga_dic["20049"] ='20811'
        orga_dic["20051"] ='20899'
        orga_dic["20052"] ='20899'
        orga_dic["20053"] ='20826'
        orga_dic["20054"] ='20825'
        orga_dic["20055"] ='20823'
        orga_dic["20058"] ='20824'
        orga_dic["20059"] ='20812'
        orga_dic["20060"] ='20812'
        orga_dic["20061"] ='20812'
        orga_dic["20062"] ='20825'
        orga_dic["20063"] ='20812'
        orga_dic["20064"] ='20899'
        orga_dic["20066"] ='20812'
        orga_dic["20067"] ='20812'
        orga_dic["20068"] ='20811'
        orga_dic["20069"] ='20811'
        orga_dic["20071"] ='20850'
        orga_dic["20073"] ='20819'
        orga_dic["20074"] ='20812'
        orga_dic["20075"] ='20819'
        orga_dic["20076"] ='20812'
        orga_dic["20077"] ='20824'
        orga_dic["20701"] ='20801'
        orga_dic["20702"] ='20802'
        orga_dic["20703"] ='20803'
        orga_dic["20704"] ='20804'
        orga_dic["20705"] ='20805'
        orga_dic["20706"] ='20806'
        orga_dic["20707"] ='20807'
        orga_dic["20708"] ='20808'
        orga_dic["20709"] ='20809'
        orga_dic["20710"] ='20810'
        orga_dic["20711"] ='20811'
        orga_dic["20712"] ='20812'
        orga_dic["20713"] ='20813'
        orga_dic["20714"] ='20814'
        orga_dic["20715"] ='20815'
        orga_dic["20716"] ='20816'
        orga_dic["20717"] ='20817'
        orga_dic["20718"] ='20818'
        orga_dic["20719"] ='20819'
        orga_dic["20720"] ='20820'
        orga_dic["20721"] ='20821'
        orga_dic["20722"] ='20822'
        orga_dic["20723"] ='20823'
        orga_dic["20801"] ='20801'
        orga_dic["20802"] ='20802'
        orga_dic["20803"] ='20803'
        orga_dic["20804"] ='20804'
        orga_dic["20805"] ='20805'
        orga_dic["20806"] ='20806'
        orga_dic["20807"] ='20807'
        orga_dic["20808"] ='20808'
        orga_dic["20809"] ='20809'
        orga_dic["20810"] ='20810'
        orga_dic["20811"] ='20811'
        orga_dic["20812"] ='20812'
        orga_dic["20813"] ='20813'
        orga_dic["20814"] ='20814'
        orga_dic["20815"] ='20815'
        orga_dic["20816"] ='20816'
        orga_dic["20817"] ='20817'
        orga_dic["20818"] ='20818'
        orga_dic["20819"] ='20819'
        orga_dic["20820"] ='20820'
        orga_dic["20821"] ='20821'
        orga_dic["20822"] ='20822'
        orga_dic["20823"] ='20823'
        orga_dic["20824"] ='20824'
        orga_dic["20825"] ='20825'
        orga_dic["20826"] ='20826'
        orga_dic["20850"] ='20850'
        orga_dic["20851"] ='20851'
        orga_dic["20852"] ='20852'
        orga_dic["20899"] ='20899'
        orga_dic["21001"] ='21805'
        orga_dic["21002"] ='21802'
        orga_dic["21003"] ='21805'
        orga_dic["21004"] ='21805'
        orga_dic["21005"] ='21805'
        orga_dic["21006"] ='21805'
        orga_dic["21007"] ='21805'
        orga_dic["21008"] ='21805'
        orga_dic["21009"] ='21805'
        orga_dic["21010"] ='21805'
        orga_dic["21011"] ='21801'
        orga_dic["21012"] ='21801'
        orga_dic["21013"] ='21805'
        orga_dic["21014"] ='21802'
        orga_dic["21015"] ='21803'
        orga_dic["21016"] ='21803'
        orga_dic["21017"] ='21805'
        orga_dic["21018"] ='21803'
        orga_dic["21019"] ='21802'
        orga_dic["21020"] ='21802'
        orga_dic["21021"] ='21805'
        orga_dic["21022"] ='21801'
        orga_dic["21023"] ='21801'
        orga_dic["21024"] ='21805'
        orga_dic["21025"] ='21806'
        orga_dic["21026"] ='21808'
        orga_dic["21027"] ='21807'
        orga_dic["21028"] ='21804'
        orga_dic["21029"] ='21804'
        orga_dic["21030"] ='21804'
        orga_dic["21031"] ='21805'
        orga_dic["21032"] ='21801'
        orga_dic["21033"] ='21802'
        orga_dic["21034"] ='21803'
        orga_dic["21035"] ='21806'
        orga_dic["21036"] ='21808'
        orga_dic["21037"] ='21808'
        orga_dic["21038"] ='21805'
        orga_dic["21039"] ='21805'
        orga_dic["21040"] ='21806'
        orga_dic["21041"] ='21806'
        orga_dic["21042"] ='21806'
        orga_dic["21043"] ='21806'
        orga_dic["21044"] ='21806'
        orga_dic["21045"] ='21806'
        orga_dic["21046"] ='21806'
        orga_dic["21047"] ='21806'
        orga_dic["21048"] ='21806'
        orga_dic["21049"] ='21806'
        orga_dic["21050"] ='21806'
        orga_dic["21051"] ='21806'
        orga_dic["21052"] ='21806'
        orga_dic["21053"] ='21806'
        orga_dic["21054"] ='21806'
        orga_dic["21055"] ='21806'
        orga_dic["21056"] ='21804'
        orga_dic["21057"] ='21804'
        orga_dic["21058"] ='21804'
        orga_dic["21059"] ='21804'
        orga_dic["21060"] ='21807'
        orga_dic["21061"] ='21806'
        orga_dic["21062"] ='21806'
        orga_dic["21063"] ='21806'
        orga_dic["21064"] ='21806'
        orga_dic["21065"] ='21806'
        orga_dic["21066"] ='21806'
        orga_dic["21067"] ='21805'
        orga_dic["21068"] ='21805'
        orga_dic["21069"] ='21805'
        orga_dic["21070"] ='21802'
        orga_dic["21071"] ='21801'
        orga_dic["21072"] ='21802'
        orga_dic["21073"] ='21802'
        orga_dic["21074"] ='21801'
        orga_dic["21075"] ='21806'
        orga_dic["21076"] ='21806'
        orga_dic["21077"] ='21805'
        orga_dic["21078"] ='21805'
        orga_dic["21079"] ='21805'
        orga_dic["21080"] ='21805'
        orga_dic["21081"] ='21807'
        orga_dic["21082"] ='21807'
        orga_dic["21083"] ='21807'
        orga_dic["21084"] ='21807'
        orga_dic["21085"] ='21803'
        orga_dic["21086"] ='21803'
        orga_dic["21087"] ='21803'
        orga_dic["21088"] ='21803'
        orga_dic["21089"] ='21801'
        orga_dic["21090"] ='21801'
        orga_dic["21091"] ='21802'
        orga_dic["21092"] ='21802'
        orga_dic["21093"] ='21802'
        orga_dic["21094"] ='21802'
        orga_dic["21701"] ='21807'
        orga_dic["21702"] ='21801'
        orga_dic["21703"] ='21802'
        orga_dic["21704"] ='21803'
        orga_dic["21705"] ='21804'
        orga_dic["21706"] ='21805'
        orga_dic["21707"] ='21806'
        orga_dic["21708"] ='21807'
        orga_dic["21709"] ='21808'
        orga_dic["21801"] ='21801'
        orga_dic["21802"] ='21802'
        orga_dic["21803"] ='21803'
        orga_dic["21804"] ='21804'
        orga_dic["21805"] ='21805'
        orga_dic["21806"] ='21806'
        orga_dic["21807"] ='21807'
        orga_dic["21808"] ='21808'
        orga_dic["21850"] ='21850'
        orga_dic["21851"] ='21851'
        orga_dic["21852"] ='21852'
        orga_dic["21899"] ='21899'
        orga_dic["22001"] ='22806'
        orga_dic["22002"] ='22806'
        orga_dic["22003"] ='22806'
        orga_dic["22004"] ='22806'
        orga_dic["22005"] ='22806'
        orga_dic["22006"] ='22806'
        orga_dic["22007"] ='22806'
        orga_dic["22008"] ='22806'
        orga_dic["22009"] ='22806'
        orga_dic["22010"] ='22806'
        orga_dic["22011"] ='22806'
        orga_dic["22012"] ='22806'
        orga_dic["22013"] ='22806'
        orga_dic["22014"] ='22806'
        orga_dic["22015"] ='22806'
        orga_dic["22016"] ='22806'
        orga_dic["22017"] ='22806'
        orga_dic["22018"] ='22806'
        orga_dic["22019"] ='22806'
        orga_dic["22020"] ='22806'
        orga_dic["22021"] ='22806'
        orga_dic["22022"] ='22806'
        orga_dic["22023"] ='22806'
        orga_dic["22024"] ='22806'
        orga_dic["22025"] ='22806'
        orga_dic["22026"] ='22806'
        orga_dic["22027"] ='22806'
        orga_dic["22028"] ='22806'
        orga_dic["22029"] ='22806'
        orga_dic["22030"] ='22806'
        orga_dic["22031"] ='22801'
        orga_dic["22032"] ='22801'
        orga_dic["22033"] ='22801'
        orga_dic["22034"] ='22801'
        orga_dic["22035"] ='22801'
        orga_dic["22036"] ='22801'
        orga_dic["22037"] ='22801'
        orga_dic["22038"] ='22801'
        orga_dic["22039"] ='22802'
        orga_dic["22040"] ='22802'
        orga_dic["22041"] ='22802'
        orga_dic["22042"] ='22802'
        orga_dic["22043"] ='22802'
        orga_dic["22044"] ='22802'
        orga_dic["22045"] ='22802'
        orga_dic["22046"] ='22802'
        orga_dic["22047"] ='22802'
        orga_dic["22048"] ='22802'
        orga_dic["22049"] ='22802'
        orga_dic["22050"] ='22802'
        orga_dic["22051"] ='22802'
        orga_dic["22052"] ='22802'
        orga_dic["22053"] ='22802'
        orga_dic["22054"] ='22802'
        orga_dic["22055"] ='22802'
        orga_dic["22056"] ='22802'
        orga_dic["22057"] ='22802'
        orga_dic["22058"] ='22802'
        orga_dic["22059"] ='22802'
        orga_dic["22060"] ='22802'
        orga_dic["22061"] ='22802'
        orga_dic["22062"] ='22802'
        orga_dic["22063"] ='22802'
        orga_dic["22064"] ='22802'
        orga_dic["22065"] ='22802'
        orga_dic["22066"] ='22802'
        orga_dic["22067"] ='22802'
        orga_dic["22068"] ='22802'
        orga_dic["22069"] ='22802'
        orga_dic["22070"] ='22806'
        orga_dic["22071"] ='22806'
        orga_dic["22072"] ='22806'
        orga_dic["22073"] ='22806'
        orga_dic["22074"] ='22806'
        orga_dic["22075"] ='22801'
        orga_dic["22076"] ='22801'
        orga_dic["22077"] ='22806'
        orga_dic["22079"] ='22805'
        orga_dic["22080"] ='22807'
        orga_dic["22081"] ='22807'
        orga_dic["22082"] ='22807'
        orga_dic["22083"] ='22803'
        orga_dic["22084"] ='22803'
        orga_dic["22085"] ='22804'
        orga_dic["22086"] ='22803'
        orga_dic["22087"] ='22805'
        orga_dic["22088"] ='22807'
        orga_dic["22089"] ='22802'
        orga_dic["22090"] ='22802'
        orga_dic["22091"] ='22807'
        orga_dic["22092"] ='22806'
        orga_dic["22093"] ='22899'
        orga_dic["22094"] ='22801'
        orga_dic["22095"] ='22806'
        orga_dic["22096"] ='22806'
        orga_dic["22097"] ='22806'
        orga_dic["22098"] ='22802'
        orga_dic["22099"] ='22802'
        orga_dic["22100"] ='22811'
        orga_dic["22101"] ='22803'
        orga_dic["22102"] ='22803'
        orga_dic["22103"] ='22801'
        orga_dic["22104"] ='22806'
        orga_dic["22105"] ='22806'
        orga_dic["22106"] ='22806'
        orga_dic["22107"] ='22806'
        orga_dic["22108"] ='22806'
        orga_dic["22109"] ='22806'
        orga_dic["22110"] ='22806'
        orga_dic["22111"] ='22802'
        orga_dic["22112"] ='22802'
        orga_dic["22113"] ='22802'
        orga_dic["22114"] ='22810'
        orga_dic["22115"] ='22810'
        orga_dic["22116"] ='22808'
        orga_dic["22117"] ='22803'
        orga_dic["22118"] ='22804'
        orga_dic["22119"] ='22804'
        orga_dic["22120"] ='22806'
        orga_dic["22121"] ='22806'
        orga_dic["22122"] ='22806'
        orga_dic["22123"] ='22802'
        orga_dic["22124"] ='22802'
        orga_dic["22125"] ='22808'
        orga_dic["22126"] ='22806'
        orga_dic["22127"] ='22802'
        orga_dic["22128"] ='22803'
        orga_dic["22129"] ='22810'
        orga_dic["22130"] ='22899'
        orga_dic["22131"] ='22806'
        orga_dic["22132"] ='22806'
        orga_dic["22133"] ='22802'
        orga_dic["22134"] ='22801'
        orga_dic["22135"] ='22807'
        orga_dic["22136"] ='22803'
        orga_dic["22137"] ='22801'
        orga_dic["22701"] ='22801'
        orga_dic["22702"] ='22802'
        orga_dic["22703"] ='22803'
        orga_dic["22704"] ='22804'
        orga_dic["22705"] ='22805'
        orga_dic["22706"] ='22806'
        orga_dic["22707"] ='22807'
        orga_dic["22708"] ='22808'
        orga_dic["22709"] ='22809'
        orga_dic["22710"] ='22810'
        orga_dic["22711"] ='22811'
        orga_dic["22801"] ='22801'
        orga_dic["22802"] ='22802'
        orga_dic["22803"] ='22803'
        orga_dic["22804"] ='22804'
        orga_dic["22805"] ='22805'
        orga_dic["22806"] ='22806'
        orga_dic["22807"] ='22807'
        orga_dic["22808"] ='22808'
        orga_dic["22809"] ='22809'
        orga_dic["22810"] ='22810'
        orga_dic["22811"] ='22811'
        orga_dic["23001"] ='23802'
        orga_dic["23002"] ='23802'
        orga_dic["23003"] ='23802'
        orga_dic["23004"] ='23802'
        orga_dic["23005"] ='23801'
        orga_dic["23006"] ='23802'
        orga_dic["23007"] ='23802'
        orga_dic["23008"] ='23802'
        orga_dic["23009"] ='23802'
        orga_dic["23010"] ='23802'
        orga_dic["23011"] ='23802'
        orga_dic["23012"] ='23802'
        orga_dic["23013"] ='23802'
        orga_dic["23014"] ='23802'
        orga_dic["23015"] ='23802'
        orga_dic["23016"] ='23802'
        orga_dic["23017"] ='23802'
        orga_dic["23018"] ='23802'
        orga_dic["23019"] ='23802'
        orga_dic["23020"] ='23802'
        orga_dic["23021"] ='23802'
        orga_dic["23022"] ='23802'
        orga_dic["23023"] ='23802'
        orga_dic["23024"] ='23802'
        orga_dic["23025"] ='23801'
        orga_dic["23026"] ='23802'
        orga_dic["23027"] ='23803'
        orga_dic["23028"] ='23805'
        orga_dic["23029"] ='23801'
        orga_dic["23030"] ='23801'
        orga_dic["23031"] ='23801'
        orga_dic["23032"] ='23801'
        orga_dic["23033"] ='23801'
        orga_dic["23034"] ='23801'
        orga_dic["23035"] ='23801'
        orga_dic["23036"] ='23801'
        orga_dic["23037"] ='23802'
        orga_dic["23038"] ='23802'
        orga_dic["23039"] ='23802'
        orga_dic["23040"] ='23801'
        orga_dic["23041"] ='23801'
        orga_dic["23042"] ='23802'
        orga_dic["23043"] ='23802'
        orga_dic["23044"] ='23802'
        orga_dic["23045"] ='23802'
        orga_dic["23046"] ='23801'
        orga_dic["23047"] ='23801'
        orga_dic["23049"] ='23802'
        orga_dic["23051"] ='23801'
        orga_dic["23052"] ='23802'
        orga_dic["23053"] ='23802'
        orga_dic["23054"] ='23802'
        orga_dic["23055"] ='23801'
        orga_dic["23056"] ='23802'
        orga_dic["23057"] ='23801'
        orga_dic["23058"] ='23801'
        orga_dic["23059"] ='23802'
        orga_dic["23060"] ='23801'
        orga_dic["23061"] ='23801'
        orga_dic["23062"] ='23801'
        orga_dic["23063"] ='23802'
        orga_dic["23064"] ='23802'
        orga_dic["23065"] ='23802'
        orga_dic["23066"] ='23802'
        orga_dic["23067"] ='23802'
        orga_dic["23068"] ='23802'
        orga_dic["23069"] ='23802'
        orga_dic["23070"] ='23801'
        orga_dic["23071"] ='23801'
        orga_dic["23072"] ='23801'
        orga_dic["23073"] ='23801'
        orga_dic["23074"] ='23802'
        orga_dic["23075"] ='23802'
        orga_dic["23076"] ='23801'
        orga_dic["23077"] ='23802'
        orga_dic["23078"] ='23802'
        orga_dic["23079"] ='23802'
        orga_dic["23080"] ='23801'
        orga_dic["23081"] ='23801'
        orga_dic["23082"] ='23801'
        orga_dic["23083"] ='23801'
        orga_dic["23084"] ='23801'
        orga_dic["23085"] ='23802'
        orga_dic["23086"] ='23802'
        orga_dic["23087"] ='23802'
        orga_dic["23088"] ='23801'
        orga_dic["23089"] ='23801'
        orga_dic["23090"] ='23802'
        orga_dic["23091"] ='23801'
        orga_dic["23092"] ='23801'
        orga_dic["23093"] ='23802'
        orga_dic["23094"] ='23802'
        orga_dic["23096"] ='23801'
        orga_dic["23097"] ='23801'
        orga_dic["23098"] ='23802'
        orga_dic["23099"] ='23802'
        orga_dic["23701"] ='23801'
        orga_dic["23702"] ='23802'
        orga_dic["23703"] ='23803'
        orga_dic["23704"] ='23804'
        orga_dic["23705"] ='23805'
        orga_dic["23801"] ='23801'
        orga_dic["23802"] ='23802'
        orga_dic["23803"] ='23803'
        orga_dic["23804"] ='23804'
        orga_dic["23805"] ='23805'
        orga_dic["24001"] ='24801'
        orga_dic["24002"] ='24801'
        orga_dic["24003"] ='24801'
        orga_dic["24004"] ='24801'
        orga_dic["24005"] ='24802'
        orga_dic["24006"] ='24802'
        orga_dic["24007"] ='24802'
        orga_dic["24008"] ='24803'
        orga_dic["24009"] ='24803'
        orga_dic["24010"] ='24804'
        orga_dic["24011"] ='24805'
        orga_dic["24012"] ='24806'
        orga_dic["24013"] ='24807'
        orga_dic["24014"] ='24807'
        orga_dic["24015"] ='24801'
        orga_dic["24016"] ='24802'
        orga_dic["24017"] ='24803'
        orga_dic["24018"] ='24804'
        orga_dic["24019"] ='24805'
        orga_dic["24020"] ='24806'
        orga_dic["24021"] ='24807'
        orga_dic["24023"] ='24899'
        orga_dic["24708"] ='24808'
        orga_dic["24709"] ='24809'
        orga_dic["24801"] ='24801'
        orga_dic["24802"] ='24802'
        orga_dic["24803"] ='24803'
        orga_dic["24804"] ='24804'
        orga_dic["24805"] ='24805'
        orga_dic["24806"] ='24806'
        orga_dic["24807"] ='24807'
        orga_dic["24808"] ='24808'
        orga_dic["24809"] ='24809'
        orga_dic["24850"] ='24850'
        orga_dic["24851"] ='24851'
        orga_dic["24852"] ='24852'
        orga_dic["24899"] ='24899'


        ###Elitendefinitionen:
        ##Elite1 = Politische Elite (unspezifisch)
        ##Elite2 = Irgend eine Elite (unspezifisch)
        ##Elite3 = namentlich genannte Regierungsmitglieder
        ##Elite4 = namentlich genannte Minister
        ##Elite5 = namentlich genannte Koalition oder Parteifamilie
        ##Elite6 = namentlich genannte Parteiführer oder Sprecher
        ##Elite7 = namentlich genannte Parlamentsmitglieder von unserer Liste
        ##Elite8 = namentlich genannte Politiker auf unserer Liste
        ##Elite9 = namentlich genannte Politiker, die nicht auf unserer Liste sind.


        tgt = ''
        tgt_org = ''
        if data['Tgt_ID'][i] == 'MPer':
            mp = data['Def_MPer'][i]
            try:
                data['Def_Actor'][i] = eval(mp)[0]
            except:
                mp = 0
        if data['Def_Elit'][i] in ['0','11','12','13','14']:
            tgt = 'Elite1'
        elif data['Tgt_ID'][i] == 'Elit':
            tgt = 'Elite2'
        elif data['Tgt_ID'][i] == 'OwnP':
            tgt = 'Self'
        elif data['Def_Actor'][i] == data['Spr_ID'][i]:
            tgt = 'Self'
        elif data['Tgt_ID'][i] == 'ForC':
            tgt = 'Foreign'
        elif data['Tgt_ID'][i] == 'SupI':
            tgt = 'Elite2'
        elif data['Def_Volk'][i] in ['0','1','2','3','4','5','6','7']:
            tgt = 'People' ## Only abstract people in different formulations
        elif data['Tgt_ID'][i] == 'Volk' and data['Auto_Coding'][i]=='1':
            tgt = 'Part' ## Group of the population, Readers, others
        elif data['Tgt_ID'][i] == 'Volk':
            tgt = 'Part_People' ## Group of the population, Readers, others
        elif data['Tgt_ID'][i] == 'Othr':
            tgt = 'Other'
        elif len(data['Def_Actor'][i])==5:
            if data['Def_Actor'][i] in orga_dic.keys():
                tgt_org = orga_dic[data['Def_Actor'][i]]
            if data['Def_Actor'][i] in current_fct1:
                tgt = 'Elite3'
            elif data['Def_Actor'][i] in current_fct3:
                tgt = 'Elite4'
            elif data['Def_Actor'][i] in current_fct6:
                tgt = 'Elite6'
            elif data['Def_Actor'][i] in current_fct4:
                tgt = 'Elite7'
            elif data['Def_Actor'][i][2] in ['0','1','2']:
                tgt = 'Elite8'
            elif data['Def_Actor'][i][2] == '7':
                tgt = 'Elite9'
            elif data['Def_Actor'][i][-3:] in ['880','885','890','895']:
                tgt = 'Elite1'
            elif data['Def_Actor'][i][-3:] in ['899']:
                tgt = 'Party'
            elif data['Def_Actor'][i][-3:-1] in ['85']:
                tgt = 'Elite5'
            elif data['Def_Actor'][i][2] == '8':
                tgt = 'Party'
            elif data['Def_Actor'][i][-3:] in ['902','903','904','905']:
                tgt = 'Economy'
            elif data['Def_Actor'][i][-3:] in ['905','907']:
                tgt = 'NGO'
            elif data['Def_Actor'][i][-3:] in ['901']:
                tgt = 'Union'
            elif data['Def_Actor'][i][-3:] in ['910','911']:
                tgt = 'Foreign'
            elif data['Def_Actor'][i][-3:] in ['950']:
                tgt = 'Expert'
            elif data['Def_Actor'][i][-3:] in ['990','953'] and not data['Def_Actor'][i]=='16953':
                tgt = 'Private' ##Private Person, Celebrity, 
            elif data['Def_Actor'][i][-3:] in ['951','952','953']:
                tgt = 'Media'
            elif data['Def_Actor'][i] == '99999':
                tgt = 'Other'

        spr = ''
        spr_org = ''
        if len(data['Spr_ID'][i]) < 4: ### For invalid Spr_IDs
            data['Spr_ID'][i] = '99999'
        if data['Spr_ID'][i] in orga_dic.keys():
            spr_org = orga_dic[data['Spr_ID'][i]]
        if data['Spr_ID'][i] in current_fct1:
            spr = 'Elite3'
        elif data['Spr_ID'][i] in current_fct3:
            spr = 'Elite4'
        elif data['Spr_ID'][i] in current_fct6:
            spr = 'Elite6'
        elif data['Spr_ID'][i] in current_fct4:
            spr = 'Elite7'
        elif data['Spr_ID'][i][:2] in ['91','92']:
            spr = 'Journ'
            spr_org = data['Spr_ID'][i][1:]
        elif data['Spr_ID'][i][2] in ['0','1','2']:
            spr = 'Elite8'
        elif data['Spr_ID'][i][2] == '7':
            spr = 'Elite9'
        elif data['Spr_ID'][i][-3:] in ['880','885','890','895']:
            spr = 'Elite1'
        elif data['Spr_ID'][i][-3:] in ['899']:
            spr = 'Party'
        elif data['Spr_ID'][i][-3:-1] in ['85']:
            spr = 'Elite5'
        elif data['Spr_ID'][i][2] == '8':
            spr = 'Party'
        elif data['Spr_ID'][i][-3:] in ['902','903','904','905']:
            spr = 'Economy'
        elif data['Spr_ID'][i][-3:] in ['905','907']:
            spr = 'NGO'
        elif data['Spr_ID'][i][-3:] in ['901']:
            spr = 'Union'
        elif data['Spr_ID'][i][-3:] in ['910','911']:
            spr = 'Foreign'
        elif data['Spr_ID'][i][-3:] in ['950']:
            spr = 'Expert'
        elif data['Spr_ID'][i][-3:] in ['990','953'] and not data['Spr_ID'][i]=='16953':
            spr = 'Private'
        elif data['Spr_ID'][i][-3:] in ['951','952','953']:
            spr = 'Media'
        elif data['Spr_ID'][i] == '99999':
            spr = 'Other'
        

        data['Tgt_Category'].append(tgt)
        data['Spr_Category'].append(spr)

        data['Spr_Orga'].append(spr_org)
        data['Tgt_Orga'].append(tgt_org)

        if tgt in ['Elite1','Elite2','Elite3','Elite4','Elite5']:
            data['Tgt_Group'].append('Elite')
        elif tgt in ['Elite6','Elite7','Elite8','Elite9','Party','Journ'] and len(tgt_org)>1:
            data['Tgt_Group'].append(tgt_org)
        elif tgt in ['Elite6','Elite7','Elite8','Elite9','Party']:
            data['Tgt_Group'].append('Other')
        elif tgt in ['Journ','Media']:
            data['Tgt_Group'].append('Media')
        elif tgt in ['People','Part_People']:
            data['Tgt_Group'].append('People')
        else:
            data['Tgt_Group'].append(tgt)
                    

        if spr in ['Elite1','Elite2','Elite3','Elite4','Elite5']:
            data['Spr_Group'].append('Elite')
        elif spr in ['Elite6','Elite7','Elite8','Elite9','Party','Journ'] and len(spr_org)>1:
            data['Spr_Group'].append(spr_org)
        elif spr in ['Elite6','Elite7','Elite8','Elite9','Party']:
            data['Spr_Group'].append('Other')
        elif spr in ['Journ','Media']:
            data['Spr_Group'].append('Media')
        else:
            data['Spr_Group'].append(spr)


        if tgt in ['Elite1','Elite2']: ##Recodierung für Populism
            tgt = 'Elite'
        elif tgt in ['Elite3','Elite4']: ##Recodierung für extended Populism
            tgt = 'Elite_ext'

        if tgt in ['People']:  ##Recodierung für Populism
            tgt = 'People'
##            elif tgt in ['Part_People','Private']: ##Recodierung für extended Populism
##                tgt = 'People_ext'
        
        if tgt == 'Elite':
            data['Tgt_Elite'].append(1)
        elif tgt == 'Elite_ext':
            data['Tgt_Elite'].append(0.5)
        else:
            data['Tgt_Elite'].append(0)
            
        if tgt == 'People':
            data['Tgt_People'].append(1)
        elif tgt == 'People_ext': ##Gibts nicht mehr
            data['Tgt_People'].append(0.5)
        else:
            data['Tgt_People'].append(0)

        st_on_people = 0
        if tgt == 'People':
            st_on_people = 1
        elif data['Impact_Tgt'][i] in ['4']: ##Impact on People (filters for Monolith)
            st_on_people = 1
        elif data['Monolith'][i] in ['9','1','0']: ##Any other condition filtering for Monolith
            st_on_people = 1

        if st_on_people == 1:
            data['St_On_People'].append(1)
        else:
            data['St_On_People'].append(0)

        st_on_people_ext = 0
        if tgt in ['People_ext','People']:
            st_on_people_ext = 1
        elif data['Impact_Tgt'][i] in ['4']:
            st_on_people_ext = 1
        elif data['Monolith'][i] in ['9','1','0']:
            st_on_people_ext = 1

        if st_on_people_ext == 1:
            data['St_On_People_ext'].append(1)
        else:
            data['St_On_People_ext'].append(0)


        if st_on_people == 1 or tgt == 'Elite' or data['Att_Act_every'][i] in ['1','-1']:
            populism_possible = 1
        else:
            populism_possible = 0

        if st_on_people_ext == 1 or tgt in ['Elite','Elite_ext'] or data['Att_Act_every'][i] in ['1','-1']:
            populism_possible_ext = 1
        else:
            populism_possible_ext = 0


        st_on_power = 0
        power_att = ['Att_Power_gain','Att_Power_lose','Att_Power_have']
        for attribute in power_att:
            if data[attribute][i] in ['2','1','0','-1','-2']:
                st_on_power = 1
                populism_possible = 1
                populism_possible_ext = 1
        if tgt == 'Self':
            populism_possible = 1
            populism_possible_ext = 1

        data['St_On_Power'].append(st_on_power)

        if data['Auto_Coding'][i] == '1' and st_on_power == 0:
            data['Filter_Auto'].append(0)
        else:
            data['Filter_Auto'].append(1)
        

        pop = 0
        pop1 = 0 #People Centrist
        pop2 = 0 #Anti_Elitist
        pop3 = 0 #Sovereignty
        pop_ad = 0 #Advocative
        pop_con = 0 #Conflictive

        apop = 0
        apop1 = 0 #ANTI People Centrist
        apop2 = 0 #ANTI Anti_Elitist
        apop3 = 0 #ANTI Sovereignty
        apop_ad = 0 #ANTI Advocative
        apop_con = 0 #ANTI Conflictive


        pop_ext = 0
        pop1_ext = 0 #People Centrist
        pop2_ext = 0 #Anti_Elitist
        pop3_ext = 0 #Sovereignty
        pop_ad_ext = 0 #Advocative
        pop_con_ext = 0 #Conflictive

        apop_ext = 0
        apop1_ext = 0 #ANTI People Centrist
        apop2_ext = 0 #ANTI Anti_Elitist
        apop3_ext = 0 #ANTI Sovereignty
        apop_ad_ext = 0 #ANTI Advocative
        apop_con_ext = 0 #ANTI Conflictive


        
        if data['STRAT_Blame'][i] == '1' and tgt == 'Elite':
            data['POP_Blame'].append(1)
            pop = 1
            pop2 = 1
            pop_con = 1
        else:
            data['POP_Blame'].append(0)

        if data['STRAT_Blame'][i] == '1' and tgt == 'People':
            apop = 1
            apop1 = 1
            apop_con = 1
            data['APOP_Blame'].append(1)
        else:
            data['APOP_Blame'].append(0)

        if data['STRAT_Achiev'][i] == '1' and tgt == 'People':
            data['POP_Achiev'].append(1)
            pop = 1
            pop1 = 1
            pop_ad = 1
        else:
            data['POP_Achiev'].append(0)

        if data['STRAT_Achiev'][i] == '1' and tgt == 'Elite':
            apop = 1
            apop2 = 1
            apop_ad = 1
            data['APOP_Achiev'].append(1)
        else:
            data['APOP_Achiev'].append(0)
            
        if data['STRAT_Denouncing'][i] == '1' and tgt == 'Elite':
            data['POP_Denouncing'].append(1)
            pop = 1
            pop2 = 1
            pop_con = 1
        else:
            data['POP_Denouncing'].append(0)

        if data['STRAT_Denouncing'][i] == '1' and tgt == 'People':
            apop = 1
            apop1 = 1
            apop_con = 1
            data['APOP_Denouncing'].append(1)
        else:
            data['APOP_Denouncing'].append(0)
            
            
        if data['STRAT_Virtues'][i] == '1' and tgt == 'People':
            data['POP_Virtues'].append(1)
            pop = 1
            pop1 = 1
            pop_ad = 1
        else:
            data['POP_Virtues'].append(0)


        if data['STRAT_Virtues'][i] == '1' and tgt == 'Elite':
            apop = 1
            apop2 = 1
            apop_ad = 1
            data['APOP_Virtues'].append(1)
        else:
            data['APOP_Virtues'].append(0)

        if data['STRAT_Exclusion'][i] == '1' and tgt == 'Elite':
            data['POP_Exclusion_Elite'].append(1)
            pop = 1
            pop2 = 1
            pop_con = 1
        else:
            data['POP_Exclusion_Elite'].append(0)

        if data['STRAT_Exclusion'][i] == '1' and not tgt == 'Elite':
            data['POP_Exclusion_Someone'].append(1)
        else:
            data['POP_Exclusion_Someone'].append(0)


        if (data['STRAT_Closeness'][i] == '1' and tgt == 'Self') or data['Embod'][i] in ['2','3'] or (data['Agreement']=='1' and tgt=='People'):
            data['POP_Closeness_Self'].append(1)
            pop = 1
            pop1 = 1
            pop_ad = 1
        else:
            data['POP_Closeness_Self'].append(0)

        if data['STRAT_Closeness'][i] == '1' and tgt in ['Spec_Person','Unsp_Person','Party']:
            data['POP_Closeness_Someone'].append(1)
        else:
            data['POP_Closeness_Someone'].append(0)

        if data['STRAT_Closeness'][i] == '1' and tgt == 'Elite':
            apop = 1
            apop2 = 1
            apop_ad = 1
            data['APOP_Closeness_Elite'].append(1)
        else:
            data['APOP_Closeness_Elite'].append(0)

        if data['Monolith'][i] == '1':
            if tgt == 'people' and tonepos < toneneg:
                data['POP_Monolith'].append(0)
            else:
                data['POP_Monolith'].append(1)
                pop = 1
                pop1 = 1
                pop_ad = 1
        else:
            data['POP_Monolith'].append(0)
        
        if data['Monolith'][i] == '0':
            apop = 1
            apop1 = 1
            apop_ad = 1
            data['APOP_Pluralist'].append(1)
        else:
            data['APOP_Pluralist'].append(0)


        if data['STRAT_Sovereign_Pro'][i] == '1' and tgt == 'People':
            data['POP_Sovereign_adv'].append(1)
            pop = 1
            pop3 = 1
            pop_ad = 1
        else:
            data['POP_Sovereign_adv'].append(0)

        if data['STRAT_Sovereign_Con'][i] == '1' and tgt == 'Elite':
            data['POP_Sovereign_con'].append(1)
            pop = 1
            pop3 = 1
            pop_con = 1
        else:
            data['POP_Sovereign_con'].append(0)

        if data['STRAT_Sovereign_Pro'][i] == '1' and tgt == 'Elite' and tonepos >= toneneg:
            apop = 1
            apop3 = 1
            apop_ad = 1
            data['APOP_Sovereign_adv'].append(1)
        else:
            data['APOP_Sovereign_adv'].append(0)

        if data['STRAT_Sovereign_Con'][i] == '1' and tgt == 'People' and toneneg >= tonepos:
            apop = 1
            apop3 = 1
            apop_con = 1
            data['APOP_Sovereign_con'].append(1)
        else:
            data['APOP_Sovereign_con'].append(0)

        data['POPULIST'].append(pop)
        data['POPULIST_PeopleCent'].append(pop1)
        data['POPULIST_AntiElite'].append(pop2)
        data['POPULIST_Sovereign'].append(pop3)
        data['POPULIST_Advocative'].append(pop_ad)
        data['POPULIST_Conflictive'].append(pop_con)
        data['ANTIPOPULIST'].append(apop)
        data['APOPULIST_PeopleCent'].append(apop1)
        data['APOPULIST_AntiElite'].append(apop2)
        data['APOPULIST_Sovereign'].append(apop3)
        data['APOPULIST_Advocative'].append(apop_ad)
        data['APOPULIST_Conflictive'].append(apop_con)

        if st_on_people == 1:
            data['POPSHARE_PeopleCent'].append(pop1)
            data['POPULIST_PC_BIAS'].append(pop1-apop1)
        else:
            data['POPSHARE_PeopleCent'].append('')
            data['POPULIST_PC_BIAS'].append('')

        if tgt == 'Elite':
            data['POPSHARE_AntiElite'].append(pop2)
            data['POPULIST_AE_BIAS'].append(pop2-apop2)
        else:
            data['POPSHARE_AntiElite'].append('')
            data['POPULIST_AE_BIAS'].append('')

        if st_on_power == 1:
            data['POPSHARE_Sovereign'].append(pop3)
            data['POPULIST_PS_BIAS'].append(pop3-apop3)
        else:
            data['POPSHARE_Sovereign'].append('')
            data['POPULIST_PS_BIAS'].append('')

        if tgt == 'People':
            data['POPSHARE_Advocative'].append(pop_ad)
        else:
            data['POPSHARE_Advocative'].append('')

        if tgt == 'Elite':
            data['POPSHARE_Conflictive'].append(pop_con)
        else:
            data['POPSHARE_Conflictive'].append('')
            
        if populism_possible == 1:
            data['POPULIST_BIAS'].append(pop-apop)
            data['Filter_Pop_possible'].append(1)
        else:
            data['POPULIST_BIAS'].append('')
            data['Filter_Pop_possible'].append(0)


#### Extended Populism

        if data['STRAT_Blame'][i] == '1' and tgt in ['Elite','Elite_ext']:
            data['POP_Blame_ext'].append(1)
            pop_ext = 1
            pop2_ext = 1
            pop_con_ext = 1
        else:
            data['POP_Blame_ext'].append(0)

        if data['STRAT_Blame'][i] == '1' and tgt in ['People','People_ext']:
            apop_ext = 1
            apop1_ext = 1
            apop_con_ext = 1
            data['APOP_Blame_ext'].append(1)
        else:
            data['APOP_Blame_ext'].append(0)

        if data['STRAT_Achiev'][i] == '1' and tgt in ['People','People_ext']:
            data['POP_Achiev_ext'].append(1)
            pop_ext = 1
            pop1_ext = 1
            pop_ad_ext = 1
        else:
            data['POP_Achiev_ext'].append(0)

        if data['STRAT_Achiev'][i] == '1' and tgt in ['Elite','Elite_ext']:
            apop_ext = 1
            apop2_ext = 1
            apop_ad_ext = 1
            data['APOP_Achiev_ext'].append(1)
        else:
            data['APOP_Achiev_ext'].append(0)
            
        if data['STRAT_Denouncing'][i] == '1' and tgt in ['Elite','Elite_ext']:
            data['POP_Denouncing_ext'].append(1)
            pop_ext = 1
            pop2_ext = 1
            pop_con_ext = 1
        else:
            data['POP_Denouncing_ext'].append(0)

        if data['STRAT_Denouncing'][i] == '1' and tgt in ['People','People_ext']:
            apop_ext = 1
            apop1_ext = 1
            apop_con_ext = 1
            data['APOP_Denouncing_ext'].append(1)
        else:
            data['APOP_Denouncing_ext'].append(0)
            
            
        if data['STRAT_Virtues'][i] == '1' and tgt in ['People','People_ext']:
            data['POP_Virtues_ext'].append(1)
            pop_ext = 1
            pop1_ext = 1
            pop_ad_ext = 1
        else:
            data['POP_Virtues_ext'].append(0)


        if data['STRAT_Virtues'][i] == '1' and tgt in ['Elite','Elite_ext']:
            apop_ext = 1
            apop2_ext = 1
            apop_ad_ext = 1
            data['APOP_Virtues_ext'].append(1)
        else:
            data['APOP_Virtues_ext'].append(0)

        if data['STRAT_Exclusion'][i] == '1' and tgt in ['Elite','Elite_ext']:
            data['POP_Exclusion_Elite_ext'].append(1)
            pop_ext = 1
            pop2_ext = 1
            pop_con_ext = 1
        else:
            data['POP_Exclusion_Elite_ext'].append(0)

        if data['STRAT_Exclusion'][i] == '1' and not tgt in ['Elite','Elite_ext']:
            data['POP_Exclusion_Someone_ext'].append(1)
        else:
            data['POP_Exclusion_Someone_ext'].append(0)


        if (data['STRAT_Closeness'][i] == '1' and tgt == 'Self') or data['Embod'][i] in ['1','2','3'] or (data['Agreement']=='1' and tgt=='People'):
            data['POP_Closeness_Self_ext'].append(1)
            pop_ext = 1
            pop1_ext = 1
            pop_ad_ext = 1
        else:
            data['POP_Closeness_Self_ext'].append(0)

        if data['STRAT_Closeness'][i] == '1' and tgt in ['Spec_Person','Unsp_Person','Party']:
            data['POP_Closeness_Someone_ext'].append(1)
        else:
            data['POP_Closeness_Someone_ext'].append(0)

        if data['STRAT_Closeness'][i] == '1' and tgt in ['Elite','Elite_ext']:
            apop_ext = 1
            apop2_ext = 1
            apop_ad_ext = 1
            data['APOP_Closeness_Elite_ext'].append(1)
        else:
            data['APOP_Closeness_Elite_ext'].append(0)

        if data['Monolith'][i] == '1':
            data['POP_Monolith_ext'].append(1)
            pop_ext = 1
            pop1_ext = 1
            pop_ad_ext = 1
        else:
            data['POP_Monolith_ext'].append(0)
        
        if data['Monolith'][i] == '0':
            apop_ext = 1
            apop1_ext = 1
            apop_ad_ext = 1
            data['APOP_Pluralist_ext'].append(1)
        else:
            data['APOP_Pluralist_ext'].append(0)


        if data['STRAT_Sovereign_Pro'][i] == '1' and tgt in ['People','People_ext']:
            data['POP_Sovereign_adv_ext'].append(1)
            pop_ext = 1
            pop3_ext = 1
            pop_ad_ext = 1
        else:
            data['POP_Sovereign_adv_ext'].append(0)

        if data['STRAT_Sovereign_Con'][i] == '1' and tgt in ['Elite','Elite_ext']:
            data['POP_Sovereign_con_ext'].append(1)
            pop_ext = 1
            pop3_ext = 1
            pop_con_ext = 1
        else:
            data['POP_Sovereign_con_ext'].append(0)

        if data['STRAT_Sovereign_Pro'][i] == '1' and tgt in ['Elite','Elite_ext']:
            apop_ext = 1
            apop3_ext = 1
            apop_ad_ext = 1
            data['APOP_Sovereign_adv_ext'].append(1)
        else:
            data['APOP_Sovereign_adv_ext'].append(0)

        if data['STRAT_Sovereign_Con'][i] == '1' and tgt in ['People','People_ext']:
            apop_ext = 1
            apop3_ext = 1
            apop_con_ext = 1
            data['APOP_Sovereign_con_ext'].append(1)
        else:
            data['APOP_Sovereign_con_ext'].append(0)

        data['POPULIST_ext'].append(pop_ext)
        data['POPULIST_PeopleCent_ext'].append(pop1_ext)
        data['POPULIST_AntiElite_ext'].append(pop2_ext)
        data['POPULIST_Sovereign_ext'].append(pop3_ext)
        data['POPULIST_Advocative_ext'].append(pop_ad_ext)
        data['POPULIST_Conflictive_ext'].append(pop_con_ext)
        data['ANTIPOPULIST_ext'].append(apop_ext)
        data['APOPULIST_PeopleCent_ext'].append(apop1_ext)
        data['APOPULIST_AntiElite_ext'].append(apop2_ext)
        data['APOPULIST_Sovereign_ext'].append(apop3_ext)
        data['APOPULIST_Advocative_ext'].append(apop_ad_ext)
        data['APOPULIST_Conflictive_ext'].append(apop_con_ext)

        if st_on_people_ext == 1:
            data['POPSHARE_PeopleCent_ext'].append(pop1_ext)
            data['POPULIST_PC_BIAS_ext'].append(pop1_ext-apop1_ext)
        else:
            data['POPSHARE_PeopleCent_ext'].append('')
            data['POPULIST_PC_BIAS_ext'].append('')

        if tgt in ['Elite','Elite_ext']:
            data['POPSHARE_AntiElite_ext'].append(pop2_ext)
            data['POPULIST_AE_BIAS_ext'].append(pop2_ext-apop2_ext)
        else:
            data['POPSHARE_AntiElite_ext'].append('')
            data['POPULIST_AE_BIAS_ext'].append('')

        if st_on_power == 1:
            data['POPSHARE_Sovereign_ext'].append(pop3_ext)
            data['POPULIST_PS_BIAS_ext'].append(pop3_ext-apop3_ext)
        else:
            data['POPSHARE_Sovereign_ext'].append('')
            data['POPULIST_PS_BIAS_ext'].append('')

        if st_on_people_ext == 1:
            data['POPSHARE_Advocative_ext'].append(pop_ad_ext)
        else:
            data['POPSHARE_Advocative_ext'].append('')

        if tgt in ['Elite','Elite_ext']:
            data['POPSHARE_Conflictive_ext'].append(pop_con_ext)
        else:
            data['POPSHARE_Conflictive_ext'].append('')
            
        if populism_possible_ext == 1:
            data['POPULIST_BIAS_ext'].append(pop_ext-apop_ext)
            data['Filter_Pop_possible_ext'].append(1)
        else:
            data['POPULIST_BIAS_ext'].append('')
            data['Filter_Pop_possible_ext'].append(0)


        if data['Def_Volk'][i] in ['901','17','1001','0901']:
            migneg = 0
            migpos = 0
            if toneneg > 0:
                migneg = 1
            if tonepos > 0:
                migpos = 1
            data['Migrants_Neg'].append(migneg)
            data['Migrants_Pos'].append(migpos)
            data['Migrants_Mentioned'].append(1)
        else:
            data['Migrants_Neg'].append(0)
            data['Migrants_Pos'].append(0)
            data['Migrants_Mentioned'].append(0)
            

        if toneneg > 0 and data['Def_Elit'][i] in ['21','22','31','32']:
            data['LW_Attack'].append(1)
        else:
            data['LW_Attack'].append(0)


    verbout('\nDone.\nPreparing output...',master=master)

    return (data,['STRAT_Blame','STRAT_Achiev','STRAT_Sovereign_Pro','STRAT_Sovereign_Con',
                  'Filter_Auto','Spr_Category','Spr_Orga','Spr_Group','Tgt_Category','Tgt_Orga','Tgt_Group',
                  'Tgt_Elite','Tgt_People','St_On_People','St_On_Power',
                  'Tone_Pos_Grp','Tone_Pos_Mean',
                  'POP_Blame','POP_Achiev','POP_Denouncing','POP_Virtues','POP_Exclusion_Elite',
                  'POP_Exclusion_Someone','POP_Closeness_Self','POP_Closeness_Someone',
                  'POP_Sovereign_adv','POP_Sovereign_con','POP_Monolith',
                  'APOP_Blame','APOP_Achiev','APOP_Denouncing','APOP_Virtues',
                  'APOP_Closeness_Elite','APOP_Sovereign_adv','APOP_Sovereign_con','APOP_Pluralist',
                  'POPULIST','POPULIST_PeopleCent','POPULIST_AntiElite','POPULIST_Sovereign',
                  'POPULIST_Advocative','POPULIST_Conflictive',
                  'ANTIPOPULIST','APOPULIST_PeopleCent','APOPULIST_AntiElite','APOPULIST_Sovereign',
                  'APOPULIST_Advocative','APOPULIST_Conflictive',
                  'POPSHARE_Advocative','POPSHARE_Conflictive','POPSHARE_PeopleCent',
                  'POPSHARE_AntiElite','POPSHARE_Sovereign','POPULIST_BIAS',
                  'POPULIST_PC_BIAS','POPULIST_AE_BIAS','POPULIST_PS_BIAS',

                  'POP_Blame_ext','POP_Achiev_ext','POP_Denouncing_ext','POP_Virtues_ext','POP_Exclusion_Elite_ext',
                  'POP_Exclusion_Someone_ext','POP_Closeness_Self_ext','POP_Closeness_Someone_ext',
                  'POP_Sovereign_adv_ext','POP_Sovereign_con_ext','POP_Monolith_ext',
                  'APOP_Blame_ext','APOP_Achiev_ext','APOP_Denouncing_ext','APOP_Virtues_ext',
                  'APOP_Closeness_Elite_ext','APOP_Sovereign_adv_ext','APOP_Sovereign_con_ext','APOP_Pluralist_ext',
                  'POPULIST_ext','POPULIST_PeopleCent_ext','POPULIST_AntiElite_ext','POPULIST_Sovereign_ext',
                  'POPULIST_Advocative_ext','POPULIST_Conflictive_ext',
                  'ANTIPOPULIST_ext','APOPULIST_PeopleCent_ext','APOPULIST_AntiElite_ext','APOPULIST_Sovereign_ext',
                  'APOPULIST_Advocative_ext','APOPULIST_Conflictive_ext',
                  'POPSHARE_Advocative_ext','POPSHARE_Conflictive_ext','POPSHARE_PeopleCent_ext',
                  'POPSHARE_AntiElite_ext','POPSHARE_Sovereign_ext','POPULIST_BIAS_ext',
                  'POPULIST_PC_BIAS_ext','POPULIST_AE_BIAS_ext','POPULIST_PS_BIAS_ext',
                  
                  'Migrants_Mentioned','Migrants_Neg','Migrants_Pos','LW_Attack','Filter_Pop_possible','Filter_Pop_possible_ext'])




def calc_timecoef(x,y,dweight,currdate=0,master=''): ##Matching support function: Calculate weighting by time
    if len(x) == len(y):
        xcorr = []
        ycorr = []
        for i in range(len(x)):
            if type(x[i]) in [float, int] and type(y[i]) in [float,int] and dweight[i] > 0.01:
                xcorr.append(x[i])
                ycorr.append(y[i])

        if len(xcorr) > 2:
            reg = calc_regression(xcorr,ycorr)
            r = calc_correlation(xcorr,ycorr)
            st = stat_desc(ycorr, verbose=0)
            recent = recentpeak(xcorr,ycorr,st['M']+1.97*st['SD'])
            werte = (reg[0],r,st['Vari'],currdate-recent,xcorr,ycorr)
        else:
            werte = ('','','','',xcorr,ycorr)
    else:
        werte = ('','','','',xcorr,ycorr)
        
    return werte



def match_nccr(sdata,mdata,wmode,dmode,cmode,suffix='',master=''):
    verbout('\nWeighting Mode: '+str(wmode),master=master)
    verbout('\nDate Mode: '+str(dmode),master=master)
    verbout('\nCalculation Mode: '+str(cmode),master=master)

    available_media = sorted(mdata[list(mdata.keys())[0]].keys())
    mvariables = sorted(mdata[list(mdata.keys())[0]][available_media[0]].keys())
    if 'res_Weight' in mvariables:
        mvariables.remove('res_Weight')
    verbout('\nMedia in Content Data: '+str(available_media),master=master)

    possible_media = []
    for a in available_media:
        if a in sdata.keys():
            possible_media.append(a)

    verbout('\n..of which in Survey: '+str(possible_media),master=master)

    verbout('\n\nMatching media content variables: '+str(mvariables)+'\n',master=master)
    add_dic = {}
    table_part_date = {}
    for v in mvariables:
        add_dic[v] = []
        table_part_date[v] = {}
        for d in mdata.keys():
            table_part_date[v][d] = {}
        

    step = int(len(sdata['res_SDate'])/40)
    if step<1: step = 1
    verbout('\nMatching up media content for each case: \n0%-------25%-------50%-------75%-------100%\n','progress',master=master)

    for i in range(len(sdata['res_SDate'])):
        if i%step == 0: verbout('.','progress',master=master)

        part_date = sdata['res_SDate'][i]
        part_weight = sdata['res_GWeight'][i]

        media_diet = {}
        for v in mvariables:
            media_diet[v] = []
        
        date_weight = [] ##Weight of dates
        date_num = []

        personal_media = []
        for m in possible_media:
            if sdata[m][i]>0:
                personal_media.append(m)

        if len(personal_media) > 0:
            for d in sorted(mdata.keys()):
                for v in mvariables:
                    table_part_date[v][d][i] = ''
                    daily_diet = []
                    intradaily_weight = []
                    for m in personal_media:
                        daily_diet.append(mdata[d][m][v])

                        if wmode == 'none':
                            if sdata['res_GWeight'][i] > 0:
                                idw = 1
                            else:
                                idw = 0                                
                        elif wmode == 'linear':
                            idw = sdata['res_GWeight'][i]
                        elif wmode == 'grouped':
                            if sdata['res_GWeight'][i] == 0:
                                idw = 0
                            elif sdata['res_GWeight'][i] <= 3:
                                idw = 1
                            else:
                                idw = 2
                        else:
                            verbout('\nStrange error: Not able to find the setting for weight mode "'+wmode+'". Not weighting by medium.',master=master)
                            idw = 1
                        intradaily_weight.append(idw)

                    intrasum = 0.0
                    intraanz = 0.0
                    for k in range(len(intradaily_weight)):
                        if not daily_diet[k] == '':
                            intrasum = intrasum + float(daily_diet[k]) * intradaily_weight[k]
                            intraanz = intraanz + intradaily_weight[k]

                    if intraanz > 0:
                        if cmode in ['sum_sum','mean_sum']:
                            intraday = intrasum
                        elif cmode in ['sum_mean','mean_mean','timeser']:
                            intraday = intrasum / intraanz
                    else:
                        intraday = ''

                    media_diet[v].append(intraday)
                    

                dist = float(part_date - d)
                dw_final = ''
                if dmode == 'nodate':
                    dw_final = 1
                elif dmode == 'before':
                    if dist > 0:
                        dw_final = 1
                    else:
                        dw_final = 0
                elif dmode == 'after':
                    if dist < 0:
                        dw_final = 1
                    else:
                        dw_final = 0
                elif dmode == '1d':
                    if dist > 0 and dist < 10:
                        dw_final = 1.0/2**(dist-1) ## Nicht der Befragungstag, sondern der Vortag wird genommen
                    else:
                        dw_final = 0
                elif dmode == '7d':
                    if dist > 0 and dist < 70:
                        dw_final = 1.0/2**(dist/7)
                    else:
                        dw_final = 0
                elif dmode == 'div':
                    if dist > 0:
                        dw_final = 1.0/dist
                elif dmode == '1mbefore':
                    if dist > 0 and dist < 30:
                        dw_final = 1.0
                    else:
                        dw_final = 0.0
                elif dmode == '2mbefore':
                    if dist > 0 and dist < 60:
                        dw_final = 1.0
                    else:
                        dw_final = 0.0
                else:
                    dw_final = 1
                    verbout('\nStrange error: Not able to find the setting for date mode "'+dmode+'". Not weighting by date.',master=master)

                date_weight.append(dw_final) ## Gewicht des Tages
                date_num.append(d) ##Nummer des Tages. Excel Date

        ## Calculating Score over all Days

        for v in mvariables:
            content_sum = 0.0
            content_anz = 0.0

            if cmode in ['sum_sum','sum_mean','mean_sum','mean_mean'] or v == 'N_Texts': ##Falls Mittel oder Summe oder N
                for k in range(len(date_weight)):
                    if type(media_diet[v][k]) in [float,int]:
                        content_sum = content_sum + media_diet[v][k]*date_weight[k]
                        content_anz = content_anz + date_weight[k]

                if content_anz > 0:
                    if cmode in ['sum_sum','sum_mean']:
                        add_dic[v].append(content_sum)
                    elif cmode in ['mean_sum','mean_mean']:
                        add_dic[v].append(content_sum/content_anz)
                    elif cmode == 'timeser': ## Only N_Texts would be in this category
                        add_dic[v].append(content_sum)
                    else:
                        add_dic[v].append('')                    
                else:
                    add_dic[v].append('')

            elif cmode == 'timeser':
                tc = calc_timecoef(date_num,media_diet[v],date_weight,part_date)
                if settings['Debugging'] in [1,'1']:
                    title = 'ID: '+str(i)+'; V: '+v+'; Correlation = '+str(tc[1])
                    master.display_line_plot({'X':tc[4],'Y':tc[5],'Title':title})
                add_dic[v].append(tc[:4])

            for resp in range(len(date_num)):
                d = date_num[resp]
                table_part_date[v][d][i] = media_diet[v][resp]


    ### Attaching added values to survey data

    add_var = []

    for v in sorted(add_dic.keys()):
        if cmode in ['sum_sum','sum_mean','mean_sum','mean_mean'] or v == 'N_Texts':
            varlab = 'MATCHED_' + v + suffix
            varnr = 0
            while varlab in sdata.keys():
                varnr = varnr + 1
                varlab = 'MATCHED_' + v + suffix + "{0:03}".format(varnr)
            sdata[varlab] = add_dic[v]
            add_var.append(varlab)
        else:
            v1 = 'MATCH_CALC_B_' + v + suffix
            v2 = 'MATCH_CALC_Beta_' + v + suffix
            v3 = 'MATCH_CALC_Var_' + v + suffix
            v4 = 'MATCH_CALC_LPeak_' + v + suffix
            varnr = 0
            while v1 in sdata.keys() or v2 in sdata.keys() or v3 in sdata.keys() or v4 in sdata.keys():
                varnr = varnr + 1
                v1 = 'MATCH_CALC_B_' + v + suffix+ "{0:03}".format(varnr)
                v2 = 'MATCH_CALC_Beta_' + v + suffix+ "{0:03}".format(varnr)
                v3 = 'MATCH_CALC_Var_' + v + suffix+ "{0:03}".format(varnr)
                v4 = 'MATCH_CALC_LPeak_' + v + suffix+ "{0:03}".format(varnr)

            sdata[v1] = list(list(zip(*add_dic[v]))[0])
            sdata[v2] = list(list(zip(*add_dic[v]))[1])
            sdata[v3] = list(list(zip(*add_dic[v]))[2])
            sdata[v4] = list(list(zip(*add_dic[v]))[3])
            
            add_var = add_var + [v1,v2,v3,v4]

##        if cmode in ['sum_sum','sum_mean','mean_sum','mean_mean']:
##            for v in sorted(add_dic.keys()):
##                varlab = 'MATCHED_' + v + suffix
##                if varlab in sdata.keys():
##                    varnr = 1
##                    varlab = 'MATCHED_' + v + suffix + "{0:03}".format(varnr)
##                    while varlab in sdata.keys():
##                        varnr = varnr + 1
##                        varlab = 'MATCHED_' + v + suffix + "{0:03}".format(varnr)
##                sdata[varlab] = add_dic[v]
##                add_var.append(varlab)
##                
##        elif cmode == 'timeser':
##            for v in sorted(add_dic.keys()):
##                v1 = 'MATCH_CALC_B_' + v + suffix
##                v2 = 'MATCH_CALC_SE_' + v + suffix
##                v3 = 'MATCH_CALC_Var_' + v + suffix
##                v4 = 'MATCH_CALC_Peak_' + v + suffix
##
##                varnr = 0
##                while v1 in sdata.keys() or v2 in sdata.keys() or v3 in sdata.keys() or v4 in sdata.keys():
##                    varnr = varnr + 1
##                    v1 = 'MATCH_CALC_B_' + v + suffix+ "{0:03}".format(varnr)
##                    v2 = 'MATCH_CALC_SE_' + v + suffix+ "{0:03}".format(varnr)
##                    v3 = 'MATCH_CALC_Var_' + v + suffix+ "{0:03}".format(varnr)
##                    v4 = 'MATCH_CALC_Peak_' + v + suffix+ "{0:03}".format(varnr)
##                
##                sdata[v1] = list(zip(*add_dic[v])[0])
##                sdata[v2] = list(zip(*add_dic[v])[1])
##                sdata[v3] = list(zip(*add_dic[v])[2])
##                sdata[v4] = list(zip(*add_dic[v])[3])
##                
##                add_var = add_var + [v1,v2,v3,v4]


    if settings['Debugging'] in [1,'1']:
        verbout('\nWriting respondent daywise dumps.\n',master=master)

        for v in mvariables:
            dump_dic = {'Date':sorted(mdata.keys())}
            for i in range(len(sdata['res_SDate'])):
                rid = sdata['responseidL1'][i]
                rid = 'Resp_'+str(rid)
                dump_dic[rid] = []
                for d in dump_dic['Date']:
                    try:
                        dump_dic[rid].append(table_part_date[v][d][i])
                    except:
                        dump_dic[rid].append('')
            t = write_data(dump_dic,sorted(dump_dic.keys()),'Dump_'+v+'.txt')
            verbout(t[0],master=self)
            

    verbout('\ndone.\n',master=master)          
    
    return sdata, add_var
    


def reformat_mediause(data,master=''):
    varlist = list(data.keys())
    ovar = {} ##Online Media Use variables
    mvar = {} ##Print Media Use variables
    mvar['1373'] = {'Panel_Dicho':'ID71_CH_1', 'Panel_Freq':'#Not known'}
    mvar['1374'] = {'Panel_Dicho':'ID71_CH_2', 'Panel_Freq':'#Not known'}
    mvar['1573'] = {'Panel_Dicho':'ID71_DE_1', 'Panel_Freq':'#Not known'}
    mvar['1574'] = {'Panel_Dicho':'ID71_DE_3', 'Panel_Freq':'#Not known'}
    mvar['2273'] = {'Panel_Dicho':'ID71_UK_1', 'Panel_Freq':'#Not known'}
    mvar['2274'] = {'Panel_Dicho':'ID71_UK_3', 'Panel_Freq':'#Not known'}
    mvar['1773'] = {'Panel_Dicho':'ID71_FR_2', 'Panel_Freq':'#Not known'}
    mvar['1774'] = {'Panel_Dicho':'ID71_FR_3', 'Panel_Freq':'#Not known'}
    mvar['1301'] = {'Panel_Dicho':'ID67_CH_1', 'Panel_Freq':'#Not known'}
    mvar['1302'] = {'Panel_Dicho':'ID67_CH_2', 'Panel_Freq':'#Not known'}
    mvar['1303'] = {'Panel_Dicho':'ID67_CH_3', 'Panel_Freq':'#Not known'}
    mvar['1304'] = {'Panel_Dicho':'ID67_CH_4', 'Panel_Freq':'#Not known'}
    mvar['1309'] = {'Panel_Dicho':'ID67_CH_6', 'Panel_Freq':'#Not known'}
    mvar['1310'] = {'Panel_Dicho':'ID67_CH_7', 'Panel_Freq':'#Not known'}
    mvar['1402'] = {'Panel_Dicho':'ID67_CH_10', 'Panel_Freq':'#Not known'}
    mvar['1401'] = {'Panel_Dicho':'ID67_CH_11', 'Panel_Freq':'#Not known'}
    mvar['1502'] = {'Panel_Dicho':'ID67_DE_1', 'Panel_Freq':'#Not known'}
    mvar['1501'] = {'Panel_Dicho':'ID67_DE_2', 'Panel_Freq':'#Not known'}
    mvar['1503'] = {'Panel_Dicho':'ID67_DE_5', 'Panel_Freq':'#Not known'}
    mvar['1507'] = {'Panel_Dicho':'ID67_DE_6', 'Panel_Freq':'#Not known'}
    mvar['1509'] = {'Panel_Dicho':'ID67_DE_7', 'Panel_Freq':'#Not known'}
    mvar['1510'] = {'Panel_Dicho':'ID67_DE_8', 'Panel_Freq':'#Not known'}
    mvar['1504'] = {'Panel_Dicho':'ID67_DE_9', 'Panel_Freq':'#Not known'}
    mvar['1511'] = {'Panel_Dicho':'ID67_DE_10', 'Panel_Freq':'#Not known'}
    mvar['1512'] = {'Panel_Dicho':'ID67_DE_11', 'Panel_Freq':'#Not known'}
    mvar['2203'] = {'Panel_Dicho':'ID67_UK_2', 'Panel_Freq':'#Not known'}
    mvar['2204'] = {'Panel_Dicho':'ID67_UK_4', 'Panel_Freq':'#Not known'}
    mvar['2201'] = {'Panel_Dicho':'ID67_UK_7', 'Panel_Freq':'#Not known'}
    mvar['2202'] = {'Panel_Dicho':'ID67_UK_9', 'Panel_Freq':'#Not known'}
    mvar['2207'] = {'Panel_Dicho':'ID67_UK_11', 'Panel_Freq':'#Not known'}
    mvar['2209'] = {'Panel_Dicho':'ID67_UK_12', 'Panel_Freq':'#Not known'}
    mvar['2208'] = {'Panel_Dicho':'ID67_UK_13', 'Panel_Freq':'#Not known'}
    mvar['1701'] = {'Panel_Dicho':'ID67_FR_1', 'Panel_Freq':'#Not known'}
    mvar['1702'] = {'Panel_Dicho':'ID67_FR_2', 'Panel_Freq':'#Not known'}
    mvar['1703'] = {'Panel_Dicho':'ID67_FR_3', 'Panel_Freq':'#Not known'}
    mvar['1704'] = {'Panel_Dicho':'ID67_FR_4', 'Panel_Freq':'#Not known'}
    mvar['1707'] = {'Panel_Dicho':'ID67_FR_10', 'Panel_Freq':'#Not known'}
    mvar['1708'] = {'Panel_Dicho':'ID67_FR_11', 'Panel_Freq':'#Not known'}
    mvar['1709'] = {'Panel_Dicho':'ID67_FR_12', 'Panel_Freq':'#Not known'}
    mvar['1306'] = {'Panel_Dicho':'ID68_CH_1', 'Panel_Freq':'#Not known'}
    mvar['1305'] = {'Panel_Dicho':'ID68_CH_2', 'Panel_Freq':'#Not known'}
    mvar['1307'] = {'Panel_Dicho':'ID68_CH_3', 'Panel_Freq':'#Not known'}
    mvar['1308'] = {'Panel_Dicho':'ID68_CH_4', 'Panel_Freq':'#Not known'}
    mvar['1505'] = {'Panel_Dicho':'ID68_DE_5', 'Panel_Freq':'#Not known'}
    mvar['1506'] = {'Panel_Dicho':'ID68_DE_6', 'Panel_Freq':'#Not known'}
    mvar['2206'] = {'Panel_Dicho':'ID68_UK_2', 'Panel_Freq':'#Not known'}
    mvar['2205'] = {'Panel_Dicho':'ID68_UK_7', 'Panel_Freq':'#Not known'}
    mvar['2210'] = {'Panel_Dicho':'ID68_UK_8', 'Panel_Freq':'#Not known'}
    mvar['2211'] = {'Panel_Dicho':'ID68_UK_9', 'Panel_Freq':'#Not known'}
    mvar['1705'] = {'Panel_Dicho':'ID68_FR_1', 'Panel_Freq':'#Not known'}
    mvar['1706'] = {'Panel_Dicho':'ID68_FR_2', 'Panel_Freq':'#Not known'}
    mvar['1301']['Panel_Freq'] = 'ID69_CH_1'
    mvar['1302']['Panel_Freq'] = 'ID69_CH_2'
    mvar['1303']['Panel_Freq'] = 'ID69_CH_3'
    mvar['1304']['Panel_Freq'] = 'ID69_CH_4'
    mvar['1309']['Panel_Freq'] = 'ID69_CH_6'
    mvar['1310']['Panel_Freq'] = 'ID69_CH_7'
    mvar['1402']['Panel_Freq'] = 'ID69_CH_10'
    mvar['1401']['Panel_Freq'] = 'ID69_CH_11'
    mvar['1502']['Panel_Freq'] = 'ID69_DE_1'
    mvar['1501']['Panel_Freq'] = 'ID69_DE_2'
    mvar['1503']['Panel_Freq'] = 'ID69_DE_5'
    mvar['1507']['Panel_Freq'] = 'ID69_DE_6'
    mvar['1509']['Panel_Freq'] = 'ID69_DE_7'
    mvar['1510']['Panel_Freq'] = 'ID69_DE_8'
    mvar['1504']['Panel_Freq'] = 'ID69_DE_9'
    mvar['1511']['Panel_Freq'] = 'ID69_DE_10'
    mvar['1512']['Panel_Freq'] = 'ID69_DE_11'
    mvar['2203']['Panel_Freq'] = 'ID69_UK_2'
    mvar['2204']['Panel_Freq'] = 'ID69_UK_4'
    mvar['2201']['Panel_Freq'] = 'ID69_UK_7'
    mvar['2202']['Panel_Freq'] = 'ID69_UK_9'
    mvar['2207']['Panel_Freq'] = 'ID69_UK_11'
    mvar['2209']['Panel_Freq'] = 'ID69_UK_12'
    mvar['2208']['Panel_Freq'] = 'ID69_UK_13'
    mvar['1701']['Panel_Freq'] = 'ID69_FR_1'
    mvar['1702']['Panel_Freq'] = 'ID69_FR_2'
    mvar['1703']['Panel_Freq'] = 'ID69_FR_3'
    mvar['1704']['Panel_Freq'] = 'ID69_FR_4'
    mvar['1707']['Panel_Freq'] = 'ID69_FR_10'
    mvar['1708']['Panel_Freq'] = 'ID69_FR_11'
    mvar['1709']['Panel_Freq'] = 'ID69_FR_12'
    mvar['1306']['Panel_Freq'] = 'ID77_CH_1'
    mvar['1305']['Panel_Freq'] = 'ID77_CH_2'
    mvar['1307']['Panel_Freq'] = 'ID77_CH_3'
    mvar['1308']['Panel_Freq'] = 'ID77_CH_4'
    mvar['1505']['Panel_Freq'] = 'ID77_DE_5'
    mvar['1506']['Panel_Freq'] = 'ID77_DE_6'
    mvar['2206']['Panel_Freq'] = 'ID77_UK_2'
    mvar['2205']['Panel_Freq'] = 'ID77_UK_7'
    mvar['2210']['Panel_Freq'] = 'ID77_UK_8'
    mvar['2211']['Panel_Freq'] = 'ID77_UK_9'
    mvar['1705']['Panel_Freq'] = 'ID77_FR_1'
    mvar['1706']['Panel_Freq'] = 'ID77_FR_2'

    ##Online_Nutzungen
    ovar['1302'] = {'Panel_Dicho':'ID62_CH_1', 'Panel_Freq':'ID63_CH_1'}
    ovar['1304'] = {'Panel_Dicho':'ID62_CH_2', 'Panel_Freq':'ID63_CH_2'}
    ovar['1303'] = {'Panel_Dicho':'ID62_CH_3', 'Panel_Freq':'ID63_CH_3'}
    ovar['1301'] = {'Panel_Dicho':'ID62_CH_4', 'Panel_Freq':'ID63_CH_4'}      
    ovar['1505'] = {'Panel_Dicho':'ID62_DE_1', 'Panel_Freq':'ID63_DE_1'}
    ovar['1506'] = {'Panel_Dicho':'ID62_DE_2', 'Panel_Freq':'ID63_DE_2'}
    ovar['1503'] = {'Panel_Dicho':'ID62_DE_6', 'Panel_Freq':'ID63_DE_6'}
    ovar['1501'] = {'Panel_Dicho':'ID62_DE_7', 'Panel_Freq':'ID63_DE_7'}
    ovar['1502'] = {'Panel_Dicho':'ID62_DE_9', 'Panel_Freq':'ID63_DE_9'}
    ovar['2202'] = {'Panel_Dicho':'ID62_UK_2', 'Panel_Freq':'ID63_UK_2'}
    ovar['2201'] = {'Panel_Dicho':'ID62_UK_3', 'Panel_Freq':'ID63_UK_3'}
    ovar['2204'] = {'Panel_Dicho':'ID62_UK_6', 'Panel_Freq':'ID63_UK_6'}
    ovar['2203'] = {'Panel_Dicho':'ID62_UK_8', 'Panel_Freq':'ID63_UK_8'}
    ovar['2205'] = {'Panel_Dicho':'ID62_UK_10', 'Panel_Freq':'ID63_UK_10'}
    ovar['2206'] = {'Panel_Dicho':'ID62_UK_11', 'Panel_Freq':'ID63_UK_11'}
    ovar['1702'] = {'Panel_Dicho':'ID62_FR_2', 'Panel_Freq':'ID63_FR_2'}
    ovar['1703'] = {'Panel_Dicho':'ID62_FR_4', 'Panel_Freq':'ID63_FR_4'}
    ovar['1701'] = {'Panel_Dicho':'ID62_FR_5', 'Panel_Freq':'ID63_FR_5'}
    ovar['1707'] = {'Panel_Dicho':'ID62_FR_6', 'Panel_Freq':'ID63_FR_6'}
    ovar['1705'] = {'Panel_Dicho':'ID62_FR_7', 'Panel_Freq':'ID63_FR_7'}
    ovar['1704'] = {'Panel_Dicho':'ID62_FR_9', 'Panel_Freq':'ID63_FR_9'}
    

    pdich = 0
    pfreq = 0
    present_media = []

    for p in sorted(mvar.keys()):
        if mvar[p]['Panel_Dicho'] in varlist:
            pdich = pdich + 1
            if mvar[p]['Panel_Freq'] in varlist:
                pfreq = pfreq+1
                present_media.append(p)
            elif mvar[p]['Panel_Freq']=='#Not known':
                verbout('\nWarning: No frequency variable defined for medium: '+str(p)+ '. The medium is ignored',master=master)

    for p in sorted(ovar.keys()):
        if not p in present_media: ##Online nur zählen, wenn Offline nicht vorhanden
            if ovar[p]['Panel_Dicho'] in varlist:
                pdich = pdich + 1
                if ovar[p]['Panel_Freq'] in varlist:
                    pfreq = pfreq+1
                    present_media.append(p)
            elif ovar[p]['Panel_Freq']=='#Not known':
                verbout('\nWarning: No frequency variable defined for medium: '+str(p)+ '. The medium is ignored',master=master)

    verbout('\nThe data contain '+str(pfreq)+' of '+str(len(mvar.keys()))+' Media Use variables from panel survey',master=master)
    verbout('\nThe matching will also include self-reports of using the Online version of a newspaper or magazine!',master=master)
    if pfreq > 10:
        panel = 1
    else:
        panel = 0

    for m in present_media:
        data[m] = [] ##Add new columns to panel data, varnames are the media codes

    anz_valid = 0

    for i in range(len(data[varlist[0]])):
        anymed = 0
        for m in present_media:
            present = data[mvar[m]['Panel_Dicho']][i]
            mfreq = data[mvar[m]['Panel_Freq']][i]
            try:
                mfreq = float(mfreq)
                anymed = 1
            except:
                mfreq = 0.0
            
            try:
                mfreq2 = data[ovar[m]['Panel_Freq']][i] ##Online news frequency. Fail if no online version or if the variable is missing
                mfreq2 = float(mfreq2)
                anymed = 1
            except:
                mfreq2 = 0.0

            fullfreq = mfreq + mfreq2
            if fullfreq > 7: fullfreq = 7.0
            if mfreq2 > 0:
                pass

            data[m].append(fullfreq)
        anz_valid = anz_valid + anymed

    write_data(data,present_media,'mediatest.txt')

    verbout('\nData successfully transformed. '+str(anz_valid)+' cases (of '+str(i+1)+') have been assigned media usage frequencies for matching',master=master)

    return data,[1,panel,0]


def reformat_mediause_css(data,master=''):
    varlist = data.keys()
    mvar = {}
    mvar['1101'] = {'CSS_Dicho':'ID67_at_1', 'CSS_Freq':'ID69_at_1'}
    mvar['1102'] = {'CSS_Dicho':'ID67_at_2', 'CSS_Freq':'ID69_at_2'}
    mvar['1103'] = {'CSS_Dicho':'ID67_at_3', 'CSS_Freq':'ID69_at_3'}
    mvar['1104'] = {'CSS_Dicho':'ID67_at_4', 'CSS_Freq':'ID69_at_4'}
    mvar['1105'] = {'CSS_Dicho':'ID68_at_1', 'CSS_Freq':'ID77_at_1'}
    mvar['1106'] = {'CSS_Dicho':'ID68_at_2', 'CSS_Freq':'ID77_at_2'}
    mvar['1171'] = {'CSS_Dicho':'ID64_at_1', 'CSS_Freq':'ID65_at_1'}
    mvar['1172'] = {'CSS_Dicho':'ID64_at_5', 'CSS_Freq':'ID65_at_5'}
    mvar['1383'] = {'CSS_Dicho':'ID62_ch_1', 'CSS_Freq':'ID63_ch_1'}
    mvar['1381'] = {'CSS_Dicho':'ID62_ch_2', 'CSS_Freq':'ID63_ch_2'}
    mvar['1384'] = {'CSS_Dicho':'ID62_ch_3', 'CSS_Freq':'ID63_ch_3'}
    mvar['1382'] = {'CSS_Dicho':'ID62_ch_6', 'CSS_Freq':'ID63_ch_6'}
    mvar['1371'] = {'CSS_Dicho':'ID64_ch_1', 'CSS_Freq':'ID65_ch_1'}
    mvar['1372'] = {'CSS_Dicho':'ID64_ch_2', 'CSS_Freq':'ID65_ch_2'}
    mvar['1301'] = {'CSS_Dicho':'ID67_ch_1', 'CSS_Freq':'ID69_ch_1'}
    mvar['1302'] = {'CSS_Dicho':'ID67_ch_2', 'CSS_Freq':'ID69_ch_2'}
    mvar['1303'] = {'CSS_Dicho':'ID67_ch_3', 'CSS_Freq':'ID69_ch_3'}
    mvar['1304'] = {'CSS_Dicho':'ID67_ch_5', 'CSS_Freq':'ID69_ch_4'}
    mvar['1309'] = {'CSS_Dicho':'', 'CSS_Freq':'ID69_ch_16'}
    mvar['1310'] = {'CSS_Dicho':'', 'CSS_Freq':'ID69_ch_9'}
    mvar['1306'] = {'CSS_Dicho':'ID68_ch_1', 'CSS_Freq':'ID77_ch_1'}
    mvar['1305'] = {'CSS_Dicho':'ID68_ch_2', 'CSS_Freq':'ID77_ch_2'}
    mvar['1471'] = {'CSS_Dicho':'ID64_ch_5', 'CSS_Freq':'ID65_ch_5'}
    mvar['1402'] = {'CSS_Dicho':'ID67_ch_6', 'CSS_Freq':'ID69_ch_12'}
    mvar['1401'] = {'CSS_Dicho':'ID67_ch_7', 'CSS_Freq':'ID69_ch_13'}
    mvar['1583'] = {'CSS_Dicho':'ID62_de_1', 'CSS_Freq':'ID63_de_1'}
    mvar['1584'] = {'CSS_Dicho':'ID62_de_10', 'CSS_Freq':'ID63_de_12'}
    mvar['1582'] = {'CSS_Dicho':'ID62_de_12', 'CSS_Freq':'ID63_de_14'}
    mvar['1581'] = {'CSS_Dicho':'ID62_de_6', 'CSS_Freq':'ID63_de_6'}
    mvar['1571'] = {'CSS_Dicho':'ID64_de_1', 'CSS_Freq':'ID65_de_1'}
    mvar['1572'] = {'CSS_Dicho':'ID64_de_2', 'CSS_Freq':'ID65_de_2'}
    mvar['1576'] = {'CSS_Dicho':'ID64_de_3', 'CSS_Freq':'ID65_de_3'}
    mvar['1502'] = {'CSS_Dicho':'ID67_de_1', 'CSS_Freq':'ID69_de_1'}
    mvar['1501'] = {'CSS_Dicho':'ID67_de_2', 'CSS_Freq':'ID69_de_2'}
    mvar['1503'] = {'CSS_Dicho':'ID67_de_5', 'CSS_Freq':'ID69_de_5'}
    mvar['1504'] = {'CSS_Dicho':'ID67_de_6', 'CSS_Freq':'ID69_de_6'}
    mvar['1505'] = {'CSS_Dicho':'ID68_de_5', 'CSS_Freq':'ID77_de_5'}
    mvar['1506'] = {'CSS_Dicho':'ID68_de_6', 'CSS_Freq':'ID77_de_6'}
    mvar['1771'] = {'CSS_Dicho':'ID64_fr_12', 'CSS_Freq':'ID65_fr_14'}
    mvar['1772'] = {'CSS_Dicho':'ID64_fr_13', 'CSS_Freq':'ID65_fr_15'}
    mvar['1701'] = {'CSS_Dicho':'ID67_fr_16', 'CSS_Freq':'ID69_fr_16'}
    mvar['1707'] = {'CSS_Dicho':'ID67_fr_25', 'CSS_Freq':'ID69_fr_26'}
    mvar['1708'] = {'CSS_Dicho':'ID67_fr_26', 'CSS_Freq':'ID69_fr_27'}
    mvar['1709'] = {'CSS_Dicho':'ID67_fr_27', 'CSS_Freq':'ID69_fr_28'}
    mvar['1702'] = {'CSS_Dicho':'ID67_fr_17', 'CSS_Freq':'ID69_fr_17'}
    mvar['1703'] = {'CSS_Dicho':'ID67_fr_18', 'CSS_Freq':'ID69_fr_19'}
    mvar['1704'] = {'CSS_Dicho':'ID67_fr_19', 'CSS_Freq':'ID69_fr_20'}
    mvar['1705'] = {'CSS_Dicho':'ID68_fr_19', 'CSS_Freq':'ID77_fr_20'}
    mvar['1706'] = {'CSS_Dicho':'ID68_fr_20', 'CSS_Freq':'ID77_fr_21'}
    mvar['1801'] = {'CSS_Dicho':'ID67_it_1', 'CSS_Freq':'ID65_it_14'}
    mvar['1802'] = {'CSS_Dicho':'ID67_it_2', 'CSS_Freq':'ID69_it_1'}
    mvar['1803'] = {'CSS_Dicho':'ID67_it_3', 'CSS_Freq':'ID69_it_2'}
    mvar['1804'] = {'CSS_Dicho':'ID67_it_4', 'CSS_Freq':'ID69_it_3'}
    mvar['1805'] = {'CSS_Dicho':'ID68_it_1', 'CSS_Freq':'ID77_it_1'}
    mvar['1806'] = {'CSS_Dicho':'ID68_it_2', 'CSS_Freq':'ID77_it_2'}
    mvar['1871'] = {'CSS_Dicho':'ID64_it_1', 'CSS_Freq':'ID65_it_1'}
    mvar['1872'] = {'CSS_Dicho':'ID64_it_2', 'CSS_Freq':'ID65_it_2'}
    mvar['1874'] = {'CSS_Dicho':'ID64_it_9', 'CSS_Freq':'ID65_it_11'}
    mvar['1901'] = {'CSS_Dicho':'ID67_nl_1', 'CSS_Freq':'ID69_nl_1'}
    mvar['1902'] = {'CSS_Dicho':'ID67_nl_2', 'CSS_Freq':'ID69_nl_2'}
    mvar['1903'] = {'CSS_Dicho':'ID67_nl_3', 'CSS_Freq':'ID69_nl_3'}
    mvar['1904'] = {'CSS_Dicho':'ID67_nl_5', 'CSS_Freq':'ID69_nl_5'}
    mvar['1905'] = {'CSS_Dicho':'ID68_nl_1', 'CSS_Freq':'ID77_nl_1'}
    mvar['1906'] = {'CSS_Dicho':'ID68_nl_2', 'CSS_Freq':'ID77_nl_2'}
    mvar['1971'] = {'CSS_Dicho':'ID64_nl_1', 'CSS_Freq':'ID65_nl_1'}
    mvar['1972'] = {'CSS_Dicho':'ID64_nl_2', 'CSS_Freq':'ID65_nl_2'}
    mvar['1973'] = {'CSS_Dicho':'ID64_nl_3', 'CSS_Freq':'ID65_nl_3'}
    mvar['2001'] = {'CSS_Dicho':'ID67_pl_16', 'CSS_Freq':'ID69_pl_17'}
    mvar['2002'] = {'CSS_Dicho':'ID67_pl_17', 'CSS_Freq':'ID69_pl_18'}
    mvar['2003'] = {'CSS_Dicho':'ID67_pl_18', 'CSS_Freq':'ID69_pl_19'}
    mvar['2004'] = {'CSS_Dicho':'ID67_pl_19', 'CSS_Freq':'ID69_pl_20'}
    mvar['2005'] = {'CSS_Dicho':'ID68_pl_11', 'CSS_Freq':'ID77_pl_12'}
    mvar['2006'] = {'CSS_Dicho':'ID68_pl_10', 'CSS_Freq':'ID77_pl_11'}
    mvar['2071'] = {'CSS_Dicho':'ID64_pl_7', 'CSS_Freq':'ID65_pl_9'}
    mvar['2072'] = {'CSS_Dicho':'ID64_pl_8', 'CSS_Freq':'ID65_pl_10'}
    mvar['2101'] = {'CSS_Dicho':'ID67_swe_1', 'CSS_Freq':'ID69_swe_1'}
    mvar['2102'] = {'CSS_Dicho':'ID67_swe_2', 'CSS_Freq':'ID69_swe_2'}
    mvar['2103'] = {'CSS_Dicho':'ID67_swe_3', 'CSS_Freq':'ID69_swe_3'}
    mvar['2104'] = {'CSS_Dicho':'ID67_swe_5', 'CSS_Freq':'ID69_swe_5'}
    mvar['2105'] = {'CSS_Dicho':'ID68_swe_1', 'CSS_Freq':'ID77_swe_1'}
    mvar['2106'] = {'CSS_Dicho':'ID68_swe_3', 'CSS_Freq':'ID77_swe_3'}
    mvar['2171'] = {'CSS_Dicho':'ID64_swe_1', 'CSS_Freq':'ID65_swe_1'}
    mvar['2172'] = {'CSS_Dicho':'ID64_swe_2', 'CSS_Freq':'ID65_swe_2'}
    mvar['2271'] = {'CSS_Dicho':'ID64_uk_1', 'CSS_Freq':'ID65_uk_1'}
    mvar['2272'] = {'CSS_Dicho':'ID64_uk_2', 'CSS_Freq':'ID65_uk_2'}
    mvar['2281'] = {'CSS_Dicho':'ID62_uk_1', 'CSS_Freq':'ID63_uk_1'}
    mvar['2283'] = {'CSS_Dicho':'ID62_uk_2', 'CSS_Freq':'ID63_uk_2'}
    mvar['2282'] = {'CSS_Dicho':'ID62_uk_4', 'CSS_Freq':'ID63_uk_4'}
    mvar['2284'] = {'CSS_Dicho':'ID62_uk_5', 'CSS_Freq':'ID63_uk_5'}
    mvar['2207'] = {'CSS_Dicho':'ID67_uk_11', 'CSS_Freq':'ID69_uk_11'}
    mvar['2208'] = {'CSS_Dicho':'ID67_uk_12', 'CSS_Freq':'ID69_uk_12'}
    mvar['2203'] = {'CSS_Dicho':'ID67_uk_2', 'CSS_Freq':'ID69_uk_2'}
    mvar['2204'] = {'CSS_Dicho':'ID67_uk_4', 'CSS_Freq':'ID69_uk_4'}
    mvar['2201'] = {'CSS_Dicho':'ID67_uk_7', 'CSS_Freq':'ID69_uk_7'}
    mvar['2202'] = {'CSS_Dicho':'ID67_uk_9', 'CSS_Freq':'ID69_uk_9'}
    mvar['2206'] = {'CSS_Dicho':'ID68_uk_2', 'CSS_Freq':'ID77_uk_2'}
    mvar['2205'] = {'CSS_Dicho':'ID68_uk_5', 'CSS_Freq':'ID77_uk_5'}
    mvar['2210'] = {'CSS_Dicho':'ID68_uk_6', 'CSS_Freq':'ID77_uk_6'}
    mvar['2211'] = {'CSS_Dicho':'ID68_uk_7', 'CSS_Freq':'ID77_uk_7'}
    mvar['2301'] = {'CSS_Dicho':'ID67_usa_1', 'CSS_Freq':'ID69_usa_1'}
    mvar['2303'] = {'CSS_Dicho':'ID67_usa_3', 'CSS_Freq':'ID69_usa_3'}
    mvar['2306'] = {'CSS_Dicho':'ID68_usa_1', 'CSS_Freq':'ID77_usa_1'}
    mvar['2307'] = {'CSS_Dicho':'ID67_usa_4', 'CSS_Freq':'ID69_usa_4'}
    mvar['2308'] = {'CSS_Dicho':'ID67_usa_2', 'CSS_Freq':'ID69_usa_2'}
    mvar['2309'] = {'CSS_Dicho':'ID68_usa_2', 'CSS_Freq':'ID77_usa_2'}
    mvar['2371'] = {'CSS_Dicho':'ID64_usa_1', 'CSS_Freq':'ID65_usa_1'}
    mvar['2372'] = {'CSS_Dicho':'ID64_usa_11', 'CSS_Freq':'ID65_usa_11'}
    mvar['2373'] = {'CSS_Dicho':'ID64_usa_5', 'CSS_Freq':'ID65_usa_5'}

    pdich = 0
    pfreq = 0
    present_media = []

    verbout('\nChecking the availability of '+str(len(mvar.keys()))+' media outlets in the CSS data',master=master)

    for p in sorted(mvar.keys()):
        if mvar[p]['CSS_Dicho'] in varlist:
            pdich = pdich + 1
            if mvar[p]['CSS_Freq'] in varlist:
                pfreq = pfreq+1
                present_media.append(p)
            else:
                verbout('\nWarning: No frequency variable present for medium: '+str(p)+ '. The medium is ignored',master=master)

    verbout('\nThe data contain '+str(pfreq)+' of '+str(len(mvar.keys()))+' Media Use variables from cross sectional survey',master=master)
                 
    if pfreq > 10:
        css = 1 ##Check for continuing matching procedure
    else:
        css = 0

    for m in present_media: ##Setting up new columns
        data[m] = []

    anz_valid = 0

    for i in range(len(data[varlist[0]])):
        anymed = 0
        for m in present_media:
            present = data[mvar[m]['CSS_Dicho']][i]
            mfreq = data[mvar[m]['CSS_Freq']][i]
            try:
                mfreq = float(mfreq)
                anymed = 1
            except:
                mfreq = 0.0
            data[m].append(mfreq)
        anz_valid = anz_valid + anymed

    write_data(data,present_media,'mediatest.txt')

    verbout('\nData successfully transformed. '+str(anz_valid)+' cases (of '+str(i+1)+') have been assigned media usage frequencies for matching',master=master)

    return data,[1,0,css]

    


def get_dta_level(liste,master=''):
    sumdic = {}
    sumdic['Float'] = 0
    sumdic['Int']=0
    sumdic['Mis']=0
    sumdic['Str']=0

    for e in liste:
        if e in ['',' ']:
            typ = 'Mis'
        else:
            try:
                a = int(e)
                b = float(e)
                if a == b:
                    typ = 'Int'
            except:
                try:
                    b = float(e)
                    typ = 'Float'
                except:
                    typ = 'Str'

        sumdic[typ] = sumdic[typ] + 1

    outstr = str(sumdic['Float']+sumdic['Int'])+' Numeric values; '+str(sumdic['Str'])+' Other/String; '+str(sumdic['Mis'])+' Missing'
    return outstr 
    


def check_data(data,varlist,master=''):
    log('Calling Function: Check_Data')
    check = 1

    if data == 0 or varlist == 0:
        verb('No File found')
        verbout('\nFile not found. The filename you entered does not exist. Please select an existing file.',master=master)
        check = 0
    elif len(varlist) > 1:
        verb('File contains '+str(len(varlist))+' variables')
        verbout('\nFile contains '+str(len(varlist))+' variables',master=master)
        verbout('\nFile contains '+str(len(data[varlist[0]]))+' cases',master=master)
        slen = len(data[varlist[0]])
        unlang = 0
        outs='\nWarning: Not all columns have equal amout of cases:\n'
        for v in varlist:
            outs = outs+v + ': '+str(len(data[v]))+'\n'
            if not len(data[v]) == slen:
                unlang = 1
        if unlang ==1:
            verbout(outs,master=master)
            master.message('File Incomplete')
            check = 0         
    else:
        check = 0
        verb('ERROR: Too few variables for this operation')
        master.message('No Variables')
        
    return check





##########################
##
## GGCRISI Special Functions
##
##########################


def merge_ggcrisi(data1,data2,key1,key2,master=''):
    verbout('Getting informations from key table...',master=master)
    addvar = []
    for v in data2.keys():
        if not v in key1 and len(data2[v])>0:
            addvar.append(v)
    schldic = {}
    for i in range(len(data2[key2[0]])):
        key = []
        for v in key2:
            key.append(data2[v][i])
        keylab = str(key)

        schldic[keylab] = {}
        for v in addvar:
            schldic[keylab][v] = data2[v][i]
    verbout('Done.',master=master)
    for v in addvar:
        data1[v] = []
    verbout('\nAdding variables: '+str(addvar),master=master)
    for i in range(len(data1[key1[0]])):
        key = []
        for v in key1:
            key.append(data1[v][i])
        keylab = str(key)

        if keylab in schldic.keys():
            for v in addvar:
                data1[v].append(schldic[keylab][v])
        else:
            for v in addvar:
                data1[v].append('')
    return data1


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


###
### Initial definition of essential settings (Settings which are called within routines and have to be defined somewhere)
###
###

global settings
settings= {}
settings['Coder'] = 'default' ##May be overwritten by setting some other value in a_settings.ini
settings['Font'] = 'Arial' ##Font in all displayed text (Questions and Text)
settings['Fontsize']="10" ##Font Size within the text display. No effect on question size.
settings['Layout'] = 'Lefty' ##Layout of Angrist. 'Lefty' sets the left-handed design in which the check button is bottom-left. 'Righty' sets right-handed layout with Check on bottom-right.
settings['Styleset'] = 'Default' ##Text display settings. Loaded in define_styleset
settings['Python_Version'] = py_version ##Text display settings. Loaded in define_styleset
settings['Excel']= xlsx ## Is 1 if xlsx is supported and 0 if it is not.
settings['Curr_Page'] = [['el1','var1'],['el2','var2'],['el3','var3']] ##Temporary storage of current page order. Essential for displaying and storing values.
settings['Page_History'] = [] ##Will contain the whole list of visited pages during program execution
settings['Input'] = ['','',''] ##Temporary Storage of input values. Essential for moste question types.
settings['Path_Log'] = '\nLog of called functions:\n-----------------\n' ##String containing a detailed log-File if debugging is set 1
settings['Verb_Log'] = '\nVerbose progress:\n-----------------\n' ##String containing all comments of the verbose program
settings['Text_Aktiv'] = 0

settings['Verbose'] = '2' ##Verbosity of the program. 0 does not make any notes to the console, 1 is used for coding, 2 is used for debugging.
settings['Debugging'] = '0' ##If set to '1', invalid entries are possible and verbose is set to 1.
settings['Assurance'] = '0' ##May be used to prompt questions at critical points.
settings['AEGLOS'] = '0' ##If set to '1', the aeglos-Module will be attached and used. Only useful after training of this module.
settings['Insecure']='0' ##If set to '1', the coder may report insecurity for certain variables.

settings['Multi_Items']=[] ##Contains variables which give rise to a number of dummies. Variable names are the keys, the number of dummies are the values.
settings['Break'] = 0 ##Currently at break.
settings['Break_Time'] = 0 ##Total break time in this session
settings['Coding_Time'] = 0 ##Total coding time in this session
settings['Country']='ch' ##Current country
settings['Hotwords']={} ##List of hotwords (for different countries).
settings['Highlight_Buttons']=[]
settings['Auto_Highlight'] = 1 ##Mark the hotwords upon loading the text.
settings['Datasets'] = {}

settings['Language'] = 'en' ##Language of the codebook (may be used if different codebooks are at disposal. No effect otherwise.
settings['Todo'] = 'to_do.txt' ##Text-File which contains the list of text-files to be coded (with or without extension)
settings['Package_Todo'] = '' ##Text-File containing a list of folders containing articles and a todo-file.
settings['Codebook'] = 'n_codebuch.ini'
settings['Settings'] = 'n_settings.ini'

settings['Text_Folder'] = 'Probecodierung\\' ##Folder which contains the text-files to be coded.
settings['Out_Track'] = '' ##File to write the tracking-report (no report if set to '')
settings['Out_Tree']='trees.txt' ##File to write the tree files in indented layout (no output if set to '')
settings['Out_JSON']='' ##File to write the JSON-Outputs (no output if set to '')
settings['Out_Tmp']='' ##File to write the temporary savings (no savings if set to '')


if __name__ == "__main__":  ##If Nogrod is executed, execute these lines. If it is just called as a module, don't start the GUI
    root = Tk()
    fenster = Anzeige(root)
    root.mainloop()
